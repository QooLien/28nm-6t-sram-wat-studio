#!/usr/bin/env python3
"""6T SRAM / WAT compact-model correlation and cell-level SNM analyzer.

This is an engineering exploration model.  It calibrates simple square-law MOS
devices from WAT Vt and Ids and is intentionally independent of a PDK/ngspice.
Use a foundry BSIM deck for sign-off numbers.
"""

from __future__ import annotations

import argparse
from bisect import bisect_left
import csv
import html
import json
import math
import os
import re
import shutil
import statistics as statlib
import subprocess
import sys
import webbrowser
from dataclasses import asdict, dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable


@dataclass(frozen=True)
class WatPoint:
    corner: str = "TT"
    pu_vt: float = 0.38       # |Vtp|, V
    pu_ids: float = 45.0      # uA, WAT on-current
    pg_vt: float = 0.37       # V
    pg_ids: float = 80.0      # uA
    pd_vt: float = 0.36       # V
    pd_ids: float = 120.0     # uA


@dataclass(frozen=True)
class MosWat:
    """WAT parameters owned by one physical MOS in the bitcell."""
    vt: float
    ids: float


@dataclass(frozen=True)
class DatasheetTargets:
    """PU/PG/PD WAT targets used only for measured-WAT comparison."""
    pu: MosWat
    pg: MosWat
    pd: MosWat


@dataclass(frozen=True)
class JudgmentTargets:
    """User-owned design/datasheet limits for model-derived cell metrics."""
    cell_ratio_min: float = 1.20
    pull_up_ratio_min: float = 1.50
    hold_snm_min_mv: float = 300.0
    read_snm_min_mv: float = 200.0
    write_snm_min_mv: float = 100.0
    marginal_band_pct: float = 10.0


@dataclass(frozen=True)
class TargetValidationSettings:
    """User-owned bands/specs used to validate WAT targets against measured WT."""
    vt_tolerance_mv: float = 15.0
    idsat_tolerance_pct: float = 8.0
    scan4n_vmin_max: float = 0.650
    select_write_vmin_max: float = 0.600
    select_read_vmin_max: float = 0.620
    minimum_statistical_n: int = 10


DISPLAY_MOS_NAMES = {
    "pu1": "PUL", "pu2": "PUR",
    "pg1": "PGL", "pg2": "PGR",
    "pd1": "PDL", "pd2": "PDR",
}

# One numerical resolution is used whenever a reported Read/Write SNM value
# is compared across the single-cell, VDD-sweep, and wafer multi-chip flows.
# The square extractor operates on a sampled VTC, so differing point counts
# would otherwise create a small grid-quantization difference even for the
# exact same six MOS inputs.
SNM_FIT_GRID_POINTS = 1201


def _snm_fit_points(cfg: "Config", requested: int | None = None) -> int:
    """Return the common numerical resolution for all reported SNM metrics."""
    return max(SNM_FIT_GRID_POINTS, int(cfg.grid_points),
               0 if requested is None else int(requested))


def open_output_directory(path: str | os.PathLike[str]) -> Path:
    """Create and open the selected output directory in the system file manager."""
    raw_path = str(path).strip()
    if not raw_path:
        raise ValueError("Choose an output folder first.")
    output_dir = Path(raw_path).expanduser()
    output_dir.mkdir(parents=True, exist_ok=True)
    output_dir = output_dir.resolve()
    if sys.platform == "win32":
        os.startfile(str(output_dir))
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(output_dir)])
    else:
        subprocess.Popen(["xdg-open", str(output_dir)])
    return output_dir


def _safe_path_component(value: object, fallback: str) -> str:
    """Return a Windows-safe, readable directory component."""
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(value).strip())
    text = re.sub(r"\s+", "_", text).strip(" ._")
    return (text[:80] or fallback)


def create_run_output_dir(base_dir: str | os.PathLike[str], wafer_id: object,
                          analysis_name: str,
                          timestamp: datetime | None = None) -> Path:
    """Create a unique date/Wafer/time analysis directory and run manifest."""
    instant = timestamp or datetime.now().astimezone()
    wafer = _safe_path_component(wafer_id, "Unknown_Wafer")
    analysis = _safe_path_component(analysis_name, "analysis")
    parent = Path(base_dir) / instant.strftime("%Y-%m-%d") / wafer
    stem = f'{instant.strftime("%H%M%S")}_{analysis}'
    suffix = 1
    while True:
        candidate = parent / (stem if suffix == 1 else f"{stem}_{suffix:02d}")
        try:
            candidate.mkdir(parents=True, exist_ok=False)
            break
        except FileExistsError:
            suffix += 1
    manifest = {
        "wafer_id": str(wafer_id).strip() or "Unknown Wafer",
        "analysis": analysis_name,
        "created_local": instant.isoformat(timespec="seconds"),
        "output_directory": str(candidate.resolve()),
    }
    (candidate / "run_info.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return candidate


def _stagger_label_rows(labels: list[tuple[float, str]],
                        character_width: float = 8.0,
                        minimum_gap: float = 8.0) -> list[int]:
    """Assign compact non-overlapping rows to labels centered at pixel X positions."""
    row_right_edges: list[float] = []
    assigned: list[int] = []
    for center_x, text in labels:
        estimated_width = max(42.0, len(text) * character_width)
        left_edge = center_x - estimated_width / 2.0
        right_edge = center_x + estimated_width / 2.0
        for row_index, previous_right in enumerate(row_right_edges):
            if left_edge >= previous_right + minimum_gap:
                row_right_edges[row_index] = right_edge
                assigned.append(row_index)
                break
        else:
            row_right_edges.append(right_edge)
            assigned.append(len(row_right_edges) - 1)
    return assigned


def _place_chart_labels(
        labels: list[tuple[float, float, str]],
        point_obstacles: list[tuple[float, float]],
        bounds: tuple[float, float, float, float],
        font_size: float = 14.0) -> list[tuple[float, float, str]]:
    """Place SVG labels near marks while avoiding points, labels and plot edges."""
    left, top, right, bottom = bounds
    candidates = ((0, -25, "middle"), (0, 34, "middle"),
                  (20, -18, "start"), (-20, -18, "end"),
                  (22, 26, "start"), (-22, 26, "end"),
                  (34, 5, "start"), (-34, 5, "end"))
    occupied: list[tuple[float, float, float, float]] = []
    placed: list[tuple[float, float, str]] = []

    def intersects(first: tuple[float, float, float, float],
                   second: tuple[float, float, float, float], gap: float = 4.0) -> bool:
        return not (first[2] + gap < second[0] or second[2] + gap < first[0]
                    or first[3] + gap < second[1] or second[3] + gap < first[1])

    for point_x, point_y, text in labels:
        width = max(42.0, len(text) * font_size * .56)
        height = font_size * 1.35
        choice = None
        choice_penalty = math.inf
        for dx, dy, anchor in candidates:
            label_x, label_y = point_x + dx, point_y + dy
            if anchor == "middle":
                box_left = label_x - width / 2
            elif anchor == "end":
                box_left = label_x - width
            else:
                box_left = label_x
            box = (box_left, label_y - height, box_left + width, label_y + 3)
            penalty = 0.0
            if box[0] < left or box[2] > right or box[1] < top or box[3] > bottom:
                penalty += 10000.0
            penalty += 3000.0 * sum(intersects(box, other) for other in occupied)
            for obstacle_x, obstacle_y in point_obstacles:
                point_box = (obstacle_x - 8, obstacle_y - 8,
                             obstacle_x + 8, obstacle_y + 8)
                if intersects(box, point_box, gap=3.0):
                    penalty += 600.0
            penalty += abs(dx) + abs(dy)
            if penalty < choice_penalty:
                choice_penalty = penalty
                choice = (label_x, label_y, anchor, box)
        assert choice is not None
        label_x, label_y, anchor, box = choice
        occupied.append(box)
        placed.append((label_x, label_y, anchor))
    return placed


def _place_endpoint_label_columns(
        labels: list[tuple[float, float, str]],
        bounds: tuple[float, float, float, float],
        font_size: float = 11.0) -> list[tuple[float, float, str]]:
    """Place multi-curve endpoint labels in non-overlapping edge columns."""
    left, top, right, bottom = bounds
    midpoint = (left + right) / 2
    placed: list[tuple[float, float, str] | None] = [None] * len(labels)
    groups = {
        "left": [(index, item) for index, item in enumerate(labels)
                 if item[0] <= midpoint],
        "right": [(index, item) for index, item in enumerate(labels)
                  if item[0] > midpoint],
    }
    for side, group in groups.items():
        if not group:
            continue
        ordered = sorted(group, key=lambda entry: entry[1][1])
        minimum_center = top + 12.0
        maximum_center = bottom - 12.0
        if len(ordered) == 1:
            gap = 0.0
        else:
            gap = min(25.0, (maximum_center - minimum_center) / (len(ordered) - 1))
        centers: list[float] = []
        for _index, (_point_x, point_y, _text) in ordered:
            desired = min(max(point_y, minimum_center), maximum_center)
            centers.append(max(desired, centers[-1] + gap) if centers else desired)
        if centers[-1] > maximum_center:
            shift = centers[-1] - maximum_center
            centers = [value - shift for value in centers]
        if centers[0] < minimum_center:
            shift = minimum_center - centers[0]
            centers = [value + shift for value in centers]
        for (original_index, (_point_x, _point_y, text)), center_y in zip(
                ordered, centers):
            label_width = max(92.0, len(text) * font_size * .56 + 12.0)
            center_x = (left + 8.0 + label_width / 2.0 if side == "left"
                        else right - 8.0 - label_width / 2.0)
            placed[original_index] = (center_x, center_y + 5.0, "middle")
    assert all(item is not None for item in placed)
    return [item for item in placed if item is not None]


@dataclass(frozen=True)
class SixTWatCell:
    """Object-oriented WAT description of all six physical bitcell devices."""
    corner: str
    pu1: MosWat
    pu2: MosWat
    pg1: MosWat
    pg2: MosWat
    pd1: MosWat
    pd2: MosWat

    def side(self, index: int) -> WatPoint:
        if index not in (1, 2):
            raise ValueError("side must be 1 or 2")
        pu, pg, pd = (getattr(self, f"{name}{index}") for name in ("pu", "pg", "pd"))
        return WatPoint(self.corner, pu.vt, pu.ids, pg.vt, pg.ids, pd.vt, pd.ids)

    def representative(self) -> WatPoint:
        def avg(kind: str, attr: str) -> float:
            return (getattr(getattr(self, kind+"1"), attr) + getattr(getattr(self, kind+"2"), attr)) / 2
        return WatPoint(self.corner, avg("pu","vt"), avg("pu","ids"), avg("pg","vt"),
                        avg("pg","ids"), avg("pd","vt"), avg("pd","ids"))

    def replace_mos(self, name: str, **changes: float) -> "SixTWatCell":
        return replace(self, **{name: replace(getattr(self, name), **changes)})


@dataclass(frozen=True)
class ExcelMosStatistics:
    """Wafer-site statistics used to collapse repeated WAT rows into one MOS model input."""
    valid_count: int
    total_count: int
    vt_mean: float
    vt_median: float
    vt_stdev: float
    vt_min: float
    vt_max: float
    ids_mean: float
    ids_median: float
    ids_stdev: float
    ids_min: float
    ids_max: float


@dataclass(frozen=True)
class ExcelWatSweepPoint:
    """One 6T WAT sample measured/evaluated at one model supply voltage."""
    lot_wafer: str
    model_vdd_v: float
    cell: SixTWatCell
    statistics: dict[str, ExcelMosStatistics]


@dataclass(frozen=True)
class IvCurveExtraction:
    """One raw Id-Vg curve reduced to the Idsat used at its model VDD."""
    family: str
    model_vdd_v: float
    vt_v: float
    sampled_vg_v: float
    idsat_ua: float
    source_point_count: int


@dataclass(frozen=True)
class WaferChipWat:
    """One independently measured physical 6T cell on a wafer."""
    lot_wafer: str
    chip_id: str
    model_vdd_v: float
    cell: SixTWatCell
    # Preserve the original WAT current sign for report traceability.  The
    # compact NMOS/PMOS drive equations use the magnitude held in ``cell``.
    raw_idsat_ua: dict[str, float] | None = None


@dataclass(frozen=True)
class ThreeTWatCell:
    """Merged object mode: one PU, PG and PD object shared by both cell sides."""
    corner: str
    pu: MosWat
    pg: MosWat
    pd: MosWat

    def to_six_t(self) -> SixTWatCell:
        return SixTWatCell(self.corner, self.pu, self.pu, self.pg, self.pg, self.pd, self.pd)

    def representative(self) -> WatPoint:
        return WatPoint(self.corner, self.pu.vt, self.pu.ids, self.pg.vt,
                        self.pg.ids, self.pd.vt, self.pd.ids)


@dataclass(frozen=True)
class RsnmVccPoint:
    """Grouped PU/PG/PD WAT inputs measured at one operating VDD."""
    vcc_v: float
    pu: MosWat
    pg: MosWat
    pd: MosWat


@dataclass(frozen=True)
class ManualVmin:
    """Measured WT values entered by the user; never generated by the model."""
    scan4n: float
    select_write: float
    select_read: float

    def rows(self) -> list[dict]:
        return [
            {"test": "Scan4N", "vmin_v": self.scan4n, "source": "Manual measured value"},
            {"test": "Select_Write", "vmin_v": self.select_write, "source": "Manual measured value"},
            {"test": "Select_Read", "vmin_v": self.select_read, "source": "Manual measured value"},
        ]


@dataclass(frozen=True)
class Config:
    # Fixed generic 28 nm low-power core assumptions.
    wat_vdd: float = 0.90
    nominal_vdd: float = 0.90
    vt_step: float = 0.030
    ids_step_pct: float = 10.0
    vmin_start: float = 0.40
    vmin_stop: float = 0.90
    vmin_step: float = 0.01
    read_snm_limit: float = 0.030
    grid_points: int = 5001
    technology_node_nm: float = 28.0
    channel_length_nm: float = 28.0
    pu_width_nm: float = 70.0
    pg_width_nm: float = 100.0
    pd_width_nm: float = 140.0
    nominal_temperature_c: float = 25.0
    read_wordline_over_vdd: float = 1.0
    read_bitline_over_vdd: float = 1.0
    write_wordline_over_vdd: float = 1.0
    write_low_bitline_over_vdd: float = 0.0
    write_high_bitline_over_vdd: float = 1.0


@dataclass(frozen=True)
class Tech28nm:
    """Transparent non-foundry defaults for the generic 28 nm SRAM model."""
    node_nm: int = 28
    process_family: str = "generic planar bulk CMOS, LP-like"
    topology: str = "6T: 2×PU PMOS + 2×PG NMOS + 2×PD NMOS"
    channel_length_nm: float = 28.0
    pu_width_nm: float = 70.0
    pg_width_nm: float = 100.0
    pd_width_nm: float = 140.0
    nominal_temperature_c: float = 25.0
    read_wordline_over_vdd: float = 1.0
    read_bitline_over_vdd: float = 1.0
    write_wordline_over_vdd: float = 1.0
    write_low_bitline_over_vdd: float = 0.0
    write_high_bitline_over_vdd: float = 1.0
    hold_wordline_over_vdd: float = 0.0
    beta_policy: str = "WAT-calibrated: 2*Idsat/(WAT_VDD-|Vt|)^2"
    common_vth_policy: str = "mean(|Vt_PU|, Vt_PG, Vt_PD) for analytical Read SNM"


TECH_28NM = Tech28nm()


def tech_from_config(cfg: Config) -> Tech28nm:
    """Resolve editable GUI assumptions, falling back to generic defaults."""
    geometry = {
        "Channel length L": cfg.channel_length_nm,
        "PU width": cfg.pu_width_nm,
        "PG width": cfg.pg_width_nm,
        "PD width": cfg.pd_width_nm,
    }
    for label, value in geometry.items():
        if not math.isfinite(value) or value <= 0:
            raise ValueError(f"{label} must be a positive finite value")
    return replace(
        TECH_28NM,
        node_nm=int(round(cfg.technology_node_nm)),
        channel_length_nm=cfg.channel_length_nm,
        pu_width_nm=cfg.pu_width_nm,
        pg_width_nm=cfg.pg_width_nm,
        pd_width_nm=cfg.pd_width_nm,
        nominal_temperature_c=cfg.nominal_temperature_c,
        read_wordline_over_vdd=cfg.read_wordline_over_vdd,
        read_bitline_over_vdd=cfg.read_bitline_over_vdd,
        write_wordline_over_vdd=cfg.write_wordline_over_vdd,
        write_low_bitline_over_vdd=cfg.write_low_bitline_over_vdd,
        write_high_bitline_over_vdd=cfg.write_high_bitline_over_vdd,
    )


def _positive(value: str | float, label: str) -> float:
    x = abs(float(value))
    if not math.isfinite(x) or x <= 0:
        raise ValueError(f"{label} must be a positive finite number")
    return x


def read_wat_csv(path: str | os.PathLike[str]) -> list[WatPoint]:
    required = {"corner", "pu_vt", "pu_ids", "pg_vt", "pg_ids", "pd_vt", "pd_ids"}
    rows: list[WatPoint] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError("WAT CSV missing columns: " + ", ".join(sorted(missing)))
        for line, row in enumerate(reader, 2):
            try:
                rows.append(WatPoint(
                    corner=(row["corner"] or f"row_{line}").strip(),
                    pu_vt=_positive(row["pu_vt"], "pu_vt"),
                    pu_ids=_positive(row["pu_ids"], "pu_ids"),
                    pg_vt=_positive(row["pg_vt"], "pg_vt"),
                    pg_ids=_positive(row["pg_ids"], "pg_ids"),
                    pd_vt=_positive(row["pd_vt"], "pd_vt"),
                    pd_ids=_positive(row["pd_ids"], "pd_ids"),
                ))
            except (TypeError, ValueError) as exc:
                raise ValueError(f"WAT CSV line {line}: {exc}") from exc
    if not rows:
        raise ValueError("WAT CSV has no data rows")
    return rows


class Device:
    """WAT-calibrated square-law device with a smooth threshold transition."""

    # Generic 28 nm compact-model smoothing voltage.  This is not a PDK
    # subthreshold model; it simply avoids the discontinuous on/off threshold
    # of the previous educational square-law implementation.
    VOV_SMOOTHING_V = 0.035

    def __init__(self, vt: float, ids_ua: float, wat_vdd: float):
        self.vt = abs(vt)
        self.ids = ids_ua
        overdrive = max(self._effective_overdrive(wat_vdd - self.vt), 0.05)
        self.beta = 2.0 * ids_ua / (overdrive * overdrive)  # uA/V^2

    @classmethod
    def _effective_overdrive(cls, raw_overdrive: float) -> float:
        """Numerically stable softplus version of max(VGS - VT, 0)."""
        scaled = raw_overdrive / cls.VOV_SMOOTHING_V
        if scaled >= 40.0:
            return raw_overdrive
        if scaled <= -40.0:
            return cls.VOV_SMOOTHING_V * math.exp(scaled)
        return cls.VOV_SMOOTHING_V * math.log1p(math.exp(scaled))

    def current(self, vgs: float, vds: float) -> float:
        vov = self._effective_overdrive(vgs - self.vt)
        if vds <= 0:
            return 0.0
        if vds < vov:
            return self.beta * (vov * vds - 0.5 * vds * vds)
        return 0.5 * self.beta * vov * vov


def _inverse_vtc(curve: list[tuple[float, float]]) -> list[tuple[float, float]]:
    """Return y=f^-1(x) on the original x grid for a decreasing VTC."""
    xs = [point[0] for point in curve]
    ys = [point[1] for point in curve]
    reversed_ys = list(reversed(ys))

    def inverse_y(x_value: float) -> float:
        index = bisect_left(reversed_ys, x_value)
        if index <= 0:
            return xs[-1]
        if index >= len(reversed_ys):
            return xs[0]
        y0, y1 = reversed_ys[index - 1], reversed_ys[index]
        x0, x1 = xs[-index], xs[-index - 1]
        if abs(y1 - y0) < 1e-15:
            return (x0 + x1) / 2.0
        return x0 + (x_value - y0) * (x1 - x0) / (y1 - y0)

    return [(x_value, inverse_y(x_value)) for x_value in xs]


def _fit_butterfly_squares(direct_curve: list[tuple[float, float]],
                            mirrored_curve: list[tuple[float, float]],
                            vdd: float, mode: str) -> dict:
    """Fit independent maximum squares between two VTCs on a shared Vin grid."""
    if len(direct_curve) != len(mirrored_curve) or len(direct_curve) < 21 or vdd <= 0:
        return {"valid": False, "reason": "VDD must be positive and matched curves need >= 21 points",
                "snm_v": None, "snm_mv": None, "squares": []}
    xs = [point[0] for point in direct_curve]
    if any(abs(x - mirrored_curve[index][0]) > 1e-12 for index, x in enumerate(xs)):
        return {"valid": False, "reason": "Direct and mirrored VTC grids do not match",
                "snm_v": None, "snm_mv": None, "squares": []}
    bounds = [(x, min(direct_curve[index][1], mirrored_curve[index][1]),
               max(direct_curve[index][1], mirrored_curve[index][1]))
              for index, x in enumerate(xs)]

    # The central VTC crossing is the metastable boundary between the two eyes.
    differences = [direct_curve[index][1] - mirrored_curve[index][1]
                   for index in range(len(xs))]
    crossing_candidates = []
    for index in range(len(xs) - 1):
        d0, d1 = differences[index], differences[index + 1]
        if d0 == 0:
            crossing_candidates.append(xs[index])
        elif d0 * d1 < 0:
            fraction = abs(d0) / (abs(d0) + abs(d1))
            crossing_candidates.append(xs[index] + fraction * (xs[index + 1] - xs[index]))
    crossing = (min(crossing_candidates, key=lambda value: abs(value - vdd / 2.0))
                if crossing_candidates else
                xs[min(range(len(xs)), key=lambda index: abs(differences[index]))])
    split = min(range(len(xs)), key=lambda index: abs(xs[index] - crossing))

    states = (("upper_left", "Left node=0 / Right node=1"),
              ("lower_right", "Left node=1 / Right node=0"))
    squares = []
    failed_lobes: list[str] = []
    for lobe, ((start, stop), (state_key, state_label)) in enumerate(
            zip(((0, split), (split, len(xs) - 1)), states), 1):
        best_side = 0.0
        best_position = None
        for left_index in range(start, stop + 1):
            minimum_upper = math.inf
            maximum_lower = -math.inf
            for right_index in range(left_index, stop + 1):
                minimum_upper = min(minimum_upper, bounds[right_index][2])
                maximum_lower = max(maximum_lower, bounds[right_index][1])
                side = xs[right_index] - xs[left_index]
                clearance = minimum_upper - maximum_lower
                if side <= clearance + 1e-12:
                    if side > best_side:
                        y0 = maximum_lower + max(0.0, clearance - side) / 2.0
                        best_side = side
                        best_position = (xs[left_index], y0)
                elif right_index > left_index:
                    break
        if best_position is None:
            failed_lobes.append(state_key)
            best_position = (xs[start], bounds[start][1])
        squares.append({"lobe": lobe, "state_key": state_key, "state": state_label,
                        "x_v": best_position[0], "y_v": best_position[1],
                        "side_v": best_side, "side_mv": 1000.0 * best_side,
                        "fit_valid": state_key not in failed_lobes})

    upper_v, lower_v = squares[0]["side_v"], squares[1]["side_v"]
    mean_v = (upper_v + lower_v) / 2.0
    delta_v = upper_v - lower_v
    snm_v = min(upper_v, lower_v)
    return {
        "valid": not failed_lobes,
        "reason": ("Independent two-eye maximum-square fit" if not failed_lobes else
                   "No positive square in: " + ", ".join(failed_lobes)),
        "definition": "Cell Read SNM is the smaller of the two state-dependent square sides",
        "mode": mode,
        "grid_points": len(xs),
        "crossing_v": crossing,
        "trip_v": crossing,
        "snm_v": snm_v,
        "snm_mv": 1000.0 * snm_v,
        "snm_upper_left_v": upper_v,
        "snm_upper_left_mv": 1000.0 * upper_v,
        "snm_lower_right_v": lower_v,
        "snm_lower_right_mv": 1000.0 * lower_v,
        "delta_snm_v": delta_v,
        "delta_snm_mv": 1000.0 * delta_v,
        "mismatch_index_pct": (abs(delta_v) / mean_v * 100.0 if mean_v > 0 else None),
        "failed_state_keys": failed_lobes,
        "squares": squares,
    }


def _fit_diagonal_write_square(upper_curve: list[tuple[float, float]],
                               lower_curve: list[tuple[float, float]],
                               vdd: float) -> dict:
    """Fit the largest axis-aligned write square with its +45-degree diagonal on y=x.

    The square lower-left and upper-right corners are (a, a) and (b, b).
    Therefore its Vin=Vout diagonal is the calculation reference while the
    complete square remains inside the upper W=1 and lower W=0 write VTCs.
    """
    if (len(upper_curve) != len(lower_curve) or len(upper_curve) < 21 or
            vdd <= 0):
        return {"valid": False, "reason": "matched VTC grids with >=21 points are required",
                "square": None, "snm_v": None, "snm_mv": None}
    xs = [point[0] for point in upper_curve]
    if any(abs(x - lower_curve[index][0]) > 1e-12 for index, x in enumerate(xs)):
        return {"valid": False, "reason": "W=1 and W=0 VTC grids do not match",
                "square": None, "snm_v": None, "snm_mv": None}
    upper = [point[1] for point in upper_curve]
    lower = [point[1] for point in lower_curve]
    best: dict | None = None
    for start in range(len(xs) - 1):
        min_upper, max_lower = math.inf, -math.inf
        for stop in range(start, len(xs)):
            min_upper = min(min_upper, upper[stop])
            max_lower = max(max_lower, lower[stop])
            side = xs[stop] - xs[start]
            if side <= 0:
                continue
            if min_upper + 1e-12 >= xs[stop] and max_lower <= xs[start] + 1e-12:
                if best is None or side > best["side_v"]:
                    best = {"x_v": xs[start], "y_v": xs[start], "side_v": side,
                            "side_mv": 1000.0 * side,
                            "constraint": "square diagonal follows Vin=Vout"}
            elif max_lower > xs[start] + 1e-12:
                # The lower curve is monotonically decreasing for the modeled
                # VTCs; extending this candidate cannot recover the constraint.
                break
    if best is None:
        return {"valid": False, "reason": "no positive Vin=Vout-diagonal write square",
                "square": None, "snm_v": 0.0, "snm_mv": 0.0}
    return {"valid": True, "reason": "maximum write square constrained to Vin=Vout diagonal",
            "square": best, "snm_v": best["side_v"], "snm_mv": best["side_mv"]}


def _fit_write_closing_eye(direct_curve: list[tuple[float, float]],
                           mirrored_curve: list[tuple[float, float]],
                           fallback: dict) -> dict:
    """Fit the write-destabilized eye bounded by its two DC crossings.

    During a BL sweep, the eye belonging to the state being overwritten is
    bounded by a stable and a metastable crossing.  Restricting the square to
    that interval prevents the fitter from jumping to the surviving state as
    the write eye becomes small.  With one non-rail crossing the cell is still
    in the undisturbed/read-like region, so the ordinary limiting square is
    used.  With no sign-changing crossing, the write eye has closed.
    """
    xs = [point[0] for point in direct_curve]
    differences = [direct_curve[index][1] - mirrored_curve[index][1]
                   for index in range(len(xs))]
    crossing_intervals = [
        index for index in range(len(xs) - 1)
        if differences[index] * differences[index + 1] < 0
    ]
    if not crossing_intervals:
        return {"side_v": 0.0, "side_mv": 0.0, "fit_valid": False,
                "crossing_count": 0, "reason": "Write-state eye is closed"}
    if len(crossing_intervals) == 1:
        side_v = float(fallback.get("snm_v") or 0.0)
        return {"side_v": side_v, "side_mv": 1000.0 * side_v,
                "fit_valid": side_v > 0, "crossing_count": 1,
                "reason": "Read-like region; ordinary limiting square applies"}

    # The first adjacent crossing pair bounds the write-destabilized eye for
    # the BL-low / BLB-high polarity used here.
    start = crossing_intervals[0]
    stop = crossing_intervals[1] + 1
    bounds = [(min(direct_curve[index][1], mirrored_curve[index][1]),
               max(direct_curve[index][1], mirrored_curve[index][1]))
              for index in range(len(xs))]
    best_side = 0.0
    best_position = None
    for left_index in range(start, stop + 1):
        minimum_upper = math.inf
        maximum_lower = -math.inf
        for right_index in range(left_index, stop + 1):
            minimum_upper = min(minimum_upper, bounds[right_index][1])
            maximum_lower = max(maximum_lower, bounds[right_index][0])
            side = xs[right_index] - xs[left_index]
            clearance = minimum_upper - maximum_lower
            if side <= clearance + 1e-12:
                if side > best_side:
                    best_side = side
                    best_position = (xs[left_index], maximum_lower)
            elif right_index > left_index:
                break
    return {"side_v": best_side, "side_mv": 1000.0 * best_side,
            "fit_valid": best_side > 0, "crossing_count": len(crossing_intervals),
            "x_v": None if best_position is None else best_position[0],
            "y_v": None if best_position is None else best_position[1],
            "reason": "Maximum square inside the write-destabilized eye"}


class Sram6T:
    def __init__(self, wat: WatPoint, cfg: Config):
        self.wat, self.cfg = wat, cfg
        self.pu = Device(wat.pu_vt, wat.pu_ids, cfg.wat_vdd)
        self.pg = Device(wat.pg_vt, wat.pg_ids, cfg.wat_vdd)
        self.pd = Device(wat.pd_vt, wat.pd_ids, cfg.wat_vdd)

    def _balance(self, vin: float, vout: float, vdd: float, mode: str) -> float:
        # positive means the node is charged upward
        up = self.pu.current(vdd - vin, vdd - vout)
        down = self.pd.current(vin, vout)
        if mode == "read":
            wl = self.cfg.read_wordline_over_vdd * vdd
            bitline = self.cfg.read_bitline_over_vdd * vdd
            up += self._access_current(self.pg, vout, bitline, wl)
        return up - down

    def transfer(self, vin: float, vdd: float, mode: str = "hold") -> float:
        lo, hi = 0.0, vdd
        f_lo = self._balance(vin, lo, vdd, mode)
        f_hi = self._balance(vin, hi, vdd, mode)
        if f_lo <= 0:
            return 0.0
        if f_hi >= 0:
            return vdd
        for _ in range(48):
            mid = (lo + hi) / 2
            if self._balance(vin, mid, vdd, mode) > 0:
                lo = mid
            else:
                hi = mid
        return (lo + hi) / 2

    def transfer_with_bitline(self, vin: float, vdd: float, bitline: float,
                              wordline: float | None = None) -> float:
        """Inverter transfer with WL high and the access device tied to bitline."""
        wl = vdd if wordline is None else wordline
        def balance(vout: float) -> float:
            return (self.pu.current(vdd - vin, vdd - vout) - self.pd.current(vin, vout) +
                    self._access_current(self.pg, vout, bitline, wl))
        lo, hi = 0.0, vdd
        if balance(lo) <= 0:
            return 0.0
        if balance(hi) >= 0:
            return vdd
        for _ in range(48):
            mid = (lo + hi) / 2
            if balance(mid) > 0:
                lo = mid
            else:
                hi = mid
        return (lo + hi) / 2

    def write_vtc_pair(self, vdd: float, points: int = 201) -> tuple[list[tuple[float, float]], list[tuple[float, float]]]:
        """Write-condition VTC pair for low BL and high BLB at the configured WL."""
        low, high = [], []
        wl = self.cfg.write_wordline_over_vdd * vdd
        low_bl = self.cfg.write_low_bitline_over_vdd * vdd
        high_bl = self.cfg.write_high_bitline_over_vdd * vdd
        for i in range(points):
            vin = vdd * i / (points - 1)
            low.append((vin, self.transfer_with_bitline(vin, vdd, low_bl, wl)))
            high.append((vin, self.transfer_with_bitline(vin, vdd, high_bl, wl)))
        return low, high

    def write_butterfly(self, vdd: float, low_bitline: float,
                        high_bitline: float | None = None,
                        points: int = 401) -> dict:
        """Return the full-cell write butterfly at one low-bitline voltage.

        Coordinates are x=high-BLB storage node and y=low-BL storage node.
        The low-BL inverter directly produces y=f_low(x); the high-BLB
        inverter produces x=f_high(y), whose inverse is plotted on the
        common y(x) axes.
        """
        high_bl = vdd if high_bitline is None else high_bitline
        wl = self.cfg.write_wordline_over_vdd * vdd
        n = max(21, int(points))
        low_curve = [
            (vdd * index / (n - 1),
             self.transfer_with_bitline(vdd * index / (n - 1), vdd, low_bitline, wl))
            for index in range(n)
        ]
        high_curve = [
            (vdd * index / (n - 1),
             self.transfer_with_bitline(vdd * index / (n - 1), vdd, high_bl, wl))
            for index in range(n)
        ]
        mirrored_high = _inverse_vtc(high_curve)
        fitted = _fit_butterfly_squares(low_curve, mirrored_high, vdd, "write")
        closing_eye = _fit_write_closing_eye(low_curve, mirrored_high, fitted)
        fitted.update({"snm_v": closing_eye["side_v"],
                       "snm_mv": closing_eye["side_mv"],
                       "write_closing_eye": closing_eye})
        fitted.update({
            "low_bitline_v": low_bitline,
            "high_bitline_v": high_bl,
            "wordline_v": wl,
            "coordinate_definition": {
                "x": "storage-node voltage on the high-BLB side",
                "y": "storage-node voltage on the low-BL side",
                "direct_vtc": "low-BL inverter: y=f_low(x)",
                "mirrored_vtc": "inverse high-BLB inverter: y=f_high^-1(x)",
            },
        })
        return fitted

    def vtc(self, vdd: float, mode: str = "hold", points: int | None = None) -> list[tuple[float, float]]:
        n = points or self.cfg.grid_points
        return [(vdd * i / (n - 1), self.transfer(vdd * i / (n - 1), vdd, mode)) for i in range(n)]

    def trip_point(self, vdd: float, mode: str) -> float:
        lo, hi = 0.0, vdd
        for _ in range(50):
            mid = (lo + hi) / 2
            if self.transfer(mid, vdd, mode) > mid:
                lo = mid
            else:
                hi = mid
        return (lo + hi) / 2

    def snm(self, vdd: float, mode: str = "read") -> float:
        """Legacy metastable-centered SNM proxy (V); retained for traceability."""
        if vdd <= 0:
            return 0.0
        m = self.trip_point(vdd, mode)

        def fits(side: float) -> bool:
            if m - side < 0 or m + side > vdd:
                return False
            # Square touches the metastable point and must remain inside both
            # monotonic VTC boundaries of the upper butterfly lobe.
            return (self.transfer(m - side, vdd, mode) >= m + side and
                    self.transfer(m + side, vdd, mode) <= m - side)

        lo, hi = 0.0, min(m, vdd - m)
        for _ in range(42):
            mid = (lo + hi) / 2
            if fits(mid):
                lo = mid
            else:
                hi = mid
        return lo

    def butterfly_squares(self, vdd: float, mode: str = "read",
                          points: int = 1201) -> dict:
        """Fit the two squares of a symmetric cell from one VTC and its inverse."""
        if vdd <= 0 or points < 21:
            return {"valid": False, "reason": "VDD must be positive and points >= 21",
                    "snm_v": None, "snm_mv": None, "squares": []}
        curve = self.vtc(vdd, mode, points)
        return _fit_butterfly_squares(curve, _inverse_vtc(curve), vdd, mode)

    def analytical_read_snm_eq_3_36(self, vdd: float) -> dict:
        """Evaluate the 6T read-SNM expression from Section 3.4.2, Eq. 3.36.

        Reference: *High-Speed CMOS Circuit Technology*, Section 3.4.2,
        "Analytical SNM Expression for a 6T SRAM Cell", printed page 51.

        The source assumes one common threshold voltage for PU/PG/PD and
        long-channel square-law devices.  WAT supplies three distinct values,
        so this implementation explicitly maps them to their arithmetic mean.
        Domain failures are reported instead of forcing a non-physical result.
        """
        q = self.pu.beta / self.pg.beta  # beta_p / beta_a
        r = self.pd.beta / self.pg.beta  # beta_d / beta_a
        vth_eff = (self.wat.pu_vt + self.wat.pg_vt + self.wat.pd_vt) / 3.0
        result = {
            "source": "High-Speed CMOS Circuit Technology, Section 3.4.2, Equation 3.36",
            "mode": "read-accessed 6T SRAM",
            "valid": False,
            "reason": "",
            "snm_v": None,
            "snm_mv": None,
            "vdd_v": vdd,
            "vth_eff_v": vth_eff,
            "vth_mapping": "arithmetic mean of |Vt_PU|, Vt_PG and Vt_PD",
            "q_beta_p_over_beta_a": q,
            "r_beta_d_over_beta_a": r,
            "vs_v": None,
            "vr_v": None,
            "k": None,
            "term_a_v": None,
            "term_b_v": None,
            "assumptions": [
                "PU, PG and PD share one threshold voltage",
                "long-channel square-law MOS model",
                "Q1/Q4 saturated and Q2/Q5 linear as defined by the source",
                "local linearity around the read operating point",
                "short-channel effects are neglected",
            ],
        }

        def invalid(reason: str) -> dict:
            result["reason"] = reason
            return result

        if not all(math.isfinite(x) and x > 0 for x in (vdd, vth_eff, q, r)):
            return invalid("VDD, effective VTH, q and r must be positive finite values")

        vs = vdd - vth_eff
        vr = vs - (r / (r + 1.0)) * vth_eff
        result["vs_v"], result["vr_v"] = vs, vr
        if vs <= 0:
            return invalid("VDD must exceed the effective threshold voltage")
        if abs(vr) < 1e-15:
            return invalid("Vr is zero in the analytical model")

        k_domain = (r + 1.0) - (vs * vs) / (vr * vr)
        result["k_domain"] = k_domain
        if k_domain <= 0:
            return invalid("Analytical square-root domain is non-positive")

        k = (r / (r + 1.0)) * (math.sqrt((r + 1.0) / k_domain) - 1.0)
        result["k"] = k
        if not math.isfinite(k) or abs(k) < 1e-15 or abs(k + 1.0) < 1e-15:
            return invalid("Analytical model produced an invalid k")

        denom_a = 1.0 + r / (k * (r + 1.0))
        root_arg = (r / q) * (1.0 + 2.0 * k + (r / q) * k * k)
        result["root_argument"] = root_arg
        if root_arg < 0 or abs(denom_a) < 1e-15:
            return invalid("Analytical denominator or square-root domain is invalid")
        denom_b = 1.0 + k * r / q + math.sqrt(root_arg)
        if abs(denom_b) < 1e-15:
            return invalid("Analytical second denominator is zero")

        term_a = (vdd - ((2.0 * r + 1.0) / (r + 1.0)) * vth_eff) / denom_a
        term_b = (vdd - 2.0 * vth_eff) / denom_b
        snm_v = vth_eff - (term_a - term_b) / (k + 1.0)
        result["term_a_v"], result["term_b_v"] = term_a, term_b
        if not math.isfinite(snm_v) or snm_v < 0 or snm_v > vdd:
            return invalid("Analytical result is outside the physical range 0 to VDD")

        result.update({"valid": True, "reason": "Within the real-valued analytical domain",
                       "snm_v": snm_v, "snm_mv": 1000.0 * snm_v})
        return result

    def strength_ratios(self) -> dict[str, float]:
        """Return Vt-aware model-beta ratios plus direct WAT-current proxies."""
        return {
            "cell_ratio_beta": self.pd.beta / self.pg.beta,
            "pull_up_ratio_beta": self.pg.beta / self.pu.beta,
            "cell_ratio_ids_proxy": self.wat.pd_ids / self.wat.pg_ids,
            "pull_up_ratio_ids_proxy": self.wat.pg_ids / self.wat.pu_ids,
        }

    def write_snm(self, vdd: float) -> float:
        """Compact-model write bitline-noise margin (V).

        The result is the largest voltage allowed on the nominally-low write
        bitline while PG can still overcome PU at the hold inverter trip point.
        This is a directional WAT-calibrated proxy, not a foundry sign-off WSNM.
        """
        if vdd <= 0:
            return 0.0
        trip = self.trip_point(vdd, "hold")
        wl = self.cfg.write_wordline_over_vdd * vdd
        nominal_low = self.cfg.write_low_bitline_over_vdd * vdd
        pull_up = self.pu.current(vdd, vdd - trip)

        def writable(low_bitline: float) -> bool:
            if low_bitline >= trip:
                return False
            access = self.pg.current(wl - low_bitline, trip - low_bitline)
            return access >= pull_up

        if not writable(nominal_low):
            return 0.0
        lo, hi = nominal_low, trip
        for _ in range(42):
            mid = (lo + hi) / 2
            if writable(mid):
                lo = mid
            else:
                hi = mid
        return lo - nominal_low

    @staticmethod
    def _access_current(pg: Device, node: float, bitline: float, wl: float) -> float:
        # Signed current entering node through a symmetric NMOS access device.
        if bitline >= node:
            return pg.current(wl - node, bitline - node)
        return -pg.current(wl - bitline, node - bitline)

    def operate(self, vdd: float, operation: str) -> tuple[float, float]:
        """Damped DC relaxation of the coupled 6T cell; returns (Q, QB)."""
        if operation == "read":
            q, qb, bl, blb = 0.0, vdd, vdd, vdd
        elif operation == "write":
            q, qb, bl, blb = vdd, 0.0, 0.0, vdd  # write Q=0
        else:
            raise ValueError(operation)
        max_i = max(self.wat.pu_ids, self.wat.pg_ids, self.wat.pd_ids, 1.0)
        gain = 0.025 * max(vdd, 0.15) / max_i
        for _ in range(5000):
            iq = (self.pu.current(vdd - qb, vdd - q) - self.pd.current(qb, q) +
                  self._access_current(self.pg, q, bl, vdd))
            iqb = (self.pu.current(vdd - q, vdd - qb) - self.pd.current(q, qb) +
                   self._access_current(self.pg, qb, blb, vdd))
            nq = min(vdd, max(0.0, q + gain * iq))
            nqb = min(vdd, max(0.0, qb + gain * iqb))
            if max(abs(nq - q), abs(nqb - qb)) < max(1e-10, vdd * 1e-8):
                q, qb = nq, nqb
                break
            q, qb = nq, nqb
        return q, qb

    def read_vmin(self) -> float | None:
        for vdd in frange(self.cfg.vmin_start, self.cfg.vmin_stop, self.cfg.vmin_step):
            q, qb = self.operate(vdd, "read")
            if q < 0.35 * vdd and qb > 0.65 * vdd and self.snm(vdd, "read") >= self.cfg.read_snm_limit:
                return vdd
        return None

    def write_vmin(self) -> float | None:
        for vdd in frange(self.cfg.vmin_start, self.cfg.vmin_stop, self.cfg.vmin_step):
            q, qb = self.operate(vdd, "write")
            if q < 0.20 * vdd and qb > 0.80 * vdd:
                return vdd
        return None


def drive_monitor_metrics(wat: WatPoint, vdd: float) -> dict[str, float | list[tuple[float, float]]]:
    """Return compact-model values for the interactive 6T Drive Monitor.

    This intentionally uses the same WAT-calibrated 6T model as the main
    application.  It is a learning aid: the metrics show directional device
    trade-offs, not a replacement for PDK/array sign-off.
    """
    if not math.isfinite(vdd) or vdd <= 0:
        raise ValueError("Drive Monitor VDD must be a positive finite value")
    cfg = Config(wat_vdd=vdd, nominal_vdd=vdd, grid_points=401)
    model = Sram6T(wat, cfg)
    read_fit = model.butterfly_squares(vdd, "read", points=401)
    ratios = model.strength_ratios()
    return {
        "beta_pu": model.pu.beta,
        "beta_pg": model.pg.beta,
        "beta_pd": model.pd.beta,
        "cell_ratio": ratios["cell_ratio_beta"],
        "pull_up_ratio": ratios["pull_up_ratio_beta"],
        "read_snm_mv": float(read_fit.get("snm_mv") or 0.0),
        "write_margin_mv": 1000.0 * model.write_snm(vdd),
        "read_vtc": model.vtc(vdd, "read", points=81),
    }


def drive_monitor_shmoo_reference(vdd: float,
                                   drive_sigma: float = .08) -> dict[str, object]:
    """Build a deterministic statistical CR/PR reference for the monitor.

    The target population is anchored to the nominal 28 nm WAT reference and
    uses independent log-normal PU/PG/PD effective-drive variation. Its median
    moves with Model VDD, but it does not chase the live device sliders.
    """
    if not math.isfinite(vdd) or vdd <= 0:
        raise ValueError("Drive Monitor VDD must be a positive finite value")
    sigma = float(drive_sigma)
    if not math.isfinite(sigma) or sigma <= 0:
        raise ValueError("Drive variation sigma must be positive")
    nominal = WatPoint("DriveReference", .385, 44.0, .365, 82.0, .355, 124.0)
    model = Sram6T(nominal, Config(wat_vdd=vdd, nominal_vdd=vdd, grid_points=101))
    # Mid-quantile normal scores avoid random seeds while approximating a
    # reproducible log-normal process-strength population (9^3 = 729 cells).
    normal = statlib.NormalDist()
    factors = [math.exp(sigma * normal.inv_cdf((index + .5) / 9.0))
               for index in range(9)]
    cr_values: list[float] = []
    pr_values: list[float] = []
    for pu_factor in factors:
        for pg_factor in factors:
            for pd_factor in factors:
                beta_pu = model.pu.beta * pu_factor
                beta_pg = model.pg.beta * pg_factor
                beta_pd = model.pd.beta * pd_factor
                cr_values.append(beta_pd / beta_pg)
                pr_values.append(beta_pg / beta_pu)
    return {
        "vdd_v": vdd,
        "sample_count": len(cr_values),
        "drive_sigma": sigma,
        "cr": _robust_distribution(cr_values),
        "pr": _robust_distribution(pr_values),
        "definition": ("Nominal-reference log-normal PU/PG/PD drive population; "
                       "green requires CR and PR >= median, yellow requires both >= Q1."),
    }


class AsymmetricSram6T:
    """Cross-coupled Read-SNM model retaining all six independent WAT objects."""

    def __init__(self, cell: SixTWatCell, cfg: Config):
        self.cell = cell
        self.cfg = cfg
        self.left = Sram6T(cell.side(1), cfg)   # PUL / PGL / PDL
        self.right = Sram6T(cell.side(2), cfg)  # PUR / PGR / PDR

    def read_butterfly(self, vdd: float, points: int = 1201) -> dict:
        # Plot coordinates are (left storage node, right storage node).
        # The right inverter directly gives y=f_right(x).  The left inverter
        # gives x=f_left(y), therefore its inverse is the second curve y(x).
        direct_fit = self.right.vtc(vdd, "read", points)
        mirrored_fit = _inverse_vtc(self.left.vtc(vdd, "read", points))
        fitted = _fit_butterfly_squares(direct_fit, mirrored_fit, vdd, "read")
        fitted["coordinate_definition"] = {
            "x": "left storage-node voltage (PUL/PGL/PDL side)",
            "y": "right storage-node voltage (PUR/PGR/PDR side)",
            "direct_vtc": "right inverter: y=f_right(x)",
            "mirrored_vtc": "inverse left inverter: y=f_left^-1(x)",
        }
        direct_plot = self.right.vtc(vdd, "read", 201)
        mirrored_plot = _inverse_vtc(self.left.vtc(vdd, "read", 201))
        return {"read_butterfly": fitted, "read_vtc": direct_plot,
                "read_vtc_mirrored": mirrored_plot}

    def write_butterfly(self, vdd: float, low_bitline: float,
                        high_bitline: float | None = None,
                        points: int = 401,
                        wordline_low: float | None = None,
                        wordline_high: float | None = None) -> dict:
        """Write butterfly retaining all six left/right WAT objects."""
        high_bl = vdd if high_bitline is None else high_bitline
        wl_low = self.cfg.write_wordline_over_vdd * vdd if wordline_low is None else wordline_low
        wl_high = self.cfg.write_wordline_over_vdd * vdd if wordline_high is None else wordline_high
        n = max(21, int(points))
        low_curve = [
            (vdd * index / (n - 1),
             self.left.transfer_with_bitline(
                 vdd * index / (n - 1), vdd, low_bitline, wl_low))
            for index in range(n)
        ]
        high_curve = [
            (vdd * index / (n - 1),
             self.right.transfer_with_bitline(
                 vdd * index / (n - 1), vdd, high_bl, wl_high))
            for index in range(n)
        ]
        mirrored_high = _inverse_vtc(high_curve)
        fitted = _fit_butterfly_squares(low_curve, mirrored_high, vdd, "write")
        closing_eye = _fit_write_closing_eye(low_curve, mirrored_high, fitted)
        fitted.update({"snm_v": closing_eye["side_v"],
                       "snm_mv": closing_eye["side_mv"],
                       "write_closing_eye": closing_eye})
        fitted.update({
            "low_bitline_v": low_bitline,
            "high_bitline_v": high_bl,
            "wordline_low_v": wl_low,
            "wordline_high_v": wl_high,
            "coordinate_definition": {
                "x": "right storage-node voltage (PUR/PGR/PDR, high-BLB side)",
                "y": "left storage-node voltage (PUL/PGL/PDL, low-BL side)",
                "direct_vtc": "left inverter: y=f_left(x), PGL tied to BL low",
                "mirrored_vtc": "inverse right inverter: y=f_right^-1(x), PGR tied to BLB high",
            },
        })
        return fitted


def write_wsnm_states(vdd: float, left: Sram6T, right: Sram6T,
                      cfg: Config, points: int = 1201) -> dict:
    """Build a paper-style Write-SNM window from the W=1/W=0 VTC pair.

    W=1 is the VTC on the BLB-high side and is the upper curve; W=0 is the
    VTC on the BL-low side and is the lower curve.  The usable write lobe is
    the lobe with the largest inscribed square, following the one-window WSNM
    presentation commonly used in SRAM literature.  This is WAT-calibrated
    compact-model analysis, not a SPICE sign-off result.
    """
    if vdd <= 0:
        raise ValueError("VDD must be positive for WSNM analysis")
    n = max(201, int(points))
    wl = cfg.write_wordline_over_vdd * vdd
    low_bl = cfg.write_low_bitline_over_vdd * vdd
    high_bl = cfg.write_high_bitline_over_vdd * vdd

    def vtc(model: Sram6T, bitline: float) -> list[tuple[float, float]]:
        return [
            (vdd * index / (n - 1),
             model.transfer_with_bitline(vdd * index / (n - 1), vdd, bitline, wl))
            for index in range(n)
        ]

    write_0_curve = vtc(left, low_bl)
    write_1_curve = vtc(right, high_bl)
    fitted = _fit_diagonal_write_square(write_1_curve, write_0_curve, vdd)
    write_square = fitted.get("square")
    wsnm_v = fitted.get("snm_v")
    bias = {"wordline_v": wl, "bl_v": low_bl, "blb_v": high_bl}
    return {
        "method": "W=1 upper and W=0 lower write-VTC maximum-square extraction",
        "vdd_v": vdd,
        "write_0": {"label": "W=0 (lower VTC)", "curve": write_0_curve,
                    "write_bias": bias},
        "write_1": {"label": "W=1 (upper VTC)", "curve": write_1_curve,
                    "write_bias": bias},
        "write_square": write_square,
        "snm_v": wsnm_v,
        "snm_mv": None if wsnm_v is None else 1000.0 * wsnm_v,
        "valid": wsnm_v is not None and wsnm_v > 0,
    }


def write_snm_vs_bitline(vdd: float,
                         butterfly_at_bl: Callable[[float, int], dict],
                         sweep_points: int = 37,
                         fit_points: int = 401) -> dict:
    """Sweep write BL and find the textbook SNM=0 write boundary.

    BLB and WL are held at VDD by the supplied butterfly solver.  The write
    direction is BL: VDD -> 0 V.  Cell WSNM at each point is the smaller of
    the two state-dependent butterfly-square sides.  The write-trip BL
    voltage is the open/closed-eye boundary; the required BL swing is
    VDD-write_trip_bl.
    """
    if not math.isfinite(vdd) or vdd <= 0:
        raise ValueError("VDD must be a positive finite value")
    sweep_points = max(9, int(sweep_points))
    fit_points = max(101, int(fit_points))
    cache: dict[float, dict] = {}

    def solve(bitline_v: float) -> dict:
        key = round(min(vdd, max(0.0, bitline_v)), 12)
        if key not in cache:
            cache[key] = butterfly_at_bl(key, fit_points)
        return cache[key]

    def is_open(result: dict) -> bool:
        return bool(result.get("snm_v", 0.0) > max(1e-9, vdd / fit_points / 2.0))

    rows = []
    for index in range(sweep_points):
        bl = vdd * index / (sweep_points - 1)
        fitted = solve(bl)
        rows.append({
            "write_bl_v": bl,
            "cell_write_snm_mv": fitted.get("snm_mv", 0.0),
            "upper_left_snm_mv": fitted.get("snm_upper_left_mv", 0.0),
            "lower_right_snm_mv": fitted.get("snm_lower_right_mv", 0.0),
            "eye_open": is_open(fitted),
        })

    closed_at_zero = not is_open(solve(0.0))
    open_at_vdd = is_open(solve(vdd))
    if closed_at_zero and open_at_vdd:
        closed_bl, open_bl = 0.0, vdd
        for _ in range(26):
            mid = (closed_bl + open_bl) / 2.0
            if is_open(solve(mid)):
                open_bl = mid
            else:
                closed_bl = mid
        write_trip = (closed_bl + open_bl) / 2.0
        status = "VALID"
        reason = "SNM=0 boundary bracketed between closed and open butterfly eyes"
    elif not closed_at_zero:
        write_trip = 0.0
        status = "NO CLOSURE"
        reason = "Butterfly eye remains open at BL=0 V"
    else:
        write_trip = None
        status = "UNSTABLE"
        reason = "Butterfly eye is already closed at BL=VDD"

    return {
        "method": "Textbook writeability: sweep BL from VDD to 0 V and locate SNM=0",
        "vdd_v": vdd,
        "wordline_v": vdd,
        "high_bitline_v": vdd,
        "write_direction": "BL: VDD to 0 V",
        "status": status,
        "reason": reason,
        "write_trip_bl_v": write_trip,
        "required_bl_swing_v": (vdd - write_trip if write_trip is not None else None),
        "points": rows,
    }


def _vtc_diagonal_intersection(curve: list[tuple[float, float]]) -> tuple[float, float] | None:
    """Return the interpolated crossing of a monotonic VTC and Vout=Vin."""
    if len(curve) < 2:
        return None
    previous_x, previous_y = curve[0]
    previous_delta = previous_y - previous_x
    if abs(previous_delta) <= 1e-15:
        return previous_x, previous_y
    for x_value, y_value in curve[1:]:
        delta = y_value - x_value
        if abs(delta) <= 1e-15:
            return x_value, y_value
        if delta * previous_delta < 0:
            denominator = previous_delta - delta
            fraction = previous_delta / denominator if abs(denominator) > 1e-15 else 0.5
            crossing_x = previous_x + fraction * (x_value - previous_x)
            crossing_y = previous_y + fraction * (y_value - previous_y)
            crossing = (crossing_x + crossing_y) / 2.0
            return crossing, crossing
        previous_x, previous_y, previous_delta = x_value, y_value, delta
    return None


def single_wat_write_snm_geometry(vdd: float,
                                  high_side_curves: list[tuple[str, list[tuple[float, float]]]]) -> dict:
    """Build the single-WAT graphical WSNM view used by the reference diagram.

    With WL asserted and the retained storage-node side connected to its high
    bitline, each broken-loop write VTC is intersected with Vout=Vin.  The
    intersection voltage is drawn as the side of an origin-anchored square.
    For independent left/right WAT objects, the smaller of the two write
    polarities is reported and plotted.

    This is intentionally kept separate from the SNM-versus-BL eye-closure
    calculation: it is a compact graphical WAT estimate, not measured write
    trip voltage or foundry sign-off WSNM.
    """
    candidates = []
    for polarity, curve in high_side_curves:
        crossing = _vtc_diagonal_intersection(curve)
        candidates.append({
            "write_polarity": polarity,
            "curve": curve,
            "intersection": crossing,
            "wsnm_v": None if crossing is None else crossing[0],
            "wsnm_mv": None if crossing is None else 1000.0 * crossing[0],
        })
    valid = [candidate for candidate in candidates if candidate["wsnm_v"] is not None]
    limiting = min(valid, key=lambda candidate: candidate["wsnm_v"]) if valid else None
    return {
        "method": "Single-WAT write VTC diagonal-intersection geometry",
        "vdd_v": vdd,
        "valid": limiting is not None,
        "reason": ("Limiting write polarity selected from the two 6T storage-node sides"
                   if limiting is not None else "No VTC crossing with Vout=Vin"),
        "write_polarity": None if limiting is None else limiting["write_polarity"],
        "curve": [] if limiting is None else limiting["curve"],
        "intersection": None if limiting is None else limiting["intersection"],
        "wsnm_v": None if limiting is None else limiting["wsnm_v"],
        "wsnm_mv": None if limiting is None else limiting["wsnm_mv"],
        "polarity_results": candidates,
        "square_definition": "origin-anchored square; upper-right corner lies on Vout=Vin and the selected write VTC",
    }


class WtZeroBitVminTest:
    """Object-oriented WT 0-bit Vmin flow for one mismatched 6T bitcell."""

    TEST_NAMES = ("Scan4N", "Select_Write", "Select_Read")

    def __init__(self, cell: SixTWatCell, cfg: Config):
        self.cell = cell
        self.cfg = cfg
        self.sides = [Sram6T(cell.side(i), cfg) for i in (1, 2)]

    def _write_pass(self, model: Sram6T, vdd: float) -> bool:
        q, qb = model.operate(vdd, "write")
        return q < .20*vdd and qb > .80*vdd

    def _read_pass(self, model: Sram6T, vdd: float) -> bool:
        q, qb = model.operate(vdd, "read")
        return (q < .35*vdd and qb > .65*vdd and
                model.snm(vdd, "read") >= self.cfg.read_snm_limit)

    def evaluate(self, test_name: str, vdd: float) -> dict:
        # Side 1 and side 2 represent the two logical data polarities in the
        # cross-coupled cell. A 0-bit pass requires every required phase pass.
        write = [self._write_pass(model, vdd) for model in self.sides]
        read = [self._read_pass(model, vdd) for model in self.sides]
        if test_name == "Select_Write":
            phases = {"Write-0": write[0], "Write-1": write[1]}
        elif test_name == "Select_Read":
            phases = {"Read-0": read[0], "Read-1": read[1]}
        elif test_name == "Scan4N":
            phases = {
                "↑ Write-0": write[0],
                "↑ Read-0 / Write-1": read[0] and write[1],
                "↓ Read-1 / Write-0": read[1] and write[0],
                "↓ Read-0": read[0],
            }
        else:
            raise ValueError(f"unknown WT test: {test_name}")
        failed = [name for name, passed in phases.items() if not passed]
        return {"pass": not failed, "failed_phase_count": len(failed),
                "failed_phases": failed, "phases": phases}

    def vmin(self, test_name: str) -> dict:
        last_failure: dict | None = None
        for vdd in frange(self.cfg.vmin_start, self.cfg.vmin_stop, self.cfg.vmin_step):
            status = self.evaluate(test_name, vdd)
            if status["pass"]:
                return {"test": test_name, "vmin_v": vdd, "zero_bit_pass": True,
                        "failed_phase_count": 0, "phases_at_vmin": status["phases"]}
            last_failure = status
        return {"test": test_name, "vmin_v": None, "zero_bit_pass": False,
                "failed_phase_count": None if last_failure is None else last_failure["failed_phase_count"],
                "phases_at_vmin": {}}

    def run(self) -> list[dict]:
        return [self.vmin(name) for name in self.TEST_NAMES]


def frange(start: float, stop: float, step: float) -> Iterable[float]:
    count = int(math.floor((stop - start) / step + 1e-9))
    for i in range(count + 1):
        yield round(start + i * step, 10)


def variants(wat: WatPoint, cfg: Config, device: str) -> list[tuple[str, WatPoint]]:
    vt_name, ids_name = f"{device}_vt", f"{device}_ids"
    base_vt, base_ids = getattr(wat, vt_name), getattr(wat, ids_name)
    frac = cfg.ids_step_pct / 100.0
    return [
        ("Baseline", wat),
        (f"Vt -{cfg.vt_step*1000:.0f}mV", replace(wat, **{vt_name: max(0.01, base_vt - cfg.vt_step)})),
        (f"Vt +{cfg.vt_step*1000:.0f}mV", replace(wat, **{vt_name: base_vt + cfg.vt_step})),
        (f"Ids -{cfg.ids_step_pct:g}%", replace(wat, **{ids_name: base_ids * (1-frac)})),
        (f"Ids +{cfg.ids_step_pct:g}%", replace(wat, **{ids_name: base_ids * (1+frac)})),
    ]


def metric(model: Sram6T, cfg: Config,
           read_butterfly: dict | None = None) -> dict[str, float | None]:
    square_points = _snm_fit_points(cfg)
    read_butterfly = read_butterfly or model.butterfly_squares(
        cfg.nominal_vdd, "read", square_points)
    analytical = model.analytical_read_snm_eq_3_36(cfg.nominal_vdd)
    return {
        "read_snm_mv": read_butterfly["snm_mv"],
        "read_snm_trip_proxy_mv": 1000 * model.snm(cfg.nominal_vdd, "read"),
        "analytical_read_snm_mv": analytical["snm_mv"],
        **model.strength_ratios(),
    }


def cell_metric(cell: SixTWatCell, cfg: Config) -> dict[str, float | None]:
    """Conservative half-cell mismatch metric: the lower Read SNM wins."""
    sides = [metric(Sram6T(cell.side(i), cfg), cfg) for i in (1, 2)]
    return {
        "read_snm_mv": min(s["read_snm_mv"] for s in sides),
        "cell_ratio_beta": min(s["cell_ratio_beta"] for s in sides),
        "pull_up_ratio_beta": min(s["pull_up_ratio_beta"] for s in sides),
        "cell_ratio_ids_proxy": min(s["cell_ratio_ids_proxy"] for s in sides),
        "pull_up_ratio_ids_proxy": min(s["pull_up_ratio_ids_proxy"] for s in sides),
    }


def evaluate_judgment(metrics: dict[str, float | None],
                      targets: JudgmentTargets) -> dict:
    """Evaluate four higher-is-better SRAM metrics against user-owned limits."""
    definitions = (
        ("cell_ratio", "Cell Ratio", "cell_ratio_beta", targets.cell_ratio_min, "ratio",
         "Increase PD strength or reduce PG strength; then re-check write margin."),
        ("pull_up_ratio", "Pull-up Ratio", "pull_up_ratio_beta", targets.pull_up_ratio_min, "ratio",
         "Increase PG strength or reduce PU strength to improve writeability."),
        ("hold_snm", "Hold SNM", "hold_snm_mv", targets.hold_snm_min_mv, "mV",
         "Check PU/PD balance and left-right mismatch; PG is off during hold."),
        ("read_snm", "Read SNM", "read_snm_mv", targets.read_snm_min_mv, "mV",
         "Increase PD relative to PG, then verify the writeability trade-off."),
        ("write_snm", "Write SNM", "write_snm_mv", targets.write_snm_min_mv, "mV",
         "Increase PG strength or reduce PU strength; verify Read SNM trade-off."),
    )
    rows = []
    for key, label, metric_key, target, unit, action in definitions:
        value = metrics.get(metric_key)
        if value is None:
            status, margin, margin_pct = "N/A", None, None
        else:
            margin = value - target
            margin_pct = margin / target * 100 if target else None
            if value >= target:
                status = "PASS"
            elif value >= target * (1 - targets.marginal_band_pct / 100):
                status = "MARGINAL"
            else:
                status = "FAIL"
        displayed_action = ("Meets target; no corrective adjustment is required."
                            if status == "PASS" else action)
        rows.append({"key": key, "label": label, "metric_key": metric_key,
                     "value": value, "target": target, "unit": unit,
                     "margin": margin, "margin_pct": margin_pct,
                     "status": status, "recommended_action": displayed_action})
    statuses = [row["status"] for row in rows]
    overall = "FAIL" if "FAIL" in statuses or "N/A" in statuses else (
        "MARGINAL" if "MARGINAL" in statuses else "PASS")
    return {"overall_status": overall, "marginal_band_pct": targets.marginal_band_pct,
            "targets": asdict(targets), "items": rows,
            "rule": "All five metrics must PASS; values below target but within the marginal band are MARGINAL."}


def analyze(wat: WatPoint, cfg: Config) -> dict:
    model = Sram6T(wat, cfg)
    square_points = _snm_fit_points(cfg)
    read_butterfly = model.butterfly_squares(cfg.nominal_vdd, "read", square_points)
    write_states = write_wsnm_states(cfg.nominal_vdd, model, model, cfg, square_points)
    read_vtc = model.vtc(cfg.nominal_vdd, "read", 201)
    baseline_metrics = metric(model, cfg, read_butterfly)
    baseline_metrics.update({
        "write_snm_mv": write_states["snm_mv"],
    })
    baseline = {"metrics": baseline_metrics,
                "analytical_read_snm_eq_3_36": model.analytical_read_snm_eq_3_36(cfg.nominal_vdd),
                "read_butterfly": read_butterfly,
                "read_vtc": read_vtc,
                "read_vtc_mirrored": _inverse_vtc(read_vtc),
                "write_wsnm": write_states,
                "read_trip_v": model.trip_point(cfg.nominal_vdd, "read")}
    report_config = {
        "wat_vdd": cfg.wat_vdd, "nominal_vdd": cfg.nominal_vdd,
        "grid_points": cfg.grid_points,
        "technology_node_nm": cfg.technology_node_nm,
        "channel_length_nm": cfg.channel_length_nm,
        "pu_width_nm": cfg.pu_width_nm, "pg_width_nm": cfg.pg_width_nm,
        "pd_width_nm": cfg.pd_width_nm,
        "nominal_temperature_c": cfg.nominal_temperature_c,
        "read_wordline_over_vdd": cfg.read_wordline_over_vdd,
        "read_bitline_over_vdd": cfg.read_bitline_over_vdd,
        "write_wordline_over_vdd": cfg.write_wordline_over_vdd,
        "write_low_bitline_over_vdd": cfg.write_low_bitline_over_vdd,
        "write_high_bitline_over_vdd": cfg.write_high_bitline_over_vdd,
    }
    return {"technology": asdict(tech_from_config(cfg)), "wat": asdict(wat), "config": report_config,
            "strength_ratios": model.strength_ratios(), "baseline_6t": baseline, "groups": {}}


def _target_comparisons(cell: SixTWatCell | ThreeTWatCell,
                        targets: DatasheetTargets | None) -> list[dict]:
    """Compare each measured object against its matching datasheet device target."""
    if targets is None:
        return []
    if isinstance(cell, ThreeTWatCell):
        measured = ((name.upper(), name, getattr(cell, name)) for name in ("pu", "pg", "pd"))
    else:
        measured = ((DISPLAY_MOS_NAMES[name], name[:2], getattr(cell, name))
                    for name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2"))
    rows = []
    for object_name, device, actual in measured:
        target = getattr(targets, device)
        rows.append({
            "object": object_name,
            "device": device.upper(),
            "target_vt_v": target.vt,
            "measured_vt_v": actual.vt,
            "delta_vt_mv": round((actual.vt - target.vt) * 1000, 6),
            "target_isat_ua": target.ids,
            "measured_isat_ua": actual.ids,
            "delta_isat_ua": round(actual.ids - target.ids, 6),
            "delta_isat_pct": round((actual.ids / target.ids - 1) * 100, 6),
        })
    return rows


_MOS_EXCEL_ALIASES = {
    "pul": "pu1", "pu1": "pu1", "m2": "pu1",
    "pur": "pu2", "pu2": "pu2", "m4": "pu2",
    "pgl": "pg1", "pg1": "pg1", "m5": "pg1",
    "pgr": "pg2", "pg2": "pg2", "m6": "pg2",
    "pdl": "pd1", "pd1": "pd1", "m1": "pd1",
    "pdr": "pd2", "pd2": "pd2", "m3": "pd2",
}
_VOLTAGE_UNITS = {"v": 1.0, "mv": 1e-3, "uv": 1e-6, "nv": 1e-9}
_CURRENT_TO_UA = {"a": 1e6, "ma": 1e3, "ua": 1.0, "na": 1e-3, "pa": 1e-6}
GUI_STATE_FILENAME = ".hv28_sram_analysis_state.json"


def _normalized_excel_key(value: object) -> str:
    raw = "" if value is None else str(value).strip().lower()
    raw = raw.replace("μ", "u").replace("µ", "u")
    raw = re.sub(r"\([^)]*\)|\[[^]]*\]", "", raw)
    return re.sub(r"[^a-z0-9]", "", raw)


def _unit_text(value: object, default: str) -> str:
    unit = default if value is None or str(value).strip() == "" else str(value)
    return _normalized_excel_key(unit)


def _excel_number(value: object, unit: object, unit_map: dict[str, float], default_unit: str,
                  label: str) -> float:
    if isinstance(value, str):
        match = re.fullmatch(r"\s*([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)\s*([A-Za-zμµ]*)\s*", value)
        if not match:
            raise ValueError(f"{label} must be numeric")
        magnitude, embedded_unit = float(match.group(1)), match.group(2)
        unit = embedded_unit or unit
    else:
        magnitude = float(value)
    if not math.isfinite(magnitude):
        raise ValueError(f"{label} must be finite")
    normalized_unit = _unit_text(unit, default_unit)
    if normalized_unit not in unit_map:
        raise ValueError(f"{label} unit '{unit}' is unsupported; use {', '.join(unit_map)}")
    return magnitude * unit_map[normalized_unit]


def _excel_header_map(headers: list[object]) -> tuple[dict[str, int], dict[str, str]]:
    keys, header_units = {}, {}
    for index, header in enumerate(headers):
        key = _normalized_excel_key(header)
        if key:
            keys[key] = index
            unit_match = re.search(r"[\[(]\s*([^\])]+)\s*[\])]", str(header)) if header else None
            if unit_match:
                header_units[key] = unit_match.group(1)
    return keys, header_units


def _first_header(header_map: dict[str, int], *names: str) -> str | None:
    return next((name for name in names if name in header_map), None)


def _cell(row: list[object], headers: dict[str, int], key: str | None) -> object | None:
    return None if key is None or key not in headers or headers[key] >= len(row) else row[headers[key]]


def _read_wat_excel_rows(headers_raw: list[object], data_rows: list[list[object]],
                         default_model_vdd_v: float) -> list[ExcelWatSweepPoint]:
    """Parse either a long or a wide six-MOS WAT worksheet into SI-normalized values.

    Long form: Lot/Wafer, Model VDD, MOS, Vt, Idsat, plus optional *_Unit columns.
    Wide form: Lot/Wafer, Model VDD, PUL_Vt, PUL_Idsat ... PDR_Vt, PDR_Idsat.
    Blank unit cells default to V for Vt/VDD and uA for Idsat.
    """
    headers, header_units = _excel_header_map(headers_raw)
    lot_key = _first_header(headers, "lotwafer", "lot", "wafer", "waferid", "corner")
    vdd_key = _first_header(headers, "modelvdd", "modelvoltage", "vdd", "supplyvoltage", "testvoltage")
    vdd_unit_key = _first_header(headers, "modelvddunit", "modelvoltageunit", "vddunit", "voltageunit")
    mos_key = _first_header(headers, "mos", "mosname", "device", "transistor")
    vt_key = _first_header(headers, "vt", "vth", "thresholdvoltage")
    ids_key = _first_header(headers, "idsat", "isat", "ids", "ion")
    vt_unit_key = _first_header(headers, "vtunit", "vthunit", "voltageunit")
    ids_unit_key = _first_header(headers, "idsatunit", "isatunit", "idsunit", "ionunit", "currentunit")
    if not lot_key:
        raise ValueError("Excel requires a Lot/Wafer (or Corner) column")

    def vdd_from(row: list[object], row_number: int) -> float:
        raw_vdd = _cell(row, headers, vdd_key)
        if raw_vdd is None or str(raw_vdd).strip() == "":
            return default_model_vdd_v
        unit = _cell(row, headers, vdd_unit_key) or header_units.get(vdd_key or "", "V")
        return _excel_number(raw_vdd, unit, _VOLTAGE_UNITS, "V", f"Excel row {row_number} Model VDD")

    measurements: dict[tuple[str, float], dict[str, list[MosWat]]] = {}
    total_rows: dict[tuple[str, float], dict[str, int]] = {}

    def register(lot: str, vdd: float, mos_name: str, mos: MosWat | None) -> None:
        key = (lot, vdd)
        total_rows.setdefault(key, {})[mos_name] = total_rows.setdefault(key, {}).get(mos_name, 0) + 1
        if mos is not None:
            measurements.setdefault(key, {}).setdefault(mos_name, []).append(mos)

    if mos_key and vt_key and ids_key:  # Long form
        for row_number, row in enumerate(data_rows, 2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            alias = _normalized_excel_key(_cell(row, headers, mos_key))
            if alias not in _MOS_EXCEL_ALIASES:
                raise ValueError(f"Excel row {row_number}: unknown MOS '{_cell(row, headers, mos_key)}'")
            vdd = vdd_from(row, row_number)
            if vdd <= 0:
                continue  # SNM has no physical definition at zero supply.
            lot = str(_cell(row, headers, lot_key) or f"row_{row_number}").strip()
            mos_name = _MOS_EXCEL_ALIASES[alias]
            raw_vt = _cell(row, headers, vt_key)
            raw_ids = _cell(row, headers, ids_key)
            vt_blank = raw_vt is None or str(raw_vt).strip() == ""
            ids_blank = raw_ids is None or str(raw_ids).strip() == ""
            if vt_blank and ids_blank:
                register(lot, vdd, mos_name, None)
                continue
            if vt_blank != ids_blank:
                raise ValueError(f"Excel row {row_number}: Vt and Idsat must both be filled or both be blank")
            vt_unit = _cell(row, headers, vt_unit_key) or header_units.get(vt_key, "V")
            ids_unit = _cell(row, headers, ids_unit_key) or header_units.get(ids_key, "uA")
            mos = MosWat(
                _positive(_excel_number(raw_vt, vt_unit, _VOLTAGE_UNITS, "V", f"Excel row {row_number} Vt"), "Vt"),
                _positive(_excel_number(raw_ids, ids_unit, _CURRENT_TO_UA, "uA", f"Excel row {row_number} Idsat"), "Idsat"),
            )
            register(lot, vdd, mos_name, mos)
    else:  # Wide form
        for row_number, row in enumerate(data_rows, 2):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            vdd = vdd_from(row, row_number)
            if vdd <= 0:
                continue
            lot = str(_cell(row, headers, lot_key) or f"row_{row_number}").strip()
            for mos_name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2"):
                aliases = [alias for alias, mapped in _MOS_EXCEL_ALIASES.items() if mapped == mos_name]
                vt_header = _first_header(headers, *(f"{alias}vt" for alias in aliases),
                                          *(f"{alias}vth" for alias in aliases))
                ids_header = _first_header(headers, *(f"{alias}idsat" for alias in aliases),
                                           *(f"{alias}isat" for alias in aliases),
                                           *(f"{alias}ids" for alias in aliases),
                                           *(f"{alias}ion" for alias in aliases))
                if not vt_header or not ids_header:
                    continue
                raw_vt = _cell(row, headers, vt_header)
                raw_ids = _cell(row, headers, ids_header)
                vt_blank = raw_vt is None or str(raw_vt).strip() == ""
                ids_blank = raw_ids is None or str(raw_ids).strip() == ""
                if vt_blank and ids_blank:
                    register(lot, vdd, mos_name, None)
                    continue
                if vt_blank != ids_blank:
                    raise ValueError(f"Excel row {row_number}: {mos_name} Vt and Idsat must both be filled or both be blank")
                vt_unit = _cell(row, headers, f"{vt_header}unit") or header_units.get(vt_header, "V")
                ids_unit = _cell(row, headers, f"{ids_header}unit") or header_units.get(ids_header, "uA")
                mos = MosWat(
                    _positive(_excel_number(raw_vt, vt_unit, _VOLTAGE_UNITS, "V", f"Excel row {row_number} {mos_name} Vt"), "Vt"),
                    _positive(_excel_number(raw_ids, ids_unit, _CURRENT_TO_UA, "uA", f"Excel row {row_number} {mos_name} Idsat"), "Idsat"),
                )
                register(lot, vdd, mos_name, mos)

    parsed: list[ExcelWatSweepPoint] = []
    required = {"pu1", "pu2", "pg1", "pg2", "pd1", "pd2"}
    for (lot, vdd), mos_measurements in sorted(measurements.items(), key=lambda item: (item[0][0], item[0][1])):
        missing = required - {name for name, values in mos_measurements.items() if values}
        if missing:
            raise ValueError(f"Excel {lot} at {vdd:.3f} V is missing: {', '.join(sorted(missing))}")
        devices: dict[str, MosWat] = {}
        statistics: dict[str, ExcelMosStatistics] = {}
        for mos_name in sorted(required):
            values = mos_measurements[mos_name]
            vt_values = [item.vt for item in values]
            ids_values = [item.ids for item in values]
            devices[mos_name] = MosWat(statlib.fmean(vt_values), statlib.fmean(ids_values))
            statistics[mos_name] = ExcelMosStatistics(
                valid_count=len(values),
                total_count=total_rows[(lot, vdd)].get(mos_name, len(values)),
                vt_mean=statlib.fmean(vt_values),
                vt_median=statlib.median(vt_values),
                vt_stdev=statlib.stdev(vt_values) if len(vt_values) > 1 else 0.0,
                vt_min=min(vt_values),
                vt_max=max(vt_values),
                ids_mean=statlib.fmean(ids_values),
                ids_median=statlib.median(ids_values),
                ids_stdev=statlib.stdev(ids_values) if len(ids_values) > 1 else 0.0,
                ids_min=min(ids_values),
                ids_max=max(ids_values),
            )
        parsed.append(ExcelWatSweepPoint(lot, vdd, SixTWatCell(lot, **devices), statistics))
    if not parsed:
        raise ValueError("Excel contains no analyzable model VDD above 0 V")
    return parsed


def read_wat_excel(path: str | os.PathLike[str], default_model_vdd_v: float = 0.90,
                   sheet_name: str | None = None) -> list[ExcelWatSweepPoint]:
    """Read a six-MOS WAT Excel workbook (.xlsx) and normalize all units."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel import requires openpyxl. Run: python -m pip install -r requirements.txt") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name:
            worksheets = [workbook[sheet_name]]
        else:
            grouped = {
                _normalized_excel_key(sheet.title): sheet
                for sheet in workbook.worksheets
                if _normalized_excel_key(sheet.title) in {"pu", "pg", "pd", "puwat", "pgwat", "pdwat"}
            }
            selected = []
            for family in ("pu", "pg", "pd"):
                sheet = grouped.get(family) or grouped.get(f"{family}wat")
                if sheet is not None:
                    selected.append(sheet)
            worksheets = selected if len(selected) == 3 else [workbook.active]

        headers_raw: list[object] | None = None
        data_rows: list[list[object]] = []
        normalized_headers: list[str] | None = None
        for worksheet in worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError(f"Excel worksheet '{worksheet.title}' is empty")
            current_headers = list(rows[0])
            current_normalized = [_normalized_excel_key(value) for value in current_headers]
            if normalized_headers is not None and current_normalized != normalized_headers:
                raise ValueError("PU/PG/PD worksheets must use the same column headers")
            headers_raw = headers_raw or current_headers
            normalized_headers = normalized_headers or current_normalized
            data_rows.extend([list(row) for row in rows[1:]])
        if headers_raw is None:
            raise ValueError("Excel contains no readable worksheet")
        return _read_wat_excel_rows(headers_raw, data_rows, default_model_vdd_v)
    finally:
        workbook.close()


def _interpolate_iv_current(samples: list[tuple[float, float]], vg_v: float,
                            label: str) -> float:
    """Linearly interpolate |Ids| at Vg; require the raw curve to span Vg."""
    ordered = sorted(samples)
    if len(ordered) < 2:
        raise ValueError(f"{label}: at least two Vg/Idsat raw points are required")
    if vg_v < ordered[0][0] - 1e-12 or vg_v > ordered[-1][0] + 1e-12:
        raise ValueError(
            f"{label}: Vg sweep {ordered[0][0]:.4g} to {ordered[-1][0]:.4g} V does not span Model VDD {vg_v:.4g} V")
    xs = [item[0] for item in ordered]
    index = bisect_left(xs, vg_v)
    if index < len(ordered) and abs(ordered[index][0] - vg_v) < 1e-12:
        return ordered[index][1]
    left_vg, left_i = ordered[index - 1]
    right_vg, right_i = ordered[index]
    if right_vg <= left_vg:
        raise ValueError(f"{label}: duplicate Vg values are not allowed")
    return left_i + (vg_v - left_vg) * (right_i - left_i) / (right_vg - left_vg)


def read_iv_curve_excel(path: str | os.PathLike[str],
                        fallback_vt_v: dict[str, float] | None = None) -> tuple[str, list[RsnmVccPoint], list[IvCurveExtraction]]:
    """Read PU/PG/PD raw Id-Vg Excel sheets and create RSNM-VDD inputs.

    Each sheet (PU, PG, PD) uses one row per raw point.  Idsat is extracted at
    ``Vg = Model VDD`` (the usual compact-model VGS=VDS=VDD convention), not
    from the final plotted point.  The raw Vg sweep must therefore bracket the
    stated Model VDD.  A per-row Vt column is optional when fallback Vt values
    are supplied from the 6T GUI.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("IV curve import requires openpyxl. Run: python -m pip install -r requirements.txt") from exc
    fallback_vt_v = {key.lower(): value for key, value in (fallback_vt_v or {}).items()}
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = {_normalized_excel_key(ws.title): ws for ws in workbook.worksheets}
        raw_by_family: dict[str, dict[float, list[tuple[float, float, float | None, str]]]] = {}
        lots: set[str] = set()
        for family in ("pu", "pg", "pd"):
            worksheet = sheets.get(family) or sheets.get(f"{family}iv") or sheets.get(f"{family}curve")
            if worksheet is None:
                raise ValueError("IV curve workbook requires PU, PG and PD worksheets")
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                raise ValueError(f"IV curve worksheet '{worksheet.title}' is empty")
            headers, header_units = _excel_header_map(list(rows[0]))
            lot_key = _first_header(headers, "lotwafer", "lot", "wafer", "waferid")
            vdd_key = _first_header(headers, "modelvdd", "vdd", "supplyvoltage", "testvoltage")
            vg_key = _first_header(headers, "vg", "vgs", "gatevoltage")
            ids_key = _first_header(headers, "idsat", "isat", "ids", "id", "draincurrent")
            vt_key = _first_header(headers, "vt", "vth", "thresholdvoltage")
            if not all((vdd_key, vg_key, ids_key)):
                raise ValueError(f"IV curve worksheet '{worksheet.title}' needs Model VDD, Vg and Idsat columns")
            vdd_unit_key = _first_header(headers, "modelvddunit", "vddunit")
            vg_unit_key = _first_header(headers, "vgunit", "vgsunit", "gatevoltageunit")
            ids_unit_key = _first_header(headers, "idsatunit", "isatunit", "idsunit", "idunit", "currentunit")
            vt_unit_key = _first_header(headers, "vtunit", "vthunit")
            family_curves: dict[float, list[tuple[float, float, float | None, str]]] = {}
            for row_number, row_tuple in enumerate(rows[1:], 2):
                row = list(row_tuple)
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                vdd = _excel_number(_cell(row, headers, vdd_key),
                                    _cell(row, headers, vdd_unit_key) or header_units.get(vdd_key, "V"),
                                    _VOLTAGE_UNITS, "V", f"{family.upper()} row {row_number} Model VDD")
                vg = _excel_number(_cell(row, headers, vg_key),
                                   _cell(row, headers, vg_unit_key) or header_units.get(vg_key, "V"),
                                   _VOLTAGE_UNITS, "V", f"{family.upper()} row {row_number} Vg")
                ids = abs(_excel_number(_cell(row, headers, ids_key),
                                        _cell(row, headers, ids_unit_key) or header_units.get(ids_key, "uA"),
                                        _CURRENT_TO_UA, "uA", f"{family.upper()} row {row_number} Idsat"))
                raw_vt = _cell(row, headers, vt_key)
                vt = None if raw_vt is None or str(raw_vt).strip() == "" else abs(_excel_number(
                    raw_vt, _cell(row, headers, vt_unit_key) or header_units.get(vt_key or "", "V"),
                    _VOLTAGE_UNITS, "V", f"{family.upper()} row {row_number} Vt"))
                lot = str(_cell(row, headers, lot_key) or "IV_Curve").strip()
                lots.add(lot)
                family_curves.setdefault(vdd, []).append((vg, ids, vt, lot))
            raw_by_family[family] = family_curves

        common_vdds = set.intersection(*(set(raw_by_family[family]) for family in ("pu", "pg", "pd")))
        if len(common_vdds) < 2:
            raise ValueError("IV curve data needs at least two common Model VDD values across PU, PG and PD")
        points: list[RsnmVccPoint] = []
        extraction: list[IvCurveExtraction] = []
        for vdd in sorted(common_vdds):
            mos: dict[str, MosWat] = {}
            for family in ("pu", "pg", "pd"):
                samples = raw_by_family[family][vdd]
                vt_values = [item[2] for item in samples if item[2] is not None]
                vt = statlib.fmean(vt_values) if vt_values else fallback_vt_v.get(family)
                if vt is None or vt <= 0:
                    raise ValueError(f"{family.upper()} at {vdd:.4g} V: enter Vt in Excel or provide a positive 6T fallback Vt")
                pairs = [(item[0], item[1]) for item in samples]
                ids = _interpolate_iv_current(pairs, vdd, f"{family.upper()} at {vdd:.4g} V")
                mos[family] = MosWat(vt, ids)
                extraction.append(IvCurveExtraction(family.upper(), vdd, vt, vdd, ids, len(pairs)))
            points.append(RsnmVccPoint(vdd, mos["pu"], mos["pg"], mos["pd"]))
        lot = next(iter(lots), "IV_Curve") if len(lots) == 1 else "IV_Curve_Mixed_Lots"
        return lot, points, extraction
    finally:
        workbook.close()


def write_iv_curve_excel_template(path: str | os.PathLike[str]) -> Path:
    """Create a raw Id-Vg workbook accepted by :func:`read_iv_curve_excel`."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("IV curve template export requires openpyxl. Run: python -m pip install -r requirements.txt") from exc
    destination = Path(path).with_suffix(".xlsx")
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    headers = ["Lot/Wafer", "Model VDD", "VDD Unit", "Vg", "Vg Unit", "Idsat", "Idsat Unit", "Vt", "Vt Unit", "Notes"]
    colors = {"PU": "FFF1F0", "PG": "ECFDF3", "PD": "EEF6FF"}
    for family in ("PU", "PG", "PD"):
        sheet = workbook.create_sheet(family)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="0B6EF3")
            cell.font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for vdd in (.40, .50, .60, .70, .80, .90, 1.00, 1.10, 1.20):
            for index in range(13):
                vg = index * .10
                sheet.append(["DEMO28_TT_W01", vdd, "V", vg, "V", None, "uA", None, "V",
                              "Paste raw |Ids|; Vg sweep must include Model VDD"])
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=colors[family])
        for column, width in enumerate((20, 13, 10, 10, 10, 14, 12, 10, 10, 48), 1):
            sheet.column_dimensions[chr(64 + column)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:J{sheet.max_row}"
    instructions = workbook.create_sheet("Instructions")
    instructions.append(["HV28 IV Curve Import Template"])
    instructions.append(["1. Fill one raw Id-Vg point per row on PU, PG and PD sheets."])
    instructions.append(["2. Model VDD identifies the curve; Vg is the swept gate voltage; Idsat is the magnitude of drain current."])
    instructions.append(["3. The import extracts Idsat by linear interpolation at Vg = Model VDD. Do not use the last plotted point unless it is Vg = VDD."])
    instructions.append(["4. Enter Vt in the sheet, or leave it blank to use the current GUI PU/PG/PD average Vt."])
    instructions.column_dimensions["A"].width = 125
    workbook.save(destination)
    return destination


def _model_vdd_from_sheet_name(sheet_name: object) -> float | None:
    """Return a Model VDD encoded by names such as ``0.90V`` or ``VDD_0.90V``."""
    match = re.fullmatch(
        r"\s*(?:(?:model\s*)?vdd[\s_-]*)?([0-9]+(?:\.[0-9]+)?)\s*v?\s*",
        str(sheet_name), flags=re.IGNORECASE)
    if not match:
        return None
    vdd = float(match.group(1))
    if not 0 < vdd <= SNM_PLOT_AXIS_MAX_V:
        raise ValueError(
            f'Worksheet "{sheet_name}" Model VDD must be within 0–{SNM_PLOT_AXIS_MAX_V:.2f} V')
    return vdd


def _read_multi_chip_6t_sheet(sheet, default_model_vdd_v: float,
                              sheet_model_vdd_v: float | None = None) -> list[WaferChipWat]:
    """Parse one wide-form 6T Multi-Cell worksheet."""
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError(
            f'Multi-chip Excel worksheet "{sheet.title}" needs a header and at least one chip row')
    headers, header_units = _excel_header_map(list(rows[0]))
    lot_key = _first_header(headers, "lotwafer", "lot", "wafer", "waferid")
    chip_key = _first_header(headers, "chipid", "chip", "dieid", "site", "siteid")
    vdd_key = _first_header(headers, "modelvdd", "vdd", "supplyvoltage")
    vdd_unit_key = _first_header(headers, "modelvddunit", "vddunit")
    if not chip_key:
        raise ValueError(
            f'Multi-chip Excel worksheet "{sheet.title}" requires a Chip ID column')
    devices: dict[str, tuple[str, str]] = {}
    for device in ("pul", "pur", "pgl", "pgr", "pdl", "pdr"):
        vt_key = _first_header(headers, f"{device}vt", f"{device}vth")
        ids_key = _first_header(headers, f"{device}idsat", f"{device}isat", f"{device}ids")
        if not vt_key or not ids_key:
            raise ValueError(
                f'Worksheet "{sheet.title}" needs {device.upper()} Vt and Idsat columns')
        devices[device] = (vt_key, ids_key)
    parsed: list[WaferChipWat] = []
    for number, row_tuple in enumerate(rows[1:], 2):
        row = list(row_tuple)
        if not any(value is not None and str(value).strip() for value in row):
            continue
        lot = str(_cell(row, headers, lot_key) or "Wafer").strip()
        chip = str(_cell(row, headers, chip_key) or "").strip()
        if not chip:
            raise ValueError(
                f'Worksheet "{sheet.title}" row {number}: Chip ID is blank')
        if sheet_model_vdd_v is not None:
            # In the multi-VDD workflow the worksheet name is authoritative,
            # even when an older copied sheet still contains a Model VDD column.
            vdd = sheet_model_vdd_v
        else:
            raw_vdd = _cell(row, headers, vdd_key)
            vdd = default_model_vdd_v if raw_vdd is None or str(raw_vdd).strip() == "" else _excel_number(
                raw_vdd, _cell(row, headers, vdd_unit_key) or header_units.get(vdd_key or "", "V"),
                _VOLTAGE_UNITS, "V", f"Multi-chip row {number} Model VDD")
        mos: dict[str, MosWat] = {}
        raw_idsat_ua: dict[str, float] = {}
        for device, (vt_key, ids_key) in devices.items():
            vt = _excel_number(_cell(row, headers, vt_key),
                               _cell(row, headers, f"{vt_key}unit") or header_units.get(vt_key, "V"),
                               _VOLTAGE_UNITS, "V", f"Multi-chip row {number} {device.upper()} Vt")
            ids = _excel_number(_cell(row, headers, ids_key),
                                _cell(row, headers, f"{ids_key}unit") or header_units.get(ids_key, "uA"),
                                _CURRENT_TO_UA, "uA", f"Multi-chip row {number} {device.upper()} Idsat")
            raw_idsat_ua[device] = ids
            mos[device] = MosWat(abs(_positive(vt, f"{device.upper()} Vt")),
                                 abs(_positive(ids, f"{device.upper()} Idsat")))
        parsed.append(WaferChipWat(lot, chip, vdd, SixTWatCell(f"{lot}_{chip}",
            mos["pul"], mos["pur"], mos["pgl"], mos["pgr"], mos["pdl"], mos["pdr"]), raw_idsat_ua))
    if not parsed:
        raise ValueError(f'Worksheet "{sheet.title}" contains no chip rows')
    return parsed


def read_multi_chip_6t_excel_vdd_sheets(
        path: str | os.PathLike[str], default_model_vdd_v: float = .90,
        allow_no_vdd_sheets: bool = False) -> list[dict[str, object]]:
    """Read every worksheet whose name encodes its Model VDD."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Multi-chip Excel import requires openpyxl. Run: python -m pip install -r requirements.txt") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        grouped: dict[float, dict[str, object]] = {}
        for sheet in workbook.worksheets:
            vdd = _model_vdd_from_sheet_name(sheet.title)
            if vdd is None:
                continue
            entry = grouped.setdefault(vdd, {"vdd_v": vdd, "sheet_names": [], "chips": []})
            entry["sheet_names"].append(sheet.title)
            entry["chips"].extend(_read_multi_chip_6t_sheet(
                sheet, default_model_vdd_v, sheet_model_vdd_v=vdd))
        if not grouped and not allow_no_vdd_sheets:
            raise ValueError(
                "No Model VDD worksheet was found. Name data sheets like 0.90V, 0.80V, or VDD_0.70V.")
        return [grouped[vdd] for vdd in sorted(grouped)]
    finally:
        workbook.close()


def read_multi_chip_6t_excel(path: str | os.PathLike[str],
                             default_model_vdd_v: float = .90,
                             require_common_vdd: bool = True) -> list[WaferChipWat]:
    """Read one conventional wide-form wafer Multi-Cell worksheet."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Multi-chip Excel import requires openpyxl. Run: python -m pip install -r requirements.txt") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if "6T Multi-Cell" in workbook.sheetnames:
            sheet = workbook["6T Multi-Cell"]
        elif "6T Multi-Chip" in workbook.sheetnames:
            sheet = workbook["6T Multi-Chip"]
        else:
            sheet = workbook.active
        parsed = _read_multi_chip_6t_sheet(sheet, default_model_vdd_v)
        vdds = {round(item.model_vdd_v, 12) for item in parsed}
        if require_common_vdd and len(vdds) != 1:
            raise ValueError("Multi-chip VTC overlay requires one common Model VDD for all chip rows")
        return parsed
    finally:
        workbook.close()


def write_multi_chip_6t_excel_template(path: str | os.PathLike[str], chip_count: int = 64) -> Path:
    """Create a simple unit-free 6T wafer multi-cell template accepted by batch import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Multi-chip template export requires openpyxl. Run: python -m pip install -r requirements.txt") from exc
    destination = Path(path).with_suffix(".xlsx")
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(); sheet = workbook.active; sheet.title = "6T Multi-Cell"
    headers = ["Lot/Wafer", "Chip ID"]
    for name in ("PUL", "PUR", "PGL", "PGR", "PDL", "PDR"):
        headers.extend([f"{name} Vt", f"{name} Idsat"])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0B6EF3")
        cell.font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    defaults = {"PU": (.385, 44.0), "PG": (.365, 82.0), "PD": (.355, 124.0)}
    for index in range(1, max(1, int(chip_count)) + 1):
        row = ["DEMO28_TT_W01", f"CHIP_{index:02d}"]
        for name in ("PUL", "PUR", "PGL", "PGR", "PDL", "PDR"):
            vt, ids = defaults[name[:2]]
            row.extend([vt, ids])
        sheet.append(row)
    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 14
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    info = workbook.create_sheet("Instructions")
    info.append(["One row = one measured 6T cell/chip. Fill Vt in V and Idsat in uA; no unit columns or per-row VDD are required. The import uses the Model VDD currently selected in the GUI. The batch report overlays all cell VTC/mirror VTC curves and reports the minimum RSNM and WSNM as wafer references."])
    info.column_dimensions["A"].width = 140
    workbook.save(destination)
    return destination


def write_single_6t_wat_excel(path: str | os.PathLike[str], cell: SixTWatCell,
                              model_vdd_v: float) -> Path:
    """Write one editable six-MOS WAT set in the same long form accepted by import.

    The workbook intentionally uses a plain filtered range instead of an Excel
    Table object so it remains compatible with older/company-managed Excel
    installations that may repair table metadata on open.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError(
            "Excel export requires openpyxl. Run: python -m pip install -r requirements.txt"
        ) from exc

    model_vdd = _positive(float(model_vdd_v), "Model VDD")
    destination = Path(path)
    if destination.suffix.lower() != ".xlsx":
        destination = destination.with_suffix(".xlsx")
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "6T WAT Input"
    headers = [
        "Lot/Wafer", "Model VDD", "VDD Unit", "MOS", "Vt", "Vt Unit",
        "Idsat", "Idsat Unit", "Notes",
    ]
    sheet.append(headers)
    devices = (
        ("PUL", cell.pu1), ("PUR", cell.pu2),
        ("PGL", cell.pg1), ("PGR", cell.pg2),
        ("PDL", cell.pd1), ("PDR", cell.pd2),
    )
    for name, mos in devices:
        sheet.append([
            cell.corner, model_vdd, "V", name, mos.vt, "V", mos.ids, "uA",
            "Editable 6T WAT input",
        ])

    header_fill = PatternFill("solid", fgColor="0B6EF3")
    family_fills = {
        "PU": PatternFill("solid", fgColor="FFF1F0"),
        "PG": PatternFill("solid", fgColor="ECFDF3"),
        "PD": PatternFill("solid", fgColor="EEF6FF"),
    }
    thin_gray = Side(style="thin", color="D7DCE5")
    for cell_obj in sheet[1]:
        cell_obj.fill = header_fill
        cell_obj.font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
        cell_obj.alignment = Alignment(horizontal="center", vertical="center")
        cell_obj.border = Border(bottom=thin_gray)
    sheet.row_dimensions[1].height = 26
    for row in range(2, 8):
        family = str(sheet.cell(row=row, column=4).value)[:2]
        for column in range(1, 10):
            item = sheet.cell(row=row, column=column)
            item.font = Font(name="Calibri", size=11)
            item.fill = family_fills[family]
            item.border = Border(bottom=thin_gray)
            item.alignment = Alignment(
                horizontal="right" if column in (2, 5, 7) else "left",
                vertical="center",
            )
        sheet.cell(row=row, column=2).number_format = "0.000"
        sheet.cell(row=row, column=5).number_format = "0.0000"
        sheet.cell(row=row, column=7).number_format = "0.000"
        sheet.row_dimensions[row].height = 23

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:I7"
    sheet.sheet_view.showGridLines = False
    widths = {"A": 21, "B": 13, "C": 11, "D": 10, "E": 11,
              "F": 11, "G": 13, "H": 13, "I": 28}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    vdd_units = DataValidation(type="list", formula1='"V,mV"', allow_blank=False)
    vt_units = DataValidation(type="list", formula1='"V,mV"', allow_blank=False)
    ids_units = DataValidation(type="list", formula1='"A,mA,uA,nA"', allow_blank=False)
    sheet.add_data_validation(vdd_units); vdd_units.add("C2:C7")
    sheet.add_data_validation(vt_units); vt_units.add("F2:F7")
    sheet.add_data_validation(ids_units); ids_units.add("H2:H7")
    sheet.conditional_formatting.add(
        "E2:E7", FormulaRule(formula=["E2<=0"], fill=PatternFill("solid", fgColor="FFD6D6")))
    sheet.conditional_formatting.add(
        "G2:G7", FormulaRule(formula=["G2<=0"], fill=PatternFill("solid", fgColor="FFD6D6")))

    instructions = workbook.create_sheet("Instructions")
    instructions.sheet_view.showGridLines = False
    instruction_rows = [
        ("HV28 SRAM — Single 6T WAT Input", ""),
        ("Purpose", "Enter one Lot/Wafer and one operating VDD for all six physical MOS devices."),
        ("Required MOS", "PUL, PUR, PGL, PGR, PDL, PDR"),
        ("Units", "Vt supports V/mV; Idsat supports A/mA/uA/nA. The program converts to V and uA."),
        ("Import", "In the 6T Bitcell Analysis tab, choose Import Excel...; the first valid set fills the GUI."),
        ("Manual entry", "You may ignore Excel and type the six Vt/Idsat pairs directly beside the MOS devices."),
        ("Export", "Save Current... writes the current GUI values back into this compatible format."),
    ]
    for row in instruction_rows:
        instructions.append(row)
    instructions["A1"].font = Font(name="Microsoft JhengHei", size=16, bold=True, color="0B6EF3")
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 32
    for row in range(2, len(instruction_rows) + 1):
        instructions.cell(row=row, column=1).font = Font(name="Microsoft JhengHei", size=11, bold=True)
        instructions.cell(row=row, column=2).font = Font(name="Calibri", size=11)
        instructions.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        instructions.row_dimensions[row].height = 30
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 92

    workbook.save(destination)
    workbook.close()
    return destination


def load_gui_state(state_path: str | os.PathLike[str] | None = None) -> dict[str, object]:
    """Load portable GUI inputs; malformed state is safely ignored."""
    path = Path(state_path) if state_path else Path.cwd() / GUI_STATE_FILENAME
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def save_gui_state(state: dict[str, object], state_path: str | os.PathLike[str] | None = None) -> None:
    """Persist only user-entered GUI values next to the application workspace."""
    path = Path(state_path) if state_path else Path.cwd() / GUI_STATE_FILENAME
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


VALIDATION_COLUMNS = (
    "lot_wafer", "pu_vt", "pu_idsat", "pg_vt", "pg_idsat", "pd_vt", "pd_idsat",
    "scan4n_vmin", "select_write_vmin", "select_read_vmin",
)


def read_validation_csv(path: str | os.PathLike[str]) -> list[dict]:
    """Read lot/wafer history. Blank WAT fields are allowed because WT is often full-test."""
    aliases = {"corner": "lot_wafer", "pu_ids": "pu_idsat", "pg_ids": "pg_idsat",
               "pd_ids": "pd_idsat", "scan4n": "scan4n_vmin",
               "select_write": "select_write_vmin", "select_read": "select_read_vmin"}
    records: list[dict] = []
    with open(path, newline="", encoding="utf-8-sig") as source:
        reader = csv.DictReader(source)
        headers = {aliases.get(name.strip().lower(), name.strip().lower())
                   for name in (reader.fieldnames or [])}
        required = {"lot_wafer", "scan4n_vmin", "select_write_vmin", "select_read_vmin"}
        if missing := required - headers:
            raise ValueError("Validation CSV missing columns: " + ", ".join(sorted(missing)))
        for line, raw in enumerate(reader, 2):
            canonical = {aliases.get(str(key).strip().lower(), str(key).strip().lower()): value
                         for key, value in raw.items() if key is not None}
            record: dict[str, str | float | None] = {
                "lot_wafer": str(canonical.get("lot_wafer", "")).strip() or f"row_{line}"
            }
            for name in VALIDATION_COLUMNS[1:]:
                value = str(canonical.get(name, "") or "").strip()
                if not value:
                    record[name] = None
                    continue
                try:
                    number = float(value)
                except ValueError as exc:
                    raise ValueError(f"Validation CSV line {line}: {name} is not numeric") from exc
                if not math.isfinite(number) or number <= 0:
                    raise ValueError(f"Validation CSV line {line}: {name} must be positive")
                record[name] = number
            records.append(record)
    return records


def _pearson(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) < 3 or len(xs) != len(ys):
        return None
    mx, my = sum(xs) / len(xs), sum(ys) / len(ys)
    dx, dy = [x - mx for x in xs], [y - my for y in ys]
    den = math.sqrt(sum(x*x for x in dx) * sum(y*y for y in dy))
    return None if den == 0 else sum(x*y for x, y in zip(dx, dy)) / den


def _ranks(values: list[float]) -> list[float]:
    order = sorted(range(len(values)), key=values.__getitem__)
    ranks = [0.0] * len(values)
    index = 0
    while index < len(order):
        end = index + 1
        while end < len(order) and values[order[end]] == values[order[index]]:
            end += 1
        average_rank = (index + 1 + end) / 2
        for position in order[index:end]:
            ranks[position] = average_rank
        index = end
    return ranks


def _validation_record(cell: SixTWatCell | ThreeTWatCell, measured: ManualVmin) -> dict:
    wat = cell.representative()
    return {"lot_wafer": cell.corner, "pu_vt": wat.pu_vt, "pu_idsat": wat.pu_ids,
            "pg_vt": wat.pg_vt, "pg_idsat": wat.pg_ids, "pd_vt": wat.pd_vt,
            "pd_idsat": wat.pd_ids, "scan4n_vmin": measured.scan4n,
            "select_write_vmin": measured.select_write, "select_read_vmin": measured.select_read}


def validate_datasheet_wat_target(cell: SixTWatCell | ThreeTWatCell,
                                  measured: ManualVmin,
                                  targets: DatasheetTargets,
                                  settings: TargetValidationSettings,
                                  historical_rows: list[dict] | None = None) -> dict:
    """Validate whether closeness to WAT targets is associated with acceptable measured WT."""
    current = _validation_record(cell, measured)
    history = [dict(row) for row in (historical_rows or [])
               if str(row.get("lot_wafer", "")) != cell.corner]
    records = history + [current]
    target_values = {
        "pu_vt": targets.pu.vt, "pu_idsat": targets.pu.ids,
        "pg_vt": targets.pg.vt, "pg_idsat": targets.pg.ids,
        "pd_vt": targets.pd.vt, "pd_idsat": targets.pd.ids,
    }


def _attach_target_model(result: dict, targets: DatasheetTargets | None, cfg: Config) -> None:
    """Attach a same-condition 6T model built from datasheet PU/PG/PD targets."""
    if targets is None:
        result.pop("target_6t", None)
        result["snm_target_comparison"] = []
        analytical = result["baseline_6t"]["analytical_read_snm_eq_3_36"]
        result["analytical_read_snm_comparison"] = {
            "method": "High-Speed CMOS Circuit Technology, Section 3.4.2, Equation 3.36",
            "current": analytical, "target": None,
            "current_snm_mv": analytical["snm_mv"], "target_snm_mv": None,
            "delta_mv": None, "delta_pct": None,
        }
        return
    target_wat = WatPoint(
        corner="WAT Target",
        pu_vt=targets.pu.vt, pu_ids=targets.pu.ids,
        pg_vt=targets.pg.vt, pg_ids=targets.pg.ids,
        pd_vt=targets.pd.vt, pd_ids=targets.pd.ids,
    )
    target = analyze(target_wat, cfg)["baseline_6t"]
    result["target_6t"] = target
    current_metrics = result["baseline_6t"]["metrics"]
    result["snm_target_comparison"] = [
        {"mode": label, "current_snm_mv": current_metrics[key],
         "target_snm_mv": target["metrics"][key],
         "delta_mv": current_metrics[key] - target["metrics"][key],
         "delta_pct": ((current_metrics[key] / target["metrics"][key] - 1) * 100
                       if target["metrics"][key] else None)}
        for label, key in (("Read SNM", "read_snm_mv"),)
    ]
    current_analytical = result["baseline_6t"]["analytical_read_snm_eq_3_36"]
    target_analytical = target["analytical_read_snm_eq_3_36"]
    current_value, target_value = current_analytical["snm_mv"], target_analytical["snm_mv"]
    result["analytical_read_snm_comparison"] = {
        "method": "High-Speed CMOS Circuit Technology, Section 3.4.2, Equation 3.36",
        "current": current_analytical,
        "target": target_analytical,
        "current_snm_mv": current_value,
        "target_snm_mv": target_value,
        "delta_mv": (current_value - target_value
                     if current_value is not None and target_value is not None else None),
        "delta_pct": ((current_value / target_value - 1.0) * 100.0
                      if current_value is not None and target_value not in (None, 0) else None),
    }
    return
    limits = {"scan4n_vmin": settings.scan4n_vmin_max,
              "select_write_vmin": settings.select_write_vmin_max,
              "select_read_vmin": settings.select_read_vmin_max}

    analyzed: list[dict] = []
    for record in records:
        wt_complete = all(record.get(name) is not None for name in limits)
        wat_complete = all(record.get(name) is not None for name in target_values)
        wt_pass = (all(float(record[name]) <= limit for name, limit in limits.items())
                   if wt_complete else None)
        worst_normalized = (max(float(record[name]) / limit for name, limit in limits.items())
                            if wt_complete else None)
        deviations: list[float] = []
        within_flags: list[bool] = []
        if wat_complete:
            for name, target in target_values.items():
                value = float(record[name])
                if name.endswith("_vt"):
                    normalized = abs(value - target) * 1000 / settings.vt_tolerance_mv
                else:
                    normalized = abs(value / target - 1) * 100 / settings.idsat_tolerance_pct
                deviations.append(normalized)
                within_flags.append(normalized <= 1.0)
        distance = math.sqrt(sum(x*x for x in deviations) / len(deviations)) if deviations else None
        target_band_pass = all(within_flags) if within_flags else None
        if target_band_pass is None or wt_pass is None:
            consistency = "N/A"
        elif target_band_pass and wt_pass:
            consistency = "TRUE ACCEPT"
        elif target_band_pass and not wt_pass:
            consistency = "FALSE ACCEPT"
        elif not target_band_pass and wt_pass:
            consistency = "FALSE REJECT"
        else:
            consistency = "TRUE REJECT"
        analyzed.append({**record, "wat_complete": wat_complete, "wt_complete": wt_complete,
                         "target_distance": distance, "target_band_pass": target_band_pass,
                         "wt_pass": wt_pass, "worst_normalized_vmin": worst_normalized,
                         "consistency": consistency})

    paired = [row for row in analyzed if row["wat_complete"] and row["wt_complete"]]
    inside = [row for row in paired if row["target_band_pass"]]
    outside = [row for row in paired if not row["target_band_pass"]]
    pass_rate = lambda group: (100 * sum(bool(row["wt_pass"]) for row in group) / len(group)
                               if group else None)
    inside_rate, outside_rate = pass_rate(inside), pass_rate(outside)
    gap = None if inside_rate is None or outside_rate is None else inside_rate - outside_rate
    xs = [float(row["target_distance"]) for row in paired]
    ys = [float(row["worst_normalized_vmin"]) for row in paired]
    pearson = _pearson(xs, ys)
    spearman = _pearson(_ranks(xs), _ranks(ys)) if len(xs) >= 3 else None

    enough = (len(paired) >= settings.minimum_statistical_n and len(inside) >= 3 and len(outside) >= 3)
    if not enough:
        verdict = "INSUFFICIENT DATA"
        explanation = (f"Need at least {settings.minimum_statistical_n} paired rows and at least 3 rows "
                       "inside and outside the WAT target band.")
    elif gap is not None and gap >= 10 and spearman is not None and spearman >= 0.20:
        verdict = "SUPPORTED"
        explanation = "Target-band lots show a materially higher WT pass rate and target deviation tracks worse Vmin."
    elif gap is not None and gap <= -10:
        verdict = "CONTRADICTED"
        explanation = "Target-band lots have a lower WT pass rate; review target center, tolerance, and sampling alignment."
    else:
        verdict = "INCONCLUSIVE"
        explanation = "Available data does not show enough separation to prove this WAT target is an effective WT screen."

    parameter_evidence = []
    parameter_labels = {"pu_vt": "PU Vt", "pu_idsat": "PU Isat", "pg_vt": "PG Vt",
                        "pg_idsat": "PG Isat", "pd_vt": "PD Vt", "pd_idsat": "PD Isat"}
    for name, target in target_values.items():
        feature_rows = [row for row in analyzed if row.get(name) is not None and row["wt_complete"]]
        tolerance = (settings.vt_tolerance_mv / 1000 if name.endswith("_vt")
                     else target * settings.idsat_tolerance_pct / 100)
        feature_distance = [abs(float(row[name]) - target) / tolerance for row in feature_rows]
        feature_worst = [float(row["worst_normalized_vmin"]) for row in feature_rows]
        feature_inside = [row for row, distance in zip(feature_rows, feature_distance) if distance <= 1]
        feature_outside = [row for row, distance in zip(feature_rows, feature_distance) if distance > 1]
        feature_inside_rate, feature_outside_rate = pass_rate(feature_inside), pass_rate(feature_outside)
        feature_lift = (None if feature_inside_rate is None or feature_outside_rate is None
                        else feature_inside_rate - feature_outside_rate)
        feature_spearman = (_pearson(_ranks(feature_distance), _ranks(feature_worst))
                            if len(feature_rows) >= 3 else None)
        feature_enough = (len(feature_rows) >= settings.minimum_statistical_n and
                          len(feature_inside) >= 3 and len(feature_outside) >= 3)
        if not feature_enough:
            feature_verdict = "INSUFFICIENT DATA"
        elif feature_lift is not None and feature_lift >= 10 and feature_spearman is not None and feature_spearman >= .20:
            feature_verdict = "SUPPORTED"
        elif feature_lift is not None and feature_lift <= -10:
            feature_verdict = "CONTRADICTED"
        else:
            feature_verdict = "INCONCLUSIVE"
        parameter_evidence.append({"parameter": parameter_labels[name], "field": name,
                                   "target": target, "paired_n": len(feature_rows),
                                   "inside_n": len(feature_inside), "outside_n": len(feature_outside),
                                   "inside_wt_pass_rate_pct": feature_inside_rate,
                                   "outside_wt_pass_rate_pct": feature_outside_rate,
                                   "pass_rate_lift_pct_points": feature_lift,
                                   "deviation_vs_worst_vmin_spearman": feature_spearman,
                                   "verdict": feature_verdict})

    return {
        "objective": "Validate WAT targets using measured WT Vmin outcomes",
        "verdict": verdict, "explanation": explanation, "settings": asdict(settings),
        "counts": {"all_rows": len(analyzed), "wt_complete": sum(row["wt_complete"] for row in analyzed),
                   "paired_wat_wt": len(paired), "inside_target_band": len(inside),
                   "outside_target_band": len(outside)},
        "statistics": {"inside_wt_pass_rate_pct": inside_rate,
                       "outside_wt_pass_rate_pct": outside_rate,
                       "pass_rate_lift_pct_points": gap,
                       "target_distance_vs_worst_vmin_pearson": pearson,
                       "target_distance_vs_worst_vmin_spearman": spearman},
        "current_row": analyzed[-1], "rows": analyzed, "parameter_evidence": parameter_evidence,
        "interpretation": "Positive correlation means farther from the WAT target tends to accompany higher (worse) WT Vmin.",
    }


def analyze_six_mos(cell: SixTWatCell, cfg: Config,
                    datasheet_targets: DatasheetTargets | None = None) -> dict:
    """Full cell-level report for an OO six-device cell."""
    result = analyze(cell.representative(), cfg)
    result["object_mode"] = "6T Independent"
    result["cell"] = {
        "corner": cell.corner,
        "mos": {DISPLAY_MOS_NAMES[name]: asdict(getattr(cell, name))
                for name in ("pu1","pu2","pg1","pg2","pd1","pd2")},
        "method": "asymmetric cross-coupled 6T Read butterfly; cell RSNM uses the smaller state margin",
    }
    half_cell = cell_metric(cell, cfg)
    asymmetric_model = AsymmetricSram6T(cell, cfg)
    fit_points = _snm_fit_points(cfg)
    asymmetric = asymmetric_model.read_butterfly(cfg.nominal_vdd, fit_points)
    vdd = cfg.nominal_vdd
    baseline = result["baseline_6t"]
    baseline.update(asymmetric)
    baseline["write_wsnm"] = write_wsnm_states(
        vdd, asymmetric_model.left, asymmetric_model.right, cfg, fit_points)
    butterfly = asymmetric["read_butterfly"]
    baseline["metrics"].update({
        "read_snm_mv": butterfly["snm_mv"],
        "read_snm_upper_left_mv": butterfly["snm_upper_left_mv"],
        "read_snm_lower_right_mv": butterfly["snm_lower_right_mv"],
        "read_snm_delta_mv": butterfly["delta_snm_mv"],
        "read_snm_mismatch_index_pct": butterfly["mismatch_index_pct"],
        "write_snm_mv": baseline["write_wsnm"]["snm_mv"],
        "cell_ratio_beta": half_cell["cell_ratio_beta"],
        "pull_up_ratio_beta": half_cell["pull_up_ratio_beta"],
        "cell_ratio_ids_proxy": half_cell["cell_ratio_ids_proxy"],
        "pull_up_ratio_ids_proxy": half_cell["pull_up_ratio_ids_proxy"],
    })
    result["cell"]["baseline_metrics"] = baseline["metrics"]
    result["cell"]["read_snm_state_definition"] = butterfly["coordinate_definition"]
    result["datasheet_targets"] = asdict(datasheet_targets) if datasheet_targets else None
    result["target_comparisons"] = _target_comparisons(cell, datasheet_targets)
    _attach_target_model(result, datasheet_targets, cfg)
    return result


def _multi_cell_metrics(cell: SixTWatCell, cfg: Config, vdd: float,
                        fit_points: int) -> dict[str, object]:
    """Evaluate one independent 6T cell for the wafer multi-cell flow."""
    model = AsymmetricSram6T(cell, cfg)
    read = model.read_butterfly(vdd, fit_points)
    write = write_wsnm_states(vdd, model.left, model.right, cfg, fit_points)
    side_ratios = (model.left.strength_ratios(), model.right.strength_ratios())
    # This is the compact write-trip/bitline-tolerance proxy, deliberately
    # reported separately from geometrical WSNM.
    write_margin_mv = 1000.0 * min(model.left.write_snm(vdd), model.right.write_snm(vdd))
    return {
        "cell": cell,
        "read": read,
        "rsnm_mv": read["read_butterfly"].get("snm_mv"),
        "upper_rsnm_mv": read["read_butterfly"].get("snm_upper_left_mv"),
        "lower_rsnm_mv": read["read_butterfly"].get("snm_lower_right_mv"),
        "write": write,
        "wsnm_mv": write.get("snm_mv"),
        "write_margin_mv": write_margin_mv,
        "cell_ratio_beta": min(item["cell_ratio_beta"] for item in side_ratios),
        "pull_up_ratio_beta": min(item["pull_up_ratio_beta"] for item in side_ratios),
    }


def _dominant_snm_degradation_parameter(cell: SixTWatCell,
                                        median_cell: SixTWatCell,
                                        mode: str) -> dict[str, object] | None:
    """Find the strongest one-parameter adverse shift relative to wafer median.

    This is an explanatory marker for the multi-cell overlay, not a causal
    decomposition.  Read SNM is mainly weakened by weaker PD or stronger PG;
    Write SNM is mainly weakened by stronger PU or weaker PG.
    """
    if mode not in {"read", "write"}:
        raise ValueError("mode must be read or write")
    candidates: list[tuple[float, str, str, str]] = []
    for attr, label in (("pu1", "PUL"), ("pu2", "PUR"),
                        ("pg1", "PGL"), ("pg2", "PGR"),
                        ("pd1", "PDL"), ("pd2", "PDR")):
        measured, median = getattr(cell, attr), getattr(median_cell, attr)
        vt_scale = max(abs(median.vt), 1e-12)
        ids_scale = max(abs(median.ids), 1e-12)
        family = attr[:2]
        # sign = +1 means an upward shift is adverse; -1 means downward is.
        if mode == "read":
            sign = 1 if family == "pd" else (-1 if family == "pg" else 0)
        else:
            sign = -1 if family == "pu" else (1 if family == "pg" else 0)
        if sign == 0:
            continue
        vt_score = max(0.0, sign * (measured.vt - median.vt) / vt_scale)
        # Idsat has the opposite relation to Vt for effective drive strength.
        ids_score = max(0.0, -sign * (abs(measured.ids) - abs(median.ids)) / ids_scale)
        candidates.extend(((vt_score, label, "vt", "Vt"),
                           (ids_score, label, "idsat", "Idsat")))
    if not candidates:
        return None
    score, device, parameter, parameter_label = max(candidates, key=lambda item: item[0])
    if score <= 1e-12:
        return None
    return {"device": device, "parameter": parameter,
            "parameter_label": parameter_label, "score": score}


def _median_multi_cell(chips: list[WaferChipWat]) -> SixTWatCell:
    """Build a synthetic median 6T cell from every physical MOS parameter."""
    def median_mos(name: str) -> MosWat:
        return MosWat(
            statlib.median(getattr(item.cell, name).vt for item in chips),
            statlib.median(getattr(item.cell, name).ids for item in chips),
        )
    return SixTWatCell(
        "MEDIAN_CELL", *(median_mos(name) for name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2")))


def _move_family_toward_median(current: SixTWatCell, median: SixTWatCell,
                               family: str, attribute: str, fraction: float) -> SixTWatCell:
    """Move both sides of one PU/PG/PD family toward its physical-MOS median."""
    if family not in {"pu", "pg", "pd"} or attribute not in {"vt", "ids"}:
        raise ValueError("family must be PU/PG/PD and attribute must be Vt/Idsat")
    updates: dict[str, MosWat] = {}
    for name in (f"{family}1", f"{family}2"):
        source, target = getattr(current, name), getattr(median, name)
        updates[name] = replace(source, **{
            attribute: getattr(source, attribute) + fraction * (getattr(target, attribute) - getattr(source, attribute))
        })
    return replace(current, **updates)


def _median_target_shmoo(worst: dict[str, object], median: dict[str, object],
                         cfg: Config, vdd: float, objective: str) -> dict[str, object]:
    """One-factor 0..100% tuning sweep from one worst cell toward the median.

    Each step shifts both devices of the selected PU/PG/PD family toward their
    own physical-MOS median.  It is intentionally a screening shmoo, not a
    simultaneous process correction or a PDK optimisation.
    """
    if objective not in {"rsnm_mv", "write_margin_mv"}:
        raise ValueError("Unsupported median-target objective")
    target = float(median[objective])
    coarse_points = 151
    rows: list[dict[str, object]] = []
    recommendations: list[dict[str, object]] = []
    for family in ("pu", "pg", "pd"):
        for attribute in ("vt", "ids"):
            family_rows = []
            for percent in range(0, 101, 10):
                cell = _move_family_toward_median(
                    worst["cell"], median["cell"], family, attribute, percent / 100.0)
                values = _multi_cell_metrics(cell, cfg, vdd, coarse_points)
                row = {
                    "family": family.upper(),
                    "parameter": "Vt" if attribute == "vt" else "Idsat",
                    "toward_median_pct": percent,
                    "rsnm_mv": values["rsnm_mv"],
                    "write_margin_mv": values["write_margin_mv"],
                    "cell_ratio_beta": values["cell_ratio_beta"],
                    "pull_up_ratio_beta": values["pull_up_ratio_beta"],
                    "reaches_target": float(values[objective]) >= target,
                }
                rows.append(row); family_rows.append(row)
            reached = next((row for row in family_rows if row["reaches_target"]), None)
            if reached:
                recommendations.append(reached)
    recommendations.sort(key=lambda row: (int(row["toward_median_pct"]),
                                           -float(row[objective])))
    return {
        "objective": objective,
        "target_value_mv": target,
        "rows": rows,
        "recommendations": recommendations,
        "method": "one-factor 10% step sweep toward physical-MOS medians",
    }


def analyze_multi_chip_wafer(chips: list[WaferChipWat], cfg: Config,
                             fit_points: int | None = None,
                             include_shmoo: bool = True) -> dict:
    """Evaluate all chip rows and retain every Read/Write VTC for wafer overlay."""
    if not chips:
        raise ValueError("At least one chip is required")
    vdd = chips[0].model_vdd_v
    if any(abs(item.model_vdd_v - vdd) > 1e-12 for item in chips):
        raise ValueError("All chips must use the same Model VDD")
    # Do not use a lighter wafer-only grid: the same physical 6T input must
    # return exactly the same RSNM/WSNM as the single-bitcell analysis.
    common_fit_points = _snm_fit_points(cfg, fit_points)
    point_cfg = replace(cfg, nominal_vdd=vdd, wat_vdd=vdd,
                        grid_points=common_fit_points)
    rows = []
    for item in chips:
        rows.append({"lot_wafer": item.lot_wafer, "chip_id": item.chip_id,
                     "raw_idsat_ua": item.raw_idsat_ua or {}, **_multi_cell_metrics(
            item.cell, point_cfg, vdd, point_cfg.grid_points)})
    valid_read = [row for row in rows if row["rsnm_mv"] is not None]
    valid_write = [row for row in rows if row["wsnm_mv"] is not None]
    if not valid_read or not valid_write:
        raise ValueError("No valid RSNM or WSNM value was produced from the imported chips")
    worst_read = min(valid_read, key=lambda row: row["rsnm_mv"])
    worst_write = min(valid_write, key=lambda row: row["wsnm_mv"])
    worst_write_margin = min(rows, key=lambda row: row["write_margin_mv"])
    worst_upper = min((row for row in rows if row["upper_rsnm_mv"] is not None),
                      key=lambda row: row["upper_rsnm_mv"])
    worst_lower = min((row for row in rows if row["lower_rsnm_mv"] is not None),
                      key=lambda row: row["lower_rsnm_mv"])
    lot_wafers = sorted({item.lot_wafer for item in chips})
    lot_wafer_display = (lot_wafers[0] if len(lot_wafers) == 1
                         else f'{len(lot_wafers)} Lot/Wafer groups')
    result = {"lot_wafer": lot_wafer_display, "lot_wafers": lot_wafers,
              "vdd_v": vdd, "rows": rows,
              "worst_rsnm": worst_read, "worst_rsnm_upper": worst_upper,
              "worst_rsnm_lower": worst_lower, "worst_wsnm": worst_write,
              "worst_write_margin": worst_write_margin,
              "shmoo_enabled": bool(include_shmoo),
              "fit_points": point_cfg.grid_points}
    if not include_shmoo:
        return result

    median_cell = _median_multi_cell(chips)
    median = {"chip_id": "MEDIAN_CELL", **_multi_cell_metrics(
        median_cell, point_cfg, vdd, point_cfg.grid_points)}
    shmoo_samples = []
    for row in rows:
        source_cell = row["cell"]
        sample = {
            "lot_wafer": row["lot_wafer"], "chip_id": row["chip_id"],
            "rsnm_mv": row["rsnm_mv"], "wsnm_mv": row["wsnm_mv"],
            "write_margin_mv": row["write_margin_mv"],
            "cell_ratio_beta": row["cell_ratio_beta"],
            "pull_up_ratio_beta": row["pull_up_ratio_beta"],
        }
        for family in ("pu", "pg", "pd"):
            left_mos = getattr(source_cell, f"{family}1")
            right_mos = getattr(source_cell, f"{family}2")
            sample[f"{family}_vt_v"] = (float(left_mos.vt) + float(right_mos.vt)) / 2.0
            sample[f"{family}_idsat_ua"] = (float(left_mos.ids) + float(right_mos.ids)) / 2.0
        shmoo_samples.append(sample)
    result.update({
        "median_cell": median,
        "dominant_read_driver": _dominant_snm_degradation_parameter(
            worst_read["cell"], median_cell, "read"),
        "dominant_write_driver": _dominant_snm_degradation_parameter(
            worst_write["cell"], median_cell, "write"),
        "median_target_read_shmoo": _median_target_shmoo(
            worst_read, median, point_cfg, vdd, "rsnm_mv"),
        "median_target_write_shmoo": _median_target_shmoo(
            worst_write_margin, median, point_cfg, vdd, "write_margin_mv"),
        "relative_shmoo": _build_estimate_vmin_ratio_shmoos([
            {"vdd_v": vdd, "samples": shmoo_samples}])[0],
    })
    return result


def analyze_three_mos(cell: ThreeTWatCell, cfg: Config,
                      datasheet_targets: DatasheetTargets | None = None) -> dict:
    """Analyze a shared PU/PG/PD object set mapped onto the physical 6T cell."""
    result = analyze(cell.representative(), cfg)
    result["object_mode"] = "3T Merged"
    result["cell"] = {
        "corner": cell.corner,
        "mos": {name.upper(): asdict(getattr(cell, name)) for name in ("pu", "pg", "pd")},
        "method": "three shared objects mapped symmetrically to the physical 6T bitcell",
        "baseline_metrics": result["baseline_6t"]["metrics"],
    }
    result["datasheet_targets"] = asdict(datasheet_targets) if datasheet_targets else None
    result["target_comparisons"] = _target_comparisons(cell, datasheet_targets)
    _attach_target_model(result, datasheet_targets, cfg)
    return result


def validate_config(cfg: Config) -> None:
    positive = ("wat_vdd", "nominal_vdd", "vt_step", "ids_step_pct", "vmin_step",
                "read_snm_limit", "technology_node_nm", "channel_length_nm",
                "pu_width_nm", "pg_width_nm", "pd_width_nm",
                "read_wordline_over_vdd", "read_bitline_over_vdd",
                "write_wordline_over_vdd", "write_high_bitline_over_vdd")
    for name in positive:
        value = getattr(cfg, name)
        if not math.isfinite(value) or value <= 0:
            raise ValueError(f"{name} must be greater than zero")
    if cfg.ids_step_pct >= 100:
        raise ValueError("ids_step_pct must be below 100")
    if not math.isfinite(cfg.nominal_temperature_c) or not -100 <= cfg.nominal_temperature_c <= 200:
        raise ValueError("nominal_temperature_c must be between -100 and 200")
    if (cfg.read_wordline_over_vdd > 1.5 or cfg.read_bitline_over_vdd > 1.5 or
            cfg.write_wordline_over_vdd > 1.5 or cfg.write_high_bitline_over_vdd > 1.5):
        raise ValueError("Read/write WL and BL ratios must not exceed 1.5")
    if not math.isfinite(cfg.write_low_bitline_over_vdd) or not 0 <= cfg.write_low_bitline_over_vdd <= 1.5:
        raise ValueError("write_low_bitline_over_vdd must be between 0 and 1.5")
    if cfg.vmin_start <= 0 or cfg.vmin_stop < cfg.vmin_start:
        raise ValueError("Vmin range must satisfy 0 < start <= stop")


COLORS = ["#111827", "#2563eb", "#dc2626", "#7c3aed", "#059669"]
# Chart-only scale. This does not alter the electrical VDD used by the model.
SNM_PLOT_AXIS_MAX_V = 1.20


def _read_vtc_pair(data: dict) -> tuple[list[tuple[float, float]], list[tuple[float, float]]]:
    """Return both curves in common (Vin, Vout) plot coordinates."""
    direct = data["read_vtc"]
    mirrored = data.get("read_vtc_mirrored")
    if mirrored is None:
        mirrored = _inverse_vtc(direct)
    return direct, mirrored


def _fmt(value: float | None, digits: int = 3) -> str:
    return "N/A" if value is None else f"{value:.{digits}f}"


def _interpolate_rsnm_vcc_point(low: RsnmVccPoint, high: RsnmVccPoint,
                                vcc_v: float) -> RsnmVccPoint:
    """Linearly interpolate measured electrical inputs between two VDD rows."""
    if high.vcc_v <= low.vcc_v:
        raise ValueError("VDD interpolation requires increasing endpoints")
    fraction = (vcc_v - low.vcc_v) / (high.vcc_v - low.vcc_v)

    def mos(a: MosWat, b: MosWat) -> MosWat:
        return MosWat(a.vt + fraction * (b.vt - a.vt),
                      a.ids + fraction * (b.ids - a.ids))

    return RsnmVccPoint(vcc_v, mos(low.pu, high.pu), mos(low.pg, high.pg),
                        mos(low.pd, high.pd))


def analyze_rsnm_vcc_curve(points: list[RsnmVccPoint], cfg: Config,
                           fit_points: int | None = None) -> dict:
    """Calculate grouped-6T Read SNM at each manually entered VDD point.

    Each Idsat value is calibrated at its own row VDD.  Eye closure is only
    estimated when the supplied rows bracket an invalid-to-valid butterfly-eye
    transition; electrical values inside that bracket are linearly interpolated.

    A grouped row is mapped to a symmetric physical cell (PUL=PUR, PGL=PGR,
    PDL=PDR) and then evaluated through the same ``AsymmetricSram6T`` read
    butterfly path as the main 6T analysis.  Therefore a symmetric 6T row at
    the same VDD and model settings produces the same RSNM value.  A grouped
    row cannot reproduce a main-cell result when its left/right devices differ.
    """
    if len(points) < 2:
        raise ValueError("Enter at least two VDD rows")
    # Keep the default numerically synchronized with the main 6T report.
    # Callers that need a lighter exploratory sweep can still pass fit_points.
    fit_points = _snm_fit_points(cfg, fit_points)
    ordered = sorted(points, key=lambda point: point.vcc_v)
    for index, point in enumerate(ordered):
        if not math.isfinite(point.vcc_v) or not 0 < point.vcc_v <= SNM_PLOT_AXIS_MAX_V:
            raise ValueError(f"Row {index + 1}: VDD must be > 0 and <= {SNM_PLOT_AXIS_MAX_V:.2f} V")
        if index and abs(point.vcc_v - ordered[index - 1].vcc_v) < 1e-12:
            raise ValueError(f"Duplicate VDD row: {point.vcc_v:.6g} V")
        for name, mos in (("PU", point.pu), ("PG", point.pg), ("PD", point.pd)):
            if not math.isfinite(mos.vt) or mos.vt <= 0:
                raise ValueError(f"Row {index + 1} {name}: Vt must be greater than zero")
            if not math.isfinite(mos.ids) or mos.ids < 0:
                raise ValueError(f"Row {index + 1} {name}: Idsat must be zero or greater")

    def evaluate(point: RsnmVccPoint) -> dict:
        point_cfg = replace(cfg, nominal_vdd=point.vcc_v, wat_vdd=point.vcc_v,
                            grid_points=fit_points)
        cell = SixTWatCell(
            f"VDD_{point.vcc_v:.6g}",
            point.pu, point.pu, point.pg, point.pg, point.pd, point.pd,
        )
        butterfly = AsymmetricSram6T(cell, point_cfg).read_butterfly(
            point.vcc_v, fit_points)["read_butterfly"]
        valid = bool(butterfly.get("valid") and butterfly.get("snm_mv") is not None and
                     butterfly["snm_mv"] > 0)
        return {"valid": valid, "snm_mv": butterfly.get("snm_mv") if valid else None,
                "reason": butterfly.get("reason", ""), "butterfly": butterfly}

    evaluated = []
    for point in ordered:
        result = evaluate(point)
        evaluated.append({
            "vcc_v": point.vcc_v,
            "pu_vt_v": point.pu.vt, "pu_idsat_ua": point.pu.ids,
            "pg_vt_v": point.pg.vt, "pg_idsat_ua": point.pg.ids,
            "pd_vt_v": point.pd.vt, "pd_idsat_ua": point.pd.ids,
            "rsnm_mv": result["snm_mv"], "valid_eye": result["valid"],
            "status": "VALID" if result["valid"] else "NO VALID EYE",
            "reason": result["reason"],
        })

    closure = None
    for index in range(len(ordered) - 1):
        if evaluated[index]["valid_eye"] or not evaluated[index + 1]["valid_eye"]:
            continue
        low_point, high_point = ordered[index], ordered[index + 1]
        low_v, high_v = low_point.vcc_v, high_point.vcc_v
        for _ in range(16):
            middle_v = (low_v + high_v) / 2.0
            middle_point = _interpolate_rsnm_vcc_point(low_point, high_point, middle_v)
            if evaluate(middle_point)["valid"]:
                high_v = middle_v
            else:
                low_v = middle_v
        closure = {
            "estimated_vcc_v": (low_v + high_v) / 2.0,
            "lower_invalid_vcc_v": low_v,
            "upper_valid_vcc_v": high_v,
            "source_low_vcc_v": low_point.vcc_v,
            "source_high_vcc_v": high_point.vcc_v,
            "method": "Bisection with linear interpolation of Vt and Idsat between bracketing rows",
        }
        break

    return {
        "rows": evaluated,
        "eye_closure": closure,
        "axis_max_v": SNM_PLOT_AXIS_MAX_V,
        "fit_points": fit_points,
        "definition": "Estimated eye-closure VDD is a compact-model boundary, not measured WT Vmin",
    }


def analyze_write_trip_margin_curve(points: list[RsnmVccPoint], cfg: Config,
                                    fit_points: int = 801) -> dict:
    """Estimate write-trip margin versus VDD from grouped PU/PG/PD WAT inputs.

    WTM is the rise permitted on the nominally-low write bitline while the
    access device can still overcome the pull-up at the inverter trip point.
    It is a compact-model trend metric rather than measured Select_Write Vmin.
    """
    if len(points) < 2:
        raise ValueError("Enter at least two VDD rows")
    fit_points = max(201, int(fit_points))
    ordered = sorted(points, key=lambda point: point.vcc_v)
    for index, point in enumerate(ordered):
        if not math.isfinite(point.vcc_v) or not 0 < point.vcc_v <= SNM_PLOT_AXIS_MAX_V:
            raise ValueError(
                f"Row {index + 1}: VDD must be > 0 and <= {SNM_PLOT_AXIS_MAX_V:.2f} V")
        if index and abs(point.vcc_v - ordered[index - 1].vcc_v) < 1e-12:
            raise ValueError(f"Duplicate VDD row: {point.vcc_v:.6g} V")
        for name, mos in (("PU", point.pu), ("PG", point.pg), ("PD", point.pd)):
            if not math.isfinite(mos.vt) or mos.vt <= 0:
                raise ValueError(f"Row {index + 1} {name}: Vt must be greater than zero")
            if not math.isfinite(mos.ids) or mos.ids < 0:
                raise ValueError(f"Row {index + 1} {name}: Idsat must be zero or greater")

    def evaluate(point: RsnmVccPoint) -> dict:
        point_cfg = replace(cfg, nominal_vdd=point.vcc_v, wat_vdd=point.vcc_v,
                            grid_points=fit_points)
        wat = WatPoint(
            f"VDD_{point.vcc_v:.6g}", point.pu.vt, point.pu.ids,
            point.pg.vt, point.pg.ids, point.pd.vt, point.pd.ids,
        )
        margin_v = Sram6T(wat, point_cfg).write_snm(point.vcc_v)
        valid = bool(math.isfinite(margin_v) and margin_v > 1e-9)
        return {"valid": valid, "wtm_mv": 1000.0 * margin_v if valid else None}

    evaluated = []
    for point in ordered:
        result = evaluate(point)
        evaluated.append({
            "vdd_v": point.vcc_v,
            "pu_vt_v": point.pu.vt, "pu_idsat_ua": point.pu.ids,
            "pg_vt_v": point.pg.vt, "pg_idsat_ua": point.pg.ids,
            "pd_vt_v": point.pd.vt, "pd_idsat_ua": point.pd.ids,
            "wtm_mv": result["wtm_mv"], "writable": result["valid"],
            "status": "WRITABLE" if result["valid"] else "NO WRITE MARGIN",
        })

    boundary = None
    for index in range(len(ordered) - 1):
        if evaluated[index]["writable"] or not evaluated[index + 1]["writable"]:
            continue
        low_point, high_point = ordered[index], ordered[index + 1]
        low_v, high_v = low_point.vcc_v, high_point.vcc_v
        for _ in range(16):
            middle_v = (low_v + high_v) / 2.0
            middle_point = _interpolate_rsnm_vcc_point(low_point, high_point, middle_v)
            if evaluate(middle_point)["valid"]:
                high_v = middle_v
            else:
                low_v = middle_v
        boundary = {
            "estimated_vdd_v": (low_v + high_v) / 2.0,
            "lower_no_margin_vdd_v": low_v,
            "upper_writable_vdd_v": high_v,
            "source_low_vdd_v": low_point.vcc_v,
            "source_high_vdd_v": high_point.vcc_v,
            "method": "Bisection with linear interpolation of Vt and Idsat between bracketing rows",
        }
        break

    return {
        "rows": evaluated,
        "write_boundary": boundary,
        "axis_max_v": SNM_PLOT_AXIS_MAX_V,
        "fit_points": fit_points,
        "definition": (
            "Estimated write-trip margin is low-bitline voltage tolerance; "
            "the boundary is not measured Select_Write Vmin"),
    }


def rsnm_vcc_curve_svg(analysis: dict, width: int = 1280, height: int = 780) -> str:
    """Render Read SNM versus manually supplied operating VDD values."""
    rows = analysis["rows"]
    left, top, plot_w, plot_h = 110, 105, 1050, 470
    max_rsnm = max((row["rsnm_mv"] for row in rows if row["rsnm_mv"] is not None), default=50.0)
    y_max = max(50.0, math.ceil(max_rsnm / 50.0) * 50.0)

    def xy(vcc_v: float, rsnm_mv: float) -> tuple[float, float]:
        return (left + vcc_v / SNM_PLOT_AXIS_MAX_V * plot_w,
                top + (1.0 - rsnm_mv / y_max) * plot_h)

    valid_rows = [row for row in rows if row["rsnm_mv"] is not None]
    axis_labels = [(xy(row["vcc_v"], 0.0)[0], f'{row["vcc_v"]:.2f} V')
                   for row in valid_rows]
    axis_label_rows = _stagger_label_rows(axis_labels, character_width=8.5)
    baseline_y = top + plot_h
    axis_label_y = [baseline_y + 54 + row_index * 23 for row_index in axis_label_rows]
    axis_title_y = max(baseline_y + 122,
                       (max(axis_label_y, default=baseline_y + 54) + 48))
    height = max(height, int(axis_title_y + 78))

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Estimated Read SNM versus Model VDD" style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="52" y="54" fill="#1D1D1F" font-size="38" font-weight="700">Estimated RSNM versus Model VDD</text>',
        '<path d="M52 79 h38" stroke="#007AFF" stroke-width="4"/><text x="101" y="85" fill="#3A3A3C" font-size="16">Manual VDD / PU / PG / PD WAT inputs</text>',
    ]
    for step in range(7):
        voltage = step * 0.2
        x, _ = xy(voltage, 0.0)
        parts += [f'<path d="M{x:.1f} {top} V{top+plot_h}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{x:.1f}" y="{top+plot_h+27}" text-anchor="middle" fill="#6E6E73" font-size="14">{voltage:.1f}</text>']
    for step in range(6):
        value = y_max * step / 5.0
        _, y = xy(0.0, value)
        parts += [f'<path d="M{left} {y:.1f} H{left+plot_w}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{left-12}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="14">{value:.0f}</text>']

    closure = analysis.get("eye_closure")
    curve_points = [xy(row["vcc_v"], row["rsnm_mv"]) for row in valid_rows]
    for row, voltage_y in zip(valid_rows, axis_label_y):
        x, y = xy(row["vcc_v"], row["rsnm_mv"])
        parts += [
            f'<path data-vdd-guide="{row["vcc_v"]:.2f}" d="M{x:.1f} {y+6:.1f} V{voltage_y-18:.1f}" '
            'stroke="#B9D7FF" stroke-width="1.5" stroke-dasharray="4 5"/>',
            f'<text x="{x:.1f}" y="{voltage_y:.1f}" text-anchor="middle" fill="#0062CC" '
            f'font-size="15" font-weight="700">{row["vcc_v"]:.2f} V</text>',
        ]
    if curve_points:
        points_text = " ".join(f"{x:.1f},{y:.1f}" for x, y in curve_points)
        parts.append(f'<polyline points="{points_text}" fill="none" stroke="#007AFF" stroke-width="4" stroke-linejoin="round" stroke-linecap="round"/>')
    if closure and valid_rows:
        boundary_point = xy(float(closure["estimated_vcc_v"]), 0.0)
        first_point = curve_points[0]
        if boundary_point[0] < first_point[0]:
            parts.append(
                f'<path data-extrapolated-to-zero="true" d="M{boundary_point[0]:.1f} {boundary_point[1]:.1f} '
                f'L{first_point[0]:.1f} {first_point[1]:.1f}" fill="none" '
                'stroke="#007AFF" stroke-width="4" stroke-dasharray="8 6" '
                'stroke-linecap="round"/>')
    point_labels = [(x, y, f'{row["rsnm_mv"]:.1f} mV')
                    for row, (x, y) in zip(valid_rows, curve_points)]
    label_positions = _place_chart_labels(
        point_labels, curve_points,
        (left + 4, top + 8, left + plot_w - 4, baseline_y - 6), 16.0)
    for row, (x, y), (label_x, label_y, label_anchor) in zip(
            valid_rows, curve_points, label_positions):
        parts += [f'<circle cx="{x:.1f}" cy="{y:.1f}" r="5" fill="#FFFFFF" stroke="#007AFF" stroke-width="3"/>',
                  f'<path d="M{x:.1f} {y:.1f}L{label_x:.1f} {label_y - 7:.1f}" stroke="#AAB4BE" stroke-width="1"/>',
                  f'<text x="{label_x:.1f}" y="{label_y:.1f}" text-anchor="{label_anchor}" fill="#1D1D1F" '
                  f'font-size="16" font-weight="700" style="paint-order:stroke;stroke:#FFFFFF;stroke-width:6;stroke-linejoin:round">{row["rsnm_mv"]:.1f} mV</text>']
    for row in rows:
        if row["valid_eye"]:
            continue
        x, y = xy(row["vcc_v"], 0.0)
        parts.append(f'<path d="M{x-5:.1f} {y-5:.1f} L{x+5:.1f} {y+5:.1f} M{x+5:.1f} {y-5:.1f} L{x-5:.1f} {y+5:.1f}" stroke="#8E8E93" stroke-width="2"/>')
    if closure:
        boundary_x, boundary_y = xy(closure["estimated_vcc_v"], 0.0)
        closure_text = f'Estimated eye-closure VDD {closure["estimated_vcc_v"]:.4f} V'
        closure_width = len(closure_text) * 8.7 + 16
        closure_anchor = "end" if boundary_x > left + plot_w * .70 else "start"
        closure_label_x = boundary_x - 12 if closure_anchor == "end" else boundary_x + 12
        closure_left = (closure_label_x - closure_width
                        if closure_anchor == "end" else closure_label_x)
        parts += [f'<path d="M{boundary_x:.1f} {top} V{top+plot_h}" stroke="#FF9500" stroke-width="3" stroke-dasharray="8 6"/>',
                  f'<circle cx="{boundary_x:.1f}" cy="{boundary_y:.1f}" r="7" fill="#FFFFFF" stroke="#FF9500" stroke-width="3"/>',
                  f'<rect x="{closure_left - 6:.1f}" y="{top + 8}" width="{closure_width + 12:.1f}" height="28" rx="6" fill="#FFF4DE" stroke="#F1D399"/>',
                  f'<text x="{closure_left + closure_width / 2:.1f}" y="{top+28}" text-anchor="middle" fill="#C56A00" font-size="16" font-weight="700">{closure_text}</text>']
    else:
        parts.append(f'<text x="{left+plot_w-8}" y="{top+28}" text-anchor="end" fill="#C56A00" font-size="16" font-weight="700">Eye-closure VDD not bracketed by the entered rows</text>')
    parts += [
        f'<text x="{left+plot_w/2}" y="{axis_title_y:.1f}" text-anchor="middle" fill="#1D1D1F" font-size="21" font-weight="700">Model VDD (V)</text>',
        f'<text x="38" y="{top+plot_h/2}" transform="rotate(-90 38 {top+plot_h/2})" text-anchor="middle" fill="#1D1D1F" font-size="21" font-weight="700">Read SNM (mV)</text>',
        f'<text x="{width/2}" y="{height-18}" text-anchor="middle" fill="#6E6E73" font-size="14">X = no valid butterfly eye. Boundary is a compact-model estimate and is not measured WT Vmin.</text>',
        '</svg>',
    ]
    return "".join(parts)


def write_rsnm_vcc_curve_outputs(analysis: dict, out_dir: str | os.PathLike[str]) -> Path:
    """Write CSV, SVG, PNG and an HTML report for manual RSNM/VDD analysis."""
    out = Path(out_dir)
    image_dir = out / "images"
    image_dir.mkdir(parents=True, exist_ok=True)
    svg_path = image_dir / "01_rsnm_vs_model_vcc.svg"
    png_path = image_dir / "01_rsnm_vs_model_vcc.png"
    svg_path.write_text(rsnm_vcc_curve_svg(analysis), encoding="utf-8")
    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError("PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc
    drawing = svg2rlg(str(svg_path))
    if drawing is None:
        raise RuntimeError("Could not render RSNM versus VDD chart")
    renderPM.drawToFile(drawing, str(png_path), fmt="PNG", dpi=180, backend="rlPyCairo")

    csv_fields = ["vcc_v", "pu_vt_v", "pu_idsat_ua", "pg_vt_v", "pg_idsat_ua",
                  "pd_vt_v", "pd_idsat_ua", "rsnm_mv", "valid_eye", "status", "reason"]
    with open(out / "rsnm_vcc_curve.csv", "w", newline="", encoding="utf-8-sig") as source:
        writer = csv.DictWriter(source, fieldnames=csv_fields)
        writer.writeheader()
        writer.writerows({key: row.get(key) for key in csv_fields} for row in analysis["rows"])
    (out / "rsnm_vcc_curve.json").write_text(json.dumps(analysis, indent=2), encoding="utf-8")

    closure = analysis.get("eye_closure")
    closure_text = (f'{closure["estimated_vcc_v"]:.4f} V' if closure else "Not bracketed")
    table_rows = "".join(
        f'<tr><td>{row["vcc_v"]:.3f}</td><td>{row["pu_vt_v"]:.4f}</td><td>{row["pu_idsat_ua"]:.3f}</td>'
        f'<td>{row["pg_vt_v"]:.4f}</td><td>{row["pg_idsat_ua"]:.3f}</td>'
        f'<td>{row["pd_vt_v"]:.4f}</td><td>{row["pd_idsat_ua"]:.3f}</td>'
        f'<td>{_fmt(row["rsnm_mv"], 2)}</td><td>{row["status"]}</td></tr>'
        for row in analysis["rows"])
    document = f'''<!doctype html><html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>HV28 SRAM Analysis - RSNM vs VDD</title>
    <style>:root{{font:100%/1.5 Calibri,"Microsoft JhengHei",Arial,sans-serif;color:#1d1d1f;background:#f5f5f7}}*{{box-sizing:border-box}}body{{margin:0;padding:2rem}}main{{max-width:1500px;margin:auto}}h1{{font-size:2.6rem;letter-spacing:-.03em}}section{{background:#fff;border-radius:1.25rem;padding:1.5rem;margin:1rem 0}}img{{display:block;width:100%;height:auto;border:1px solid #e5e5ea;border-radius:1rem}}table{{border-collapse:collapse;width:100%;font-variant-numeric:tabular-nums}}th,td{{padding:.7rem;border-bottom:1px solid #e5e5ea;text-align:right}}th:first-child,td:first-child{{text-align:left}}.note{{color:#6e6e73}}</style></head><body><main>
    <h1>HV28 SRAM Analysis</h1><p>Manual VDD / PU / PG / PD WAT curve analysis</p>
    <section><h2>Estimated RSNM versus Model VDD</h2><p><b>Estimated eye-closure VDD:</b> {closure_text}</p><img src="images/{png_path.name}" alt="Estimated Read SNM versus Model VDD"><p class="note">This is a compact-model estimate derived from manually entered Vt and Idsat. It is not measured WT Vmin.</p></section>
    <section><h2>Input and calculated values</h2><table><thead><tr><th>VDD (V)</th><th>PU Vt</th><th>PU Isat</th><th>PG Vt</th><th>PG Isat</th><th>PD Vt</th><th>PD Isat</th><th>RSNM (mV)</th><th>Status</th></tr></thead><tbody>{table_rows}</tbody></table></section>
    </main></body></html>'''
    report = out / "rsnm_vcc_report.html"
    report.write_text(document, encoding="utf-8")
    return report


_ESTIMATE_VMIN_METRICS = (
    ("rsnm_mv", "RSNM", "Read SNM", "#007AFF"),
    ("wsnm_mv", "WSNM", "Write SNM", "#AF52DE"),
    ("write_margin_mv", "Write Margin", "BL Write Margin", "#34C759"),
)

_ESTIMATE_VMIN_SAMPLE_FIELDS = {
    "lot_wafer", "chip_id", "rsnm_mv", "wsnm_mv", "write_margin_mv",
    "cell_ratio_beta", "pull_up_ratio_beta", "pu_vt_v", "pu_idsat_ua",
    "pg_vt_v", "pg_idsat_ua", "pd_vt_v", "pd_idsat_ua",
    "pul_vt_v", "pul_idsat_ua", "pur_vt_v", "pur_idsat_ua",
    "pgl_vt_v", "pgl_idsat_ua", "pgr_vt_v", "pgr_idsat_ua",
    "pdl_vt_v", "pdl_idsat_ua", "pdr_vt_v", "pdr_idsat_ua",
}


def _estimate_samples_from_multi_chip_analysis(analysis: dict) -> list[dict[str, object]]:
    source_samples = analysis.get("relative_shmoo", {}).get("samples")
    if source_samples is None:
        # Fast mode deliberately omits relative-Shmoo construction. Reuse the
        # common summary exporter so Estimate Vmin still receives the same
        # per-cell margins, CR/PR and family-average WAT values.
        source_samples = _multi_chip_summary_export_rows(analysis)
    return [
        {key: value for key, value in raw_sample.items()
         if key in _ESTIMATE_VMIN_SAMPLE_FIELDS}
        for raw_sample in source_samples
    ]


def _estimate_rows_from_grouped_samples(
        grouped: dict[float, list[dict[str, object]]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for vdd in sorted(grouped):
        samples = grouped[vdd]
        row: dict[str, object] = {
            "vdd_v": vdd, "sample_count": len(samples), "samples": samples,
            "lot_wafers": sorted({str(item["lot_wafer"]) for item in samples}),
        }
        for key, *_ in _ESTIMATE_VMIN_METRICS:
            winner = min(samples, key=lambda item: float(item[key]))
            row[key] = float(winner[key])
            row[f"{key}_chip_id"] = str(winner["chip_id"])
            row[f"{key}_lot_wafer"] = str(winner["lot_wafer"])
        rows.append(row)
    return rows


def estimate_rows_from_multi_chip_analyses(
        analyses: Iterable[dict]) -> list[dict[str, object]]:
    """Convert completed per-VDD Multi-Cell analyses into Vmin input rows."""
    grouped: dict[float, list[dict[str, object]]] = {}
    for analysis in analyses:
        grouped.setdefault(float(analysis["vdd_v"]), []).extend(
            _estimate_samples_from_multi_chip_analysis(analysis))
    if not grouped:
        raise ValueError("No completed Multi-Cell VDD analysis was supplied")
    return _estimate_rows_from_grouped_samples(grouped)


def read_multi_chip_snm_summary(
        paths: Iterable[str | os.PathLike[str]],
        default_model_vdd_v: float = .90,
        config: Config | None = None,
        include_shmoo: bool = True) -> list[dict[str, object]]:
    """Read Multi-Cell summary CSV or raw 6T Multi-Cell Excel inputs.

    CSV rows already contain modeled margins.  Excel rows contain the six-MOS
    Vt/Idsat inputs and are evaluated through the same Multi-Cell 6T model
    before joining the common summary path.  Values are grouped by Model VDD
    and reduced conservatively to the minimum Cell margin at that VDD.
    """
    grouped: dict[float, list[dict[str, object]]] = {}
    for raw_path in paths:
        path = Path(raw_path)
        if not path.is_file():
            raise FileNotFoundError(f"Multi-Cell summary was not found: {path}")
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            try:
                vdd_sheets = read_multi_chip_6t_excel_vdd_sheets(
                    path, default_model_vdd_v, allow_no_vdd_sheets=True)
                if vdd_sheets:
                    chip_groups = [(float(item["vdd_v"]), list(item["chips"]))
                                   for item in vdd_sheets]
                else:
                    chips = read_multi_chip_6t_excel(
                        path, default_model_vdd_v, require_common_vdd=False)
                    chips_by_vdd: dict[float, list[WaferChipWat]] = {}
                    for chip in chips:
                        chips_by_vdd.setdefault(float(chip.model_vdd_v), []).append(chip)
                    chip_groups = sorted(chips_by_vdd.items())
            except (ValueError, RuntimeError):
                raise
            except Exception as exc:
                raise ValueError(
                    f"{path.name} could not be read as a 6T Multi-Cell Excel "
                    "workbook. Check that it is a valid .xlsx/.xlsm file and "
                    "contains the required Lot/Wafer, Chip ID and six-MOS "
                    "Vt/Idsat columns.") from exc
            for vdd, vdd_chips in chip_groups:
                model_config = replace(
                    config or Config(), nominal_vdd=vdd, wat_vdd=vdd)
                analysis = analyze_multi_chip_wafer(
                    vdd_chips, model_config, include_shmoo=include_shmoo)
                for sample in _estimate_samples_from_multi_chip_analysis(analysis):
                    grouped.setdefault(vdd, []).append(sample)
            continue
        if suffix == ".xls":
            raise ValueError(
                f"{path.name} uses the legacy .xls format. Save it as .xlsx, "
                "then import it again.")
        if suffix != ".csv":
            raise ValueError(
                f"{path.name} is not a supported Multi-Cell input. Select a generated "
                "multi_chip_snm_summary.csv or a 6T Multi-Cell Excel workbook (.xlsx/.xlsm).")
        try:
            with path.open(newline="", encoding="utf-8-sig") as source:
                reader = csv.DictReader(source)
                required = {"lot_wafer", "chip_id", "model_vdd_v", "rsnm_mv", "wsnm_mv", "write_margin_mv"}
                if not reader.fieldnames or not required.issubset(reader.fieldnames):
                    raise ValueError(
                        f"{path.name} is not a Multi-Cell multi_chip_snm_summary.csv export")
                for number, raw in enumerate(reader, 2):
                    try:
                        vdd = float(raw["model_vdd_v"])
                        values = {key: float(raw[key]) if raw.get(key) not in (None, "")
                                  else float(raw["write_margin_mv"])
                                  for key, *_ in _ESTIMATE_VMIN_METRICS}
                    except (TypeError, ValueError) as exc:
                        raise ValueError(f"{path.name} row {number}: invalid VDD or margin value") from exc
                    if not 0 < vdd <= SNM_PLOT_AXIS_MAX_V:
                        raise ValueError(f"{path.name} row {number}: Model VDD must be within 0–{SNM_PLOT_AXIS_MAX_V:.2f} V")
                    sample = {
                        "lot_wafer": str(raw["lot_wafer"] or "Wafer"),
                        "chip_id": str(raw["chip_id"] or "Unknown"), **values,
                    }
                    # Optional fields added by newer Multi-Cell exports. Keeping
                    # them optional preserves compatibility with legacy summaries.
                    for field in (
                            "cell_ratio_beta", "pull_up_ratio_beta",
                            "pu_vt_v", "pu_idsat_ua", "pg_vt_v", "pg_idsat_ua",
                            "pd_vt_v", "pd_idsat_ua",
                            "pul_vt_v", "pul_idsat_ua", "pur_vt_v", "pur_idsat_ua",
                            "pgl_vt_v", "pgl_idsat_ua", "pgr_vt_v", "pgr_idsat_ua",
                            "pdl_vt_v", "pdl_idsat_ua", "pdr_vt_v", "pdr_idsat_ua"):
                        if raw.get(field) not in (None, ""):
                            try:
                                sample[field] = float(raw[field])
                            except (TypeError, ValueError) as exc:
                                raise ValueError(f"{path.name} row {number}: invalid {field}") from exc
                    grouped.setdefault(vdd, []).append(sample)
        except UnicodeDecodeError as exc:
            raise ValueError(
                f"{path.name} is not UTF-8 CSV text. Select the generated "
                "multi_chip_snm_summary.csv file instead of an Excel or binary file.") from exc
    if not grouped:
        raise ValueError("The selected Multi-Cell summary contains no usable Model VDD data")
    return _estimate_rows_from_grouped_samples(grouped)


def _linear_quantile(values: Iterable[float], probability: float) -> float:
    """Return an inclusive, linearly interpolated population quantile."""
    ordered = sorted(float(value) for value in values)
    if not ordered:
        raise ValueError("A quantile requires at least one value")
    q = min(max(float(probability), 0.0), 1.0)
    position = (len(ordered) - 1) * q
    low = int(math.floor(position))
    high = int(math.ceil(position))
    if low == high:
        return ordered[low]
    fraction = position - low
    return ordered[low] * (1.0 - fraction) + ordered[high] * fraction


def _midrank_percentile(values: Iterable[float], measured: float) -> float:
    """Return a tie-aware 0..1 percentile rank within one wafer population."""
    ordered = sorted(float(value) for value in values)
    if not ordered:
        return 0.0
    value = float(measured)
    less = sum(item < value for item in ordered)
    equal = sum(item == value for item in ordered)
    return (less + 0.5 * equal) / len(ordered)


def _robust_distribution(values: Iterable[float]) -> dict[str, float]:
    """Summarize one wafer metric without relying on outlier-sensitive extrema."""
    ordered = [float(value) for value in values]
    median = _linear_quantile(ordered, .50)
    deviations = [abs(value - median) for value in ordered]
    return {
        "p05": _linear_quantile(ordered, .05),
        "q1": _linear_quantile(ordered, .25),
        "median": median,
        "q3": _linear_quantile(ordered, .75),
        "p95": _linear_quantile(ordered, .95),
        "mad": _linear_quantile(deviations, .50),
    }


def _build_estimate_vmin_ratio_shmoos(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Build same-VDD drive-balance shmoos and actionable target deltas."""
    shmoos: list[dict[str, object]] = []
    # Read stability and write ability are independent responses.  Use RSNM
    # and the BL Write Trip Margin once each; residual WSNM is retained as a
    # diagnostic, so it is not double-counted in the balance score.
    metric_keys = ("rsnm_mv", "write_margin_mv")
    family_fields = tuple(field for family in ("pu", "pg", "pd")
                          for field in (f"{family}_vt_v", f"{family}_idsat_ua"))
    for row in rows:
        samples = [dict(sample) for sample in row.get("samples", [])
                   if "cell_ratio_beta" in sample and "pull_up_ratio_beta" in sample]
        if not samples:
            continue
        maxima = {key: max(max(float(sample.get(key, sample.get("write_margin_mv", 0.0))), 0.0)
                           for sample in samples) for key in metric_keys}
        for sample in samples:
            normalized = {
                key: (max(float(sample.get(key, 0.0)), 0.0) / maxima[key]
                      if maxima[key] > 0 else 0.0)
                for key in metric_keys
            }
            sample["read_score"] = normalized["rsnm_mv"]
            sample["write_score"] = normalized["write_margin_mv"]
            sample["balanced_score"] = min(sample["read_score"], sample["write_score"])
        distribution_fields = ("rsnm_mv", "write_margin_mv",
                               "cell_ratio_beta", "pull_up_ratio_beta")
        distributions = {
            field: _robust_distribution(float(sample[field]) for sample in samples)
            for field in distribution_fields
        }
        read_values = [float(sample["rsnm_mv"]) for sample in samples]
        write_values = [float(sample["write_margin_mv"]) for sample in samples]
        cr_values = [float(sample["cell_ratio_beta"]) for sample in samples]
        pr_values = [float(sample["pull_up_ratio_beta"]) for sample in samples]
        for sample in samples:
            sample["read_percentile"] = _midrank_percentile(
                read_values, float(sample["rsnm_mv"]))
            sample["write_percentile"] = _midrank_percentile(
                write_values, float(sample["write_margin_mv"]))
            sample["cr_percentile"] = _midrank_percentile(
                cr_values, float(sample["cell_ratio_beta"]))
            sample["pr_percentile"] = _midrank_percentile(
                pr_values, float(sample["pull_up_ratio_beta"]))
            # The Shmoo grade follows the two plotted drive ratios, so a
            # write-heavy lower-right or read-heavy upper-left cell cannot be
            # colored green solely because one axis is strong.  Performance
            # percentiles remain available to check whether this ratio proxy
            # agrees with the modeled RSNM/Vtrip response.
            sample["performance_grade_score"] = min(
                sample["read_percentile"], sample["write_percentile"])
            sample["wafer_grade_score"] = min(
                sample["cr_percentile"], sample["pr_percentile"])
            if sample["wafer_grade_score"] >= .50:
                sample["wafer_grade"] = "preferred"
            elif sample["wafer_grade_score"] >= .25:
                sample["wafer_grade"] = "monitor"
            else:
                sample["wafer_grade"] = "low"
            sample["robust_low_outlier"] = (
                float(sample["rsnm_mv"]) < distributions["rsnm_mv"]["p05"] or
                float(sample["write_margin_mv"]) < distributions["write_margin_mv"]["p05"])
        population_medians = {
            "cell_ratio_beta": statlib.median(float(sample["cell_ratio_beta"]) for sample in samples),
            "pull_up_ratio_beta": statlib.median(float(sample["pull_up_ratio_beta"]) for sample in samples),
        }
        if all(all(field in sample for field in family_fields) for sample in samples):
            population_medians.update({
                field: statlib.median(float(sample[field]) for sample in samples)
                for field in family_fields
            })
        for sample in samples:
            for field, median_value in population_medians.items():
                measured_value = float(sample[field])
                sample[f"wafer_median_{field}"] = median_value
                sample[f"delta_vs_median_{field}_pct"] = (
                    100.0 * (measured_value - median_value) / abs(median_value)
                    if abs(median_value) > 1e-15 else 0.0)
            sample["read_balance_vs_median_pct"] = sample["delta_vs_median_cell_ratio_beta_pct"]
            sample["write_balance_vs_median_pct"] = sample["delta_vs_median_pull_up_ratio_beta_pct"]
        best_score = max(float(sample["balanced_score"]) for sample in samples)
        best_sample = max(
            samples,
            key=lambda sample: (float(sample["wafer_grade_score"]),
                                float(sample["performance_grade_score"]),
                                float(sample["balanced_score"])))
        best_cutoff = best_score * 0.90
        for sample in samples:
            sample["best_region"] = float(sample["balanced_score"]) >= best_cutoff
        preferred = [sample for sample in samples
                     if sample["wafer_grade"] == "preferred"]
        if not preferred:
            preferred = samples
        target_fields = ["cell_ratio_beta", "pull_up_ratio_beta"]
        if all(all(field in sample for field in family_fields) for sample in samples):
            target_fields.extend(family_fields)
        # The center is the whole-wafer population median.  It is intentionally
        # separate from the best measured cell and is not an absolute target.
        target = {field: statlib.median(float(sample[field]) for sample in samples)
                  for field in target_fields}
        target["chip_id"] = "WAFER_MEDIAN_CENTER"
        target["balanced_score"] = statlib.median(
            float(sample["balanced_score"]) for sample in samples)
        target["wafer_grade_score"] = statlib.median(
            float(sample["wafer_grade_score"]) for sample in samples)
        weakest = min(samples, key=lambda sample: (
            float(sample["wafer_grade_score"]), float(sample["balanced_score"])))
        for sample in samples:
            sample["target_cr"] = target["cell_ratio_beta"]
            sample["target_pr"] = target["pull_up_ratio_beta"]
            sample["delta_cr"] = target["cell_ratio_beta"] - float(sample["cell_ratio_beta"])
            sample["delta_pr"] = target["pull_up_ratio_beta"] - float(sample["pull_up_ratio_beta"])
            for field in family_fields:
                if field in target and field in sample:
                    sample[f"delta_{field}"] = float(target[field]) - float(sample[field])
        shmoos.append({"vdd_v": float(row["vdd_v"]), "samples": samples,
                        "best_score": best_score, "best_cutoff": best_cutoff,
                        "target": target, "best": best_sample, "weakest": weakest,
                        "population_medians": population_medians,
                        "distributions": distributions,
                        "grade_thresholds": {"preferred": .50, "monitor": .25},
                        "has_family_wat": all(all(field in sample for field in family_fields)
                                              for sample in samples),
                        "definition": ("Wafer grade is the smaller within-wafer percentile of "
                                       "CR and PR at one Model VDD: "
                                       "green >= P50, yellow P25-P50, red < P25. "
                                       "RSNM and BL Write Trip Margin percentiles plus the balanced "
                                       "max-normalized score are retained for correlation; "
                                       "residual WSNM is reported separately. "
                                       "X=MOSdrive(PG)/MOSdrive(PU)=PR (right is easier write); "
                                       "Y=MOSdrive(PD)/MOSdrive(PG)=CR (up is stronger read). "
                                       "This is relative screening, not absolute silicon Pass/Fail.")})
    return shmoos


def build_drive_to_preferred_advice(
        shmoo: dict[str, object], chip_id: str,
        target_percentile: float = .55,
        lot_wafer: str | None = None) -> dict[str, object]:
    """Build a same-VDD, population-relative CR/PR adjustment screen.

    The default P55 guardband sits just above the Preferred P50 boundary.  PG
    is held as the common denominator so read and write gaps can be addressed
    independently: PD is strengthened only when CR is short, and PU is
    weakened only when PR is short.  Suggested Idsat values are fixed-Vt
    equivalents, not independent process prescriptions.
    """
    probability = float(target_percentile)
    if not .50 <= probability <= .95:
        raise ValueError("Advisor target percentile must be between P50 and P95")
    samples = list(shmoo.get("samples", []))
    matches = [sample for sample in samples
               if str(sample.get("chip_id")) == str(chip_id)
               and (lot_wafer is None or str(sample.get("lot_wafer")) == str(lot_wafer))]
    if not matches:
        raise ValueError(f"Cell {chip_id} was not found in this VDD population")
    sample = matches[0]
    cr_values = [float(item["cell_ratio_beta"]) for item in samples]
    pr_values = [float(item["pull_up_ratio_beta"]) for item in samples]
    target_cr = _linear_quantile(cr_values, probability)
    target_pr = _linear_quantile(pr_values, probability)
    current_cr = float(sample["cell_ratio_beta"])
    current_pr = float(sample["pull_up_ratio_beta"])
    planned_cr = max(current_cr, target_cr)
    planned_pr = max(current_pr, target_pr)

    # Normalize the PG MOS-drive proxy to 1.0. Only ratios are needed, so this
    # avoids implying that the internal beta proxy has a calibrated absolute unit.
    current_beta = {"PD": current_cr, "PG": 1.0, "PU": 1.0 / current_pr}
    target_beta = {"PD": planned_cr, "PG": 1.0, "PU": 1.0 / planned_pr}
    device_rows: list[dict[str, object]] = []
    directions = {
        "PD": ("Idsat ↑ or Vt ↓", "Strengthen read pull-down"),
        "PG": ("Hold", "Keep the shared read/write denominator fixed"),
        "PU": ("Idsat ↓ or |Vt| ↑", "Reduce write contention"),
    }
    for family in ("PD", "PG", "PU"):
        current = current_beta[family]
        target = target_beta[family]
        change_pct = 100.0 * (target / current - 1.0)
        field = family.lower()
        idsat = sample.get(f"{field}_idsat_ua")
        vt = sample.get(f"{field}_vt_v")
        idsat_target = (float(idsat) * target / current
                        if idsat is not None else None)
        if abs(change_pct) < .05:
            action, reason = "HOLD", "Already meets this ratio target"
        else:
            action, reason = directions[family]
        device_rows.append({
            "family": family,
            "beta_current_relative": current,
            "beta_target_relative": target,
            "beta_change_pct": change_pct,
            "action": action,
            "reason": reason,
            "vt_v": float(vt) if vt is not None else None,
            "idsat_current_ua": float(idsat) if idsat is not None else None,
            "idsat_target_fixed_vt_ua": idsat_target,
        })

    predicted_cr_percentile = _midrank_percentile(cr_values, planned_cr)
    predicted_pr_percentile = _midrank_percentile(pr_values, planned_pr)
    predicted_score = min(predicted_cr_percentile, predicted_pr_percentile)
    predicted_grade = ("preferred" if predicted_score >= .50 else
                       "monitor" if predicted_score >= .25 else "low")
    current_score = min(float(sample.get("cr_percentile", 0.0)),
                        float(sample.get("pr_percentile", 0.0)))
    return {
        "vdd_v": float(shmoo["vdd_v"]),
        "lot_wafer": str(sample.get("lot_wafer", "")),
        "chip_id": str(sample.get("chip_id", chip_id)),
        "target_percentile": probability,
        "current": {
            "cr": current_cr, "pr": current_pr,
            "cr_percentile": float(sample.get("cr_percentile", 0.0)),
            "pr_percentile": float(sample.get("pr_percentile", 0.0)),
            "score": current_score,
            "grade": str(sample.get("wafer_grade", "unknown")),
        },
        "target": {"cr": target_cr, "pr": target_pr},
        "predicted": {
            "cr": planned_cr, "pr": planned_pr,
            "cr_percentile": predicted_cr_percentile,
            "pr_percentile": predicted_pr_percentile,
            "score": predicted_score, "grade": predicted_grade,
        },
        "devices": device_rows,
        "method": ("P55 same-VDD population guardband; PG MOSdrive held, PD MOSdrive "
                   "raised for CR and PU MOSdrive lowered for PR only where required."),
        "caution": ("Relative compact-model sensitivity only. Vt and Idsat are correlated; "
                    "confirm any process action with Device/PDK and measured WT."),
    }


def build_batch_drive_to_preferred_advice(
        shmoo: dict[str, object],
        target_percentile: float = .55,
        lot_wafer: str | None = None) -> dict[str, object]:
    """Estimate one common drive shift for the Low/Monitor batch population.

    The P55 CR/PR targets are frozen from the current same-VDD population.
    A common PD multiplier and PU multiplier are then chosen from the most
    limiting Low/Monitor cells while PG is held.  Uniform shifts do not change
    percentile ranks if the population target is recalculated, so the reported
    coverage is explicitly against the frozen pre-adjustment target.
    """
    probability = float(target_percentile)
    if not .50 <= probability <= .95:
        raise ValueError("Batch Advisor target percentile must be between P50 and P95")
    reference_samples = list(shmoo.get("samples", []))
    if not reference_samples:
        raise ValueError("Batch Advisor requires at least one same-VDD cell")
    samples = [item for item in reference_samples
               if lot_wafer is None or str(item.get("lot_wafer")) == str(lot_wafer)]
    if not samples:
        raise ValueError(f"Lot/Wafer {lot_wafer} was not found in this VDD population")
    cr_values = [float(item["cell_ratio_beta"]) for item in reference_samples]
    pr_values = [float(item["pull_up_ratio_beta"]) for item in reference_samples]
    target_cr = _linear_quantile(cr_values, probability)
    target_pr = _linear_quantile(pr_values, probability)
    affected = [item for item in samples
                if str(item.get("wafer_grade", "low")) in {"low", "monitor"}]
    read_limit = (min(affected, key=lambda item: float(item["cell_ratio_beta"]))
                  if affected else None)
    write_limit = (min(affected, key=lambda item: float(item["pull_up_ratio_beta"]))
                   if affected else None)
    pd_multiplier = (max(1.0, target_cr / float(read_limit["cell_ratio_beta"]))
                     if read_limit else 1.0)
    pu_multiplier = (min(1.0, float(write_limit["pull_up_ratio_beta"]) / target_pr)
                     if write_limit else 1.0)
    pg_multiplier = 1.0

    def reaches_frozen_target(item: dict[str, object], adjusted: bool) -> bool:
        cr = float(item["cell_ratio_beta"]) * (pd_multiplier if adjusted else 1.0)
        pr = float(item["pull_up_ratio_beta"]) / (pu_multiplier if adjusted else 1.0)
        return cr >= target_cr - 1e-12 and pr >= target_pr - 1e-12

    before_count = sum(reaches_frozen_target(item, False) for item in samples)
    after_count = sum(reaches_frozen_target(item, True) for item in samples)
    affected_after = sum(reaches_frozen_target(item, True) for item in affected)
    device_rows: list[dict[str, object]] = []
    for family, multiplier, action, driver in (
            ("PD", pd_multiplier, "Idsat ↑ or Vt ↓", read_limit),
            ("PG", pg_multiplier, "HOLD", None),
            ("PU", pu_multiplier, "Idsat ↓ or |Vt| ↑", write_limit)):
        field = family.lower()
        vt_values = [float(item[f"{field}_vt_v"]) for item in samples
                     if item.get(f"{field}_vt_v") is not None]
        idsat_values = [float(item[f"{field}_idsat_ua"]) for item in samples
                        if item.get(f"{field}_idsat_ua") is not None]
        median_vt = statlib.median(vt_values) if vt_values else None
        median_idsat = statlib.median(idsat_values) if idsat_values else None
        change_pct = 100.0 * (multiplier - 1.0)
        if abs(change_pct) < .05:
            action = "HOLD"
        device_rows.append({
            "family": family,
            "drive_multiplier": multiplier,
            "drive_change_pct": change_pct,
            "action": action,
            "median_vt_v": median_vt,
            "median_idsat_current_ua": median_idsat,
            "median_idsat_target_fixed_vt_ua": (
                median_idsat * multiplier if median_idsat is not None else None),
            "limiting_chip_id": (str(driver.get("chip_id", "")) if driver else ""),
        })
    grade_counts = {
        grade: sum(str(item.get("wafer_grade")) == grade for item in samples)
        for grade in ("preferred", "monitor", "low")
    }
    return {
        "vdd_v": float(shmoo["vdd_v"]),
        "lot_wafer": str(lot_wafer or "ALL"),
        "target_percentile": probability,
        "target": {"cr": target_cr, "pr": target_pr},
        "sample_count": len(samples),
        "reference_sample_count": len(reference_samples),
        "affected_count": len(affected),
        "grade_counts": grade_counts,
        "frozen_target_coverage_before_pct": 100.0 * before_count / len(samples),
        "frozen_target_coverage_after_pct": 100.0 * after_count / len(samples),
        "affected_coverage_after_pct": (
            100.0 * affected_after / len(affected) if affected else 100.0),
        "devices": device_rows,
        "read_limiting_chip_id": str(read_limit.get("chip_id", "")) if read_limit else "",
        "write_limiting_chip_id": str(write_limit.get("chip_id", "")) if write_limit else "",
        "method": ("Common batch shift sized by the limiting Low/Monitor CR and PR, "
                   "using the current population P55 values as frozen references."),
        "caution": ("A uniform wafer/batch shift preserves CR/PR rank ordering. If P55 is "
                    "recalculated after the shift, relative grades do not improve; use this "
                    "only as frozen-target sensitivity guidance, not a process recipe."),
    }


_LOT_WAFER_ADVISOR_METRICS = (
    ("rsnm_mv", "Read SNM", "mV", "#007AFF"),
    ("write_margin_mv", "BL Write Trip Margin", "mV", "#FF9500"),
    ("balanced_drive_pct", "Balanced Drive Score", "%", "#AF52DE"),
)
_LOT_WAFER_COLORS = ("#007AFF", "#FF9500", "#AF52DE", "#008F5D", "#D64D73")
_LOT_WAFER_LIGHTS = ("#EAF3FF", "#FFF1DB", "#F3EAFE", "#E3F5ED", "#FBE9EF")
_LOT_WAFER_MARKERS = ("circle", "square", "triangle", "diamond")


def _tukey_box(values: Iterable[float]) -> dict[str, object]:
    """Return quartiles, Tukey whiskers and outliers for one group."""
    ordered = sorted(float(value) for value in values)
    if not ordered:
        raise ValueError("Box-plot statistics require at least one value")
    q1 = _linear_quantile(ordered, .25)
    median = _linear_quantile(ordered, .50)
    q3 = _linear_quantile(ordered, .75)
    iqr = q3 - q1
    lower_fence = q1 - 1.5 * iqr
    upper_fence = q3 + 1.5 * iqr
    inside = [value for value in ordered if lower_fence <= value <= upper_fence]
    return {
        "count": len(ordered), "minimum": ordered[0], "maximum": ordered[-1],
        "q1": q1, "median": median, "q3": q3, "iqr": iqr,
        "whisker_low": min(inside) if inside else ordered[0],
        "whisker_high": max(inside) if inside else ordered[-1],
        "mean": statlib.fmean(ordered),
        "outliers": [value for value in ordered
                     if value < lower_fence or value > upper_fence],
        "values": ordered,
    }


def analyze_lot_wafer_drive_advisor(
        summary_rows: list[dict[str, object]]) -> dict[str, object]:
    """Compare Lot/Wafer read, write and balanced drive at each Model VDD.

    Lot/Wafer names are the grouping key, even when matching names arrive from
    different imported files.  Percentiles and the frozen P55 reference are
    calculated across the complete same-VDD population; distributions are
    then summarized separately for each Lot/Wafer group.
    """
    shmoos = _build_estimate_vmin_ratio_shmoos(summary_rows)
    if not shmoos:
        raise ValueError(
            "Lot/Wafer Advisor requires CR, PR, RSNM and Write Margin columns")
    all_lots = sorted({str(sample.get("lot_wafer", "Wafer"))
                       for shmoo in shmoos for sample in shmoo["samples"]})
    style_map = {
        lot: {
            "color": _LOT_WAFER_COLORS[index % len(_LOT_WAFER_COLORS)],
            "light": _LOT_WAFER_LIGHTS[index % len(_LOT_WAFER_LIGHTS)],
            "marker": _LOT_WAFER_MARKERS[index % len(_LOT_WAFER_MARKERS)],
        }
        for index, lot in enumerate(all_lots)
    }
    vdd_groups: list[dict[str, object]] = []
    for shmoo in shmoos:
        grouped: dict[str, list[dict[str, object]]] = {}
        for sample in shmoo["samples"]:
            item = dict(sample)
            item["balanced_drive_pct"] = 100.0 * float(item["wafer_grade_score"])
            grouped.setdefault(str(item.get("lot_wafer", "Wafer")), []).append(item)
        groups: list[dict[str, object]] = []
        for lot in sorted(grouped):
            samples = grouped[lot]
            metrics = {
                key: _tukey_box(float(sample[key]) for sample in samples)
                for key, *_ in _LOT_WAFER_ADVISOR_METRICS
            }
            cr_values = [float(sample["cell_ratio_beta"]) for sample in samples]
            pr_values = [float(sample["pull_up_ratio_beta"]) for sample in samples]
            grade_counts = {
                grade: sum(str(sample.get("wafer_grade")) == grade for sample in samples)
                for grade in ("preferred", "monitor", "low")
            }
            groups.append({
                "lot_wafer": lot, "sample_count": len(samples), "samples": samples,
                "metrics": metrics, "grade_counts": grade_counts,
                "median_cr": _linear_quantile(cr_values, .50),
                "median_pr": _linear_quantile(pr_values, .50),
                "q1_cr": _linear_quantile(cr_values, .25),
                "q3_cr": _linear_quantile(cr_values, .75),
                "q1_pr": _linear_quantile(pr_values, .25),
                "q3_pr": _linear_quantile(pr_values, .75),
                "batch_advice": build_batch_drive_to_preferred_advice(
                    shmoo, .55, lot_wafer=lot),
            })
        vdd_groups.append({
            "vdd_v": float(shmoo["vdd_v"]), "sample_count": len(shmoo["samples"]),
            "lot_count": len(groups), "groups": groups, "shmoo": shmoo,
            "target_cr": _linear_quantile(
                (float(sample["cell_ratio_beta"]) for sample in shmoo["samples"]), .55),
            "target_pr": _linear_quantile(
                (float(sample["pull_up_ratio_beta"]) for sample in shmoo["samples"]), .55),
        })
    return {
        "vdds": vdd_groups, "lot_wafers": all_lots, "styles": style_map,
        "definition": (
            "Groups use exact Lot/Wafer names. Read=RSNM; Write=BL Write Trip Margin; "
            "Balanced Drive Score=100×min(CR percentile, PR percentile) within the "
            "complete same-VDD population. Box plots use Tukey 1.5×IQR whiskers."),
    }


def _lot_marker_svg(kind: str, x: float, y: float, size: float,
                    color: str, fill: str | None = None,
                    opacity: float = 1.0) -> str:
    """Draw one color-plus-shape marker so Lot/Wafer groups remain distinct."""
    inside = fill or color
    common = f'fill="{inside}" stroke="{color}" stroke-width="2" opacity="{opacity:.2f}"'
    if kind == "square":
        return (f'<rect x="{x-size:.1f}" y="{y-size:.1f}" width="{2*size:.1f}" '
                f'height="{2*size:.1f}" {common}/>')
    if kind == "triangle":
        return (f'<path d="M{x:.1f} {y-size*1.18:.1f} L{x+size*1.08:.1f} '
                f'{y+size:.1f} L{x-size*1.08:.1f} {y+size:.1f} Z" {common}/>')
    if kind == "diamond":
        return (f'<path d="M{x:.1f} {y-size*1.2:.1f} L{x+size:.1f} {y:.1f} '
                f'L{x:.1f} {y+size*1.2:.1f} L{x-size:.1f} {y:.1f} Z" {common}/>')
    return f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{size:.1f}" {common}/>'


def lot_wafer_drive_scatter_svg(vdd_group: dict[str, object],
                                 styles: dict[str, dict[str, str]],
                                 width: int = 1600, height: int = 900) -> str:
    """Render same-VDD PR/CR locations colored and shaped by Lot/Wafer."""
    groups = list(vdd_group["groups"])
    samples = [sample for group in groups for sample in group["samples"]]
    left, top, right, bottom = 110, 145, 1190, 790
    x_values = [float(sample["pull_up_ratio_beta"]) for sample in samples]
    y_values = [float(sample["cell_ratio_beta"]) for sample in samples]

    def padded_domain(values: list[float]) -> tuple[float, float]:
        low, high = min(values), max(values)
        span = high - low
        padding = max(span * .10, max(abs(low), abs(high), 1.0) * .025)
        return low - padding, high + padding

    x_min, x_max = padded_domain(x_values)
    y_min, y_max = padded_domain(y_values)
    sx = lambda value: left + (float(value) - x_min) / (x_max - x_min) * (right-left)
    sy = lambda value: bottom - (float(value) - y_min) / (y_max - y_min) * (bottom-top)
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        f'<text x="56" y="54" fill="#1D1D1F" font-size="31" font-weight="700">Lot/Wafer CR–PR Drive Distribution</text>',
        f'<text x="56" y="86" fill="#6E6E73" font-size="16">Model VDD {float(vdd_group["vdd_v"]):.3f} V · {len(samples)} Cells · {len(groups)} Lot/Wafer groups</text>',
        '<text x="56" y="111" fill="#6E6E73" font-size="14">Color + marker identify Lot/Wafer; dashed boxes show each group’s central 50% CR/PR window.</text>',
    ]
    for fraction in (0, .25, .50, .75, 1):
        x = left + fraction * (right-left)
        y = bottom - fraction * (bottom-top)
        xv = x_min + fraction * (x_max-x_min)
        yv = y_min + fraction * (y_max-y_min)
        parts += [
            f'<path d="M{x:.1f} {top} V{bottom}" stroke="#E5E5EA" stroke-width="1"/>',
            f'<path d="M{left} {y:.1f} H{right}" stroke="#E5E5EA" stroke-width="1"/>',
            f'<text x="{x:.1f}" y="{bottom+27}" text-anchor="middle" fill="#6E6E73" font-size="13">{xv:.3f}</text>',
            f'<text x="{left-14}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="13">{yv:.3f}</text>',
        ]
    parts += [f'<path d="M{left} {top} V{bottom} H{right}" fill="none" stroke="#1D1D1F" stroke-width="2"/>']
    median_pr = _linear_quantile(x_values, .50)
    median_cr = _linear_quantile(y_values, .50)
    parts += [
        f'<path d="M{sx(median_pr):.1f} {top} V{bottom}" stroke="#8E8E93" stroke-width="2" stroke-dasharray="6 6"/>',
        f'<path d="M{left} {sy(median_cr):.1f} H{right}" stroke="#8E8E93" stroke-width="2" stroke-dasharray="6 6"/>',
    ]
    for group in groups:
        lot = str(group["lot_wafer"])
        style = styles[lot]
        x1, x2 = sx(group["q1_pr"]), sx(group["q3_pr"])
        y1, y2 = sy(group["q3_cr"]), sy(group["q1_cr"])
        parts.append(
            f'<rect x="{min(x1,x2):.1f}" y="{min(y1,y2):.1f}" '
            f'width="{abs(x2-x1):.1f}" height="{abs(y2-y1):.1f}" '
            f'fill="{style["light"]}" stroke="{style["color"]}" '
            'stroke-opacity="0.65" stroke-width="2" stroke-dasharray="7 5"/>')
        for sample in group["samples"]:
            tooltip = html.escape(
                f'{lot} · {sample["chip_id"]} · PR {float(sample["pull_up_ratio_beta"]):.3f} · '
                f'CR {float(sample["cell_ratio_beta"]):.3f} · RSNM {float(sample["rsnm_mv"]):.1f} mV · '
                f'Vtrip {float(sample["write_margin_mv"]):.1f} mV')
            marker = _lot_marker_svg(
                style["marker"], sx(sample["pull_up_ratio_beta"]),
                sy(sample["cell_ratio_beta"]), 5.0, style["color"], opacity=.78)
            parts.append(f'<g><title>{tooltip}</title>{marker}</g>')
        parts.append(_lot_marker_svg(
            style["marker"], sx(group["median_pr"]), sy(group["median_cr"]),
            9.0, style["color"], fill="#FFFFFF"))
    legend_x, legend_y = 1245, 164
    parts += [
        f'<text x="{legend_x}" y="{legend_y-25}" fill="#1D1D1F" font-size="19" font-weight="700">LOT / WAFER</text>',
        f'<text x="{legend_x}" y="{legend_y}" fill="#6E6E73" font-size="13">Large open marker = group median</text>',
    ]
    for index, group in enumerate(groups):
        lot = str(group["lot_wafer"]); style = styles[lot]; y = legend_y + 42 + index*48
        parts += [
            _lot_marker_svg(style["marker"], legend_x+10, y-5, 7, style["color"]),
            f'<text x="{legend_x+30}" y="{y}" fill="#1D1D1F" font-size="15" font-weight="700">{html.escape(lot)}</text>',
            f'<text x="{legend_x+30}" y="{y+19}" fill="#6E6E73" font-size="12">n={int(group["sample_count"])} · median PR {float(group["median_pr"]):.3f} · CR {float(group["median_cr"]):.3f}</text>',
        ]
    parts += [
        f'<text x="{(left+right)/2:.1f}" y="{height-55}" text-anchor="middle" fill="#1D1D1F" font-size="17" font-weight="700">Write drive — Pull-up Ratio (PR)</text>',
        f'<text x="{(left+right)/2:.1f}" y="{height-29}" text-anchor="middle" fill="#6E6E73" font-size="14">MOSdrive(PG) / MOSdrive(PU) · right = easier write</text>',
        f'<text x="27" y="{(top+bottom)/2:.1f}" transform="rotate(-90 27 {(top+bottom)/2:.1f})" text-anchor="middle" fill="#1D1D1F" font-size="17" font-weight="700">Read drive — Cell Ratio (CR)</text>',
        f'<text x="52" y="{(top+bottom)/2:.1f}" transform="rotate(-90 52 {(top+bottom)/2:.1f})" text-anchor="middle" fill="#6E6E73" font-size="14">MOSdrive(PD) / MOSdrive(PG) · up = stronger read</text>',
        '</svg>',
    ]
    return "".join(parts)


def lot_wafer_boxplot_svg(vdd_group: dict[str, object],
                          styles: dict[str, dict[str, str]],
                          width: int = 1600, height: int = 1120) -> str:
    """Render grouped Tukey box plots for Read, Write and balanced drive."""
    groups = list(vdd_group["groups"])
    left, right = 125, width-70
    panel_top, panel_height, panel_gap = 145, 245, 70
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="56" y="54" fill="#1D1D1F" font-size="31" font-weight="700">Lot/Wafer Read, Write and Balanced Drive Distributions</text>',
        f'<text x="56" y="86" fill="#6E6E73" font-size="16">Model VDD {float(vdd_group["vdd_v"]):.3f} V · Tukey box plots (1.5×IQR whiskers) · same Lot/Wafer names are grouped</text>',
    ]
    slot = (right-left) / max(len(groups), 1)
    for panel_index, (key, label, unit, _accent) in enumerate(_LOT_WAFER_ADVISOR_METRICS):
        top = panel_top + panel_index*(panel_height+panel_gap)
        bottom = top + panel_height
        all_values = [value for group in groups for value in group["metrics"][key]["values"]]
        low, high = min(all_values), max(all_values)
        span = high-low
        pad = max(span*.12, max(abs(low), abs(high), 1.0)*.025)
        y_min, y_max = low-pad, high+pad
        if key == "balanced_drive_pct":
            y_min, y_max = max(0.0, y_min), min(100.0, y_max)
            if y_max <= y_min:
                y_min, y_max = 0.0, 100.0
        sy = lambda value: bottom - (float(value)-y_min)/(y_max-y_min)*panel_height
        parts += [
            f'<text x="{left}" y="{top-20}" fill="#1D1D1F" font-size="19" font-weight="700">{label} ({unit})</text>',
        ]
        for fraction in (0, .25, .50, .75, 1):
            y = bottom-fraction*panel_height
            value = y_min+fraction*(y_max-y_min)
            parts += [
                f'<path d="M{left} {y:.1f} H{right}" stroke="#E5E5EA" stroke-width="1"/>',
                f'<text x="{left-13}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="12">{value:.1f}</text>',
            ]
        parts.append(f'<path d="M{left} {top} V{bottom} H{right}" fill="none" stroke="#1D1D1F" stroke-width="2"/>')
        for index, group in enumerate(groups):
            lot = str(group["lot_wafer"]); style = styles[lot]
            stats = group["metrics"][key]
            x = left + slot*(index+.5); box_width = min(88.0, slot*.46)
            q1_y, q3_y = sy(stats["q1"]), sy(stats["q3"])
            low_y, high_y = sy(stats["whisker_low"]), sy(stats["whisker_high"])
            median_y = sy(stats["median"])
            parts += [
                f'<path d="M{x:.1f} {high_y:.1f} V{q3_y:.1f} M{x:.1f} {q1_y:.1f} V{low_y:.1f}" stroke="{style["color"]}" stroke-width="2"/>',
                f'<path d="M{x-box_width*.28:.1f} {high_y:.1f} H{x+box_width*.28:.1f} M{x-box_width*.28:.1f} {low_y:.1f} H{x+box_width*.28:.1f}" stroke="{style["color"]}" stroke-width="2"/>',
                f'<rect x="{x-box_width/2:.1f}" y="{min(q1_y,q3_y):.1f}" width="{box_width:.1f}" height="{max(abs(q1_y-q3_y),2):.1f}" fill="{style["light"]}" stroke="{style["color"]}" stroke-width="2"/>',
                f'<path d="M{x-box_width/2:.1f} {median_y:.1f} H{x+box_width/2:.1f}" stroke="#1D1D1F" stroke-width="3"/>',
            ]
            for outlier in stats["outliers"]:
                parts.append(_lot_marker_svg(style["marker"], x, sy(outlier), 3.5, style["color"], fill="#FFFFFF"))
            if panel_index == len(_LOT_WAFER_ADVISOR_METRICS)-1:
                parts += [
                    f'<text x="{x:.1f}" y="{bottom+27}" text-anchor="middle" fill="#1D1D1F" font-size="13" font-weight="700">{html.escape(lot)}</text>',
                    f'<text x="{x:.1f}" y="{bottom+46}" text-anchor="middle" fill="#6E6E73" font-size="12">n={int(group["sample_count"])}</text>',
                ]
    parts += [
        f'<text x="{width/2:.1f}" y="{height-28}" text-anchor="middle" fill="#6E6E73" font-size="13">Box = Q1–Q3 · center line = median · whiskers = last value within 1.5×IQR · open markers = outliers</text>',
        '</svg>',
    ]
    return "".join(parts)


def lot_wafer_grade_counts_svg(vdd_group: dict[str, object],
                                width: int = 1600, height: int = 900) -> str:
    """Render Preferred / Monitor / Low Cell counts for every Lot/Wafer."""
    groups = list(vdd_group["groups"])
    grades = (
        ("preferred", "Preferred", "#34C759"),
        ("monitor", "Monitor", "#FFB000"),
        ("low", "Low", "#FF3B30"),
    )
    left, top, right, bottom = 115, 170, width-70, 735
    maximum = max(
        (int(group["grade_counts"][key]) for group in groups for key, *_ in grades),
        default=0)
    tick_step = max(1, math.ceil(max(maximum, 1) / 5))
    axis_max = max(
        tick_step, (math.ceil(max(maximum, 1) / tick_step) + 1) * tick_step)
    sy = lambda value: bottom - float(value) / axis_max * (bottom-top)
    slot = (right-left) / max(len(groups), 1)
    bar_width = min(48.0, slot*.23)
    gap = max(4.0, min(10.0, slot*.035))
    cluster_width = 3*bar_width + 2*gap
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="56" y="54" fill="#1D1D1F" font-size="31" font-weight="700">Lot/Wafer Preferred, Monitor and Low Counts</text>',
        f'<text x="56" y="86" fill="#6E6E73" font-size="16">Model VDD {float(vdd_group["vdd_v"]):.3f} V · {int(vdd_group["sample_count"])} Cells · labels show Cell count</text>',
        '<text x="56" y="111" fill="#6E6E73" font-size="14">Relative same-VDD grading; every Lot/Wafer is evaluated against the shared population thresholds.</text>',
    ]
    legend_x = right-385
    for index, (_key, label, color) in enumerate(grades):
        x = legend_x + index*135
        parts += [
            f'<rect x="{x}" y="128" width="18" height="18" rx="3" fill="{color}"/>',
            f'<text x="{x+27}" y="143" fill="#1D1D1F" font-size="14" font-weight="700">{label}</text>',
        ]
    for tick in range(0, axis_max+1, tick_step):
        y = sy(tick)
        parts += [
            f'<path d="M{left} {y:.1f} H{right}" stroke="#E5E5EA" stroke-width="1"/>',
            f'<text x="{left-15}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="13">{tick}</text>',
        ]
    parts.append(
        f'<path d="M{left} {top} V{bottom} H{right}" fill="none" stroke="#1D1D1F" stroke-width="2"/>')
    for group_index, group in enumerate(groups):
        center = left + slot*(group_index+.5)
        start = center-cluster_width/2
        for grade_index, (key, label, color) in enumerate(grades):
            value = int(group["grade_counts"][key])
            x = start + grade_index*(bar_width+gap)
            y = sy(value)
            bar_height = max(0.0, bottom-y)
            if value:
                parts.append(
                    f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_width:.1f}" height="{bar_height:.1f}" rx="4" fill="{color}" stroke="{color}" stroke-width="1"/>')
            parts.append(
                f'<text x="{x+bar_width/2:.1f}" y="{y-10:.1f}" text-anchor="middle" fill="#1D1D1F" font-size="16" font-weight="700">{value}</text>')
        lot = html.escape(str(group["lot_wafer"]))
        parts += [
            f'<text x="{center:.1f}" y="{bottom+34}" text-anchor="end" transform="rotate(-18 {center:.1f} {bottom+34})" fill="#1D1D1F" font-size="14" font-weight="700">{lot}</text>',
            f'<text x="{center:.1f}" y="{bottom+76}" text-anchor="middle" fill="#6E6E73" font-size="12">n={int(group["sample_count"])}</text>',
        ]
    parts += [
        f'<text x="31" y="{(top+bottom)/2:.1f}" transform="rotate(-90 31 {(top+bottom)/2:.1f})" text-anchor="middle" fill="#1D1D1F" font-size="17" font-weight="700">Cell count</text>',
        f'<text x="{(left+right)/2:.1f}" y="{height-32}" text-anchor="middle" fill="#6E6E73" font-size="13">Preferred = both CR and PR at/above P50 · Monitor = both at/above P25 · Low = remaining Cells</text>',
        '</svg>',
    ]
    return "".join(parts)


def write_lot_wafer_drive_advisor_outputs(
        analysis: dict[str, object], out_dir: str | os.PathLike[str],
        source_paths: Iterable[str | os.PathLike[str]]) -> Path:
    """Export the dedicated Lot/Wafer Advisor report, charts and CSV data."""
    out = Path(out_dir); images = out / "images"; images.mkdir(parents=True, exist_ok=True)
    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError(
            "PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc

    statistics_rows: list[dict[str, object]] = []
    cell_rows: list[dict[str, object]] = []
    batch_rows: list[dict[str, object]] = []
    report_sections: list[str] = []
    styles = analysis["styles"]
    for index, vdd_group in enumerate(analysis["vdds"], 1):
        prefix = f'{index:02d}_vdd_{float(vdd_group["vdd_v"]):.3f}'
        scatter_svg_name = f"{prefix}_lot_wafer_drive_scatter.svg"
        scatter_png_name = scatter_svg_name.replace(".svg", ".png")
        box_svg_name = f"{prefix}_lot_wafer_boxplots.svg"
        box_png_name = box_svg_name.replace(".svg", ".png")
        grade_svg_name = f"{prefix}_lot_wafer_grade_counts.svg"
        grade_png_name = grade_svg_name.replace(".svg", ".png")
        shmoo_svg_name = f"{prefix}_all_cell_drive_balance_shmoo.svg"
        shmoo_png_name = shmoo_svg_name.replace(".svg", ".png")
        scatter_svg = lot_wafer_drive_scatter_svg(vdd_group, styles)
        box_svg = lot_wafer_boxplot_svg(vdd_group, styles)
        grade_svg = lot_wafer_grade_counts_svg(vdd_group)
        shmoo_svg = estimate_vmin_ratio_shmoo_svg(vdd_group["shmoo"])
        for svg_name, png_name, content in (
                (shmoo_svg_name, shmoo_png_name, shmoo_svg),
                (scatter_svg_name, scatter_png_name, scatter_svg),
                (box_svg_name, box_png_name, box_svg),
                (grade_svg_name, grade_png_name, grade_svg)):
            svg_path = images / svg_name
            svg_path.write_text(content, encoding="utf-8")
            drawing = svg2rlg(str(svg_path))
            if drawing is None:
                raise RuntimeError(f"Could not render Lot/Wafer chart: {svg_name}")
            renderPM.drawToFile(
                drawing, str(images / png_name), fmt="PNG", dpi=180,
                backend="rlPyCairo")

        summary_body: list[str] = []
        batch_body: list[str] = []
        for group in vdd_group["groups"]:
            lot = str(group["lot_wafer"])
            grade = group["grade_counts"]
            read_stats = group["metrics"]["rsnm_mv"]
            write_stats = group["metrics"]["write_margin_mv"]
            balance_stats = group["metrics"]["balanced_drive_pct"]
            summary_body.append(
                f'<tr><td>{html.escape(lot)}</td><td>{int(group["sample_count"])}</td>'
                f'<td>{float(read_stats["median"]):.1f}</td><td>{float(read_stats["iqr"]):.1f}</td>'
                f'<td>{float(write_stats["median"]):.1f}</td><td>{float(write_stats["iqr"]):.1f}</td>'
                f'<td>{float(balance_stats["median"]):.1f}%</td>'
                f'<td>{float(group["median_cr"]):.3f}</td><td>{float(group["median_pr"]):.3f}</td>'
                f'<td>{int(grade["preferred"])} / {int(grade["monitor"])} / {int(grade["low"])}</td></tr>')
            for key, label, unit, _color in _LOT_WAFER_ADVISOR_METRICS:
                statistics_rows.append({
                    "vdd_v": vdd_group["vdd_v"], "lot_wafer": lot,
                    "sample_count": group["sample_count"], "metric": key,
                    "metric_label": label, "unit": unit,
                    **{field: group["metrics"][key][field]
                       for field in ("minimum", "q1", "median", "q3", "maximum",
                                     "whisker_low", "whisker_high", "mean", "iqr")},
                    "outlier_count": len(group["metrics"][key]["outliers"]),
                })
            for sample in group["samples"]:
                cell_rows.append({
                    "vdd_v": vdd_group["vdd_v"], "lot_wafer": lot,
                    "chip_id": sample["chip_id"], "rsnm_mv": sample["rsnm_mv"],
                    "write_margin_mv": sample["write_margin_mv"],
                    "balanced_drive_pct": sample["balanced_drive_pct"],
                    "cell_ratio_beta": sample["cell_ratio_beta"],
                    "pull_up_ratio_beta": sample["pull_up_ratio_beta"],
                    "cr_percentile": sample["cr_percentile"],
                    "pr_percentile": sample["pr_percentile"],
                    "wafer_grade": sample["wafer_grade"],
                })
            batch = group["batch_advice"]
            devices = {str(device["family"]): device for device in batch["devices"]}
            batch_body.append(
                f'<tr><td>{html.escape(lot)}</td><td>{int(batch["affected_count"])}</td>'
                f'<td>{float(batch["target"]["cr"]):.3f}</td><td>{float(batch["target"]["pr"]):.3f}</td>'
                f'<td>{float(devices["PD"]["drive_change_pct"]):+.1f}%</td>'
                f'<td>{float(devices["PG"]["drive_change_pct"]):+.1f}%</td>'
                f'<td>{float(devices["PU"]["drive_change_pct"]):+.1f}%</td>'
                f'<td>{float(batch["frozen_target_coverage_before_pct"]):.1f}% → '
                f'{float(batch["frozen_target_coverage_after_pct"]):.1f}%</td></tr>')
            for device in batch["devices"]:
                batch_rows.append({
                    "vdd_v": vdd_group["vdd_v"], "lot_wafer": lot,
                    "reference_sample_count": batch["reference_sample_count"],
                    "sample_count": batch["sample_count"],
                    "affected_count": batch["affected_count"],
                    "target_percentile": batch["target_percentile"],
                    "target_cr": batch["target"]["cr"], "target_pr": batch["target"]["pr"],
                    "coverage_before_pct": batch["frozen_target_coverage_before_pct"],
                    "coverage_after_pct": batch["frozen_target_coverage_after_pct"],
                    "family": device["family"],
                    "drive_multiplier": device["drive_multiplier"],
                    "drive_change_pct": device["drive_change_pct"],
                    "action": device["action"], "median_vt_v": device["median_vt_v"],
                    "median_idsat_current_ua": device["median_idsat_current_ua"],
                    "median_idsat_target_fixed_vt_ua": device["median_idsat_target_fixed_vt_ua"],
                    "limiting_chip_id": device["limiting_chip_id"],
                })
        report_sections.append(f'''<section>
<div class="section-head"><div><p class="eyebrow">MODEL VDD</p><h2>{float(vdd_group["vdd_v"]):.3f} V Lot/Wafer Comparison</h2></div><div class="count"><strong>{int(vdd_group["sample_count"])}</strong><span>Cells<br>{int(vdd_group["lot_count"])} groups</span></div></div>
<h3>All-Cell Preferred / Monitor / Low Shmoo</h3><p class="note">All Cells at this Model VDD share one dynamic CR/PR reference population. Preferred requires both CR and PR at or above P50; Monitor requires the weaker metric at P25–P50; Low means the weaker metric is below P25. The background regions inherit the grade of the nearest measured Cell and are relative screening, not silicon Pass/Fail.</p>
<div class="chart-grid"><figure><img src="images/{shmoo_png_name}" alt="All Cell Preferred Monitor Low drive balance shmoo"><figcaption><a href="images/{shmoo_svg_name}">SVG</a> · <a href="images/{shmoo_png_name}">PNG</a></figcaption></figure><figure><img src="images/{scatter_png_name}" alt="Lot Wafer CR PR scatter"><figcaption><a href="images/{scatter_svg_name}">SVG</a> · <a href="images/{scatter_png_name}">PNG</a></figcaption></figure><figure><img src="images/{box_png_name}" alt="Lot Wafer box plots"><figcaption><a href="images/{box_svg_name}">SVG</a> · <a href="images/{box_png_name}">PNG</a></figcaption></figure><figure><img src="images/{grade_png_name}" alt="Lot Wafer Preferred Monitor Low grade counts"><figcaption><a href="images/{grade_svg_name}">SVG</a> · <a href="images/{grade_png_name}">PNG</a></figcaption></figure></div>
<h3>Lot/Wafer Distribution Summary</h3><div class="table-wrap"><table><thead><tr><th>Lot/Wafer</th><th>n</th><th>Read median</th><th>Read IQR</th><th>Write median</th><th>Write IQR</th><th>Balanced median</th><th>Median CR</th><th>Median PR</th><th>P / M / L</th></tr></thead><tbody>{''.join(summary_body)}</tbody></table></div>
<h3>Drive-to-Preferred Batch Sensitivity</h3><p class="note">Each Lot/Wafer is compared with the same frozen P55 target from all Cells at this VDD. PG is held; PD addresses CR and PU addresses PR.</p><div class="table-wrap"><table><thead><tr><th>Lot/Wafer</th><th>Low + Monitor</th><th>P55 CR</th><th>P55 PR</th><th>PD MOS<sub>drive</sub></th><th>PG MOS<sub>drive</sub></th><th>PU MOS<sub>drive</sub></th><th>Frozen-target coverage</th></tr></thead><tbody>{''.join(batch_body)}</tbody></table></div>
</section>''')

    def write_csv(filename: str, rows: list[dict[str, object]]) -> None:
        if not rows:
            return
        with (out / filename).open("w", newline="", encoding="utf-8-sig") as stream:
            writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
            writer.writeheader(); writer.writerows(rows)

    write_csv("lot_wafer_distribution_statistics.csv", statistics_rows)
    write_csv("lot_wafer_cell_drive_scores.csv", cell_rows)
    write_csv("lot_wafer_batch_drive_advisor.csv", batch_rows)
    backup_dir = out / "imported_multi_chip_summaries"; backup_dir.mkdir(exist_ok=True)
    for index, raw_path in enumerate(source_paths, 1):
        source = Path(raw_path)
        shutil.copy2(source, backup_dir / f"{index:02d}_{source.name}")
    report = out / "lot_wafer_drive_advisor.html"
    report.write_text(f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>HV28 SRAM Lot/Wafer Drive Advisor</title><style>
*{{box-sizing:border-box}}body{{margin:0;padding:clamp(12px,2vw,32px);background:#f5f5f7;color:#1d1d1f;font-family:Calibri,"Microsoft JhengHei",Arial,sans-serif}}main{{max-width:1680px;margin:auto}}h1{{font-size:clamp(30px,4vw,48px);letter-spacing:-.025em;margin-bottom:8px}}section{{background:#fff;border-radius:18px;padding:clamp(18px,2vw,28px);margin:20px 0}}.note{{color:#6e6e73;line-height:1.5}}.eyebrow{{margin:0;color:#ff385c;font-size:12px;font-weight:700;letter-spacing:.08em}}.section-head{{display:flex;align-items:flex-end;justify-content:space-between;gap:18px}}.section-head h2{{margin:5px 0}}.count{{display:flex;align-items:center;gap:10px;padding:10px 15px;border-radius:14px;background:#f5f5f7}}.count strong{{font-size:28px}}.count span{{color:#6e6e73;font-size:12px}}.chart-grid{{display:grid;grid-template-columns:1fr;gap:18px;margin:20px 0}}figure{{margin:0}}img{{display:block;width:100%;height:auto;border:1px solid #e5e5ea;border-radius:14px}}figcaption{{padding-top:7px;text-align:right}}a{{color:#0066cc}}.table-wrap{{overflow-x:auto}}table{{width:100%;border-collapse:collapse;font-variant-numeric:tabular-nums}}th,td{{padding:11px 12px;border-bottom:1px solid #e5e5ea;text-align:right;white-space:nowrap}}th:first-child,td:first-child{{text-align:left}}th{{color:#6e6e73;font-size:13px}}sub{{font-family:"Times New Roman",serif;font-style:italic}}@media(max-width:800px){{.section-head{{align-items:flex-start;flex-direction:column}}}}
</style></head><body><main><h1>HV28 SRAM Lot/Wafer Drive Advisor</h1><p class="note">{html.escape(str(analysis["definition"]))} Same Lot/Wafer names are merged as one group, including rows imported from different files.</p>{''.join(report_sections)}<section><h2>Outputs</h2><p class="note"><code>lot_wafer_distribution_statistics.csv</code> · <code>lot_wafer_cell_drive_scores.csv</code> · <code>lot_wafer_batch_drive_advisor.csv</code> · source backups in <code>imported_multi_chip_summaries/</code>.</p><p class="note">Relative compact-model screening only. Uniform batch shifts preserve rank order if the percentile target is recalculated; use the frozen target for sensitivity comparison and confirm process actions with Device/PDK and measured WT.</p></section></main></body></html>''', encoding="utf-8")
    return report


def _estimate_zero_boundary(rows: list[dict[str, object]], key: str) -> dict[str, object] | None:
    """Locate margin=0 by interpolation, or extrapolate from the low-VDD end."""
    for low, high in zip(rows, rows[1:]):
        low_value, high_value = float(low[key]), float(high[key])
        if low_value <= 0 < high_value:
            if low_value == 0:
                estimate = float(low["vdd_v"])
            else:
                estimate = float(low["vdd_v"]) + (0.0 - low_value) * (
                    float(high["vdd_v"]) - float(low["vdd_v"])) / (high_value - low_value)
            return {"estimated_vdd_v": estimate, "low_vdd_v": low["vdd_v"],
                    "high_vdd_v": high["vdd_v"], "extrapolated": False,
                    "method": "linear interpolation of imported Multi-Cell margins"}
    # When every imported SNM/margin is still positive, the two lowest-VDD
    # points are closest to the missing eye-closure boundary.  Extend their
    # local linear slope down to margin=0, but accept only a physically
    # downward result between 0 V and the lowest measured VDD.
    if len(rows) >= 2:
        first, second = rows[0], rows[1]
        first_vdd, second_vdd = float(first["vdd_v"]), float(second["vdd_v"])
        first_value, second_value = float(first[key]), float(second[key])
        delta_vdd = second_vdd - first_vdd
        if first_value > 0 and second_value > 0 and delta_vdd > 0:
            slope_mv_per_v = (second_value - first_value) / delta_vdd
            if slope_mv_per_v > 0:
                estimate = first_vdd - first_value / slope_mv_per_v
                if 0 <= estimate < first_vdd:
                    return {
                        "estimated_vdd_v": estimate,
                        "low_vdd_v": first_vdd,
                        "high_vdd_v": second_vdd,
                        "low_margin_mv": first_value,
                        "high_margin_mv": second_value,
                        "slope_mv_per_v": slope_mv_per_v,
                        "extrapolated": True,
                        "method": "linear extrapolation from the two lowest imported VDD margins",
                    }
    return None


def analyze_estimate_vmin_curves(summary_rows: list[dict[str, object]],
                                 force_shmoo_only: bool = False,
                                 include_shmoo: bool = True) -> dict:
    """Build RSNM, WSNM and Write-Margin VDD trends from Multi-Cell summaries."""
    if not summary_rows:
        raise ValueError("At least one Multi-Cell VDD summary point is required")
    ordered = sorted(summary_rows, key=lambda row: float(row["vdd_v"]))
    curves: dict[str, dict] = {}
    for key, short_label, label, color in _ESTIMATE_VMIN_METRICS:
        # Older summary files and hand-built datasets do not contain the WL
        # metric.  In this compact symmetric-cell estimate, it is the same
        # PG-overdrive write margin until dedicated WL-driver data is added.
        curve_source_rows = [{**item, key: item.get(key, item["write_margin_mv"])}
                             for item in ordered]
        rows = [{"vdd_v": float(item["vdd_v"]), "margin_mv": float(item[key]),
                 "chip_id": item.get(f"{key}_chip_id", item.get("write_margin_mv_chip_id", "Unknown")),
                 "lot_wafer": item.get(f"{key}_lot_wafer", item.get("write_margin_mv_lot_wafer", "Wafer")),
                 "sample_count": item["sample_count"], "valid": float(item[key]) > 0}
                for item in curve_source_rows]
        curves[key] = {"key": key, "short_label": short_label, "label": label,
                       "color": color, "rows": rows,
                       "eye_closure": _estimate_zero_boundary(curve_source_rows, key)}
    shmoo_only = bool(include_shmoo) and (
        bool(force_shmoo_only) or len(ordered) == 1)
    mode = ("shmoo_only" if shmoo_only else
            "single_vdd" if len(ordered) == 1 else "estimate_vmin")
    return {"rows": ordered, "curves": curves,
            "ratio_shmoos": (_build_estimate_vmin_ratio_shmoos(ordered)
                              if include_shmoo else []),
            "shmoo_enabled": bool(include_shmoo),
            "mode": mode,
            "definition": (
                "Single-file or single-VDD input: only same-VDD CR/PR Shmoo screening is produced."
                if shmoo_only else
                "Each VDD point is the minimum per-cell margin in the imported Multi-Cell summary data.")}


_COMPARISON_COLORS = ("#007AFF", "#AF52DE", "#00A844", "#FF9500",
                      "#FF375F", "#5E5CE6", "#00A6A6", "#8E5A2B")


def read_estimate_vmin_combined_files(
        paths: Iterable[str | os.PathLike[str]]) -> list[dict[str, object]]:
    """Read one or more Multi-VDD run folders or legacy combined-summary CSVs.

    A folder produced by :func:`process_multi_vdd_6t_excel` is resolved only
    inside that selected run: its per-VDD ``multi_chip_snm_summary.csv`` files
    are preferred, with ``estimate_vmin/multi_chip_snm_summary_combined.csv``
    retained as a compatibility fallback. One source becomes one curve set.
    """
    selected = [Path(item).expanduser() for item in paths]
    if not selected:
        raise ValueError("Select at least one Multi-VDD output folder or combined summary CSV")
    datasets: list[dict[str, object]] = []
    used_labels: set[str] = set()
    required = {"vdd_v", "rsnm_mv", "wsnm_mv", "write_margin_mv"}
    for index, source_path in enumerate(selected):
        path = source_path.resolve()
        per_vdd_files: list[Path] = []
        combined_path: Path | None = None
        if path.is_dir():
            search_roots = [path]
            if path.name == "estimate_vmin":
                search_roots.append(path.parent)
            for root in search_roots:
                per_vdd_files.extend(sorted(
                    root.glob("multi_cell_by_vdd/Model_VDD_*/multi_chip_snm_summary.csv")))
                if root.name == "multi_cell_by_vdd":
                    per_vdd_files.extend(sorted(
                        root.glob("Model_VDD_*/multi_chip_snm_summary.csv")))
                if root.name.startswith("Model_VDD_") and (
                        root / "multi_chip_snm_summary.csv").is_file():
                    per_vdd_files.append(root / "multi_chip_snm_summary.csv")
            per_vdd_files = list(dict.fromkeys(item.resolve() for item in per_vdd_files))
            fallback_candidates = [
                path / "estimate_vmin" / "multi_chip_snm_summary_combined.csv",
                path / "multi_chip_snm_summary_combined.csv",
            ]
            if path.name == "multi_cell_by_vdd":
                fallback_candidates.append(
                    path.parent / "estimate_vmin" / "multi_chip_snm_summary_combined.csv")
            for candidate in fallback_candidates:
                if candidate.is_file():
                    combined_path = candidate.resolve()
                    break
            if not per_vdd_files and combined_path is None:
                raise FileNotFoundError(
                    f"No Multi-VDD summaries were found inside the selected folder: {path}")
        elif path.is_file():
            if path.suffix.lower() != ".csv":
                raise ValueError(f"{path.name} is not a CSV combined summary")
            combined_path = path
        else:
            raise FileNotFoundError(f"Estimate Vmin comparison source was not found: {path}")

        rows: list[dict[str, object]]
        lot_names: list[str] = []
        if per_vdd_files:
            rows = read_multi_chip_snm_summary(per_vdd_files)
            for row in rows:
                for key, *_ in _ESTIMATE_VMIN_METRICS:
                    value = str(row.get(f"{key}_lot_wafer", "")).strip()
                    if value:
                        lot_names.append(value)
            source_files = [str(item) for item in per_vdd_files]
        else:
            assert combined_path is not None
            rows = []
            with combined_path.open(newline="", encoding="utf-8-sig") as stream:
                reader = csv.DictReader(stream)
                if not reader.fieldnames or not required.issubset(reader.fieldnames):
                    raise ValueError(
                        f"{combined_path.name} is not a valid "
                        "multi_chip_snm_summary_combined.csv export")
                for number, raw in enumerate(reader, 2):
                    try:
                        vdd = float(raw["vdd_v"])
                        if not 0 < vdd <= SNM_PLOT_AXIS_MAX_V:
                            raise ValueError
                        row = {
                            "vdd_v": vdd,
                            "sample_count": int(float(raw.get("sample_count") or 1)),
                        }
                        for key, *_ in _ESTIMATE_VMIN_METRICS:
                            row[key] = float(raw[key])
                            lot_key = f"{key}_lot_wafer"
                            chip_key = f"{key}_chip_id"
                            if raw.get(lot_key):
                                row[lot_key] = str(raw[lot_key]).strip()
                                lot_names.append(str(raw[lot_key]).strip())
                            if raw.get(chip_key):
                                row[chip_key] = str(raw[chip_key]).strip()
                        rows.append(row)
                    except (TypeError, ValueError) as exc:
                        raise ValueError(
                            f"{combined_path.name} row {number}: invalid curve value") from exc
            source_files = [str(combined_path)]
        rows.sort(key=lambda item: float(item["vdd_v"]))
        if len(rows) < 2:
            raise ValueError(f"{path.name} has fewer than two Model VDD points")
        unique_lots = sorted({item for item in lot_names if item})
        label = unique_lots[0] if len(unique_lots) == 1 else (
            path.name if path.is_dir() else path.parent.name)
        if not label:
            label = path.stem
        if label in used_labels:
            label = f"{label} · {path.parent.name}"
        used_labels.add(label)
        datasets.append({
            "lot_wafer": label,
            "rows": rows,
            "color": _COMPARISON_COLORS[index % len(_COMPARISON_COLORS)],
            "sources": source_files,
        })
    return datasets


def estimate_vmin_combined_comparison_svg(datasets: list[dict[str, object]],
                                width: int = 1500, height: int = 720,
                                transparent_background: bool = False) -> str:
    """Overlay combined-summary files in paired SNM and BL-margin panels."""
    groups = (
        ("Read / Write SNM",
         (("rsnm_mv", "R", ""), ("wsnm_mv", "W", "8 5")), "SNM (mV)"),
        ("BL Write Margin", (("write_margin_mv", "BL", ""),), "Vtrip (mV)"),
    )
    left, right, bottom, panel_gap = 92, 48, 84, 72
    legend_items = [(str(item["lot_wafer"]),
                     max(210.0, 78.0 + len(str(item["lot_wafer"])) * 9.5),
                     str(item["color"])) for item in datasets]
    legend_rows: list[list[tuple[str, float, str]]] = [[]]
    current_width = 0.0
    for item in legend_items:
        if legend_rows[-1] and current_width + item[1] > width - 700:
            legend_rows.append([]); current_width = 0.0
        legend_rows[-1].append(item); current_width += item[1]
    top = 145 + len(legend_rows) * 27
    plot_w = (width - left - right - panel_gap) / len(groups)
    panel_h = height - top - bottom
    measured_vdds = sorted({float(row["vdd_v"])
                            for dataset in datasets for row in dataset["rows"]})
    if len(measured_vdds) < 2 or measured_vdds[-1] <= measured_vdds[0]:
        raise ValueError("Comparison View requires at least two distinct Model VDD points")
    vdd_min, vdd_max = measured_vdds[0], measured_vdds[-1]
    vdd_span = vdd_max - vdd_min
    parts = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">']
    if not transparent_background:
        parts.append('<rect width="100%" height="100%" fill="#FFFFFF"/>')
    panel_fill = "none" if transparent_background else "#FFFFFF"
    parts += ['<text x="56" y="46" fill="#1D1D1F" font-size="34" font-weight="700">Estimate Vmin Curves - Comparison View</text>',
             f'<text x="56" y="73" fill="#6E6E73" font-size="16">{len(datasets)} Multi-VDD source(s) · X range {vdd_min:.3f}–{vdd_max:.3f} V</text>',
             '<path d="M56 108h30" stroke="#3A3A3C" stroke-width="4"/><text x="96" y="113" fill="#3A3A3C" font-size="14">Read SNM / BL Write Margin</text>',
             '<path d="M330 108h30" stroke="#3A3A3C" stroke-width="4" stroke-dasharray="8 5"/><text x="370" y="113" fill="#3A3A3C" font-size="14">Write SNM</text>']
    for row_index, legend_row in enumerate(legend_rows):
        legend_x = 650 if row_index == 0 else 56
        legend_y = 113 + row_index * 27
        for label_text, item_width, color in legend_row:
            label = html.escape(label_text)
            parts += [f'<path d="M{legend_x} {legend_y-5}h30" stroke="{color}" stroke-width="4"/>',
                      f'<text x="{legend_x+38}" y="{legend_y}" fill="#30343B" font-size="14" font-weight="700">{label}</text>']
            legend_x += item_width
    for panel_index, (title, series_specs, y_label) in enumerate(groups):
        panel_left = left + panel_index * (plot_w + panel_gap)
        panel_top = top
        panel_bottom = panel_top + panel_h
        maximum = max(float(row[key]) for dataset in datasets
                      for key, _prefix, _dash in series_specs for row in dataset["rows"])
        y_max = max(50.0, math.ceil(maximum / 50.0) * 50.0)

        def xy(vdd: float, value: float) -> tuple[float, float]:
            return (panel_left + (vdd - vdd_min) / vdd_span * plot_w,
                    panel_top + (1 - value / y_max) * panel_h)

        parts += [f'<rect x="{panel_left}" y="{panel_top}" width="{plot_w}" height="{panel_h:.1f}" fill="{panel_fill}" stroke="#D8DDE3"/>',
                  f'<text x="{panel_left:.1f}" y="{panel_top-16:.1f}" fill="#1D1D1F" font-size="20" font-weight="700">{title}</text>']
        for step in range(5):
            value = y_max * step / 4
            _x, y = xy(vdd_min, value)
            parts += [f'<path d="M{panel_left} {y:.1f}H{panel_left+plot_w}" stroke="#E5E5EA"/>',
                      f'<text x="{panel_left-10}" y="{y+4:.1f}" text-anchor="end" fill="#6E6E73" font-size="12">{value:.0f}</text>']
        for vdd_step in range(6):
            voltage = vdd_min + vdd_span * vdd_step / 5
            x, _y = xy(voltage, 0)
            parts.append(f'<path d="M{x:.1f} {panel_top}V{panel_bottom}" stroke="#F1F1F4"/>')
        measured_voltages: set[float] = set()
        for dataset in datasets:
            for key, prefix, dash in series_specs:
                points = [xy(float(row["vdd_v"]), float(row[key])) for row in dataset["rows"]]
                dash_attr = f' stroke-dasharray="{dash}"' if dash else ""
                parts.append(f'<polyline points="{" ".join(f"{x:.1f},{y:.1f}" for x, y in points)}" fill="none" stroke="{dataset["color"]}" stroke-width="3"{dash_attr}/>')
                for row, (x, y) in zip(dataset["rows"], points):
                    voltage = float(row["vdd_v"])
                    measured_voltages.add(voltage)
                    parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="#FFF" stroke="{dataset["color"]}" stroke-width="2"/>')
        for voltage in sorted(measured_voltages):
            x, _y = xy(voltage, 0)
            parts += [
                f'<text class="vertical-vdd-label" data-vdd="{voltage:.4f}" '
                f'x="{x:.1f}" y="{panel_bottom+24:.1f}" text-anchor="middle" '
                f'fill="#0062CC" font-size="12" font-weight="700">{voltage:.2f} V</text>',
            ]
        center_y = panel_top + panel_h / 2
        axis_x = panel_left - 54
        parts += [f'<text x="{axis_x:.1f}" y="{center_y:.1f}" transform="rotate(-90 {axis_x:.1f} {center_y:.1f})" text-anchor="middle" fill="#1D1D1F" font-size="16" font-weight="700">{y_label}</text>',
                  f'<text x="{panel_left+plot_w/2:.1f}" y="{height-18}" text-anchor="middle" fill="#1D1D1F" font-size="17" font-weight="700">Model VDD (V)</text>']
    parts.append('</svg>')
    return "".join(parts)


_COMPARISON_WORST_METRICS = (
    ("rsnm_mv", "Read SNM"),
    ("wsnm_mv", "Write SNM"),
    ("write_margin_mv", "BL Write Margin"),
)
_SIX_MOS_NAMES = ("pul", "pur", "pgl", "pgr", "pdl", "pdr")


def _comparison_row_samples(dataset: dict[str, object],
                            row: dict[str, object]) -> list[dict[str, object]]:
    """Return full per-cell samples, with a reduced-row compatibility fallback."""
    samples = [dict(item) for item in row.get("samples", [])]
    if samples:
        return samples
    fallback = dict(row)
    fallback.setdefault("lot_wafer", dataset.get("lot_wafer", "Wafer"))
    fallback.setdefault("chip_id", "Unavailable")
    return [fallback]


def _comparison_sample_value(sample: dict[str, object], key: str) -> float | None:
    """Read one metric, falling back from side-specific to family-average WAT."""
    value = sample.get(key)
    if value not in (None, ""):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if key[:3] in _SIX_MOS_NAMES:
        family = key[:2]
        fallback_key = f'{family}_{key.split("_", 1)[1]}'
        value = sample.get(fallback_key)
        if value not in (None, ""):
            try:
                return float(value)
            except (TypeError, ValueError):
                return None
    return None


def estimate_vmin_comparison_worst_rows(
        datasets: list[dict[str, object]]) -> list[dict[str, object]]:
    """Select the limiting Cell for every source, VDD and comparison metric."""
    records: list[dict[str, object]] = []
    for dataset in datasets:
        for row in dataset["rows"]:
            samples = _comparison_row_samples(dataset, row)
            for metric_key, metric_label in _COMPARISON_WORST_METRICS:
                valid = [(sample, _comparison_sample_value(sample, metric_key))
                         for sample in samples]
                valid = [(sample, value) for sample, value in valid
                         if value is not None]
                if not valid:
                    continue
                sample, margin = min(valid, key=lambda item: item[1])
                record: dict[str, object] = {
                    "source": str(dataset["lot_wafer"]),
                    "vdd_v": float(row["vdd_v"]),
                    "metric": metric_label,
                    "metric_key": metric_key,
                    "margin_mv": float(margin),
                    "lot_wafer": str(sample.get(
                        "lot_wafer", dataset["lot_wafer"])),
                    "chip_id": str(sample.get(
                        "chip_id", row.get(f"{metric_key}_chip_id", "Unavailable"))),
                    "raw_6t_available": all(
                        sample.get(f"{device}_{quantity}") not in (None, "")
                        for device in _SIX_MOS_NAMES
                        for quantity in ("vt_v", "idsat_ua")),
                }
                for key in ("cell_ratio_beta", "pull_up_ratio_beta",
                            "pu_vt_v", "pu_idsat_ua", "pg_vt_v", "pg_idsat_ua",
                            "pd_vt_v", "pd_idsat_ua"):
                    record[key] = _comparison_sample_value(sample, key)
                for device in _SIX_MOS_NAMES:
                    record[f"{device}_vt_v"] = _comparison_sample_value(
                        sample, f"{device}_vt_v")
                    record[f"{device}_idsat_ua"] = _comparison_sample_value(
                        sample, f"{device}_idsat_ua")
                records.append(record)
    return records


def estimate_vmin_distribution_boxplots_svg(
        datasets: list[dict[str, object]],
        metrics: tuple[tuple[str, str, str, str], ...],
        title: str, width: int = 1500, panel_height: int = 350) -> str:
    """Render VDD-grouped Tukey box plots from per-cell comparison samples."""
    vdds = sorted({float(row["vdd_v"])
                   for dataset in datasets for row in dataset["rows"]})
    column_count = 2
    row_count = max(1, math.ceil(len(metrics) / column_count))
    height = 105 + row_count * panel_height + 45
    left, right, gap_x = 95.0, 45.0, 75.0
    panel_w = (width - left - right - gap_x) / column_count
    plot_top_offset, plot_bottom_offset = 48.0, 62.0
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" '
        'style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        f'<text x="56" y="46" fill="#1D1D1F" font-size="34" font-weight="700">'
        f'{html.escape(title)}</text>',
        '<text x="56" y="74" fill="#6E6E73" font-size="15">'
        'All imported sources are pooled within each VDD · Box = Q1–Q3 · median line · Tukey 1.5×IQR whiskers · circles = outliers</text>',
    ]
    for metric_index, (key, label, unit, color) in enumerate(metrics):
        panel_column = metric_index % column_count
        panel_row = metric_index // column_count
        panel_left = left + panel_column * (panel_w + gap_x)
        panel_top = 105 + panel_row * panel_height
        plot_top = panel_top + plot_top_offset
        plot_bottom = panel_top + panel_height - plot_bottom_offset
        plot_h = plot_bottom - plot_top
        grouped: dict[float, list[float]] = {vdd: [] for vdd in vdds}
        for dataset in datasets:
            for row in dataset["rows"]:
                vdd = float(row["vdd_v"])
                for sample in _comparison_row_samples(dataset, row):
                    value = _comparison_sample_value(sample, key)
                    if value is not None and math.isfinite(value):
                        grouped[vdd].append(value)
        all_values = [value for values in grouped.values() for value in values]
        parts += [
            f'<text x="{panel_left:.1f}" y="{panel_top+22:.1f}" fill="#1D1D1F" '
            f'font-size="20" font-weight="700">{html.escape(label)}</text>',
        ]
        if not all_values:
            parts.append(
                f'<text x="{panel_left+panel_w/2:.1f}" y="{plot_top+plot_h/2:.1f}" '
                'text-anchor="middle" fill="#8E8E93" font-size="16">Raw per-cell data unavailable</text>')
            continue
        value_min, value_max = min(all_values), max(all_values)
        span = max(value_max - value_min, abs(value_max) * .08, 1e-6)
        y_min = min(0.0, value_min - span * .10)
        y_max = value_max + span * .12
        if y_max <= y_min:
            y_max = y_min + 1.0

        def y_of(value: float) -> float:
            return plot_top + (y_max - value) / (y_max - y_min) * plot_h

        parts.append(
            f'<rect x="{panel_left:.1f}" y="{plot_top:.1f}" width="{panel_w:.1f}" '
            f'height="{plot_h:.1f}" fill="#FFFFFF" stroke="#D8DDE3"/>')
        for step in range(5):
            value = y_min + (y_max - y_min) * step / 4
            y = y_of(value)
            parts += [
                f'<path d="M{panel_left:.1f} {y:.1f}H{panel_left+panel_w:.1f}" '
                'stroke="#E5E5EA"/>',
                f'<text x="{panel_left-10:.1f}" y="{y+4:.1f}" text-anchor="end" '
                f'fill="#6E6E73" font-size="12">{value:.2f}</text>',
            ]
        category_w = panel_w / max(len(vdds), 1)
        box_w = min(38.0, category_w * .48)
        for vdd_index, vdd in enumerate(vdds):
            x = panel_left + category_w * (vdd_index + .5)
            values = grouped[vdd]
            parts.append(
                f'<text x="{x:.1f}" y="{plot_bottom+21:.1f}" text-anchor="middle" '
                f'fill="#3A3A3C" font-size="12" font-weight="700">{vdd:.2f} V</text>')
            if not values:
                continue
            stats = _tukey_box(values)
            y_low = y_of(float(stats["whisker_low"]))
            y_high = y_of(float(stats["whisker_high"]))
            y_q1 = y_of(float(stats["q1"]))
            y_q3 = y_of(float(stats["q3"]))
            y_median = y_of(float(stats["median"]))
            median_digits = 3 if key in {
                "cell_ratio_beta", "pull_up_ratio_beta"} else 1
            median_text = f'{float(stats["median"]):.{median_digits}f}'
            parts += [
                f'<path d="M{x:.1f} {y_high:.1f}V{y_low:.1f}" stroke="#4A4A4A" stroke-width="1.5"/>',
                f'<path d="M{x-box_w*.28:.1f} {y_high:.1f}H{x+box_w*.28:.1f} '
                f'M{x-box_w*.28:.1f} {y_low:.1f}H{x+box_w*.28:.1f}" '
                'stroke="#4A4A4A" stroke-width="1.5"/>',
                f'<rect x="{x-box_w/2:.1f}" y="{min(y_q1,y_q3):.1f}" width="{box_w:.1f}" '
                f'height="{max(abs(y_q1-y_q3),1.5):.1f}" fill="{color}" fill-opacity=".18" '
                f'stroke="{color}" stroke-width="2"/>',
                f'<path d="M{x-box_w/2:.1f} {y_median:.1f}H{x+box_w/2:.1f}" '
                f'stroke="{color}" stroke-width="3"/>',
                f'<text class="boxplot-median-label" x="{x:.1f}" '
                f'y="{y_median-6:.1f}" text-anchor="middle" fill="#1D1D1F" '
                f'font-size="11" font-weight="700" '
                f'style="paint-order:stroke;stroke:#FFFFFF;stroke-width:4">'
                f'{median_text}</text>',
                f'<text x="{x:.1f}" y="{plot_bottom+39:.1f}" text-anchor="middle" '
                f'fill="#8E8E93" font-size="10">N={len(values)}</text>',
            ]
            for outlier_index, value in enumerate(stats["outliers"]):
                jitter = ((outlier_index % 5) - 2) * min(2.5, box_w / 12)
                parts.append(
                    f'<circle cx="{x+jitter:.1f}" cy="{y_of(float(value)):.1f}" r="2.6" '
                    f'fill="#FFFFFF" stroke="{color}" stroke-width="1.4"/>')
        axis_x = panel_left - 63
        center_y = plot_top + plot_h / 2
        parts += [
            f'<text x="{axis_x:.1f}" y="{center_y:.1f}" '
            f'transform="rotate(-90 {axis_x:.1f} {center_y:.1f})" text-anchor="middle" '
            f'fill="#1D1D1F" font-size="15" font-weight="700">{html.escape(unit)}</text>',
            f'<text x="{panel_left+panel_w/2:.1f}" y="{panel_top+panel_height-5:.1f}" '
            'text-anchor="middle" fill="#1D1D1F" font-size="15" '
            'font-weight="700">Model VDD (V)</text>',
        ]
    parts.append('</svg>')
    return "".join(parts)


def write_estimate_vmin_combined_comparison_outputs(datasets: list[dict[str, object]],
                                          out_dir: str | os.PathLike[str]) -> Path:
    """Export the Multi-VDD SNM / BL-margin comparison as HTML, SVG, PNG and CSV."""
    out = Path(out_dir); image_dir = out / "images"; image_dir.mkdir(parents=True, exist_ok=True)
    svg_path = image_dir / "01_estimate_vmin_combined_comparison.svg"
    png_path = image_dir / "01_estimate_vmin_combined_comparison.png"
    transparent_svg_path = image_dir / "01_estimate_vmin_combined_comparison_transparent.svg"
    transparent_png_path = image_dir / "01_estimate_vmin_combined_comparison_transparent.png"
    distribution_svg_path = image_dir / "02_vdd_margin_ratio_boxplots.svg"
    distribution_png_path = image_dir / "02_vdd_margin_ratio_boxplots.png"
    idsat_svg_path = image_dir / "03_vdd_6t_idsat_boxplots.svg"
    idsat_png_path = image_dir / "03_vdd_6t_idsat_boxplots.png"
    svg_path.write_text(estimate_vmin_combined_comparison_svg(datasets), encoding="utf-8")
    transparent_svg_path.write_text(
        estimate_vmin_combined_comparison_svg(datasets, transparent_background=True),
        encoding="utf-8")
    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError("PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc
    drawing = svg2rlg(str(svg_path))
    if drawing is None:
        raise RuntimeError("Could not render multi-Lot/Wafer comparison")
    renderPM.drawToFile(drawing, str(png_path), fmt="PNG", dpi=180, backend="rlPyCairo")
    transparent_drawing = svg2rlg(str(transparent_svg_path))
    if transparent_drawing is None:
        raise RuntimeError("Could not render transparent combined comparison")
    renderPM.drawToFile(
        transparent_drawing, str(transparent_png_path), fmt="PNG", dpi=180,
        bg=None, backend="rlPyCairo", backendFmt="RGBA")
    distribution_metrics = (
        ("rsnm_mv", "Read SNM distribution", "RSNM (mV)", "#007AFF"),
        ("write_margin_mv", "BL Write Margin distribution", "Vtrip (mV)", "#007AFF"),
        ("cell_ratio_beta", "Cell Ratio distribution", "CR = MOSdrive(PD) / MOSdrive(PG)", "#007AFF"),
        ("pull_up_ratio_beta", "Pull-up Ratio distribution", "PR = MOSdrive(PG) / MOSdrive(PU)", "#007AFF"),
    )
    distribution_svg_path.write_text(
        estimate_vmin_distribution_boxplots_svg(
            datasets, distribution_metrics,
            "Per-Cell Margin and Drive-Ratio Distributions by Model VDD"),
        encoding="utf-8")
    distribution_drawing = svg2rlg(str(distribution_svg_path))
    if distribution_drawing is None:
        raise RuntimeError("Could not render VDD margin / ratio box plots")
    renderPM.drawToFile(
        distribution_drawing, str(distribution_png_path), fmt="PNG", dpi=180,
        backend="rlPyCairo")
    idsat_metrics = (
        ("pul_idsat_ua", "PUL Idsat", "Idsat (µA)", "#D92D55"),
        ("pur_idsat_ua", "PUR Idsat", "Idsat (µA)", "#D92D55"),
        ("pgl_idsat_ua", "PGL Idsat", "Idsat (µA)", "#008F5D"),
        ("pgr_idsat_ua", "PGR Idsat", "Idsat (µA)", "#008F5D"),
        ("pdl_idsat_ua", "PDL Idsat", "Idsat (µA)", "#007AFF"),
        ("pdr_idsat_ua", "PDR Idsat", "Idsat (µA)", "#007AFF"),
    )
    idsat_svg_path.write_text(
        estimate_vmin_distribution_boxplots_svg(
            datasets, idsat_metrics,
            "Six-MOS Idsat Distributions by Model VDD", panel_height=315),
        encoding="utf-8")
    idsat_drawing = svg2rlg(str(idsat_svg_path))
    if idsat_drawing is None:
        raise RuntimeError("Could not render six-MOS Idsat box plots")
    renderPM.drawToFile(
        idsat_drawing, str(idsat_png_path), fmt="PNG", dpi=180,
        backend="rlPyCairo")
    statistic_fields = [
        "metric", "label", "unit", "vdd_v", "count", "minimum",
        "q1", "median", "q3", "maximum", "whisker_low", "whisker_high",
        "mean", "outlier_count",
    ]
    with (out / "estimate_vmin_boxplot_statistics.csv").open(
            "w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=statistic_fields)
        writer.writeheader()
        for key, label, unit, _color in distribution_metrics + idsat_metrics:
            for vdd in sorted({float(row["vdd_v"])
                               for dataset in datasets for row in dataset["rows"]}):
                values = []
                for dataset in datasets:
                    for row in dataset["rows"]:
                        if abs(float(row["vdd_v"]) - vdd) > 1e-12:
                            continue
                        for sample in _comparison_row_samples(dataset, row):
                            value = _comparison_sample_value(sample, key)
                            if value is not None and math.isfinite(value):
                                values.append(value)
                if not values:
                    continue
                stats = _tukey_box(values)
                writer.writerow({
                    "metric": key, "label": label, "unit": unit,
                    "vdd_v": vdd, "count": stats["count"],
                    "minimum": stats["minimum"], "q1": stats["q1"],
                    "median": stats["median"], "q3": stats["q3"],
                    "maximum": stats["maximum"],
                    "whisker_low": stats["whisker_low"],
                    "whisker_high": stats["whisker_high"],
                    "mean": stats["mean"],
                    "outlier_count": len(stats["outliers"]),
                })
    fields = ["lot_wafer", "vdd_v", "sample_count", "rsnm_mv", "wsnm_mv", "write_margin_mv", "source_files"]
    with (out / "estimate_vmin_combined_comparison.csv").open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields); writer.writeheader()
        for dataset in datasets:
            for row in dataset["rows"]:
                writer.writerow({"lot_wafer": dataset["lot_wafer"], "vdd_v": row["vdd_v"],
                                 "sample_count": row["sample_count"], "rsnm_mv": row["rsnm_mv"],
                                 "wsnm_mv": row["wsnm_mv"], "write_margin_mv": row["write_margin_mv"],
                                 "source_files": " | ".join(dataset["sources"])})
    worst_rows = estimate_vmin_comparison_worst_rows(datasets)
    worst_fields = [
        "source", "vdd_v", "metric", "margin_mv", "lot_wafer", "chip_id",
        "cell_ratio_beta", "pull_up_ratio_beta", "raw_6t_available",
    ] + [f"{device}_{quantity}" for device in _SIX_MOS_NAMES
         for quantity in ("vt_v", "idsat_ua")]
    with (out / "estimate_vmin_worst_cell_details.csv").open(
            "w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=worst_fields)
        writer.writeheader()
        writer.writerows({key: row.get(key, "") for key in worst_fields}
                         for row in worst_rows)
    sample_fields = [
        "source", "vdd_v", "lot_wafer", "chip_id", "rsnm_mv", "wsnm_mv",
        "write_margin_mv", "cell_ratio_beta", "pull_up_ratio_beta",
        "pu_vt_v", "pu_idsat_ua", "pg_vt_v", "pg_idsat_ua",
        "pd_vt_v", "pd_idsat_ua",
    ] + [f"{device}_{quantity}" for device in _SIX_MOS_NAMES
         for quantity in ("vt_v", "idsat_ua")]
    with (out / "estimate_vmin_comparison_per_cell_data.csv").open(
            "w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=sample_fields)
        writer.writeheader()
        for dataset in datasets:
            for row in dataset["rows"]:
                for sample in _comparison_row_samples(dataset, row):
                    record = {"source": dataset["lot_wafer"],
                              "vdd_v": row["vdd_v"], **sample}
                    writer.writerow({key: record.get(key, "")
                                     for key in sample_fields})
    def cell_text(value: object, digits: int = 4) -> str:
        if value in (None, ""):
            return "—"
        try:
            return f"{float(value):.{digits}f}"
        except (TypeError, ValueError):
            return html.escape(str(value))

    worst_table_rows = "".join(
        '<tr>'
        f'<td>{html.escape(str(row["source"]))}</td>'
        f'<td>{float(row["vdd_v"]):.3f}</td>'
        f'<td>{html.escape(str(row["metric"]))}</td>'
        f'<td>{float(row["margin_mv"]):.2f}</td>'
        f'<td>{html.escape(str(row["lot_wafer"]))}</td>'
        f'<td>{html.escape(str(row["chip_id"]))}</td>'
        f'<td>{cell_text(row.get("cell_ratio_beta"), 3)}</td>'
        f'<td>{cell_text(row.get("pull_up_ratio_beta"), 3)}</td>'
        + ''.join(
            f'<td>{cell_text(row.get(f"{device}_vt_v"))} / '
            f'{cell_text(row.get(f"{device}_idsat_ua"), 3)}</td>'
            for device in _SIX_MOS_NAMES)
        + f'<td>{"6T raw" if row.get("raw_6t_available") else "family-average fallback"}</td>'
        '</tr>' for row in worst_rows)
    device_headers = "".join(
        f'<th>{device.upper()}<br><small>Vt (V) / Idsat (µA)</small></th>'
        for device in _SIX_MOS_NAMES)
    report = out / "estimate_vmin_combined_comparison.html"
    report.write_text(f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Estimate Vmin Combined Comparison</title><style>*{{box-sizing:border-box}}body{{margin:0;padding:clamp(10px,2vw,30px);font-family:Calibri,"Microsoft JhengHei",Arial,sans-serif;background:#f5f5f7;color:#1d1d1f}}main{{max-width:1750px;margin:auto}}h1{{margin-bottom:4px}}section{{background:#fff;padding:22px;border-radius:16px;margin:18px 0}}img{{display:block;width:100%;height:auto}}.note{{color:#6e6e73}}.downloads{{margin:18px 0}}.table-wrap{{overflow-x:auto}}table{{border-collapse:collapse;width:100%;min-width:1650px;font-variant-numeric:tabular-nums}}th,td{{padding:9px 10px;border-bottom:1px solid #e5e5ea;text-align:right;white-space:nowrap}}th{{position:sticky;top:0;background:#f8f8fa;color:#3a3a3c}}th:first-child,td:first-child,th:nth-child(3),td:nth-child(3),th:nth-child(5),td:nth-child(5),th:nth-child(6),td:nth-child(6){{text-align:left}}small{{color:#6e6e73;font-weight:400}}</style></head><body><main><h1>Estimate Vmin Curves - Comparison View</h1><p class="note">Compared summaries: {" · ".join(html.escape(str(item["lot_wafer"])) for item in datasets)}. Point labels are moved to the table below to keep dense curves readable.</p><section><h2>W/R Estimate Vmin trends</h2><img src="images/{png_path.name}" alt="Estimate Vmin combined summary comparison"><p class="downloads">Original: <a href="images/{svg_path.name}">SVG</a> · <a href="images/{png_path.name}">PNG</a>　 Transparent background: <a href="images/{transparent_svg_path.name}">SVG</a> · <a href="images/{transparent_png_path.name}">PNG</a>　 Trend data: <a href="estimate_vmin_combined_comparison.csv">CSV</a></p></section><section><h2>Worst Cell at each Model VDD</h2><p class="note">One limiting Cell is listed for Read SNM, Write SNM and BL Write Margin at every source/VDD. New Multi-Cell summaries contain side-specific six-MOS WAT data; legacy summaries fall back to PU/PG/PD family averages.</p><div class="table-wrap"><table><thead><tr><th>Source</th><th>VDD (V)</th><th>Limit</th><th>Margin (mV)</th><th>Lot/Wafer</th><th>Cell/Chip</th><th>CR</th><th>PR</th>{device_headers}<th>Data detail</th></tr></thead><tbody>{worst_table_rows}</tbody></table></div><p class="downloads"><a href="estimate_vmin_worst_cell_details.csv">Download worst-Cell details</a> · <a href="estimate_vmin_comparison_per_cell_data.csv">Download per-Cell comparison data</a></p></section><section><h2>Margin and CR/PR distributions</h2><p class="note">All selected Lot/Wafer sources are pooled within each Model VDD; compare median, IQR and outliers rather than only the limiting Cell.</p><img src="images/{distribution_png_path.name}" alt="Per-cell margin and drive ratio box plots"><p class="downloads"><a href="images/{distribution_svg_path.name}">SVG</a> · <a href="images/{distribution_png_path.name}">PNG</a> · <a href="estimate_vmin_boxplot_statistics.csv">Box statistics CSV</a></p></section><section><h2>Six-MOS Idsat distributions</h2><p class="note">Side-specific PUL/PUR/PGL/PGR/PDL/PDR Idsat is retained when available. Legacy files use family-average fallback and are identified in the worst-Cell table.</p><img src="images/{idsat_png_path.name}" alt="Six-MOS Idsat box plots by VDD"><p class="downloads"><a href="images/{idsat_svg_path.name}">SVG</a> · <a href="images/{idsat_png_path.name}">PNG</a></p></section></main></body></html>''', encoding="utf-8")
    return report


def _legacy_estimate_vmin_ratio_shmoo_svg(shmoo: dict, width: int = 1400, height: int = 1040) -> str:
    """Render upper-right-good drive balance plus PU/PG/PD tuning views."""
    samples = shmoo["samples"]
    panels = [("pull_up_ratio_beta", "cell_ratio_beta", "Read / Write Drive-Balance Shmoo",
               "Pull-up Ratio = MOSdrive(PG) / MOSdrive(PU) = PR (higher is better)",
               "Read cell ratio = MOSdrive(PD) / MOSdrive(PG) = CR (higher is better)")]
    if shmoo.get("has_family_wat"):
        panels.extend([
            ("pu_idsat_ua", "pu_vt_v", "PU tuning — upper-left weakens write contention",
             "PU Idsat (uA) — lower is easier write", "PU |Vt| (V) — higher is easier write"),
            ("pg_vt_v", "pg_idsat_ua", "PG tuning — Read / Write trade-off",
             "PG Vt (V)", "PG Idsat (uA)"),
            ("pd_vt_v", "pd_idsat_ua", "PD tuning — upper-left strengthens read",
             "PD Vt (V) — lower is stronger", "PD Idsat (uA) — higher is stronger"),
        ])
    panel_w, panel_h = 590, 350
    origins = ((95, 155), (765, 155), (95, 600), (765, 600))

    def score_color(item: dict) -> str:
        return "#00C93A" if item["best_region"] else "#FF3B30"

    parts = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">',
             '<rect width="100%" height="100%" fill="#FFFFFF"/>',
             f'<text x="56" y="52" fill="#1D1D1F" font-size="34" font-weight="700">Model VDD {shmoo["vdd_v"]:.3f} V - Drive-Balance Shmoo</text>',
             '<text x="56" y="82" fill="#6E6E73" font-size="15">Upper-right = preferred direction. Red = weak; green = strong; purple ring = relative preferred region (not silicon Pass/Fail).</text>',
             '<defs><marker id="arrow" markerWidth="10" markerHeight="10" refX="8" refY="3" orient="auto"><path d="M0,0 L0,6 L9,3 z" fill="#007AFF"/></marker></defs>']
    for panel_index, ((x_key, y_key, title, x_label, y_label), (left, top)) in enumerate(zip(panels, origins)):
        xs = [float(item[x_key]) for item in samples]; ys = [float(item[y_key]) for item in samples]
        x_min, x_max = min(xs), max(xs); y_min, y_max = min(ys), max(ys)
        x_pad = max((x_max-x_min)*.10, max(abs(x_min), 1.0)*.01)
        y_pad = max((y_max-y_min)*.10, max(abs(y_min), 1.0)*.01)
        x_min -= x_pad; x_max += x_pad; y_min -= y_pad; y_max += y_pad
        def xy(x: float, y: float) -> tuple[float, float]:
            return (left+(x-x_min)/(x_max-x_min)*panel_w,
                    top+(1-(y-y_min)/(y_max-y_min))*panel_h)
        parts += [f'<rect x="{left}" y="{top}" width="{panel_w}" height="{panel_h}" fill="#FAFAFA" stroke="#D2D2D7"/>',
                  f'<text x="{left}" y="{top-24}" fill="#1D1D1F" font-size="21" font-weight="700">{title}</text>']
        if panel_index == 0:
            # Data-driven shmoo background: nearest measured-cell classification
            # of the observed balanced score.  This visualizes the current measured
            # population without pretending the upper-left extreme is always
            # a guaranteed silicon pass region.
            grid_cols, grid_rows = 18, 12
            for gy_index in range(grid_rows):
                for gx_index in range(grid_cols):
                    gx_value = x_min + (gx_index + .5) / grid_cols * (x_max - x_min)
                    gy_value = y_max - (gy_index + .5) / grid_rows * (y_max - y_min)
                    nearest = min(samples, key=lambda item: (
                        ((gx_value - float(item[x_key])) / (x_max - x_min)) ** 2 +
                        ((gy_value - float(item[y_key])) / (y_max - y_min)) ** 2))
                    fill = "#71E68B" if nearest["best_region"] else "#FFB0A8"
                    parts.append(
                        f'<rect x="{left+gx_index*panel_w/grid_cols:.1f}" '
                        f'y="{top+gy_index*panel_h/grid_rows:.1f}" '
                        f'width="{panel_w/grid_cols+0.5:.1f}" height="{panel_h/grid_rows+0.5:.1f}" '
                        f'fill="{fill}" opacity="0.38"/>')
        for step in range(5):
            gx = left+panel_w*step/4; gy = top+panel_h*step/4
            xv = x_min+(x_max-x_min)*step/4; yv = y_max-(y_max-y_min)*step/4
            parts += [f'<path d="M{gx:.1f} {top} V{top+panel_h}" stroke="#E5E5EA"/>',
                      f'<path d="M{left} {gy:.1f} H{left+panel_w}" stroke="#E5E5EA"/>',
                      f'<text x="{gx:.1f}" y="{top+panel_h+22}" text-anchor="middle" fill="#6E6E73" font-size="12">{xv:.3g}</text>',
                      f'<text x="{left-10}" y="{gy+4:.1f}" text-anchor="end" fill="#6E6E73" font-size="12">{yv:.3g}</text>']
        for item in samples:
            x, y = xy(float(item[x_key]), float(item[y_key])); score = float(item["balanced_score"])
            ring, ring_width = ("#460479", 4) if item["best_region"] else ("#FFFFFF", 1.5)
            parts += [f'<circle cx="{x:.1f}" cy="{y:.1f}" r="8" fill="{score_color(item)}" stroke="{ring}" stroke-width="{ring_width}"/>',
                      f'<title>{html.escape(str(item["lot_wafer"]))} / {html.escape(str(item["chip_id"]))}; score={score:.3f}; CR={float(item["cell_ratio_beta"]):.3f}; PR={float(item["pull_up_ratio_beta"]):.3f}</title>']
        target = shmoo["target"]
        if x_key in target and y_key in target:
            tx, ty = xy(float(target[x_key]), float(target[y_key]))
            parts += [f'<path d="M{tx-10:.1f} {ty} H{tx+10:.1f} M{tx} {ty-10:.1f} V{ty+10:.1f}" stroke="#460479" stroke-width="4"/>',
                      f'<text x="{tx+12:.1f}" y="{ty-12:.1f}" fill="#460479" font-size="12" font-weight="700">Preferred center</text>']
        if panel_index == 0:
            weak = shmoo["weakest"]
            wx, wy = xy(float(weak[x_key]), float(weak[y_key]))
            tx, ty = xy(float(target[x_key]), float(target[y_key]))
            parts += [f'<path d="M{wx:.1f} {wy:.1f} L{tx:.1f} {ty:.1f}" stroke="#007AFF" stroke-width="3" stroke-dasharray="7 5" marker-end="url(#arrow)"/>',
                      f'<rect x="{wx-7:.1f}" y="{wy-7:.1f}" width="14" height="14" fill="#FF3B30" stroke="#FFFFFF" stroke-width="2"/>',
                      f'<text x="{wx+12:.1f}" y="{wy+20:.1f}" fill="#C13515" font-size="12" font-weight="700">Weakest: {html.escape(str(weak["chip_id"]))}</text>',
                      f'<text x="{left+panel_w-12}" y="{top+22}" text-anchor="end" fill="#248A3D" font-size="14" font-weight="700">Preferred drive direction ↗</text>',
                      f'<text x="{left+panel_w-12}" y="{top+panel_h-12}" text-anchor="end" fill="#C13515" font-size="14" font-weight="700">WEAK</text>']
        parts += [f'<text x="{left+panel_w/2}" y="{top+panel_h+50}" text-anchor="middle" fill="#1D1D1F" font-size="15" font-weight="700">{x_label}</text>',
                  f'<text x="{left-62}" y="{top+panel_h/2}" transform="rotate(-90 {left-62} {top+panel_h/2})" text-anchor="middle" fill="#1D1D1F" font-size="15" font-weight="700">{y_label}</text>']
    if len(panels) == 1:
        parts.append('<text x="765" y="210" fill="#C56A00" font-size="18" font-weight="700">PU/PG/PD Vt and Idsat are unavailable in this legacy summary.</text>')
    parts += [f'<text x="{width/2}" y="{height-20}" text-anchor="middle" fill="#6E6E73" font-size="14">Arrow shows the weakest observed cell toward the median preferred center; empirical same-VDD guidance, not foundry sign-off.</text>', '</svg>']
    return "".join(parts)


def estimate_vmin_ratio_shmoo_svg(shmoo: dict, width: int = 1600,
                                  height: int = 930) -> str:
    """Render one focused drive-balance shmoo with a separate reading guide."""
    samples = shmoo["samples"]
    x_key, y_key = "pull_up_ratio_beta", "cell_ratio_beta"
    xs = [float(item[x_key]) for item in samples]
    ys = [float(item[y_key]) for item in samples]
    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)
    x_pad = max((x_max - x_min) * .12, max(abs(x_min), abs(x_max), 1.0) * .01)
    y_pad = max((y_max - y_min) * .12, max(abs(y_min), abs(y_max), 1.0) * .01)
    x_min -= x_pad; x_max += x_pad
    y_min -= y_pad; y_max += y_pad
    left, top, plot_w, plot_h = 105, 112, 1110, 690
    side_x = 1270

    def xy(x: float, y: float) -> tuple[float, float]:
        return (left + (x - x_min) / (x_max - x_min) * plot_w,
                top + (1 - (y - y_min) / (y_max - y_min)) * plot_h)

    def background_color(grade: str) -> str:
        if grade == "preferred":
            return "#DDF3E2"
        if grade == "monitor":
            return "#FFF0C2"
        return "#F7C9C2"

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        f'<text x="{left}" y="48" fill="#111111" font-size="34" font-weight="700">Model VDD {shmoo["vdd_v"]:.3f} V - Read / Write Drive-Balance Shmoo</text>',
        f'<text x="{left}" y="76" fill="#626B73" font-size="15">CR/PR population P25 / median grading within this VDD | {len(samples)} measured cells | not silicon Pass/Fail</text>',
        f'<rect x="{left}" y="{top}" width="{plot_w}" height="{plot_h}" fill="#FAFAFA" stroke="#AEB7C0"/>',
    ]

    grid_cols, grid_rows = 24, 16
    preferred_grid: list[list[bool]] = []
    for gy_index in range(grid_rows):
        preferred_row = []
        for gx_index in range(grid_cols):
            gx_value = x_min + (gx_index + .5) / grid_cols * (x_max - x_min)
            gy_value = y_max - (gy_index + .5) / grid_rows * (y_max - y_min)
            nearest = min(samples, key=lambda item: (
                ((gx_value - float(item[x_key])) / (x_max - x_min)) ** 2 +
                ((gy_value - float(item[y_key])) / (y_max - y_min)) ** 2))
            grade = str(nearest["wafer_grade"])
            preferred_row.append(grade == "preferred")
            parts.append(
                f'<rect x="{left + gx_index * plot_w / grid_cols:.1f}" '
                f'y="{top + gy_index * plot_h / grid_rows:.1f}" '
                f'width="{plot_w / grid_cols + .5:.1f}" height="{plot_h / grid_rows + .5:.1f}" '
                f'fill="{background_color(grade)}"/>')
        preferred_grid.append(preferred_row)

    # Outline transitions into/out of the P50/P50 nearest-cell region.  This
    # preserves the current empirical classification without implying a
    # physically simulated continuous contour.
    boundary_paths = []
    for gy_index in range(grid_rows):
        for gx_index in range(grid_cols):
            state = preferred_grid[gy_index][gx_index]
            x0 = left + gx_index * plot_w / grid_cols
            x1 = left + (gx_index + 1) * plot_w / grid_cols
            y0 = top + gy_index * plot_h / grid_rows
            y1 = top + (gy_index + 1) * plot_h / grid_rows
            if gx_index + 1 < grid_cols and state != preferred_grid[gy_index][gx_index + 1]:
                boundary_paths.append(f'M{x1:.1f} {y0:.1f}V{y1:.1f}')
            if gy_index + 1 < grid_rows and state != preferred_grid[gy_index + 1][gx_index]:
                boundary_paths.append(f'M{x0:.1f} {y1:.1f}H{x1:.1f}')
    if boundary_paths:
        parts.append(f'<path d="{" ".join(boundary_paths)}" fill="none" stroke="#61308C" stroke-width="2.2" stroke-dasharray="7 5"/>')

    for step in range(6):
        gx = left + plot_w * step / 5
        gy = top + plot_h * step / 5
        xv = x_min + (x_max - x_min) * step / 5
        yv = y_max - (y_max - y_min) * step / 5
        parts += [
            f'<path d="M{gx:.1f} {top}V{top + plot_h}" stroke="#D8DEE4" stroke-width="1"/>',
            f'<path d="M{left} {gy:.1f}H{left + plot_w}" stroke="#D8DEE4" stroke-width="1"/>',
            f'<text x="{gx:.1f}" y="{top + plot_h + 25}" text-anchor="middle" fill="#4F5962" font-size="13">{xv:.3f}</text>',
            f'<text x="{left - 12}" y="{gy + 5:.1f}" text-anchor="end" fill="#4F5962" font-size="13">{yv:.3f}</text>',
        ]

    sample_positions = [xy(float(item[x_key]), float(item[y_key])) for item in samples]
    for sample_index, (item, (x, y)) in enumerate(zip(samples, sample_positions), 1):
        score = float(item["balanced_score"])
        read_delta = float(item.get("read_balance_vs_median_pct", 0.0))
        write_delta = float(item.get("write_balance_vs_median_pct", 0.0))
        read_direction = "stronger" if read_delta > .05 else ("weaker" if read_delta < -.05 else "near median")
        write_direction = "stronger" if write_delta > .05 else ("weaker" if write_delta < -.05 else "near median")
        tooltip_rows = [
            f'Cell: {item["chip_id"]} ({sample_index}/{len(samples)})',
            f'Lot/Wafer: {item["lot_wafer"]}',
            f'Model VDD: {shmoo["vdd_v"]:.3f} V',
            f'RSNM: {float(item.get("rsnm_mv", 0.0)):.1f} mV',
            f'BL Write Vtrip: {float(item.get("write_margin_mv", 0.0)):.1f} mV',
            f'Balanced score: {score:.3f}',
            f'Wafer relative grade: {str(item.get("wafer_grade", "unknown")).upper()}',
            f'CR drive percentile: P{100.0 * float(item.get("cr_percentile", 0.0)):.0f}',
            f'PR drive percentile: P{100.0 * float(item.get("pr_percentile", 0.0)):.0f}',
            f'RSNM performance percentile: P{100.0 * float(item.get("read_percentile", 0.0)):.0f}',
            f'Vtrip performance percentile: P{100.0 * float(item.get("write_percentile", 0.0)):.0f}',
            f'CR = MOSdrive(PD)/MOSdrive(PG): {float(item["cell_ratio_beta"]):.3f}',
            f'PR = MOSdrive(PG)/MOSdrive(PU): {float(item["pull_up_ratio_beta"]):.3f}',
            f'Read balance vs median: {read_delta:+.1f}% ({read_direction})',
            f'Write balance vs median: {write_delta:+.1f}% ({write_direction})',
        ]
        for family, family_name in (("pu", "PU"), ("pg", "PG"), ("pd", "PD")):
            vt_key, idsat_key = f"{family}_vt_v", f"{family}_idsat_ua"
            if vt_key in item and idsat_key in item:
                tooltip_rows.append(
                    f'{family_name}: Vt {float(item[vt_key]):.4f} V '
                    f'(Δmed {float(item.get(f"delta_vs_median_{vt_key}_pct", 0.0)):+.1f}%) / '
                    f'Idsat {float(item[idsat_key]):.3f} µA '
                    f'(Δmed {float(item.get(f"delta_vs_median_{idsat_key}_pct", 0.0)):+.1f}%)')
        tooltip_text = " | ".join(tooltip_rows)
        tooltip_attr = html.escape(tooltip_text, quote=True)
        parts.append(
            f'<circle class="measured-cell" data-cell-tooltip="{tooltip_attr}" '
            f'aria-label="{tooltip_attr}" '
            f'cx="{x:.1f}" cy="{y:.1f}" r="6" fill="#26323D" '
            f'stroke="#FFFFFF" stroke-width="1.2" tabindex="0"/>')

    target = shmoo["target"]
    tx, ty = xy(float(target[x_key]), float(target[y_key]))
    best = shmoo.get("best") or max(
        samples, key=lambda item: (float(item["balanced_score"]),
                                  float(item["read_score"]) + float(item["write_score"])))
    bx, by = xy(float(best[x_key]), float(best[y_key]))
    weak = shmoo["weakest"]
    wx, wy = xy(float(weak[x_key]), float(weak[y_key]))
    preferred_text = "Wafer median center"
    best_text = f'Best measured: {html.escape(str(best["chip_id"]))}'
    weak_text = f'Weakest: {html.escape(str(weak["chip_id"]))}'
    label_positions = _place_chart_labels(
        [(tx, ty, preferred_text), (bx, by, best_text), (wx, wy, weak_text)],
        sample_positions,
        (left + 8, top + 8, left + plot_w - 8, top + plot_h - 8), 14.0)
    ((preferred_x, preferred_y, preferred_anchor),
     (best_x, best_y, best_anchor),
     (weak_x, weak_y, weak_anchor)) = label_positions

    def label_left(x: float, anchor: str, text: str) -> float:
        label_width = max(42.0, len(text) * 14.0 * .56)
        return x - label_width / 2 if anchor == "middle" else (x - label_width if anchor == "end" else x)

    preferred_left = label_left(preferred_x, preferred_anchor, preferred_text)
    preferred_width = max(42.0, len(preferred_text) * 14.0 * .56)
    best_left = label_left(best_x, best_anchor, best_text)
    best_width = max(42.0, len(best_text) * 14.0 * .56)
    weak_left = label_left(weak_x, weak_anchor, weak_text)
    weak_width = max(42.0, len(weak_text) * 14.0 * .56)
    parts += [
        '<g class="special-cell-highlights" pointer-events="none">',
        f'<path d="M{tx - 11:.1f} {ty}H{tx + 11:.1f}M{tx} {ty - 11:.1f}V{ty + 11:.1f}" stroke="#61308C" stroke-width="4"/>',
        f'<rect x="{preferred_left - 6:.1f}" y="{preferred_y - 18:.1f}" width="{preferred_width + 12:.1f}" height="24" rx="6" fill="#F5EFFA" stroke="#D8C8E6"/>',
        f'<text x="{preferred_left + preferred_width / 2:.1f}" y="{preferred_y:.1f}" text-anchor="middle" fill="#61308C" font-size="14" font-weight="700">{preferred_text}</text>',
        f'<polygon points="{bx:.1f},{by - 10:.1f} {bx + 10:.1f},{by:.1f} {bx:.1f},{by + 10:.1f} {bx - 10:.1f},{by:.1f}" fill="#FFC447" stroke="#8A5A00" stroke-width="2"/>',
        f'<rect x="{best_left - 6:.1f}" y="{best_y - 18:.1f}" width="{best_width + 12:.1f}" height="24" rx="6" fill="#FFF4DE" stroke="#E7C16B"/>',
        f'<text x="{best_left + best_width / 2:.1f}" y="{best_y:.1f}" text-anchor="middle" fill="#8A5A00" font-size="14" font-weight="700">{best_text}</text>',
        f'<rect x="{wx - 7:.1f}" y="{wy - 7:.1f}" width="14" height="14" fill="#E4513B" stroke="#FFFFFF" stroke-width="2"/>',
        f'<rect x="{weak_left - 6:.1f}" y="{weak_y - 18:.1f}" width="{weak_width + 12:.1f}" height="24" rx="6" fill="#FFF1EF" stroke="#F0C1B9"/>',
        f'<text x="{weak_left + weak_width / 2:.1f}" y="{weak_y:.1f}" text-anchor="middle" fill="#B2382B" font-size="14" font-weight="700">{weak_text}</text>',
        '</g>',
    ]

    parts += [
        f'<text x="{left + plot_w / 2}" y="{height - 54}" text-anchor="middle" fill="#111111" font-size="18" font-weight="700">Write drive — Pull-up Ratio (PR)</text>',
        f'<text x="{left + plot_w / 2}" y="{height - 27}" text-anchor="middle" fill="#535D66" font-size="14">MOSdrive: PG / PU · right = easier write</text>',
        f'<text x="22" y="{top + plot_h / 2}" transform="rotate(-90 22 {top + plot_h / 2})" text-anchor="middle" fill="#111111" font-size="18" font-weight="700">Read drive — Cell Ratio (CR)</text>',
        f'<text x="55" y="{top + plot_h / 2}" transform="rotate(-90 55 {top + plot_h / 2})" text-anchor="middle" fill="#535D66" font-size="14">MOSdrive: PD / PG · up = stronger read</text>',
    ]

    # Reading guide on the right keeps formulas and status semantics outside
    # the data region so labels cannot collide with measured samples.
    parts += [
        f'<text x="{side_x}" y="132" fill="#20262D" font-size="24" font-weight="700">HOW TO READ</text>',
        f'<rect x="{side_x}" y="154" width="280" height="210" rx="16" fill="#F5F7F9" stroke="#DCE1E6"/>',
        f'<text x="{side_x + 22}" y="198" fill="#20262D" font-size="17" font-weight="700">Wafer-relative grade</text>',
        f'<text x="{side_x + 22}" y="244" fill="#303941" font-size="16">R_pct = percentile(CR)</text>',
        f'<text x="{side_x + 22}" y="276" fill="#303941" font-size="16">W_pct = percentile(PR)</text>',
        f'<text x="{side_x + 22}" y="322" fill="#303941" font-size="17" font-style="italic">Grade = min(R_pct, W_pct)</text>',
        f'<rect x="{side_x}" y="382" width="280" height="126" rx="16" fill="#F5EFFA" stroke="#D8C8E6"/>',
        f'<text x="{side_x + 22}" y="424" fill="#61308C" font-size="17" font-weight="700">Dynamic wafer thresholds</text>',
        f'<text x="{side_x + 22}" y="456" fill="#61308C" font-size="16">Preferred: both metrics &gt;= P50</text>',
        f'<text x="{side_x + 22}" y="482" fill="#61308C" font-size="16">Monitor: weaker metric P25-P50</text>',
        f'<text x="{side_x}" y="550" fill="#20262D" font-size="17" font-weight="700">COLOR SCALE</text>',
        f'<text x="{side_x}" y="583" fill="#535D66" font-size="15">Each region inherits the score</text>',
        f'<text x="{side_x}" y="608" fill="#535D66" font-size="15">of the nearest measured cell.</text>',
        f'<text x="{side_x}" y="633" fill="#535D66" font-size="15">Ranking is relative within this VDD,</text>',
        f'<text x="{side_x}" y="658" fill="#535D66" font-size="15">not guaranteed silicon pass.</text>',
    ]
    legend_y = 690
    parts += [
        f'<rect x="{side_x + 2}" y="{legend_y - 10}" width="17" height="17" fill="#DDF3E2"/><text x="{side_x + 34}" y="{legend_y + 4}" fill="#20262D" font-size="14">Preferred (both ≥ P50)</text>',
        f'<rect x="{side_x + 2}" y="{legend_y + 17}" width="17" height="17" fill="#FFF0C2"/><text x="{side_x + 34}" y="{legend_y + 31}" fill="#20262D" font-size="14">Monitor (weaker P25-P50)</text>',
        f'<rect x="{side_x + 2}" y="{legend_y + 44}" width="17" height="17" fill="#F7C9C2"/><text x="{side_x + 34}" y="{legend_y + 58}" fill="#20262D" font-size="14">Low (weaker &lt; P25)</text>',
        f'<circle cx="{side_x + 10}" cy="{legend_y + 85}" r="6" fill="#26323D"/><text x="{side_x + 34}" y="{legend_y + 90}" fill="#20262D" font-size="14">Measured cell</text>',
        f'<polygon points="{side_x + 10},{legend_y + 103} {side_x + 19},{legend_y + 112} {side_x + 10},{legend_y + 121} {side_x + 1},{legend_y + 112}" fill="#FFC447" stroke="#8A5A00" stroke-width="1.5"/><text x="{side_x + 34}" y="{legend_y + 117}" fill="#20262D" font-size="14">Best measured cell</text>',
        f'<path d="M{side_x + 1} {legend_y + 139}H{side_x + 19}M{side_x + 10} {legend_y + 130}V{legend_y + 148}" stroke="#61308C" stroke-width="4"/><text x="{side_x + 34}" y="{legend_y + 144}" fill="#20262D" font-size="14">Wafer median center</text>',
        f'<rect x="{side_x + 3}" y="{legend_y + 159}" width="14" height="14" fill="#E4513B"/><text x="{side_x + 34}" y="{legend_y + 171}" fill="#20262D" font-size="14">Weakest sample</text>',
        f'<path d="M{side_x} {legend_y + 198}H{side_x + 22}" stroke="#61308C" stroke-width="2.5" stroke-dasharray="7 5"/><text x="{side_x + 34}" y="{legend_y + 203}" fill="#20262D" font-size="14">P50/P50 relative boundary</text>',
    ]
    parts.append('</svg>')
    return "".join(parts)


def estimate_vmin_curve_svg(curve: dict, width: int = 1280, height: int = 780) -> str:
    """Render one imported Multi-Cell conservative margin versus VDD curve."""
    rows, color, label = curve["rows"], curve["color"], curve["label"]
    left, top, plot_w, plot_h = 110, 105, 1050, 470
    measured_vdds = sorted({float(row["vdd_v"]) for row in rows})
    if len(measured_vdds) < 2:
        vdd_min = measured_vdds[0] - .01 if measured_vdds else 0.0
        vdd_max = measured_vdds[0] + .01 if measured_vdds else .02
    else:
        vdd_min, vdd_max = measured_vdds[0], measured_vdds[-1]
    vdd_span = vdd_max - vdd_min
    maximum = max((row["margin_mv"] for row in rows), default=50.0)
    y_max = max(50.0, math.ceil(maximum / 50.0) * 50.0)
    def xy(vdd: float, value: float) -> tuple[float, float]:
        return (left + (vdd - vdd_min) / vdd_span * plot_w,
                top + (1 - value / y_max) * plot_h)
    baseline_y = top + plot_h
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        f'<text x="52" y="54" fill="#1D1D1F" font-size="38" font-weight="700">Estimated {label} versus Model VDD</text>',
        f'<path d="M52 79 h38" stroke="{color}" stroke-width="4"/><text x="101" y="85" fill="#3A3A3C" font-size="16">Minimum cell value from imported Multi-Cell summaries</text>',
    ]
    for step in range(6):
        voltage = vdd_min + vdd_span * step / 5; x, _ = xy(voltage, 0)
        parts.append(
            f'<path d="M{x:.1f} {top} V{baseline_y}" stroke="#E5E5EA"/>')
    for step in range(6):
        value = y_max * step / 5; _, y = xy(vdd_min, value)
        parts += [f'<path d="M{left} {y:.1f} H{left+plot_w}" stroke="#E5E5EA"/>',
                  f'<text x="{left-12}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="14">{value:.0f}</text>']
    points = [xy(row["vdd_v"], row["margin_mv"]) for row in rows]
    parts.append(f'<polyline points="{" ".join(f"{x:.1f},{y:.1f}" for x,y in points)}" fill="none" stroke="{color}" stroke-width="4" stroke-linejoin="round"/>')
    positive_label_items = [(x, y, f'{row["margin_mv"]:.1f} mV')
                            for row, (x, y) in zip(rows, points)
                            if row["margin_mv"] > 0]
    positive_label_positions = iter(_place_chart_labels(
        positive_label_items, points,
        (left + 4, top + 8, left + plot_w - 4, baseline_y - 6), 16.0))
    for row, (x, y) in zip(rows, points):
        parts += [f'<path d="M{x:.1f} {y+6:.1f} V{baseline_y+35:.1f}" stroke="#B9D7FF" stroke-dasharray="4 5"/>',
                  f'<circle cx="{x:.1f}" cy="{y:.1f}" r="5" fill="#FFFFFF" stroke="{color}" stroke-width="3"/>']
        if row["margin_mv"] > 0:
            label_x, label_y, label_anchor = next(positive_label_positions)
            parts += [
                f'<path d="M{x:.1f} {y:.1f}L{label_x:.1f} {label_y - 7:.1f}" stroke="#AAB4BE" stroke-width="1"/>',
                f'<text x="{label_x:.1f}" y="{label_y:.1f}" text-anchor="{label_anchor}" fill="#1D1D1F" font-size="16" font-weight="700" style="paint-order:stroke;stroke:#FFFFFF;stroke-width:6">{row["margin_mv"]:.1f} mV</text>',
            ]
        parts.append(f'<text x="{x:.1f}" y="{baseline_y+55:.1f}" text-anchor="middle" fill="#0062CC" font-size="14" font-weight="700">{row["vdd_v"]:.2f} V</text>')
    if curve["key"] == "rsnm_mv" and len(rows) >= 2:
        left_row, right_row = max(
            zip(rows, rows[1:]),
            key=lambda pair: abs((pair[1]["margin_mv"] - pair[0]["margin_mv"]) /
                                 (pair[1]["vdd_v"] - pair[0]["vdd_v"])) )
        marker_row = right_row
        marker_x, marker_y = xy(marker_row["vdd_v"], marker_row["margin_mv"])
        slope_text = f'Largest RSNM slope  {marker_row["vdd_v"]:.2f} V'
        slope_width = len(slope_text) * 7.8 + 16
        slope_anchor = "end" if marker_x > left + plot_w * .72 else "start"
        slope_label_x = marker_x - 8 if slope_anchor == "end" else marker_x + 8
        slope_left = slope_label_x - slope_width if slope_anchor == "end" else slope_label_x
        parts += [f'<path d="M{marker_x:.1f} {top} V{baseline_y}" stroke="#FF385C" stroke-width="2" stroke-dasharray="5 5"/>',
                  f'<rect x="{slope_left - 5:.1f}" y="{top + 64}" width="{slope_width + 10:.1f}" height="27" rx="6" fill="#FFF1EF" stroke="#F0C1B9"/>',
                  f'<text x="{slope_left + slope_width / 2:.1f}" y="{top+83}" text-anchor="middle" fill="#C13515" font-size="14" font-weight="700">{slope_text}</text>']
    closure = curve.get("eye_closure")
    if closure:
        closure_vdd = float(closure["estimated_vdd_v"])
        x, closure_y = xy(closure_vdd, 0)
        estimate_kind = "Extrapolated" if closure.get("extrapolated") else "Estimated"
        dash = "4 5" if closure.get("extrapolated") else "8 6"
        closure_text = f'{estimate_kind} eye-closure VDD {closure["estimated_vdd_v"]:.4f} V'
        closure_width = len(closure_text) * 8.7 + 16
        closure_anchor = "end" if x > left + plot_w * .70 else "start"
        closure_label_x = x - 12 if closure_anchor == "end" else x + 12
        closure_left = closure_label_x - closure_width if closure_anchor == "end" else closure_label_x
        closure_inside = vdd_min <= closure_vdd <= vdd_max
        if closure_inside:
            parts.append(
                f'<path d="M{x:.1f} {top} V{baseline_y}" stroke="#FF9500" '
                f'stroke-width="3" stroke-dasharray="{dash}"/>')
        else:
            closure_left = left + 8
        parts += [f'<rect x="{closure_left - 6:.1f}" y="{top + 8}" width="{closure_width + 12:.1f}" height="28" rx="6" fill="#FFF4DE" stroke="#F1D399"/>',
                  f'<text x="{closure_left + closure_width / 2:.1f}" y="{top+28}" text-anchor="middle" fill="#C56A00" font-size="16" font-weight="700">{closure_text}</text>']
        if closure.get("extrapolated") and closure_inside:
            first_x, first_y = points[0]
            slope_note = f'Two-lowest-VDD slope: {closure["slope_mv_per_v"]:.2f} mV/V'
            slope_note_width = len(slope_note) * 7.6 + 16
            slope_note_left = (closure_label_x - slope_note_width
                               if closure_anchor == "end" else closure_label_x)
            parts += [
                f'<path data-extrapolated-to-zero="true" d="M{x:.1f} {closure_y:.1f}L{first_x:.1f} {first_y:.1f}" fill="none" stroke="{color}" stroke-width="4" stroke-dasharray="8 6" stroke-linecap="round"/>',
                f'<rect x="{slope_note_left - 6:.1f}" y="{top + 40}" width="{slope_note_width + 12:.1f}" height="25" rx="6" fill="#FFF4DE" stroke="#F1D399"/>',
                f'<text x="{slope_note_left + slope_note_width / 2:.1f}" y="{top+58}" text-anchor="middle" fill="#C56A00" font-size="14">{slope_note}</text>',
            ]
    else:
        parts.append(f'<text x="{left+plot_w-8}" y="{top+28}" text-anchor="end" fill="#C56A00" font-size="16" font-weight="700">Eye-closure VDD not bracketed by imported points</text>')
    parts += [f'<text x="{left+plot_w/2}" y="{baseline_y+112}" text-anchor="middle" fill="#1D1D1F" font-size="21" font-weight="700">Model VDD (V)</text>',
              f'<text x="38" y="{top+plot_h/2}" transform="rotate(-90 38 {top+plot_h/2})" text-anchor="middle" fill="#1D1D1F" font-size="21" font-weight="700">{label} (mV)</text>',
              f'<text x="{width/2}" y="{height-18}" text-anchor="middle" fill="#6E6E73" font-size="14">Each point is the lowest measured cell result. Boundary uses interpolation when bracketed, otherwise qualified low-VDD linear extrapolation; it is not measured WT Vmin.</text>', '</svg>']
    return "".join(parts)


def _drive_advisor_html(
        shmoo: dict[str, object], index: int
        ) -> tuple[str, list[dict[str, object]]]:
    """Return the interactive per-cell advisor and flat export records."""
    advice = [build_drive_to_preferred_advice(
        shmoo, str(sample["chip_id"]), .55, str(sample.get("lot_wafer", "")))
        for sample in shmoo["samples"]]
    grade_order = {"low": 0, "monitor": 1, "preferred": 2}
    advice.sort(key=lambda item: (
        grade_order.get(str(item["current"]["grade"]), 9),
        float(item["current"]["score"]), str(item["chip_id"])))
    options = "".join(
        f'<option value="{position}">{html.escape(str(item["chip_id"]))} · '
        f'{html.escape(str(item["current"]["grade"]).upper())} · '
        f'P{100*float(item["current"]["score"]):.0f}</option>'
        for position, item in enumerate(advice))
    payload = json.dumps(advice, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    mosdrive = 'MOS<sub>drive</sub>'
    section = f'''<section id="drive-advisor-{index}" class="drive-advisor" data-advisor-index="{index}">
<div class="advisor-head"><div><p class="eyebrow">SAME-VDD RELATIVE SCREEN</p><h2>Drive-to-Preferred Advisor</h2><p class="note">Select a cell to estimate the smallest decoupled {mosdrive} change needed to reach a P55 guardband above the P50 Preferred boundary.</p></div><div class="advisor-select-wrap"><label for="advisor-select-{index}">Cell / Chip</label><select id="advisor-select-{index}" class="advisor-select">{options}</select></div></div>
<div class="advisor-kpis"><div class="advisor-kpi"><span>Current grade</span><strong data-field="current-grade">—</strong><small data-field="current-score">—</small></div><div class="advisor-kpi"><span>Current balance</span><strong data-field="current-ratios">—</strong><small data-field="current-percentiles">—</small></div><div class="advisor-kpi"><span>P55 target</span><strong data-field="target-ratios">—</strong><small>Same-VDD population quantile</small></div><div class="advisor-kpi advisor-result"><span>Predicted grade</span><strong data-field="predicted-grade">—</strong><small data-field="predicted-score">—</small></div></div>
<div class="advisor-device-grid" data-field="devices"></div>
<p class="advisor-method" data-field="method"></p><p class="advisor-caution" data-field="caution"></p>
<script type="application/json" class="advisor-data">{payload}</script>
</section>'''
    flat_rows: list[dict[str, object]] = []
    for item in advice:
        for device in item["devices"]:
            flat_rows.append({
                "vdd_v": item["vdd_v"], "lot_wafer": item["lot_wafer"],
                "chip_id": item["chip_id"], "current_grade": item["current"]["grade"],
                "current_score": item["current"]["score"],
                "target_percentile": item["target_percentile"],
                "target_cr": item["target"]["cr"], "target_pr": item["target"]["pr"],
                "predicted_cr": item["predicted"]["cr"],
                "predicted_pr": item["predicted"]["pr"],
                "predicted_grade": item["predicted"]["grade"],
                "family": device["family"],
                "beta_current_relative": device["beta_current_relative"],
                "beta_target_relative": device["beta_target_relative"],
                "beta_change_pct": device["beta_change_pct"],
                "action": device["action"], "vt_v": device["vt_v"],
                "idsat_current_ua": device["idsat_current_ua"],
                "idsat_target_fixed_vt_ua": device["idsat_target_fixed_vt_ua"],
            })
    return section, flat_rows


def write_estimate_vmin_outputs(analysis: dict, out_dir: str | os.PathLike[str],
                                source_paths: Iterable[str | os.PathLike[str]]) -> Path:
    """Write all imported Multi-Cell estimate curves and an HTML selector report."""
    out = Path(out_dir); image_dir = out / "images"; image_dir.mkdir(parents=True, exist_ok=True)
    source_paths = list(source_paths)
    # The analysis mode is authoritative.  A single multi-sheet workbook may
    # contain several Model VDD populations and must still produce Vmin curves.
    shmoo_only = (analysis.get("mode") == "shmoo_only" or
                   len(analysis.get("rows", [])) == 1)
    shmoo_enabled = bool(analysis.get("shmoo_enabled", True))
    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError("PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc
    image_rows = []
    if not shmoo_only:
        for index, (key, _short, _label, _color) in enumerate(_ESTIMATE_VMIN_METRICS, 1):
            svg_name = f"{index:02d}_{key}_estimate_vmin.svg"; png_name = svg_name.replace(".svg", ".png")
            svg_path = image_dir / svg_name; svg_path.write_text(estimate_vmin_curve_svg(analysis["curves"][key]), encoding="utf-8")
            drawing = svg2rlg(str(svg_path))
            if drawing is None: raise RuntimeError("Could not render Estimate Vmin chart")
            renderPM.drawToFile(drawing, str(image_dir / png_name), fmt="PNG", dpi=180, backend="rlPyCairo")
            image_rows.append((key, png_name))
        # Multi-VDD only: a single portable comparison image for the trends.
        stacked_svg = image_dir / "05_estimate_vmin_stacked.svg"
        stacked_png = image_dir / "05_estimate_vmin_stacked.png"
        transparent_svg = image_dir / "05_estimate_vmin_stacked_transparent.svg"
        transparent_png = image_dir / "05_estimate_vmin_stacked_transparent.png"
        stacked_svg.write_text(estimate_vmin_stacked_svg(analysis), encoding="utf-8")
        drawing = svg2rlg(str(stacked_svg))
        if drawing is None: raise RuntimeError("Could not render stacked Estimate Vmin chart")
        renderPM.drawToFile(drawing, str(stacked_png), fmt="PNG", dpi=180, backend="rlPyCairo")
        transparent_svg.write_text(
            estimate_vmin_stacked_svg(analysis, transparent_background=True), encoding="utf-8")
        transparent_drawing = svg2rlg(str(transparent_svg))
        if transparent_drawing is None:
            raise RuntimeError("Could not render transparent stacked Estimate Vmin chart")
        renderPM.drawToFile(transparent_drawing, str(transparent_png), fmt="PNG", dpi=180,
                            bg=None, backend="rlPyCairo", backendFmt="RGBA")
    shmoo_sections = []
    advisor_rows: list[dict[str, object]] = []
    shmoo_fields = ["vdd_v", "lot_wafer", "chip_id", "read_score", "write_score",
                    "balanced_score", "best_region", "read_percentile", "write_percentile",
                    "cr_percentile", "pr_percentile", "performance_grade_score",
                    "wafer_grade_score", "wafer_grade",
                    "robust_low_outlier",
                    "cell_ratio_beta", "pull_up_ratio_beta",
                    "target_cr", "target_pr", "delta_cr", "delta_pr",
                    "read_balance_vs_median_pct", "write_balance_vs_median_pct",
                    "pu_vt_v", "pu_idsat_ua", "delta_pu_vt_v", "delta_pu_idsat_ua",
                    "pg_vt_v", "pg_idsat_ua", "pd_vt_v", "pd_idsat_ua",
                    "delta_pg_vt_v", "delta_pg_idsat_ua", "delta_pd_vt_v", "delta_pd_idsat_ua",
                    "delta_vs_median_pu_vt_v_pct", "delta_vs_median_pu_idsat_ua_pct",
                    "delta_vs_median_pg_vt_v_pct", "delta_vs_median_pg_idsat_ua_pct",
                    "delta_vs_median_pd_vt_v_pct", "delta_vs_median_pd_idsat_ua_pct",
                    "rsnm_mv", "wsnm_mv", "write_margin_mv"]
    statistics_fields = ["vdd_v", "metric", "p05", "q1", "median",
                         "q3", "p95", "mad"]
    statistics_stream = (out / "estimate_vmin_wafer_distribution_statistics.csv").open(
        "w", newline="", encoding="utf-8-sig")
    statistics_writer = csv.DictWriter(statistics_stream, fieldnames=statistics_fields)
    statistics_writer.writeheader()
    with (out / "estimate_vmin_cr_pr_shmoo.csv").open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=shmoo_fields); writer.writeheader()
        for index, shmoo in enumerate(analysis.get("ratio_shmoos", []), 1):
            svg_name = f"{index:02d}_vdd_{shmoo['vdd_v']:.3f}_cr_pr_shmoo.svg"
            png_name = svg_name.replace(".svg", ".png")
            svg_path = image_dir / svg_name
            shmoo_svg_text = estimate_vmin_ratio_shmoo_svg(shmoo)
            svg_path.write_text(shmoo_svg_text, encoding="utf-8")
            shmoo_drawing = svg2rlg(str(svg_path))
            if shmoo_drawing is None:
                raise RuntimeError("Could not render Estimate Vmin CR/PR shmoo")
            renderPM.drawToFile(shmoo_drawing, str(image_dir / png_name), fmt="PNG",
                                dpi=180, backend="rlPyCairo")
            advisor_section, advisor_records = _drive_advisor_html(shmoo, index)
            advisor_rows.extend(advisor_records)
            shmoo_sections.append(
                f'<section><h2>Model VDD {shmoo["vdd_v"]:.3f} V — Drive-Balance Shmoo</h2>'
                f'<p class="note">Whole-wafer median center: CR={shmoo["target"]["cell_ratio_beta"]:.3f}; '
                f'PR=MOSdrive(PG)/MOSdrive(PU)={shmoo["target"]["pull_up_ratio_beta"]:.3f}. '
                f'Best measured cell: {html.escape(str(shmoo["best"]["chip_id"]))} '
                f'(Score={float(shmoo["best"]["balanced_score"]):.3f}). '
                'Move the pointer over a measured cell to inspect its WAT and margin values. '
                'Color regions use the weaker CR/PR population percentile: green ≥P50, yellow P25–P50, red &lt;P25. RSNM and BL-Vtrip percentiles remain available for correlation. These are relative screening within this VDD, not silicon Pass/Fail.</p>'
                f'<div class="interactive-shmoo">{shmoo_svg_text}</div>'
                f'<p class="note">Downloads: <a href="images/{svg_name}">interactive SVG</a> · '
                f'<a href="images/{png_name}">PNG</a></p></section>'
                f'{advisor_section}')
            for sample in shmoo["samples"]:
                record = {key: sample.get(key, "") for key in shmoo_fields}
                record["vdd_v"] = shmoo["vdd_v"]
                writer.writerow(record)
            for metric, values in shmoo["distributions"].items():
                statistics_writer.writerow({"vdd_v": shmoo["vdd_v"], "metric": metric,
                                             **values})
    statistics_stream.close()
    advisor_fields = [
        "vdd_v", "lot_wafer", "chip_id", "current_grade", "current_score",
        "target_percentile", "target_cr", "target_pr", "predicted_cr",
        "predicted_pr", "predicted_grade", "family", "beta_current_relative",
        "beta_target_relative", "beta_change_pct", "action", "vt_v",
        "idsat_current_ua", "idsat_target_fixed_vt_ua",
    ]
    with (out / "estimate_vmin_drive_to_preferred_advisor.csv").open(
            "w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=advisor_fields)
        writer.writeheader()
        writer.writerows(advisor_rows)
    if not shmoo_only or not shmoo_enabled:
        with (out / "multi_chip_snm_summary_combined.csv").open(
                "w", newline="", encoding="utf-8-sig") as stream:
            fields = ["vdd_v", "sample_count"] + [
                field for key, *_ in _ESTIMATE_VMIN_METRICS
                for field in (key, f"{key}_lot_wafer", f"{key}_chip_id")]
            writer = csv.DictWriter(stream, fieldnames=fields)
            writer.writeheader()
            writer.writerows({key: row.get(key) for key in fields}
                             for row in analysis["rows"])
    backup_dir = out / "imported_multi_chip_summaries"; backup_dir.mkdir(exist_ok=True)
    for index, source in enumerate(source_paths, 1):
        source_path = Path(source); shutil.copy2(source_path, backup_dir / f"{index:02d}_{source_path.name}")
    if shmoo_only and shmoo_enabled:
        vdd_list = ", ".join(f'{float(row["vdd_v"]):.3f} V' for row in analysis["rows"])
        sections = (f'<section class="shmoo-only-notice"><h2>Shmoo-only mode</h2>'
                    f'<p>Analyzed Model VDD: {vdd_list}. This run was started from one source file '
                    'or contains only one VDD, so Estimate Vmin trends and eye-closure extrapolation '
                    'are omitted. Only same-VDD Shmoo screening and the Drive-to-Preferred Advisor are output.</p></section>'
                    + "".join(shmoo_sections))
    elif shmoo_only:
        vdd_list = ", ".join(
            f'{float(row["vdd_v"]):.3f} V' for row in analysis["rows"])
        sections = (f'<section class="shmoo-only-notice"><h2>Single-VDD fast mode</h2>'
                    f'<p>Analyzed Model VDD: {vdd_list}. Shmoo was disabled and '
                    'one voltage cannot form an Estimate Vmin trend. The imported '
                    'minimum SNM and BL Write Margin remain in the CSV outputs.</p></section>')
    else:
        shmoo_notice = ("" if shmoo_enabled else
                         '<section><h2>Fast curve mode</h2><p class="note">'
                         'Shmoo and Drive Advisor were disabled for this run.</p></section>')
        sections = shmoo_notice + '<section id="stacked-trends"><h2>Stacked VDD trends</h2><img src="images/05_estimate_vmin_stacked.png" alt="Stacked Estimate Vmin curves"><p class="note">PNG exports: <a href="images/05_estimate_vmin_stacked.png">white background</a> · <a href="images/05_estimate_vmin_stacked_transparent.png">transparent background</a></p></section>' + "".join(f'<section><h2>{analysis["curves"][key]["label"]}</h2><img src="images/{image}" alt="{key} Estimate Vmin curve"></section>' for key, image in image_rows) + "".join(shmoo_sections)
    report = out / "estimate_vmin_report.html"
    report.write_text(f'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>HV28 SRAM Estimate Vmin Curve</title>
<style>
*{{box-sizing:border-box}}
body{{font-family:Calibri,"Microsoft JhengHei",Arial,sans-serif;background:#f5f5f7;color:#1d1d1f;margin:0;padding:clamp(10px,2vw,32px)}}
main{{width:100%;max-width:1600px;margin:auto}}
section{{background:#fff;border-radius:16px;padding:24px;margin:18px 0}}
img,.interactive-shmoo svg{{display:block;width:100%;height:auto;border:1px solid #e5e5ea;border-radius:12px}}
.interactive-shmoo{{width:100%;min-width:0;overflow-x:auto}}
.interactive-shmoo circle.measured-cell{{cursor:help}}
.note{{color:#6e6e73}}
.eyebrow{{margin:0 0 5px;color:#ff385c;font-size:12px;font-weight:700;letter-spacing:.09em}}
.advisor-head{{display:flex;align-items:flex-end;justify-content:space-between;gap:28px}}
.advisor-head h2{{margin:.1rem 0 .35rem}}
.advisor-select-wrap{{display:grid;gap:7px;min-width:270px;color:#6e6e73;font-size:13px;font-weight:700}}
.advisor-select{{width:100%;padding:11px 38px 11px 12px;border:1px solid #d2d2d7;border-radius:10px;background:#fff;color:#1d1d1f;font:600 15px Calibri,"Microsoft JhengHei",Arial,sans-serif}}
.advisor-kpis{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin:22px 0}}
.advisor-kpi{{display:grid;gap:5px;padding:16px;border:1px solid #e5e5ea;border-radius:14px;background:#fafafa}}
.advisor-kpi span,.advisor-kpi small{{color:#6e6e73}}
.advisor-kpi strong{{font-size:21px;font-variant-numeric:tabular-nums}}
.advisor-result{{background:#f0faf3;border-color:#c8e9d1}}
.advisor-device-grid{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:12px}}
.advisor-device{{padding:16px;border-radius:14px;border:1px solid #e5e5ea;background:#fff}}
.advisor-device[data-family="PD"]{{border-top:4px solid #007aff}}.advisor-device[data-family="PG"]{{border-top:4px solid #34c759}}.advisor-device[data-family="PU"]{{border-top:4px solid #ff385c}}
.advisor-device-top{{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:12px}}
.advisor-device-name{{font-size:19px;font-weight:700}}.advisor-action{{color:#6e6e73;font-weight:700}}
.advisor-change{{font-size:25px;font-weight:700;font-variant-numeric:tabular-nums}}
.advisor-device dl{{display:grid;grid-template-columns:1fr auto;gap:7px 12px;margin:10px 0 0;font-variant-numeric:tabular-nums}}.advisor-device dt{{color:#6e6e73}}.advisor-device dd{{margin:0;text-align:right;font-weight:700}}
.advisor-method{{margin:18px 0 4px;font-weight:700}}.advisor-caution{{margin:0;color:#9a6700}}
.mosdrive{{white-space:nowrap;font-family:"Times New Roman",serif;font-style:normal}}.mosdrive sub{{font-size:.62em;line-height:0;vertical-align:-.35em;font-weight:inherit;font-style:italic}}
#cell-tooltip{{position:fixed;z-index:9999;display:none;width:520px;max-width:calc(100vw - 32px);max-height:calc(100vh - 32px);overflow-x:hidden;overflow-y:auto;padding:15px 17px;border-radius:14px;background:#1d1d1f;color:#fff;box-shadow:0 10px 34px rgba(0,0,0,.26);font-size:13px;line-height:1.35;pointer-events:auto;overscroll-behavior:contain;scrollbar-width:thin}}
.tooltip-title{{font-size:17px;font-weight:700;margin-bottom:6px}}
.tooltip-meta{{display:grid;grid-template-columns:1fr;gap:2px;color:#c8c8cc;margin-bottom:11px}}
.tooltip-section-label{{color:#aeb0b5;font-size:11px;font-weight:700;letter-spacing:.07em;margin:0 0 7px;text-transform:uppercase}}
.tooltip-metrics{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:1px;margin-bottom:11px;border-radius:9px;overflow:hidden;background:#46464b}}
.tooltip-metric{{display:flex;align-items:baseline;justify-content:space-between;gap:10px;padding:7px 9px;background:#303034;min-width:0}}
.tooltip-metric-label{{color:#b8b8bd;font-size:12px}}
.tooltip-metric-value{{color:#fff;font-weight:700;text-align:right;font-variant-numeric:tabular-nums}}
.tooltip-device-list{{display:grid;grid-template-columns:1fr;gap:6px}}
.tooltip-device-card{{display:grid;grid-template-columns:1fr;gap:3px;padding:7px 10px;border-left:3px solid #8e8e93;border-radius:9px;background:#303034;font-variant-numeric:tabular-nums}}
.tooltip-device-card[data-family="PU"]{{border-left-color:#ff453a}}
.tooltip-device-card[data-family="PG"]{{border-left-color:#30d158}}
.tooltip-device-card[data-family="PD"]{{border-left-color:#0a84ff}}
.tooltip-device-name{{font-size:14px;font-weight:700;color:#fff;padding-bottom:3px;margin-bottom:1px;border-bottom:1px solid #47474c}}
.tooltip-device-value{{display:flex;justify-content:space-between;gap:16px;color:#fff}}
.tooltip-device-label{{color:#aeb0b5}}
.tooltip-device-reading{{display:flex;align-items:baseline;gap:8px;text-align:right}}
.tooltip-device-delta{{color:#ffd60a;font-size:11px;font-weight:700}}
@media(max-width:900px){{.advisor-head{{align-items:stretch;flex-direction:column}}.advisor-select-wrap{{min-width:0}}.advisor-kpis{{grid-template-columns:repeat(2,minmax(0,1fr))}}.advisor-device-grid{{grid-template-columns:1fr}}}}
@media(max-width:600px){{.tooltip-metrics,.advisor-kpis{{grid-template-columns:1fr}}}}
</style></head><body><main><h1>HV28 SRAM Estimate Vmin Curve</h1>
<p class="note">{html.escape(analysis["definition"])}</p>
<p class="note">At each Model VDD, every cell is ranked by its CR percentile and PR percentile; the weaker drive-ratio percentile determines its wafer-relative grade. Green means both are at or above the median, yellow means the weaker ratio is P25–P50, and red means it is below P25. RSNM and BL Write Trip Margin percentiles are retained for correlation, while residual WSNM remains a separate diagnostic. X=<span class="mosdrive">MOS<sub>drive</sub></span>(PG)/<span class="mosdrive">MOS<sub>drive</sub></span>(PU)=PR (right improves write); Y=<span class="mosdrive">MOS<sub>drive</sub></span>(PD)/<span class="mosdrive">MOS<sub>drive</sub></span>(PG)=CR (up improves read). These colors are not absolute silicon Pass/Fail.</p>
{sections}
<p class="note">Source summary backups: <code>imported_multi_chip_summaries/</code>. {('Combined conservative points: <code>multi_chip_snm_summary_combined.csv</code>. ' if not shmoo_only else '')}Per-cell Shmoo results: <code>estimate_vmin_cr_pr_shmoo.csv</code>. Cell Advisor: <code>estimate_vmin_drive_to_preferred_advisor.csv</code>. Lot/Wafer batch comparison is available from the dedicated <b>Lot/Wafer Advisor</b> tab.</p>
</main><div id="cell-tooltip" role="tooltip"></div>
<script>
const cellTooltip=document.getElementById('cell-tooltip');
let cellTooltipHideTimer=null;
const makeTooltipNode=(className,text)=>{{
  const node=document.createElement('div');
  node.className=className;
  node.textContent=text;
  return node;
}};
const renderCellTooltip=(raw)=>{{
  const rows=raw.split(' | ');
  cellTooltip.replaceChildren();
  cellTooltip.appendChild(makeTooltipNode('tooltip-title',rows[0]||'Cell information'));
  const meta=document.createElement('div');
  meta.className='tooltip-meta';
  (rows.slice(1,3)).forEach((text)=>meta.appendChild(makeTooltipNode('',text)));
  cellTooltip.appendChild(meta);
  const appendMetricSection=(title,values)=>{{
    cellTooltip.appendChild(makeTooltipNode('tooltip-section-label',title));
    const metrics=document.createElement('div');
    metrics.className='tooltip-metrics';
    values.forEach((text)=>{{
      const separator=text.lastIndexOf(':');
      const metric=document.createElement('div');
      metric.className='tooltip-metric';
      metric.appendChild(makeTooltipNode('tooltip-metric-label',separator>=0?text.slice(0,separator):text));
      metric.appendChild(makeTooltipNode('tooltip-metric-value',separator>=0?text.slice(separator+1).trim():'—'));
      metrics.appendChild(metric);
    }});
    cellTooltip.appendChild(metrics);
  }};
  appendMetricSection('Performance and wafer-relative grade',rows.slice(3,11));
  appendMetricSection('Drive ratios',rows.slice(11,13));
  appendMetricSection('Balance vs same-VDD median',rows.slice(13,15));
  const devices=rows.slice(15);
  if(devices.length){{
    cellTooltip.appendChild(makeTooltipNode('tooltip-section-label','Device WAT · PU → PG → PD'));
    const list=document.createElement('div');
    list.className='tooltip-device-list';
    devices.forEach((text)=>{{
      const match=text.match(/^(PU|PG|PD): Vt ([^ ]+) V \\(Δmed ([^)]+)\\) \\/ Idsat ([^ ]+) µA \\(Δmed ([^)]+)\\)$/);
      if(match){{
        const card=document.createElement('div');
        card.className='tooltip-device-card';
        card.dataset.family=match[1];
        card.appendChild(makeTooltipNode('tooltip-device-name',match[1]));
        const vt=document.createElement('div');
        vt.className='tooltip-device-value';
        vt.appendChild(makeTooltipNode('tooltip-device-label','Vt'));
        const vtReading=document.createElement('div');
        vtReading.className='tooltip-device-reading';
        vtReading.appendChild(makeTooltipNode('',match[2]+' V'));
        vtReading.appendChild(makeTooltipNode('tooltip-device-delta',match[3]));
        vt.appendChild(vtReading);
        card.appendChild(vt);
        const idsat=document.createElement('div');
        idsat.className='tooltip-device-value';
        idsat.appendChild(makeTooltipNode('tooltip-device-label','Idsat'));
        const idsatReading=document.createElement('div');
        idsatReading.className='tooltip-device-reading';
        idsatReading.appendChild(makeTooltipNode('',match[4]+' µA'));
        idsatReading.appendChild(makeTooltipNode('tooltip-device-delta',match[5]));
        idsat.appendChild(idsatReading);
        card.appendChild(idsat);
        list.appendChild(card);
      }}else{{
        list.appendChild(makeTooltipNode('tooltip-device-card',text));
      }}
    }});
    cellTooltip.appendChild(list);
  }}
}};
const moveCellTooltip=(event)=>{{
  const gap=16;
  const box=cellTooltip.getBoundingClientRect();
  const left=Math.min(event.clientX+gap,window.innerWidth-box.width-gap);
  const top=Math.min(event.clientY+gap,window.innerHeight-box.height-gap);
  cellTooltip.style.left=Math.max(gap,left)+'px';
  cellTooltip.style.top=Math.max(gap,top)+'px';
}};
document.querySelectorAll('[data-cell-tooltip]').forEach((mark)=>{{
  const show=(event)=>{{
    clearTimeout(cellTooltipHideTimer);
    renderCellTooltip(mark.dataset.cellTooltip);
    cellTooltip.style.display='block';
    moveCellTooltip(event);
  }};
  mark.addEventListener('mouseenter',show);
  mark.addEventListener('mousemove',moveCellTooltip);
  mark.addEventListener('mouseleave',()=>{{cellTooltipHideTimer=setTimeout(()=>{{cellTooltip.style.display='none';}},160);}});
  mark.addEventListener('focus',()=>{{
    const box=mark.getBoundingClientRect();
    show({{clientX:box.left+box.width/2,clientY:box.top+box.height/2}});
  }});
  mark.addEventListener('blur',()=>{{cellTooltipHideTimer=setTimeout(()=>{{cellTooltip.style.display='none';}},160);}});
}});
cellTooltip.addEventListener('mouseenter',()=>clearTimeout(cellTooltipHideTimer));
cellTooltip.addEventListener('mouseleave',()=>{{cellTooltip.style.display='none';}});
const fmt=(value,digits=3)=>Number(value).toFixed(digits);
const pct=(value)=>'P'+Math.round(Number(value)*100);
const makeMosdrive=()=>{{const span=document.createElement('span');span.className='mosdrive';span.append('MOS');const sub=document.createElement('sub');sub.textContent='drive';span.appendChild(sub);return span;}};
document.querySelectorAll('.drive-advisor').forEach((card)=>{{
  const records=JSON.parse(card.querySelector('.advisor-data').textContent);
  const select=card.querySelector('.advisor-select');
  const setText=(field,value)=>{{card.querySelector(`[data-field="${{field}}"]`).textContent=value;}};
  const draw=()=>{{
    const item=records[Number(select.value)||0];
    setText('current-grade',String(item.current.grade).toUpperCase());
    setText('current-score',`${{item.chip_id}} · ${{pct(item.current.score)}}`);
    setText('current-ratios',`CR ${{fmt(item.current.cr)}} · PR ${{fmt(item.current.pr)}}`);
    setText('current-percentiles',`Read ${{pct(item.current.cr_percentile)}} · Write ${{pct(item.current.pr_percentile)}}`);
    setText('target-ratios',`CR ${{fmt(item.target.cr)}} · PR ${{fmt(item.target.pr)}}`);
    setText('predicted-grade',String(item.predicted.grade).toUpperCase());
    setText('predicted-score',`Read ${{pct(item.predicted.cr_percentile)}} · Write ${{pct(item.predicted.pr_percentile)}}`);
    setText('method',item.method);
    setText('caution',item.caution);
    const deviceGrid=card.querySelector('[data-field="devices"]');
    deviceGrid.replaceChildren();
    item.devices.forEach((device)=>{{
      const panel=document.createElement('article');panel.className='advisor-device';panel.dataset.family=device.family;
      const top=document.createElement('div');top.className='advisor-device-top';
      top.appendChild(makeTooltipNode('advisor-device-name',device.family));
      top.appendChild(makeTooltipNode('advisor-action',device.action));panel.appendChild(top);
      const change=makeTooltipNode('advisor-change',(Number(device.beta_change_pct)>=0?'+':'')+fmt(device.beta_change_pct,1)+'% ');
      change.appendChild(makeMosdrive());panel.appendChild(change);
      const list=document.createElement('dl');
      const add=(label,value)=>{{list.appendChild(makeTooltipNode('',label));const dd=document.createElement('dd');dd.textContent=value;list.appendChild(dd);}};
      const driveLabel=document.createElement('dt');driveLabel.append('Relative ');driveLabel.appendChild(makeMosdrive());list.appendChild(driveLabel);const driveValue=document.createElement('dd');driveValue.textContent=`${{fmt(device.beta_current_relative)}} → ${{fmt(device.beta_target_relative)}}`;list.appendChild(driveValue);
      add('Current Vt',device.vt_v==null?'—':fmt(device.vt_v,4)+' V');
      add('Idsat @ fixed Vt',device.idsat_current_ua==null?'—':`${{fmt(device.idsat_current_ua,2)}} → ${{fmt(device.idsat_target_fixed_vt_ua,2)}} µA`);
      panel.appendChild(list);deviceGrid.appendChild(panel);
    }});
  }};
  select.addEventListener('change',draw);draw();
}});
</script></body></html>''', encoding="utf-8")
    return report


def write_trip_margin_curve_svg(analysis: dict, width: int = 1280,
                                height: int = 780) -> str:
    """Render WTM versus VDD with the same visual grammar as the RSNM curve."""
    boundary = analysis.get("write_boundary")
    adapted = {
        "rows": [
            {**row, "vcc_v": row["vdd_v"], "rsnm_mv": row["wtm_mv"],
             "valid_eye": row["writable"]}
            for row in analysis["rows"]
        ],
        "eye_closure": ({"estimated_vcc_v": boundary["estimated_vdd_v"]}
                        if boundary else None),
    }
    svg = rsnm_vcc_curve_svg(adapted, width=width, height=height)
    replacements = (
        ("Estimated Read SNM versus Model VDD", "Estimated Write Trip Margin versus Model VDD"),
        ("Estimated RSNM versus Model VDD", "Estimated Write Trip Margin versus Model VDD"),
        ("Estimated eye-closure VDD", "Estimated write boundary VDD"),
        ("Eye-closure VDD not bracketed by the entered rows",
         "Write boundary VDD not bracketed by the entered rows"),
        ("Read SNM (mV)", "Write Trip Margin (mV)"),
        ("X = no valid butterfly eye. Boundary is a compact-model estimate and is not measured WT Vmin.",
         "X = no positive write margin. Boundary is a compact-model estimate and is not measured Select_Write Vmin."),
    )
    for source, target in replacements:
        svg = svg.replace(source, target)
    return svg


def write_write_trip_margin_outputs(analysis: dict,
                                    out_dir: str | os.PathLike[str]) -> Path:
    """Write HTML, PNG, SVG, CSV and JSON for manual WTM/VDD analysis."""
    out = Path(out_dir)
    image_dir = out / "images"
    image_dir.mkdir(parents=True, exist_ok=True)
    svg_path = image_dir / "01_write_trip_margin_vs_model_vdd.svg"
    png_path = image_dir / "01_write_trip_margin_vs_model_vdd.png"
    svg_path.write_text(write_trip_margin_curve_svg(analysis), encoding="utf-8")
    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError(
            "PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc
    drawing = svg2rlg(str(svg_path))
    if drawing is None:
        raise RuntimeError("Could not render Write Trip Margin versus VDD chart")
    renderPM.drawToFile(drawing, str(png_path), fmt="PNG", dpi=180, backend="rlPyCairo")

    csv_fields = ["vdd_v", "pu_vt_v", "pu_idsat_ua", "pg_vt_v", "pg_idsat_ua",
                  "pd_vt_v", "pd_idsat_ua", "wtm_mv", "writable", "status"]
    with open(out / "write_trip_margin_curve.csv", "w", newline="",
              encoding="utf-8-sig") as source:
        writer = csv.DictWriter(source, fieldnames=csv_fields)
        writer.writeheader()
        writer.writerows({key: row.get(key) for key in csv_fields}
                         for row in analysis["rows"])
    (out / "write_trip_margin_curve.json").write_text(
        json.dumps(analysis, indent=2), encoding="utf-8")

    boundary = analysis.get("write_boundary")
    boundary_text = (f'{boundary["estimated_vdd_v"]:.4f} V'
                     if boundary else "Not bracketed")
    table_rows = "".join(
        f'<tr><td>{row["vdd_v"]:.3f}</td><td>{row["pu_vt_v"]:.4f}</td><td>{row["pu_idsat_ua"]:.3f}</td>'
        f'<td>{row["pg_vt_v"]:.4f}</td><td>{row["pg_idsat_ua"]:.3f}</td>'
        f'<td>{row["pd_vt_v"]:.4f}</td><td>{row["pd_idsat_ua"]:.3f}</td>'
        f'<td>{_fmt(row["wtm_mv"], 2)}</td><td>{row["status"]}</td></tr>'
        for row in analysis["rows"])
    document = f'''<!doctype html><html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>HV28 SRAM Analysis - Write Trip Margin</title>
    <style>:root{{font:100%/1.5 Calibri,"Microsoft JhengHei",Arial,sans-serif;color:#1d1d1f;background:#f5f5f7}}*{{box-sizing:border-box}}body{{margin:0;padding:2rem}}main{{max-width:1500px;margin:auto}}h1{{font-size:2.6rem;letter-spacing:-.03em}}section{{background:#fff;border-radius:1.25rem;padding:1.5rem;margin:1rem 0}}img{{display:block;width:100%;height:auto;border:1px solid #e5e5ea;border-radius:1rem}}table{{border-collapse:collapse;width:100%;font-variant-numeric:tabular-nums}}th,td{{padding:.7rem;border-bottom:1px solid #e5e5ea;text-align:right}}th:first-child,td:first-child{{text-align:left}}.note{{color:#6e6e73}}</style></head><body><main>
    <h1>HV28 SRAM Analysis</h1><p>Manual VDD / PU / PG / PD WAT write-trip analysis</p>
    <section><h2>Estimated Write Trip Margin versus Model VDD</h2><p><b>Estimated write boundary VDD:</b> {boundary_text}</p><img src="images/{png_path.name}" alt="Estimated Write Trip Margin versus Model VDD"><p class="note">WTM is the permitted rise of the nominally-low write bitline while PG can still overcome PU at the inverter trip point. This compact-model estimate is not measured Select_Write Vmin.</p></section>
    <section><h2>Input and calculated values</h2><table><thead><tr><th>VDD (V)</th><th>PU Vt</th><th>PU Isat</th><th>PG Vt</th><th>PG Isat</th><th>PD Vt</th><th>PD Isat</th><th>WTM (mV)</th><th>Status</th></tr></thead><tbody>{table_rows}</tbody></table></section>
    </main></body></html>'''
    report = out / "write_trip_margin_report.html"
    report.write_text(document, encoding="utf-8")
    return report


def _legacy_estimate_vmin_vertical_stacked_svg(
        analysis: dict, width: int = 1280, height: int = 1120) -> str:
    """Aligned VDD trend panels, each retaining its own mV scale."""
    left, right, top, bottom = 125, 65, 92, 58
    panel_gap = 22
    panel_h = (height - top - bottom - panel_gap * (len(_ESTIMATE_VMIN_METRICS) - 1)) / len(_ESTIMATE_VMIN_METRICS)
    plot_w = width - left - right
    parts = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">',
             '<rect width="100%" height="100%" fill="#FFFFFF"/>',
             '<text x="56" y="52" fill="#1D1D1F" font-size="34" font-weight="700">Estimate Vmin Curves — Stacked View</text>',
             '<text x="56" y="78" fill="#6E6E73" font-size="16">Each panel uses its own margin scale; X-axis is Model VDD (V).</text>']
    for index, (key, _short, label, color) in enumerate(_ESTIMATE_VMIN_METRICS):
        curve = analysis["curves"][key]; rows = curve["rows"]
        panel_top = top + index * (panel_h + panel_gap); panel_bottom = panel_top + panel_h
        maximum = max((row["margin_mv"] for row in rows), default=50.0)
        y_max = max(50.0, math.ceil(maximum / 50.0) * 50.0)
        def xy(vdd: float, margin: float) -> tuple[float, float]:
            return left + vdd / SNM_PLOT_AXIS_MAX_V * plot_w, panel_top + (1 - margin / y_max) * panel_h
        parts += [f'<text x="{left}" y="{panel_top-6:.1f}" fill="#1D1D1F" font-size="18" font-weight="700">{label} (mV)</text>']
        for step in range(4):
            margin = y_max * step / 3; _x, y = xy(0, margin)
            parts += [f'<path d="M{left} {y:.1f} H{left+plot_w}" stroke="#E5E5EA"/>',
                      f'<text x="{left-12}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="12">{margin:.0f}</text>']
        points = [xy(row["vdd_v"], row["margin_mv"]) for row in rows]
        parts.append(f'<polyline points="{" ".join(f"{x:.1f},{y:.1f}" for x,y in points)}" fill="none" stroke="{color}" stroke-width="3"/>')
        for row, (x, y) in zip(rows, points):
            parts += [f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="#FFF" stroke="{color}" stroke-width="2"/>',
                      f'<text x="{x:.1f}" y="{panel_bottom+17:.1f}" text-anchor="middle" fill="#6E6E73" font-size="11">{row["vdd_v"]:.2f}</text>']
        if key == "rsnm_mv" and len(rows) >= 2:
            _low, marker = max(zip(rows, rows[1:]), key=lambda pair: abs((pair[1]["margin_mv"]-pair[0]["margin_mv"])/(pair[1]["vdd_v"]-pair[0]["vdd_v"])))
            marker_x, _marker_y = xy(marker["vdd_v"], marker["margin_mv"])
            parts += [f'<path d="M{marker_x:.1f} {panel_top} V{panel_bottom}" stroke="#FF385C" stroke-width="2" stroke-dasharray="5 5"/>',
                      f'<text x="{marker_x+7:.1f}" y="{panel_top+19:.1f}" fill="#C13515" font-size="13" font-weight="700">Largest slope: {marker["vdd_v"]:.2f} V</text>']
    parts += [f'<text x="{left+plot_w/2}" y="{height-16}" text-anchor="middle" fill="#1D1D1F" font-size="18" font-weight="700">Model VDD (V)</text>', '</svg>']
    return "".join(parts)


def estimate_vmin_stacked_svg(analysis: dict, width: int = 1280, height: int = 620,
                              transparent_background: bool = False) -> str:
    """Render the same compact Multi-VDD format used by folder comparison."""
    rows = [{"vdd_v": float(row["vdd_v"]),
             "sample_count": int(row.get("sample_count", 1)),
             **{key: float(row[key]) for key, *_ in _ESTIMATE_VMIN_METRICS}}
            for row in analysis["rows"]]
    lot_names = sorted({str(row.get(f"{key}_lot_wafer", "")).strip()
                        for row in analysis["rows"]
                        for key, *_ in _ESTIMATE_VMIN_METRICS
                        if str(row.get(f"{key}_lot_wafer", "")).strip()})
    label = lot_names[0] if len(lot_names) == 1 else "Multi-Cell conservative minimum"
    dataset = {"lot_wafer": label, "rows": rows,
               "color": "#007AFF", "sources": ["Estimate Vmin import"]}
    return estimate_vmin_combined_comparison_svg(
        [dataset], width, height, transparent_background)


def _legacy_estimate_vmin_stacked_svg(analysis: dict, width: int = 1280,
                                      height: int = 620,
                                      transparent_background: bool = False) -> str:
    """Render SNM and BL write margin as two aligned side-by-side panels."""
    groups = (
        ("Read / Write SNM", ("rsnm_mv", "wsnm_mv"), "SNM (mV)"),
        ("BL Write Margin", ("write_margin_mv",), "Vtrip (mV)"),
    )
    left, right, top, bottom, panel_gap = 92, 48, 145, 78, 72
    plot_w = (width - left - right - panel_gap) / len(groups)
    panel_h = height - top - bottom
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Microsoft JhengHei,Arial,sans-serif">',
    ]
    if not transparent_background:
        parts.append('<rect width="100%" height="100%" fill="#FFFFFF"/>')
    parts += [
        '<text x="56" y="46" fill="#1D1D1F" font-size="34" font-weight="700">Estimate Vmin Curves - Comparison View</text>',
        '<text x="56" y="73" fill="#6E6E73" font-size="16">Left: Read / Write SNM · Right: BL Write Margin · X-axis: Model VDD (V)</text>',
    ]
    for index, (group_label, keys, y_axis_label) in enumerate(groups):
        panel_left = left + index * (plot_w + panel_gap)
        panel_top = top
        panel_bottom = panel_top + panel_h
        curves = [analysis["curves"][key] for key in keys]
        maximum = max((row["margin_mv"] for curve in curves for row in curve["rows"]), default=50.0)
        y_max = max(50.0, math.ceil(maximum / 50.0) * 50.0)

        def xy(vdd: float, margin: float) -> tuple[float, float]:
            return (panel_left + vdd / SNM_PLOT_AXIS_MAX_V * plot_w,
                    panel_top + (1 - margin / y_max) * panel_h)

        header_y = panel_top - 24
        parts.append(f'<text x="{panel_left}" y="{header_y:.1f}" fill="#1D1D1F" font-size="20" font-weight="700">{group_label}</text>')
        legend_x = panel_left + (210 if index == 0 else 230)
        for curve_index, curve in enumerate(curves):
            parts += [f'<path d="M{legend_x} {header_y-6:.1f} h26" stroke="{curve["color"]}" stroke-width="4"/>',
                      f'<text x="{legend_x+34}" y="{header_y:.1f}" fill="#3A3A3C" font-size="14">{curve["label"]}</text>']
            legend_x += 145
        for step in range(5):
            margin = y_max * step / 4
            _x, y = xy(0, margin)
            parts += [f'<path d="M{panel_left} {y:.1f} H{panel_left+plot_w}" stroke="#E5E5EA"/>',
                      f'<text x="{panel_left-12}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="12">{margin:.0f}</text>']
        for vdd_step in range(7):
            x, _y = xy(vdd_step * .2, 0)
            parts.append(f'<path d="M{x:.1f} {panel_top} V{panel_bottom}" stroke="#F1F1F4"/>')
        measured_vdd_guides: dict[float, float] = {}
        for curve_index, curve in enumerate(curves):
            points = [xy(row["vdd_v"], row["margin_mv"]) for row in curve["rows"]]
            point_string = " ".join(f"{x:.1f},{y:.1f}" for x, y in points)
            parts.append(f'<polyline points="{point_string}" fill="none" stroke="{curve["color"]}" stroke-width="3.5"/>')
            for point_index, (_row, (x, y)) in enumerate(zip(curve["rows"], points)):
                vdd = float(_row["vdd_v"])
                measured_vdd_guides[vdd] = min(y, measured_vdd_guides.get(vdd, panel_bottom))
                parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="#FFF" stroke="{curve["color"]}" stroke-width="2.2"/>')
                # Keep the two SNM series readable when their points are close:
                # Read labels sit above and Write labels below.  A label near
                # the panel edge is flipped back inside the plotting area.
                label_above = curve_index == 0
                label_y = y - 10 if label_above else y + 17
                if label_y < panel_top + 13:
                    label_y = y + 17
                elif label_y > panel_bottom - 5:
                    label_y = y - 10
                label_anchor = "middle"
                # Alternate a small horizontal offset for dense adjacent data.
                label_x = x + (-3 if point_index % 2 else 3)
                parts.append(
                    f'<text class="curve-data-label" data-series="{html.escape(str(curve["label"]), quote=True)}" '
                    f'x="{label_x:.1f}" y="{label_y:.1f}" text-anchor="{label_anchor}" '
                    f'fill="{curve["color"]}" font-size="11" font-weight="700">'
                    f'{float(_row["margin_mv"]):.1f} mV</text>')
        # One measured-VDD guide per voltage and panel.  Read and Write SNM
        # share the same guide in the upper panel instead of duplicating it.
        for voltage, guide_y in sorted(measured_vdd_guides.items()):
            x, _y = xy(voltage, 0)
            parts += [
                f'<path class="measured-vdd-guide" data-vdd="{voltage:.4f}" '
                f'd="M{x:.1f} {guide_y+6:.1f}V{panel_bottom:.1f}" '
                'stroke="#8E8E93" stroke-width="1.4" stroke-dasharray="5 5" opacity=".78"/>',
                f'<text class="vertical-vdd-label" x="{x:.1f}" y="{panel_bottom+23:.1f}" '
                f'text-anchor="middle" fill="#0062CC" font-size="12" font-weight="700">'
                f'{voltage:.2f} V</text>',
            ]
        center_y = panel_top + panel_h / 2
        axis_x = panel_left - 54
        parts += [f'<text x="{axis_x:.1f}" y="{center_y:.1f}" transform="rotate(-90 {axis_x:.1f} {center_y:.1f})" text-anchor="middle" fill="#1D1D1F" font-size="16" font-weight="700">{y_axis_label}</text>',
                  f'<text x="{panel_left+plot_w/2:.1f}" y="{height-18}" text-anchor="middle" fill="#1D1D1F" font-size="17" font-weight="700">Model VDD (V)</text>']
    parts.append('</svg>')
    return "".join(parts)


def butterfly_svg(items: list[dict], vdd: float, width: int = 900, height: int = 590) -> str:
    left, top, right, bottom = 62, 25, 18, 55
    pw, ph = width-left-right, height-top-bottom
    def xy(x: float, y: float) -> tuple[float, float]:
        return left + x/vdd*pw, top + (1-y/vdd)*ph
    parts = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Read butterfly curve" style="font-family:Calibri,Arial,sans-serif">',
             '<rect width="100%" height="100%" fill="white"/>']
    for i in range(6):
        val = vdd*i/5; x, y = xy(val, val)
        parts.append(f'<path d="M {x:.1f} {top} V {top+ph} M {left} {y:.1f} H {left+pw}" stroke="#e5e7eb" stroke-width="1"/>')
        parts.append(f'<text x="{x:.1f}" y="{top+ph+25}" text-anchor="middle" font-size="15">{val:.2f}</text>')
        parts.append(f'<text x="{left-9}" y="{y+5:.1f}" text-anchor="end" font-size="15">{val:.2f}</text>')
    for idx, item in enumerate(items):
        pts = item["read_vtc"]
        p1 = " ".join(f"{xy(x,y)[0]:.1f},{xy(x,y)[1]:.1f}" for x,y in pts)
        p2 = " ".join(f"{xy(y,x)[0]:.1f},{xy(y,x)[1]:.1f}" for x,y in pts)
        c = COLORS[idx]
        parts.append(f'<polyline points="{p1}" fill="none" stroke="{c}" stroke-width="1.7"/>')
        parts.append(f'<polyline points="{p2}" fill="none" stroke="{c}" stroke-width="1.7" opacity=".82"/>')
        lx = left+10+(idx%2)*235; ly = top+17+(idx//2)*20
        parts.append(f'<path d="M {lx} {ly-5} h 25" stroke="{c}" stroke-width="4"/><text x="{lx+32}" y="{ly}" font-size="15">{html.escape(item["label"])}</text>')
    parts += [f'<text x="{left+pw/2}" y="{height-8}" text-anchor="middle" font-size="17">Q (V)</text>',
              f'<text x="18" y="{top+ph/2}" transform="rotate(-90 18 {top+ph/2})" text-anchor="middle" font-size="17">QB (V)</text>', '</svg>']
    return "".join(parts)


def bar_svg(items: list[dict], key: str, title: str, unit: str, width: int = 900, height: int = 380) -> str:
    vals = [x["metrics"][key] for x in items]
    finite = [v for v in vals if v is not None]
    ymax = max(finite or [1.0]) * 1.18 or 1.0
    left, top, bottom = 55, 45, 55
    pw, ph = width-left-15, height-top-bottom
    barw = pw/len(items)*0.58
    p = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Arial,sans-serif"><rect width="100%" height="100%" fill="white"/>',
         f'<text x="{width/2}" y="30" text-anchor="middle" font-size="19" font-weight="650">{html.escape(title)}</text>',
         f'<path d="M {left} {top} V {top+ph} H {left+pw}" fill="none" stroke="#374151"/>']
    for i, (item, val) in enumerate(zip(items, vals)):
        x = left + (i+.5)*pw/len(items); h = 0 if val is None else val/ymax*ph
        p.append(f'<rect x="{x-barw/2:.1f}" y="{top+ph-h:.1f}" width="{barw:.1f}" height="{h:.1f}" fill="{COLORS[i]}" opacity=".86"/>')
        p.append(f'<text x="{x:.1f}" y="{top+ph-h-7:.1f}" text-anchor="middle" font-size="15">{_fmt(val)}</text>')
        p.append(f'<text x="{x:.1f}" y="{top+ph+22}" transform="rotate(18 {x:.1f} {top+ph+22})" text-anchor="start" font-size="14">{html.escape(item["label"])}</text>')
    p.append(f'<text x="18" y="{top+ph/2}" transform="rotate(-90 18 {top+ph/2})" text-anchor="middle" font-size="15">{unit}</text></svg>')
    return "".join(p)


def snm_overview_svg(result: dict, width: int = 1440, height: int = 720) -> str:
    """Read VTC comparison using Vin on X and Vout on Y, both in volts."""
    current = result["baseline_6t"]
    has_target = bool(result.get("datasheet_targets") and result.get("target_6t"))
    target = result.get("target_6t") if has_target else None
    axis_max = SNM_PLOT_AXIS_MAX_V
    lot_label_raw = str(result["wat"]["corner"])
    lot_label = html.escape(lot_label_raw)
    analytical = result.get("analytical_read_snm_comparison", {})
    plot_left, plot_top, plot_w, plot_h = 145, 145, 1140, 455

    def xy(vin: float, vout: float) -> tuple[float, float]:
        return (plot_left + vin / axis_max * plot_w,
                plot_top + (1.0 - vout / axis_max) * plot_h)

    current_value = current["metrics"]["read_snm_mv"]
    target_value = target["metrics"]["read_snm_mv"] if target else None
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Read SNM Lot/Wafer versus target VTC comparison" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        f'<text x="54" y="48" fill="#1D1D1F" font-size="30" font-weight="700">{"Read SNM Target Comparison" if has_target else "Read SNM Analysis"}</text>',
        f'<path d="M54 82 h34" stroke="#007AFF" stroke-width="4"/><text x="98" y="88" fill="#3A3A3C" font-size="17">{lot_label} WAT VTC</text>',
        f'<path d="M270 82 h34" stroke="#007AFF" stroke-width="4" stroke-dasharray="10 7"/><text x="314" y="88" fill="#3A3A3C" font-size="17">{lot_label} mirrored VTC</text>',
    ]
    if has_target:
        parts += [
            '<path d="M570 82 h34" stroke="#FF9500" stroke-width="4"/><text x="614" y="88" fill="#3A3A3C" font-size="17">WAT Target VTC</text>',
            '<path d="M842 82 h34" stroke="#FF9500" stroke-width="4" stroke-dasharray="10 7"/><text x="886" y="88" fill="#3A3A3C" font-size="17">WAT Target mirrored VTC</text>',
        ]
    for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
        px, py = xy(voltage, voltage)
        parts += [
            f'<path d="M{px:.1f} {plot_top} V{plot_top+plot_h} M{plot_left} {py:.1f} H{plot_left+plot_w}" stroke="#E5E5EA" stroke-width="1"/>',
            f'<text x="{px:.1f}" y="{plot_top+plot_h+27}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.2f}</text>',
            f'<text x="{plot_left-12}" y="{py+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="15">{voltage:.2f}</text>',
        ]
    datasets = [(current, "#007AFF")] + ([(target, "#FF9500")] if target else [])
    for data, color in datasets:
        direct_curve, mirrored_curve = _read_vtc_pair(data)
        direct = " ".join(f'{xy(vin,vout)[0]:.1f},{xy(vin,vout)[1]:.1f}' for vin, vout in direct_curve)
        mirrored = " ".join(f'{xy(vin,vout)[0]:.1f},{xy(vin,vout)[1]:.1f}' for vin, vout in mirrored_curve)
        parts += [f'<polyline points="{direct}" fill="none" stroke="{color}" stroke-width="4"/>',
                  f'<polyline points="{mirrored}" fill="none" stroke="{color}" stroke-width="4" stroke-dasharray="10 7" opacity=".9"/>']
    analytical_current = analytical.get("current_snm_mv")
    analytical_target = analytical.get("target_snm_mv") if has_target else None
    analytical_text = (f'Analytical RSNM: {lot_label_raw} {analytical_current:.1f} mV · WAT Target {analytical_target:.1f} mV'
                       if analytical_current is not None and analytical_target is not None
                       else 'Analytical RSNM: N/A for this input set')
    if analytical_current is not None and not has_target:
        analytical_text = f'Analytical RSNM: {lot_label_raw} {analytical_current:.1f} mV'
    parts += [
        f'<text x="{plot_left+plot_w-8}" y="{plot_top+30}" text-anchor="end" fill="#007AFF" font-size="18" font-weight="700">{lot_label} {current_value:.1f} mV</text>',
        *([f'<text x="{plot_left+plot_w-8}" y="{plot_top+58}" text-anchor="end" fill="#C56A00" font-size="18" font-weight="700">WAT Target {target_value:.1f} mV</text>',
           f'<text x="{plot_left+plot_w-8}" y="{plot_top+86}" text-anchor="end" fill="#1D1D1F" font-size="17" font-weight="700">Delta {current_value-target_value:+.1f} mV</text>']
          if target_value is not None else []),
        f'<text x="{plot_left+plot_w-8}" y="{plot_top+114}" text-anchor="end" fill="#5856D6" font-size="16" font-weight="700">{html.escape(analytical_text)}</text>',
        f'<text x="{plot_left+plot_w/2}" y="{height-50}" text-anchor="middle" fill="#1D1D1F" font-size="19">Vin (V)</text>',
        f'<text x="48" y="{plot_top+plot_h/2}" transform="rotate(-90 48 {plot_top+plot_h/2})" text-anchor="middle" fill="#1D1D1F" font-size="19">Vout (V)</text>',
        '<text x="720" y="700" text-anchor="middle" fill="#6E6E73" font-size="15">Read bias uses the configured WL and BL/BLB levels. Axis values are actual volts.</text></svg>',
    ]
    return "".join(parts)


def write_wsnm_states_svg(result: dict, width: int = 1440, height: int = 820) -> str:
    """Render W0 and W1 write-SNM panels with their state-specific squares."""
    current = result["baseline_6t"]["write_wsnm"]
    has_target = bool(result.get("datasheet_targets") and result.get("target_6t"))
    target = result.get("target_6t", {}).get("write_wsnm") if has_target else None
    cfg, wat = result["config"], result["wat"]
    axis_max = SNM_PLOT_AXIS_MAX_V
    # Keep the drawing scale identical on both axes so the WSNM marker is a
    # physical square as well as a voltage square.
    margin_x, gap, panel_w = 95, 150, 500
    top, size = 160, 500

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="W0 W1 Write SNM analysis" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="50" fill="#1D1D1F" font-size="32" font-weight="700">W0 / W1 Write SNM Analysis</text>',
        f'<path d="M54 86 h34" stroke="#007AFF" stroke-width="4"/><text x="98" y="92" fill="#3A3A3C" font-size="17">{html.escape(str(wat["corner"]))} write VTC</text>',
        f'<path d="M300 86 h34" stroke="#5856D6" stroke-width="4" stroke-dasharray="10 7"/><text x="344" y="92" fill="#3A3A3C" font-size="17">{html.escape(str(wat["corner"]))} mirrored VTC</text>',
        '<path d="M598 86 h34" stroke="#3A3A3C" stroke-width="3" stroke-dasharray="8 6"/><text x="642" y="92" fill="#3A3A3C" font-size="17">Vout = Vin</text>',
        '<rect x="818" y="74" width="20" height="20" fill="#EFFAF2" stroke="#34C759" stroke-width="3"/><text x="850" y="92" fill="#3A3A3C" font-size="17">Maximum WSNM square</text>',
        '<path d="M1110 86 h34" stroke="#FF9500" stroke-width="4"/><text x="1154" y="92" fill="#3A3A3C" font-size="17">WAT Target pair</text>' if target else '',
    ]

    for panel_index, (state_key, title, condition) in enumerate((
            ("write_0", "W0: write Q = 0", "BL=0, BLB=VDD, WL=VDD"),
            ("write_1", "W1: write QB = 0", "BL=VDD, BLB=0, WL=VDD"))):
        panel_left = margin_x + panel_index * (panel_w + gap)
        current_state = current[state_key]
        target_state = target[state_key] if target else None

        def xy(vin: float, vout: float) -> tuple[float, float]:
            return (panel_left + vin / axis_max * panel_w,
                    top + (1.0 - vout / axis_max) * size)

        parts += [
            f'<text x="{panel_left}" y="128" fill="#1D1D1F" font-size="23" font-weight="700">{title}</text>',
            f'<text x="{panel_left}" y="151" fill="#6E6E73" font-size="15">{condition}</text>',
        ]
        for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
            px, py = xy(voltage, voltage)
            parts += [f'<path d="M{px:.1f} {top} V{top+size} M{panel_left} {py:.1f} H{panel_left+panel_w}" stroke="#E5E5EA" stroke-width="1"/>',
                      f'<text x="{px:.1f}" y="{top+size+27}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.2f}</text>']
        polyline = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in current_state["curve"])
        mirrored_polyline = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in current_state["mirrored_curve"])
        parts += [f'<polyline points="{polyline}" fill="none" stroke="#007AFF" stroke-width="4"/>',
                  f'<polyline points="{mirrored_polyline}" fill="none" stroke="#5856D6" stroke-width="4" stroke-dasharray="10 7"/>']
        if target_state:
            polyline = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in target_state["curve"])
            mirrored_polyline = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in target_state["mirrored_curve"])
            parts += [f'<polyline points="{polyline}" fill="none" stroke="#FF9500" stroke-width="3" opacity=".88"/>',
                      f'<polyline points="{mirrored_polyline}" fill="none" stroke="#FF2D55" stroke-width="3" stroke-dasharray="10 7" opacity=".88"/>']
        square = current_state.get("limiting_square")
        if square is not None and square["side_v"] > 0:
            side = square["side_v"]
            x0, y0 = xy(square["x_v"], square["y_v"] + side)
            side_px = side / axis_max * panel_w
            side_py = side / axis_max * size
            parts += [f'<rect x="{x0:.1f}" y="{y0:.1f}" width="{side_px:.1f}" height="{side_py:.1f}" fill="#EFFAF2" fill-opacity=".70" stroke="#34C759" stroke-width="3"/>',
                      f'<path d="M{x0:.1f} {y0+side_py:.1f} L{x0+side_px:.1f} {y0:.1f}" stroke="#34C759" stroke-width="2"/>',
                      f'<text x="{x0+side_px/2:.1f}" y="{y0+side_py/2+5:.1f}" text-anchor="middle" fill="#1D1D1F" font-size="15" font-weight="700">{current_state["snm_mv"]:.1f} mV</text>']
        parts.append(f'<path d="M{xy(0,0)[0]:.1f} {xy(0,0)[1]:.1f} L{xy(axis_max,axis_max)[0]:.1f} {xy(axis_max,axis_max)[1]:.1f}" stroke="#3A3A3C" stroke-width="3" stroke-dasharray="8 6"/>')
        current_value = current_state["snm_mv"]
        target_value = target_state["snm_mv"] if target_state else None
        metric_text = f'Lot/Wafer WSNM = {current_value:.1f} mV'
        if target_value is not None:
            metric_text += f'   |   Target = {target_value:.1f} mV   |   Δ = {current_value-target_value:+.1f} mV'
        parts += [f'<text x="{panel_left+panel_w/2:.1f}" y="{top+size+60}" text-anchor="middle" fill="#1D1D1F" font-size="17" font-weight="700">{metric_text}</text>',
                  f'<text x="{panel_left+panel_w/2:.1f}" y="{top+size+92}" text-anchor="middle" fill="#1D1D1F" font-size="19">Vin (V)</text>']
        parts.append(f'<text x="{panel_left-42}" y="{top+size/2}" transform="rotate(-90 {panel_left-42} {top+size/2})" text-anchor="middle" fill="#1D1D1F" font-size="19">Vout (V)</text>')
    parts += [f'<text x="720" y="780" text-anchor="middle" fill="#6E6E73" font-size="15">Cell WSNM = min(WSNM_W0, WSNM_W1). Write-state margin is evaluated separately for both data polarities.</text>', '</svg>']
    return "".join(part for part in parts if part)


def write_wsnm_window_svg(result: dict, width: int = 1440, height: int = 900) -> str:
    """Render one W=1/W=0 write-VTC window and its maximum WSNM square."""
    current = result["baseline_6t"]["write_wsnm"]
    has_target = bool(result.get("datasheet_targets") and result.get("target_6t"))
    target = result.get("target_6t", {}).get("write_wsnm") if has_target else None
    axis_max, left, top, size = SNM_PLOT_AXIS_MAX_V, 210, 190, 560

    def xy(vin: float, vout: float) -> tuple[float, float]:
        return left + vin / axis_max * size, top + (1.0 - vout / axis_max) * size

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="W1 W0 Write SNM butterfly analysis" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="50" fill="#1D1D1F" font-size="32" font-weight="700">Write SNM Butterfly Analysis</text>',
        '<path d="M54 88 h34" stroke="#007AFF" stroke-width="4"/><text x="98" y="94" fill="#3A3A3C" font-size="17">W=1 VTC (upper)</text>',
        '<path d="M300 88 h34" stroke="#5856D6" stroke-width="4"/><text x="344" y="94" fill="#3A3A3C" font-size="17">W=0 VTC (lower)</text>',
        '<rect x="54" y="112" width="20" height="20" fill="#EFFAF2" stroke="#34C759" stroke-width="3"/><text x="86" y="130" fill="#3A3A3C" font-size="17">Vin=Vout diagonal-constrained WSNM square</text>',
        '<path d="M575 122 h34" stroke="#FF9500" stroke-width="4"/><text x="619" y="128" fill="#3A3A3C" font-size="17">WAT Target pair</text>' if target else '',
        f'<text x="54" y="162" fill="#1D1D1F" font-size="19" font-weight="700">WSNM @ VDD = {current["vdd_v"]:.3f} V</text>',
    ]
    for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
        px, py = xy(voltage, voltage)
        parts += [f'<path d="M{px:.1f} {top} V{top+size} M{left} {py:.1f} H{left+size}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{px:.1f}" y="{top+size+27}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.2f}</text>',
                  f'<text x="{left-14}" y="{py+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="15">{voltage:.2f}</text>']
    for data, upper_color, lower_color in ((current, "#007AFF", "#5856D6"),
                                             *(([(target, "#FF9500", "#FF2D55")] if target else []))):
        upper = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in data["write_1"]["curve"])
        lower = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in data["write_0"]["curve"])
        parts += [f'<polyline points="{upper}" fill="none" stroke="{upper_color}" stroke-width="4"/>',
                  f'<polyline points="{lower}" fill="none" stroke="{lower_color}" stroke-width="4"/>']
    square = current.get("write_square")
    if square is not None:
        side = square["side_v"]
        x0, y0 = xy(square["x_v"], square["y_v"] + side)
        side_px = side / axis_max * size
        parts += [f'<rect x="{x0:.1f}" y="{y0:.1f}" width="{side_px:.1f}" height="{side_px:.1f}" fill="#EFFAF2" fill-opacity=".70" stroke="#34C759" stroke-width="3"/>',
                  f'<text x="{x0+side_px/2:.1f}" y="{y0+side_px/2+5:.1f}" text-anchor="middle" fill="#1D1D1F" font-size="18" font-weight="700">WSNM {current["snm_mv"]:.1f} mV</text>']
    parts += [f'<text x="{left+size/2:.1f}" y="{top+size+76}" text-anchor="middle" fill="#1D1D1F" font-size="19">Vin (V)</text>',
              f'<text x="{left-62}" y="{top+size/2}" transform="rotate(-90 {left-62} {top+size/2})" text-anchor="middle" fill="#1D1D1F" font-size="19">Vout (V)</text>',
              '<text x="720" y="875" text-anchor="middle" fill="#6E6E73" font-size="15">W=1 is the BLB-high upper VTC; W=0 is the BL-low lower VTC. The WSNM square diagonal is constrained to Vin=Vout.</text>',
              '</svg>']
    return "".join(part for part in parts if part)


def write_butterfly_svg(result: dict, width: int = 1440, height: int = 720) -> str:
    """Plot the two write-biased 6T VTCs as a butterfly, without square fitting."""
    current = result["baseline_6t"]
    has_target = bool(result.get("datasheet_targets") and result.get("target_6t"))
    target = result.get("target_6t") if has_target else None
    cfg, wat = result["config"], result["wat"]
    axis_max = SNM_PLOT_AXIS_MAX_V
    left, top, pw, ph = 145, 145, 1140, 455

    def xy(vin: float, vout: float) -> tuple[float, float]:
        return left + vin / axis_max * pw, top + (1.0 - vout / axis_max) * ph

    lot_label = html.escape(str(wat["corner"]))
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="6T Write Butterfly Curve" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="48" fill="#1D1D1F" font-size="30" font-weight="700">6T Write Butterfly Curve</text>',
        f'<path d="M54 82 h34" stroke="#007AFF" stroke-width="4"/><text x="98" y="88" fill="#3A3A3C" font-size="17">{lot_label} BL-low VTC</text>',
        f'<path d="M310 82 h34" stroke="#AF52DE" stroke-width="4" stroke-dasharray="10 7"/><text x="354" y="88" fill="#3A3A3C" font-size="17">{lot_label} mirrored BLB-high VTC</text>',
        '<path d="M690 82 h34" stroke="#3A3A3C" stroke-width="3" stroke-dasharray="8 6"/><text x="734" y="88" fill="#3A3A3C" font-size="17">Vout = Vin</text>',
    ]
    if target:
        parts += ['<path d="M925 82 h34" stroke="#FF9500" stroke-width="4"/><text x="969" y="88" fill="#3A3A3C" font-size="17">WAT Target VTC pair</text>']
    for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
        px, py = xy(voltage, voltage)
        parts += [f'<path d="M{px:.1f} {top} V{top+ph} M{left} {py:.1f} H{left+pw}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{px:.1f}" y="{top+ph+27}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.2f}</text>',
                  f'<text x="{left-12}" y="{py+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="15">{voltage:.2f}</text>']
    for data, direct_color, mirror_color in ((current, "#007AFF", "#AF52DE"),
                                               *(([(target, "#FF9500", "#FF2D55")] if target else []))):
        curves = data["write_butterfly"]
        direct = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in curves["direct_vtc"])
        mirrored = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in curves["mirrored_vtc"])
        parts += [f'<polyline points="{direct}" fill="none" stroke="{direct_color}" stroke-width="4"/>',
                  f'<polyline points="{mirrored}" fill="none" stroke="{mirror_color}" stroke-width="4" stroke-dasharray="10 7"/>']
    parts += [
        f'<path d="M{xy(0,0)[0]:.1f} {xy(0,0)[1]:.1f} L{xy(axis_max,axis_max)[0]:.1f} {xy(axis_max,axis_max)[1]:.1f}" stroke="#3A3A3C" stroke-width="3" stroke-dasharray="8 6"/>',
        f'<text x="{left+pw/2}" y="{height-50}" text-anchor="middle" fill="#1D1D1F" font-size="19">Vin (V)</text>',
        f'<text x="48" y="{top+ph/2}" transform="rotate(-90 48 {top+ph/2})" text-anchor="middle" fill="#1D1D1F" font-size="19">Vout (V)</text>',
        f'<text x="720" y="680" text-anchor="middle" fill="#6E6E73" font-size="15">Write bias: WL={cfg["write_wordline_over_vdd"]:.2f} × VDD, BL={cfg["write_low_bitline_over_vdd"]:.2f} × VDD, BLB={cfg["write_high_bitline_over_vdd"]:.2f} × VDD.</text>',
        '</svg>',
    ]
    return "".join(parts)


def _legacy_write_vtc_svg(result: dict, width: int = 1440, height: int = 720) -> str:
    """Write-condition VTC comparison: BL low side versus mirrored BLB high side."""
    current = result["baseline_6t"]
    target = result.get("target_6t", current)
    cfg, wat = result["config"], result["wat"]
    vdd, lot_raw = cfg["nominal_vdd"], str(wat["corner"])
    axis_max = SNM_PLOT_AXIS_MAX_V
    lot_label = html.escape(lot_raw)
    left, top, pw, ph = 145, 145, 1140, 455

    def xy(vin: float, vout: float) -> tuple[float, float]:
        return left + vin / axis_max * pw, top + (1.0 - vout / axis_max) * ph

    current_value = current["metrics"]["write_snm_proxy_mv"]
    target_value = target["metrics"]["write_snm_proxy_mv"]
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Write SNM target comparison" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="48" fill="#1D1D1F" font-size="30" font-weight="700">Write SNM Target Comparison</text>',
        f'<path d="M54 82 h34" stroke="#007AFF" stroke-width="4"/><text x="98" y="88" fill="#3A3A3C" font-size="17">{lot_label} BL-low VTC</text>',
        f'<path d="M270 82 h34" stroke="#AF52DE" stroke-width="4" stroke-dasharray="10 7"/><text x="314" y="88" fill="#3A3A3C" font-size="17">{lot_label} mirrored BLB-high VTC</text>',
        '<path d="M650 82 h34" stroke="#FF9500" stroke-width="4"/><text x="694" y="88" fill="#3A3A3C" font-size="17">WAT Target write VTC pair</text>',
    ]
    for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
        px, py = xy(voltage, voltage)
        parts += [f'<path d="M{px:.1f} {top} V{top+ph} M{left} {py:.1f} H{left+pw}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{px:.1f}" y="{top+ph+27}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.2f}</text>',
                  f'<text x="{left-12}" y="{py+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="15">{voltage:.2f}</text>']
    for data, color, mirror_color in ((current, "#007AFF", "#AF52DE"),
                                      (target, "#FF9500", "#FF2D55")):
        low = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in data["write_vtc_low"])
        high = " ".join(f'{xy(y, x)[0]:.1f},{xy(y, x)[1]:.1f}' for x, y in data["write_vtc_high"])
        parts += [f'<polyline points="{low}" fill="none" stroke="{color}" stroke-width="4"/>',
                  f'<polyline points="{high}" fill="none" stroke="{mirror_color}" stroke-width="4" stroke-dasharray="10 7" opacity=".95"/>']
    parts += [
        f'<text x="{left+pw-8}" y="{top+30}" text-anchor="end" fill="#007AFF" font-size="18" font-weight="700">{lot_label} WSNM proxy {current_value:.1f} mV</text>',
        f'<text x="{left+pw-8}" y="{top+58}" text-anchor="end" fill="#C56A00" font-size="18" font-weight="700">WAT Target {target_value:.1f} mV</text>',
        f'<text x="{left+pw-8}" y="{top+86}" text-anchor="end" fill="#1D1D1F" font-size="17" font-weight="700">Delta {current_value-target_value:+.1f} mV</text>',
        f'<text x="{left+pw/2}" y="{height-50}" text-anchor="middle" fill="#1D1D1F" font-size="19">Vin (V)</text>',
        f'<text x="48" y="{top+ph/2}" transform="rotate(-90 48 {top+ph/2})" text-anchor="middle" fill="#1D1D1F" font-size="19">Vout (V)</text>',
        f'<text x="720" y="700" text-anchor="middle" fill="#6E6E73" font-size="15">Write bias: WL={cfg["write_wordline_over_vdd"]:.2f}×VDD, BL={cfg["write_low_bitline_over_vdd"]:.2f}×VDD, BLB={cfg["write_high_bitline_over_vdd"]:.2f}×VDD. WSNM is a WAT-calibrated bitline-noise proxy.</text>',
        '</svg>',
    ]
    return "".join(parts)


def write_snm_overview_svg(result: dict, width: int = 1440, height: int = 720) -> str:
    """Textbook writeability chart: butterfly SNM versus swept write BL."""
    current = result["baseline_6t"]
    has_target = bool(result.get("datasheet_targets") and result.get("target_6t"))
    target = result.get("target_6t") if has_target else None
    cfg, wat = result["config"], result["wat"]
    vdd = cfg["nominal_vdd"]
    lot_label = html.escape(str(wat["corner"]))
    axis_max = SNM_PLOT_AXIS_MAX_V
    left, top, pw, ph = 145, 155, 1140, 430
    current_sweep = current["write_snm_vs_bitline"]
    target_sweep = target["write_snm_vs_bitline"] if target else None
    values = [row["cell_write_snm_mv"]
              for sweep in ([current_sweep] + ([target_sweep] if target_sweep else []))
              for row in sweep["points"]]
    y_max = max(50.0, math.ceil(max(values, default=0.0) * 1.20 / 50.0) * 50.0)

    def xy(bitline_v: float, snm_mv: float) -> tuple[float, float]:
        return (left + bitline_v / axis_max * pw,
                top + (1.0 - snm_mv / y_max) * ph)

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Write SNM versus write bitline voltage" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<defs><marker id="writeArrow" markerWidth="8" markerHeight="8" refX="6" refY="3" orient="auto"><path d="M0,0 L0,6 L6,3 z" fill="#6E6E73"/></marker></defs>',
        '<text x="54" y="48" fill="#1D1D1F" font-size="30" font-weight="700">Write SNM versus Write-Bitline Voltage</text>',
        '<text x="54" y="78" fill="#6E6E73" font-size="16">WL=VDD and BLB=VDD; BL is swept from VDD toward 0 V. The write boundary occurs when the limiting butterfly SNM reaches 0.</text>',
        f'<path d="M54 112 h34" stroke="#007AFF" stroke-width="4"/><text x="98" y="118" fill="#3A3A3C" font-size="17">{lot_label} WAT model</text>',
    ]
    if has_target:
        parts.append('<path d="M325 112 h34" stroke="#FF9500" stroke-width="4"/><text x="369" y="118" fill="#3A3A3C" font-size="17">WAT Target model</text>')
    for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
        px, _ = xy(voltage, 0.0)
        parts += [f'<path d="M{px:.1f} {top} V{top+ph}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{px:.1f}" y="{top+ph+27}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.2f}</text>']
    for index in range(6):
        value = y_max * index / 5.0
        _, py = xy(0.0, value)
        parts += [f'<path d="M{left} {py:.1f} H{left+pw}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{left-12}" y="{py+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="15">{value:.0f}</text>']

    sweep_datasets = [(current_sweep, "#007AFF")] + ([(target_sweep, "#FF9500")] if target_sweep else [])
    for sweep, color in sweep_datasets:
        curve = " ".join(
            f'{xy(row["write_bl_v"], row["cell_write_snm_mv"])[0]:.1f},'
            f'{xy(row["write_bl_v"], row["cell_write_snm_mv"])[1]:.1f}'
            for row in sweep["points"])
        parts.append(f'<polyline points="{curve}" fill="none" stroke="{color}" stroke-width="4" stroke-linejoin="round"/>')
        trip = sweep.get("write_trip_bl_v")
        if trip is not None:
            trip_x, zero_y = xy(trip, 0.0)
            parts += [
                f'<path d="M{trip_x:.1f} {top} V{top+ph}" stroke="{color}" stroke-width="2" stroke-dasharray="7 6" opacity=".85"/>',
                f'<circle cx="{trip_x:.1f}" cy="{zero_y:.1f}" r="7" fill="#FFFFFF" stroke="{color}" stroke-width="4"/>',
            ]

    def sweep_text(label: str, sweep: dict) -> str:
        trip = sweep.get("write_trip_bl_v")
        swing = sweep.get("required_bl_swing_v")
        if trip is None or swing is None:
            return f'{label} Write Trip BL N/A'
        return f'{label} Write Trip BL {trip:.4f} V; required swing {swing:.4f} V'

    vdd_x, arrow_y = xy(vdd, y_max * 0.08)
    zero_x, _ = xy(0.0, y_max * 0.08)
    parts += [
        f'<text x="{left+pw-8}" y="{top+30}" text-anchor="end" fill="#007AFF" font-size="17" font-weight="700">{sweep_text(lot_label, current_sweep)}</text>',
        *([f'<text x="{left+pw-8}" y="{top+58}" text-anchor="end" fill="#C56A00" font-size="17" font-weight="700">{sweep_text("WAT Target", target_sweep)}</text>'] if target_sweep else []),
        f'<path d="M{vdd_x:.1f} {arrow_y:.1f} H{zero_x+18:.1f}" stroke="#6E6E73" stroke-width="2" marker-end="url(#writeArrow)"/>',
        f'<text x="{(vdd_x+zero_x)/2:.1f}" y="{arrow_y-10:.1f}" text-anchor="middle" fill="#6E6E73" font-size="15">Write direction: BL decreases from VDD to 0 V</text>',
        f'<text x="{left+pw/2}" y="{height-58}" text-anchor="middle" fill="#1D1D1F" font-size="20">Write BL Voltage (V)</text>',
        f'<text x="48" y="{top+ph/2}" transform="rotate(-90 48 {top+ph/2})" text-anchor="middle" fill="#1D1D1F" font-size="20">Limiting Write SNM (mV)</text>',
        '<text x="720" y="700" text-anchor="middle" fill="#6E6E73" font-size="15">Compact WAT-calibrated estimate. The write-trip boundary is the BL voltage where the smaller 6T butterfly eye closes.</text>',
        '</svg>',
    ]
    return "".join(parts)


def single_wat_write_snm_geometry_svg(result: dict,
                                       width: int = 1120,
                                       height: int = 820) -> str:
    """Single 6T WAT write VTC with the reference-style WSNM square."""
    geometry = result["baseline_6t"]["single_wat_write_snm_geometry"]
    cfg, wat = result["config"], result["wat"]
    axis_max = SNM_PLOT_AXIS_MAX_V
    left, top, size = 190, 170, 560

    def xy(x_value: float, y_value: float) -> tuple[float, float]:
        return (left + x_value / axis_max * size,
                top + (1.0 - y_value / axis_max) * size)

    lot_label = html.escape(str(wat["corner"]))
    polarity = html.escape(str(geometry.get("write_polarity") or "N/A"))
    curve = geometry.get("curve", [])
    curve_points = " ".join(f'{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}' for x, y in curve)
    wsnm_v = geometry.get("wsnm_v")
    wsnm_mv = geometry.get("wsnm_mv")
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Single WAT Write SNM geometry" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="52" fill="#1D1D1F" font-size="32" font-weight="700">6T Single-WAT Write SNM</text>',
        f'<text x="54" y="84" fill="#6E6E73" font-size="17">{lot_label}; one write-condition VTC from the WAT-calibrated 6T model</text>',
        '<path d="M54 116 h38" stroke="#007AFF" stroke-width="5"/><text x="104" y="123" fill="#3A3A3C" font-size="18">Write-condition VTC</text>',
        '<rect x="350" y="101" width="27" height="27" fill="#FFF6E8" stroke="#FF9500" stroke-width="3"/><text x="390" y="123" fill="#3A3A3C" font-size="18">WSNM geometry</text>',
    ]
    for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
        px, py = xy(voltage, voltage)
        parts += [
            f'<path d="M{px:.1f} {top} V{top+size} M{left} {py:.1f} H{left+size}" stroke="#E5E5EA" stroke-width="1"/>',
            f'<text x="{px:.1f}" y="{top+size+30}" text-anchor="middle" fill="#6E6E73" font-size="17">{voltage:.2f}</text>',
            f'<text x="{left-14}" y="{py+6:.1f}" text-anchor="end" fill="#6E6E73" font-size="17">{voltage:.2f}</text>',
        ]
    parts.append(f'<polyline points="{curve_points}" fill="none" stroke="#007AFF" stroke-width="5" stroke-linejoin="round"/>')
    if wsnm_v is not None and wsnm_mv is not None:
        origin_x, origin_y = xy(0.0, 0.0)
        corner_x, corner_y = xy(wsnm_v, wsnm_v)
        side_px = wsnm_v / axis_max * size
        parts += [
            f'<rect x="{origin_x:.1f}" y="{corner_y:.1f}" width="{side_px:.1f}" height="{side_px:.1f}" fill="#FFF6E8" fill-opacity=".72" stroke="#FF9500" stroke-width="4"/>',
            f'<path d="M{origin_x:.1f} {origin_y:.1f} L{corner_x:.1f} {corner_y:.1f}" stroke="#FF9500" stroke-width="3"/>',
            f'<circle cx="{corner_x:.1f}" cy="{corner_y:.1f}" r="7" fill="#FFFFFF" stroke="#FF9500" stroke-width="4"/>',
            f'<text x="{origin_x+side_px/2:.1f}" y="{corner_y+side_px/2-8:.1f}" text-anchor="middle" fill="#9A4D00" font-size="19" font-weight="700">WSNM</text>',
            f'<text x="{origin_x+side_px/2:.1f}" y="{corner_y+side_px/2+18:.1f}" text-anchor="middle" fill="#9A4D00" font-size="18" font-weight="700">{wsnm_mv:.1f} mV</text>',
        ]
    else:
        parts.append(f'<text x="{left+size/2}" y="{top+size/2}" text-anchor="middle" fill="#FF3B30" font-size="24" font-weight="700">No valid VTC / diagonal crossing</text>')
    parts += [
        f'<text x="{left+size/2}" y="{height-12}" text-anchor="middle" fill="#1D1D1F" font-size="22">Vin (V)</text>',
        f'<text x="70" y="{top+size/2}" transform="rotate(-90 70 {top+size/2})" text-anchor="middle" fill="#1D1D1F" font-size="22">Vout (V)</text>',
        f'<text x="790" y="220" fill="#1D1D1F" font-size="18" font-weight="700">Write bias</text>',
        f'<text x="790" y="252" fill="#3A3A3C" font-size="17">WL = {cfg["write_wordline_over_vdd"]:.2f} x VDD</text>',
        f'<text x="790" y="282" fill="#3A3A3C" font-size="17">BL = {cfg["write_low_bitline_over_vdd"]:.2f} x VDD</text>',
        f'<text x="790" y="312" fill="#3A3A3C" font-size="17">BLB = {cfg["write_high_bitline_over_vdd"]:.2f} x VDD</text>',
        f'<text x="790" y="354" fill="#1D1D1F" font-size="18" font-weight="700">Limiting polarity</text>',
        f'<text x="790" y="385" fill="#3A3A3C" font-size="15">{polarity}</text>',
        '<text x="790" y="445" fill="#6E6E73" font-size="15">The square side is the VTC</text>',
        '<text x="790" y="468" fill="#6E6E73" font-size="15">intersection with Vout = Vin.</text>',
        '<text x="790" y="514" fill="#6E6E73" font-size="15">Compact WAT model estimate;</text>',
        '<text x="790" y="537" fill="#6E6E73" font-size="15">not measured WT sign-off.</text>',
        '</svg>',
    ]
    return "".join(parts)


def read_snm_butterfly_svg(result: dict, width: int = 1440, height: int = 820) -> str:
    """Read butterfly plots with independently fitted maximum squares."""
    current = result["baseline_6t"]
    has_target = bool(result.get("datasheet_targets") and result.get("target_6t"))
    target = result.get("target_6t") if has_target else None
    lot_label = str(result["wat"]["corner"])
    axis_max = SNM_PLOT_AXIS_MAX_V
    panels = [(lot_label, current, "#007AFF")]
    if target:
        panels.append(("WAT Target", target, "#FF9500"))
    margin_x, gap = 60, 64
    panel_w = (width - 2 * margin_x - gap) / 2
    plot_top, plot_h = 158, 480
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Read SNM butterfly maximum squares" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="48" fill="#1D1D1F" font-size="30" font-weight="700">Read SNM Butterfly Analysis</text>',
        '<path d="M54 82 h34" stroke="#3A3A3C" stroke-width="4"/><text x="98" y="88" fill="#3A3A3C" font-size="17">Right inverter VTC</text>',
        '<path d="M285 82 h34" stroke="#3A3A3C" stroke-width="4" stroke-dasharray="10 7"/><text x="329" y="88" fill="#3A3A3C" font-size="17">Inverse left inverter VTC</text>',
        '<rect x="600" y="68" width="24" height="24" fill="none" stroke="#34C759" stroke-width="3"/><text x="636" y="88" fill="#3A3A3C" font-size="17">Maximum squares 1 and 2</text>',
    ]
    for index, (title, data, color) in enumerate(panels):
        x0 = ((width - 670) / 2 if len(panels) == 1
              else margin_x + index * (panel_w + gap))
        # Equal X/Y pixel scale keeps a voltage square visually square.
        plot_left, plot_w = x0 + 95, 480

        def xy(x_value: float, y_value: float) -> tuple[float, float]:
            return (plot_left + x_value / axis_max * plot_w,
                    plot_top + (1.0 - y_value / axis_max) * plot_h)

        parts.append(
            f'<text x="{x0:.1f}" y="126" fill="#1D1D1F" font-size="25" font-weight="700">{html.escape(title)}</text>')
        for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
            px, py = xy(voltage, voltage)
            parts += [f'<path d="M{px:.1f} {plot_top} V{plot_top+plot_h} M{plot_left} {py:.1f} H{plot_left+plot_w}" stroke="#E5E5EA" stroke-width="1"/>',
                      f'<text x="{px:.1f}" y="{plot_top+plot_h+26}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.2f}</text>',
                      f'<text x="{plot_left-10}" y="{py+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="15">{voltage:.2f}</text>']
        direct_curve, mirrored_curve = _read_vtc_pair(data)
        direct = " ".join(f'{xy(x,y)[0]:.1f},{xy(x,y)[1]:.1f}' for x, y in direct_curve)
        mirrored = " ".join(f'{xy(x,y)[0]:.1f},{xy(x,y)[1]:.1f}' for x, y in mirrored_curve)
        parts += [f'<polyline points="{direct}" fill="none" stroke="{color}" stroke-width="4"/>',
                  f'<polyline points="{mirrored}" fill="none" stroke="{color}" stroke-width="4" stroke-dasharray="10 7"/>']

        squares = data["read_butterfly"]["squares"]
        limiting = min(square["side_v"] for square in squares)
        for square in squares:
            left, top = xy(square["x_v"], square["y_v"] + square["side_v"])
            side_px_x = square["side_v"] / axis_max * plot_w
            side_px_y = square["side_v"] / axis_max * plot_h
            stroke_width = 4 if abs(square["side_v"] - limiting) < 1e-12 else 3
            value_label = f'{square["side_mv"]:.1f} mV'
            parts += [f'<rect x="{left:.1f}" y="{top:.1f}" width="{side_px_x:.1f}" height="{side_px_y:.1f}" fill="#EFFAF2" stroke="#34C759" stroke-width="{stroke_width}"/>',
                      f'<text x="{left+side_px_x/2:.1f}" y="{top+side_px_y/2+5:.1f}" text-anchor="middle" fill="#1D1D1F" font-size="15" font-weight="700">{value_label}</text>']

        butterfly = data["read_butterfly"]
        state_text = (f'Upper {butterfly.get("snm_upper_left_mv", squares[0]["side_mv"]):.1f} mV · '
                      f'Lower {butterfly.get("snm_lower_right_mv", squares[1]["side_mv"]):.1f} mV · '
                      f'Asymmetry {_fmt(butterfly.get("mismatch_index_pct"), 1)}%')
        parts += [f'<text x="{x0:.1f}" y="{plot_top+plot_h+60}" fill="#3A3A3C" font-size="15" font-weight="700">{state_text}</text>',
                  f'<text x="{plot_left+plot_w/2:.1f}" y="{height-62}" text-anchor="middle" fill="#1D1D1F" font-size="18">Vin (V)</text>',
                  f'<text x="{x0+45:.1f}" y="{plot_top+plot_h/2:.1f}" transform="rotate(-90 {x0+45:.1f} {plot_top+plot_h/2:.1f})" text-anchor="middle" fill="#1D1D1F" font-size="18">Vout (V)</text>']
    parts.append(f'<text x="720" y="{height-14}" text-anchor="middle" fill="#6E6E73" font-size="15">Upper and lower margins represent opposite stored states; cell RSNM is the smaller value.</text></svg>')
    return "".join(parts)


def model_vdd_butterfly_svg(entries: list[dict], width: int = 1440) -> str:
    """Small-multiple Read butterflies with measured-WAT and WAT-target curves."""
    columns = 3
    rows = max(1, math.ceil(len(entries) / columns))
    panel_w, panel_h = 450, 405
    height = 125 + rows * panel_h + 45
    axis_max = SNM_PLOT_AXIS_MAX_V
    has_target = any(item["result"].get("datasheet_targets") for item in entries)
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="Read SNM butterflies by model VDD" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="48" fill="#1D1D1F" font-size="30" font-weight="700">Read SNM Butterfly by Model VDD</text>',
        '<path d="M54 82 h32" stroke="#007AFF" stroke-width="4"/><text x="96" y="88" fill="#3A3A3C" font-size="15">WAT Inverter VTC</text>',
        '<path d="M247 82 h32" stroke="#007AFF" stroke-width="4" stroke-dasharray="10 7"/><text x="289" y="88" fill="#3A3A3C" font-size="15">WAT Mirrored VTC</text>',
        '<rect x="452" y="69" width="21" height="21" fill="#EFFAF2" stroke="#34C759" stroke-width="3"/><text x="484" y="88" fill="#3A3A3C" font-size="15">WAT SNM square</text>',
    ]
    if has_target:
        parts += [
            '<path d="M650 82 h32" stroke="#FF9500" stroke-width="4"/><text x="692" y="88" fill="#3A3A3C" font-size="15">Target Inverter VTC</text>',
            '<path d="M865 82 h32" stroke="#FF9500" stroke-width="4" stroke-dasharray="10 7"/><text x="907" y="88" fill="#3A3A3C" font-size="15">Target Mirrored VTC</text>',
            '<rect x="1095" y="69" width="21" height="21" fill="none" stroke="#FF9500" stroke-width="3" stroke-dasharray="6 4"/><text x="1127" y="88" fill="#3A3A3C" font-size="15">Target SNM square</text>',
        ]
    for index, entry in enumerate(entries):
        col, row = index % columns, index // columns
        x0, y0 = 38 + col * 468, 125 + row * panel_h
        left, top, size = x0 + 52, y0 + 72, 250
        data = entry["result"]["baseline_6t"]
        target = entry["result"].get("target_6t", data)
        show_target = bool(entry["result"].get("datasheet_targets"))

        def xy(x: float, y: float) -> tuple[float, float]:
            return left + x / axis_max * size, top + (1 - y / axis_max) * size

        measured_snm = data["metrics"]["read_snm_mv"]
        target_snm = target["metrics"]["read_snm_mv"]
        parts += [f'<text x="{x0}" y="{y0+22}" fill="#1D1D1F" font-size="19" font-weight="700">Operating VDD = {entry["model_vdd_v"]:.3f} V</text>']
        if show_target:
            delta = measured_snm - target_snm
            parts.append(f'<text x="{x0}" y="{y0+47}" fill="#3A3A3C" font-size="14"><tspan fill="#007AFF" font-weight="700">WAT {measured_snm:.1f} mV</tspan><tspan> · </tspan><tspan fill="#FF9500" font-weight="700">Target {target_snm:.1f} mV</tspan><tspan> · Δ {delta:+.1f} mV</tspan></text>')
        else:
            parts.append(f'<text x="{x0}" y="{y0+47}" fill="#007AFF" font-size="14" font-weight="700">WAT RSNM {measured_snm:.1f} mV</text>')
        for voltage in (0.0, 0.60, axis_max):
            px, py = xy(voltage, voltage)
            parts += [f'<path d="M{px:.1f} {top} V{top+size} M{left} {py:.1f} H{left+size}" stroke="#E5E5EA" stroke-width="1"/>',
                      f'<text x="{px:.1f}" y="{top+size+19}" text-anchor="middle" fill="#6E6E73" font-size="12">{voltage:.2f}</text>',
                      f'<text x="{left-7}" y="{py+4:.1f}" text-anchor="end" fill="#6E6E73" font-size="12">{voltage:.2f}</text>']
        direct_curve, mirrored_curve = _read_vtc_pair(data)
        direct = " ".join(f"{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}" for x, y in direct_curve)
        mirrored = " ".join(f"{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}" for x, y in mirrored_curve)
        parts += [f'<polyline points="{direct}" fill="none" stroke="#007AFF" stroke-width="3"/>',
                  f'<polyline points="{mirrored}" fill="none" stroke="#007AFF" stroke-width="3" stroke-dasharray="8 5"/>']
        if show_target:
            target_direct_curve, target_mirrored_curve = _read_vtc_pair(target)
            target_direct = " ".join(f"{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}" for x, y in target_direct_curve)
            target_mirrored = " ".join(f"{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}" for x, y in target_mirrored_curve)
            parts += [f'<polyline points="{target_direct}" fill="none" stroke="#FF9500" stroke-width="2.5"/>',
                      f'<polyline points="{target_mirrored}" fill="none" stroke="#FF9500" stroke-width="2.5" stroke-dasharray="8 5"/>']
        for square in data["read_butterfly"]["squares"]:
            sx, sy = xy(square["x_v"], square["y_v"] + square["side_v"])
            side = square["side_v"] / axis_max * size
            label_y = sy - 7 if square["lobe"] == 1 else sy + side + 15
            parts += [f'<rect x="{sx:.1f}" y="{sy:.1f}" width="{side:.1f}" height="{side:.1f}" fill="#EFFAF2" stroke="#34C759" stroke-width="3"/>',
                      f'<text x="{sx+side/2:.1f}" y="{label_y:.1f}" text-anchor="middle" fill="#248A3D" font-size="11" font-weight="700">WAT {square["side_mv"]:.1f} mV</text>']
        if show_target:
            for square in target["read_butterfly"]["squares"]:
                sx, sy = xy(square["x_v"], square["y_v"] + square["side_v"])
                side = square["side_v"] / axis_max * size
                label_y = sy - 20 if square["lobe"] == 1 else sy + side + 28
                parts += [f'<rect x="{sx:.1f}" y="{sy:.1f}" width="{side:.1f}" height="{side:.1f}" fill="none" stroke="#FF9500" stroke-width="2.5" stroke-dasharray="6 4"/>',
                          f'<text x="{sx+side/2:.1f}" y="{label_y:.1f}" text-anchor="middle" fill="#C56A00" font-size="11" font-weight="700">Target {square["side_mv"]:.1f} mV</text>']
        parts += [f'<text x="{left+size/2:.1f}" y="{top+size+48}" text-anchor="middle" fill="#1D1D1F" font-size="14">Vin (V)</text>',
                  f'<text x="{x0+16}" y="{top+size/2:.1f}" transform="rotate(-90 {x0+16} {top+size/2:.1f})" text-anchor="middle" fill="#1D1D1F" font-size="14">Vout (V)</text>']
    parts.append(f'<text x="{width/2}" y="{height-14}" text-anchor="middle" fill="#6E6E73" font-size="14">Each panel is a complete 6T Read butterfly. Vin/Vout axes are fixed at 0 to 1.20 V.</text></svg>')
    return "".join(parts)


def all_model_vdd_butterfly_overlay_svg(entries: list[dict], width: int = 1440, height: int = 920) -> str:
    """Overlay every analyzed operating VDD on one Vin/Vout chart."""
    left, top, plot_size = 125, 145, 660
    axis_max = SNM_PLOT_AXIS_MAX_V
    palette = ("#007AFF", "#34C759", "#AF52DE", "#FF9500", "#FF3B30", "#00A7A7")
    has_target = any(item["result"].get("datasheet_targets") for item in entries)

    def xy(x: float, y: float) -> tuple[float, float]:
        return left + x / axis_max * plot_size, top + (1 - y / axis_max) * plot_size

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="All operating VDD Read SNM butterfly overlay" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="54" y="50" fill="#1D1D1F" font-size="30" font-weight="700">All Operating Voltages — Read SNM Butterfly Overlay</text>',
        '<path d="M55 83 h42" stroke="#1D1D1F" stroke-width="4"/><text x="108" y="89" fill="#3A3A3C" font-size="16">Measured WAT</text>',
    ]
    if has_target:
        parts += ['<path d="M260 83 h42" stroke="#1D1D1F" stroke-width="3" stroke-dasharray="10 7"/><text x="313" y="89" fill="#3A3A3C" font-size="16">WAT Target</text>']
    for step in range(7):
        voltage = step * 0.2
        px, py = xy(voltage, voltage)
        parts += [f'<path d="M{px:.1f} {top} V{top+plot_size} M{left} {py:.1f} H{left+plot_size}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{px:.1f}" y="{top+plot_size+28}" text-anchor="middle" fill="#6E6E73" font-size="15">{voltage:.1f} V</text>',
                  f'<text x="{left-12}" y="{py+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="15">{voltage:.1f} V</text>']
    legend_x, legend_y = 925, 160
    for index, entry in enumerate(entries):
        color = palette[index % len(palette)]
        data = entry["result"]["baseline_6t"]
        target = entry["result"].get("target_6t", data)
        show_target = bool(entry["result"].get("datasheet_targets"))
        for points in _read_vtc_pair(data):
            path = " ".join(f"{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}" for x, y in points)
            parts.append(f'<polyline points="{path}" fill="none" stroke="{color}" stroke-width="3.5" opacity="0.88"/>')
        if show_target:
            for points in _read_vtc_pair(target):
                path = " ".join(f"{xy(x, y)[0]:.1f},{xy(x, y)[1]:.1f}" for x, y in points)
                parts.append(f'<polyline points="{path}" fill="none" stroke="{color}" stroke-width="2.5" stroke-dasharray="10 7" opacity="0.88"/>')
        y_pos = legend_y + index * 64
        measured_snm = data["metrics"]["read_snm_mv"]
        target_snm = target["metrics"]["read_snm_mv"]
        parts += [f'<path d="M{legend_x} {y_pos} h42" stroke="{color}" stroke-width="5"/>',
                  f'<text x="{legend_x+55}" y="{y_pos+6}" fill="#1D1D1F" font-size="18" font-weight="700">VDD {entry["model_vdd_v"]:.1f} V</text>',
                  f'<text x="{legend_x+55}" y="{y_pos+29}" fill="#007AFF" font-size="14">WAT RSNM {measured_snm:.1f} mV</text>']
        if show_target:
            parts.append(f'<text x="{legend_x+220}" y="{y_pos+29}" fill="#FF9500" font-size="14">Target {target_snm:.1f} mV · Δ {measured_snm-target_snm:+.1f}</text>')
    parts += [
        f'<text x="{left+plot_size/2}" y="{height-38}" text-anchor="middle" fill="#1D1D1F" font-size="19">Vin (V) — operating VDD identified by curve color</text>',
        f'<text x="42" y="{top+plot_size/2}" transform="rotate(-90 42 {top+plot_size/2})" text-anchor="middle" fill="#1D1D1F" font-size="19">Vout (V) — 0.0 to 1.2 V</text>',
        '<text x="925" y="120" fill="#6E6E73" font-size="15">Color = operating VDD · Solid = measured WAT · Dashed = target</text>',
        '</svg>',
    ]
    return "".join(parts)


def snm_by_model_vdd_svg(entries: list[dict], width: int = 1100, height: int = 540) -> str:
    """Independent SNM versus model-VDD trend chart for an Excel import."""
    left, top, plot_w, plot_h = 115, 110, 880, 320
    axis_max = SNM_PLOT_AXIS_MAX_V
    has_target = any(item["result"].get("datasheet_targets") for item in entries)
    datasets = [item["result"]["baseline_6t"] for item in entries]
    if has_target:
        datasets.extend(item["result"].get("target_6t", item["result"]["baseline_6t"]) for item in entries)
    max_snm = max(data["metrics"]["read_snm_mv"] for data in datasets)
    y_max = max(50.0, math.ceil(max_snm / 50.0) * 50.0)

    def xy(vdd: float, snm: float) -> tuple[float, float]:
        return left + vdd / axis_max * plot_w, top + (1 - snm / y_max) * plot_h

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" role="img" aria-label="SNM versus model VDD" style="font-family:Calibri,Arial,sans-serif">',
        '<rect width="100%" height="100%" fill="#FFFFFF"/>',
        '<text x="52" y="48" fill="#1D1D1F" font-size="29" font-weight="700">Read SNM versus Model VDD</text>',
        '<path d="M52 80 h32" stroke="#007AFF" stroke-width="4"/><text x="94" y="86" fill="#3A3A3C" font-size="16">Read SNM</text>',
    ]
    if has_target:
        parts += ['<path d="M455 80 h32" stroke="#1D1D1F" stroke-width="3"/><text x="497" y="86" fill="#3A3A3C" font-size="16">Measured WAT</text>',
                  '<path d="M650 80 h32" stroke="#1D1D1F" stroke-width="3" stroke-dasharray="9 6"/><text x="692" y="86" fill="#3A3A3C" font-size="16">WAT Target</text>']
    for voltage in (0.0, 0.30, 0.60, 0.90, axis_max):
        x, _ = xy(voltage, 0)
        parts += [f'<path d="M{x:.1f} {top} V{top+plot_h}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{x:.1f}" y="{top+plot_h+25}" text-anchor="middle" fill="#6E6E73" font-size="14">{voltage:.2f}</text>']
    for snm in (0.0, y_max / 2, y_max):
        _, y = xy(0, snm)
        parts += [f'<path d="M{left} {y:.1f} H{left+plot_w}" stroke="#E5E5EA" stroke-width="1"/>',
                  f'<text x="{left-10}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="14">{snm:.0f}</text>']
    for key, color in (("read_snm_mv", "#007AFF"),):
        points = [xy(item["model_vdd_v"], item["result"]["baseline_6t"]["metrics"][key]) for item in entries]
        parts.append(f'<polyline points="{" ".join(f"{x:.1f},{y:.1f}" for x, y in points)}" fill="none" stroke="{color}" stroke-width="4"/>')
        for x, y in points:
            parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="5" fill="#FFFFFF" stroke="{color}" stroke-width="3"/>')
        if has_target:
            target_points = [xy(item["model_vdd_v"], item["result"].get("target_6t", item["result"]["baseline_6t"])["metrics"][key]) for item in entries]
            parts.append(f'<polyline points="{" ".join(f"{x:.1f},{y:.1f}" for x, y in target_points)}" fill="none" stroke="{color}" stroke-width="3" stroke-dasharray="9 6"/>')
            for x, y in target_points:
                parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4" fill="#FFFFFF" stroke="{color}" stroke-width="2"/>')
    parts += [f'<text x="{left+plot_w/2}" y="{height-47}" text-anchor="middle" fill="#1D1D1F" font-size="18">Model VDD (V)</text>',
              f'<text x="38" y="{top+plot_h/2}" transform="rotate(-90 38 {top+plot_h/2})" text-anchor="middle" fill="#1D1D1F" font-size="18">SNM (mV)</text>',
              '<text x="550" y="512" text-anchor="middle" fill="#6E6E73" font-size="14">VDD axis is fixed at 0 to 1.20 V. Zero-VDD Excel rows are omitted because SNM is undefined at 0 V.</text></svg>']
    return "".join(parts)


def wat_electrical_snm_rows(result: dict) -> list[dict]:
    """Flatten WAT-measurable electrical inputs and derived SNM outputs."""
    cfg = result["config"]
    current_wat = result["wat"]
    datasets = [("Lot/Wafer", {
        "pu_vt": current_wat["pu_vt"], "pu_ids": current_wat["pu_ids"],
        "pg_vt": current_wat["pg_vt"], "pg_ids": current_wat["pg_ids"],
        "pd_vt": current_wat["pd_vt"], "pd_ids": current_wat["pd_ids"],
    }, result["baseline_6t"],
                 result.get("analytical_read_snm_comparison", {}).get("current", {}))]
    targets = result.get("datasheet_targets")
    if targets:
        datasets.append(("WAT Target", {
            "pu_vt": targets["pu"]["vt"], "pu_ids": targets["pu"]["ids"],
            "pg_vt": targets["pg"]["vt"], "pg_ids": targets["pg"]["ids"],
            "pd_vt": targets["pd"]["vt"], "pd_ids": targets["pd"]["ids"],
        }, result["target_6t"],
                         result.get("analytical_read_snm_comparison", {}).get("target", {})))

    def beta_proxy(vt: float, ids: float) -> float:
        overdrive = max(cfg["wat_vdd"] - abs(vt), 0.05)
        return 2.0 * ids / (overdrive * overdrive)

    rows = []
    for dataset, values, modeled, analytical in datasets:
        beta_pu = beta_proxy(values["pu_vt"], values["pu_ids"])
        beta_pg = beta_proxy(values["pg_vt"], values["pg_ids"])
        beta_pd = beta_proxy(values["pd_vt"], values["pd_ids"])
        rows.append({
            "dataset": dataset,
            "wat_vdd_v": cfg["wat_vdd"],
            "sram_vdd_v": cfg["nominal_vdd"],
            "pu_vt_v": values["pu_vt"], "pu_idsat_ua": values["pu_ids"],
            "pg_vt_v": values["pg_vt"], "pg_idsat_ua": values["pg_ids"],
            "pd_vt_v": values["pd_vt"], "pd_idsat_ua": values["pd_ids"],
            "beta_pu_proxy_ua_per_v2": beta_pu,
            "beta_pg_proxy_ua_per_v2": beta_pg,
            "beta_pd_proxy_ua_per_v2": beta_pd,
            "q_beta_pu_over_pg": beta_pu / beta_pg,
            "r_beta_pd_over_pg": beta_pd / beta_pg,
            "idsat_pd_over_pg": values["pd_ids"] / values["pg_ids"],
            "idsat_pg_over_pu": values["pg_ids"] / values["pu_ids"],
            "read_snm_geometric_mv": modeled["metrics"]["read_snm_mv"],
            "read_snm_eq_3_36_mv": analytical.get("snm_mv"),
            "eq_3_36_status": "VALID" if analytical.get("valid") else "N/A",
            "eq_3_36_reason": analytical.get("reason", ""),
            "evidence_scope": "WAT Vt + Idsat; no PDK/model-card-only parameters",
        })
    return rows


def generic_28nm_assumption_rows(result: dict) -> list[dict]:
    """Describe the maintained bitcell geometry and its ratio references."""
    tech = result["technology"]
    cell_ratio = tech["pd_width_nm"] / tech["pg_width_nm"]
    pull_up_ratio = tech["pg_width_nm"] / tech["pu_width_nm"]
    return [
        {"parameter": "Channel length L", "value": tech["channel_length_nm"], "unit": "nm",
         "source": "User input; generic default 28 nm", "used_by": "6T geometry reference", "active": "REFERENCE"},
        {"parameter": "PU width", "value": tech["pu_width_nm"], "unit": "nm",
         "source": "User input; generic default 70 nm", "used_by": "Pull-up geometry ratio", "active": "REFERENCE"},
        {"parameter": "PG width", "value": tech["pg_width_nm"], "unit": "nm",
         "source": "User input; generic default 100 nm", "used_by": "Cell / pull-up geometry ratios", "active": "REFERENCE"},
        {"parameter": "PD width", "value": tech["pd_width_nm"], "unit": "nm",
         "source": "User input; generic default 140 nm", "used_by": "Cell geometry ratio", "active": "REFERENCE"},
        {"parameter": "Geometry Cell Ratio", "value": round(cell_ratio, 4), "unit": "WPD/WPG",
         "source": "Derived from entered widths", "used_by": "Reference beside WAT-calibrated beta ratio", "active": "REFERENCE"},
        {"parameter": "Geometry Pull-up Ratio", "value": round(pull_up_ratio, 4), "unit": "WPG/WPU",
         "source": "Derived from entered widths", "used_by": "Reference beside WAT-calibrated beta ratio", "active": "REFERENCE"},
    ]


def architecture_svg(width: int = 1080, height: int = 445) -> str:
    """Self-contained schematic-style view of the fixed 28 nm 6T topology."""
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 900 500" width="{width}" height="{height}" role="img" aria-label="28 nm 6T SRAM architecture" style="font-family:Calibri,Arial,sans-serif">
    <rect width="100%" height="100%" fill="white"/>
    <text x="450" y="28" text-anchor="middle" font-size="19" font-weight="700">Generic 28 nm 6T SRAM bitcell</text>
    <path d="M95 72 H805 M125 72 V418 M775 72 V418" stroke="#94a3b8" stroke-width="2" fill="none"/>
    <text x="450" y="62" text-anchor="middle" font-size="17" font-weight="700">WL</text>
    <text x="125" y="447" text-anchor="middle" font-size="17" font-weight="700">BLB</text>
    <text x="775" y="447" text-anchor="middle" font-size="17" font-weight="700">BL</text>
    <path d="M315 105 H585 M450 105 V82 M315 390 H585 M450 390 V414" stroke="#334155" stroke-width="2.5" fill="none"/>
    <text x="450" y="97" text-anchor="middle" font-size="15" font-weight="700">VDD</text>
    <text x="450" y="438" text-anchor="middle" font-size="15" font-weight="700">GND</text>
    <path d="M315 105 V145 M585 105 V145 M315 205 V255 M585 205 V255 M315 315 V390 M585 315 V390" stroke="#334155" stroke-width="2.5"/>
    <rect x="270" y="145" width="90" height="60" rx="12" fill="#fff1f0" stroke="#ff3b30" stroke-width="2"/>
    <rect x="540" y="145" width="90" height="60" rx="12" fill="#fff1f0" stroke="#ff3b30" stroke-width="2"/>
    <rect x="270" y="255" width="90" height="60" rx="12" fill="#eef6ff" stroke="#007aff" stroke-width="2"/>
    <rect x="540" y="255" width="90" height="60" rx="12" fill="#eef6ff" stroke="#007aff" stroke-width="2"/>
    <rect x="145" y="220" width="100" height="60" rx="12" fill="#effcf2" stroke="#34c759" stroke-width="2"/>
    <rect x="655" y="220" width="100" height="60" rx="12" fill="#effcf2" stroke="#34c759" stroke-width="2"/>
    <text x="315" y="171" text-anchor="middle" font-size="15" font-weight="700">PUL</text><text x="315" y="190" text-anchor="middle" font-size="12">PMOS</text>
    <text x="585" y="171" text-anchor="middle" font-size="15" font-weight="700">PUR</text><text x="585" y="190" text-anchor="middle" font-size="12">PMOS</text>
    <text x="315" y="281" text-anchor="middle" font-size="15" font-weight="700">PDL</text><text x="315" y="300" text-anchor="middle" font-size="12">NMOS</text>
    <text x="585" y="281" text-anchor="middle" font-size="15" font-weight="700">PDR</text><text x="585" y="300" text-anchor="middle" font-size="12">NMOS</text>
    <text x="195" y="246" text-anchor="middle" font-size="15" font-weight="700">PGL</text><text x="195" y="265" text-anchor="middle" font-size="12">ACCESS</text>
    <text x="705" y="246" text-anchor="middle" font-size="15" font-weight="700">PGR</text><text x="705" y="265" text-anchor="middle" font-size="12">ACCESS</text>
    <path d="M125 250 H145 M245 250 H315 M585 250 H655 M755 250 H775" stroke="#334155" stroke-width="2.5" fill="none"/>
    <circle cx="315" cy="250" r="6" fill="#1d1d1f"/><circle cx="585" cy="250" r="6" fill="#1d1d1f"/>
    <text x="294" y="238" font-size="15" font-weight="700">QB</text><text x="598" y="238" font-size="15" font-weight="700">Q</text>
    <path d="M315 250 H420 V175 H540 M585 250 H480 V285 H360" stroke="#af52de" stroke-width="2" fill="none" stroke-dasharray="7 5"/>
    <path d="M195 220 V72 M705 220 V72" stroke="#34c759" stroke-width="2" fill="none"/>
    <text x="450" y="475" text-anchor="middle" font-size="12" fill="#64748b">L=28 nm · WPU=70 nm · WPG=100 nm · WPD=140 nm</text>
    </svg>'''


def export_pdf_report(result: dict, out: Path, image_dir: Path) -> Path:
    """Build a sharp, browser-independent PDF from the analysis and SVG charts."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (KeepTogether, PageBreak, Paragraph, SimpleDocTemplate,
                                       Spacer, Table, TableStyle)
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError("PDF export packages are missing. Run: python -m pip install -r requirements.txt") from exc

    pdf_dir = out / "pdf"; pdf_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = pdf_dir / "HV28_SRAM_Analysis.pdf"
    regular, bold = "Helvetica", "Helvetica-Bold"
    calibri = Path(r"C:\Windows\Fonts\calibri.ttf")
    calibrib = Path(r"C:\Windows\Fonts\calibrib.ttf")
    if calibri.exists() and calibrib.exists():
        pdfmetrics.registerFont(TTFont("Calibri", str(calibri)))
        pdfmetrics.registerFont(TTFont("Calibri-Bold", str(calibrib)))
        regular, bold = "Calibri", "Calibri-Bold"

    styles = getSampleStyleSheet()
    title = ParagraphStyle("HVTitle", parent=styles["Title"], fontName=bold, fontSize=29,
                           leading=32, textColor=colors.HexColor("#1D1D1F"), spaceAfter=5*mm)
    h1 = ParagraphStyle("HVH1", parent=styles["Heading1"], fontName=bold, fontSize=18,
                        leading=21, textColor=colors.HexColor("#1D1D1F"), spaceBefore=3*mm, spaceAfter=3*mm)
    h2 = ParagraphStyle("HVH2", parent=styles["Heading2"], fontName=bold, fontSize=12,
                        leading=14, textColor=colors.HexColor("#1D1D1F"), spaceBefore=2*mm, spaceAfter=2*mm)
    body = ParagraphStyle("HVBody", parent=styles["BodyText"], fontName=regular, fontSize=9.5,
                          leading=13, textColor=colors.HexColor("#3A3A3C"))
    small = ParagraphStyle("HVSmall", parent=body, fontSize=8, leading=10)

    def svg_drawing(filename: str, max_width: float, max_height: float):
        drawing = svg2rlg(str(image_dir / filename))
        if drawing is None or not drawing.width or not drawing.height:
            raise RuntimeError(f"Could not render SVG chart: {filename}")
        factor = min(max_width / drawing.width, max_height / drawing.height)
        drawing.width *= factor; drawing.height *= factor
        drawing.scale(factor, factor)
        return drawing

    table_header = colors.HexColor("#EEF4FC")
    border = colors.HexColor("#D2D2D7")
    def data_table(data: list[list], widths=None, font_size: float = 8):
        table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), bold), ("FONTNAME", (0, 1), (-1, -1), regular),
            ("FONTSIZE", (0, 0), (-1, -1), font_size), ("LEADING", (0, 0), (-1, -1), font_size+2),
            ("BACKGROUND", (0, 0), (-1, 0), table_header), ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1D1D1F")),
            ("GRID", (0, 0), (-1, -1), .35, border), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"), ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return table

    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont(regular, 8); canvas.setFillColor(colors.HexColor("#6E6E73"))
        canvas.drawString(12*mm, 7*mm, "HV28 SRAM Analysis")
        canvas.drawRightString(landscape(A4)[0]-12*mm, 7*mm, f"Page {doc.page}")
        canvas.restoreState()

    cfg, wat = result["config"], result["wat"]
    ratios = result.get("strength_ratios", Sram6T(WatPoint(**wat), Config(**cfg)).strength_ratios())
    doc = SimpleDocTemplate(str(pdf_path), pagesize=landscape(A4), leftMargin=11*mm, rightMargin=11*mm,
                            topMargin=10*mm, bottomMargin=12*mm, title="HV28 SRAM Analysis")
    story = [Paragraph("HV28 SRAM Analysis", title),
             Paragraph(f'Lot/Wafer: <b>{html.escape(wat["corner"])}</b> &nbsp;&nbsp; Object mode: <b>{html.escape(result.get("object_mode", "Grouped"))}</b> &nbsp;&nbsp; SRAM VDD: <b>{cfg["nominal_vdd"]:.3f} V</b>', body),
             Spacer(1, 4*mm), svg_drawing("00_28nm_6t_bitcell_architecture.svg", 250*mm, 116*mm),
             Spacer(1, 3*mm), Paragraph("Generic 28 nm 6T compact model. WAT targets are comparison references; model estimates remain calibrated from measured WAT inputs.", small)]

    comparisons = result.get("target_comparisons", [])
    if comparisons:
        story += [PageBreak(), Paragraph("WAT Target vs WAT Measured", h1)]
        rows = [["Object", "Type", "Target Vt (V)", "WAT Vt (V)", "Delta Vt (mV)",
                 "Target Isat (uA)", "WAT Isat (uA)", "Delta Isat (uA)", "Delta Isat (%)"]]
        rows += [[x["object"], x["device"], f'{x["target_vt_v"]:.3f}', f'{x["measured_vt_v"]:.3f}',
                  f'{x["delta_vt_mv"]:+.1f}', f'{x["target_isat_ua"]:.2f}', f'{x["measured_isat_ua"]:.2f}',
                  f'{x["delta_isat_ua"]:+.2f}', f'{x["delta_isat_pct"]:+.2f}%'] for x in comparisons]
        story += [data_table(rows, font_size=7.5), Spacer(1, 5*mm)]

    if result.get("wt_test_0bit"):
        story += [Paragraph("WT Test 0-Bit Vmin", h1)]
        rows = [["WT mode", "Measured Vmin (V)", "Source"]] + [
            [x["test"], _fmt(x["vmin_v"], 3), x.get("source", "Model estimate")] for x in result["wt_test_0bit"]]
        story += [data_table(rows, widths=[55*mm, 48*mm, 82*mm]), Spacer(1, 4*mm),
                  Paragraph(f'WT sweep setup: Start={cfg["vmin_start"]:.3f} V, Stop={cfg["vmin_stop"]:.3f} V, Step={cfg["vmin_step"]:.3f} V.', body)]

    if result.get("cell"):
        story += [Paragraph(f'{result.get("object_mode", "6T Independent")} WAT Objects', h1)]
        rows = [["MOS object", "Vt (V)", "Isat / Ids (uA)"]] + [
            [name, f'{values["vt"]:.3f}', f'{values["ids"]:.2f}'] for name, values in result["cell"]["mos"].items()]
        story += [data_table(rows, widths=[50*mm, 40*mm, 50*mm])]

    story += [PageBreak(), Paragraph("WAT Vt / Isat Overview", h1)]
    rows = [["Device", "Vt (V)", "Isat / Ids (uA)"]] + [
        [dev, f'{wat[dev.lower()+"_vt"]:.3f}', f'{wat[dev.lower()+"_ids"]:.2f}'] for dev in ("PU", "PG", "PD")]
    ratio_rows = [["Ratio", "Model beta (Vt-aware)", "WAT Isat proxy", "Definition"],
                  ["Cell Ratio", f'{ratios["cell_ratio_beta"]:.3f}', f'{ratios["cell_ratio_ids_proxy"]:.3f}', "PD / PG"],
                  ["Pull-up Ratio", f'{ratios["pull_up_ratio_beta"]:.3f}', f'{ratios["pull_up_ratio_ids_proxy"]:.3f}', "PG / PU"]]
    story += [data_table(rows, widths=[50*mm, 40*mm, 50*mm]), Spacer(1, 4*mm),
              Paragraph("Cell Strength Ratios", h2),
              data_table(ratio_rows, widths=[48*mm, 48*mm, 45*mm, 40*mm]), Spacer(1, 3*mm),
              Paragraph(f'Model beta ratios include Vt through WAT calibration; Isat ratios are direct measured-current proxies. Assumed sensitivity sweep: Vt +/-{cfg["vt_step"]*1000:.0f} mV and Isat +/-{cfg["ids_step_pct"]:g}%.', body)]

    judgment = result.get("judgment")
    if judgment:
        story += [Spacer(1, 5*mm), Paragraph(f'Model Judgment: {judgment["overall_status"]}', h1)]
        judgment_rows = [["Parameter", "Model value", "Design target", "Margin", "Status", "Recommended action"]]
        for item in judgment["items"]:
            unit = item["unit"]
            digits = 3 if unit == "ratio" else 2
            suffix = "" if unit == "ratio" else f" {unit}"
            judgment_rows.append([
                item["label"], f'{item["value"]:.{digits}f}{suffix}',
                f'{item["target"]:.{digits}f}{suffix}', f'{item["margin"]:+.{digits}f}{suffix}',
                item["status"], Paragraph(item["recommended_action"], small)])
        story += [data_table(judgment_rows, widths=[31*mm, 30*mm, 30*mm, 28*mm, 24*mm, 116*mm], font_size=7),
                  Spacer(1, 2*mm), Paragraph(f'MARGINAL band: {judgment["marginal_band_pct"]:.1f}% below the user-entered target. These limits are design/datasheet inputs, not universal 28 nm criteria.', small)]

    for index, (dev, items) in enumerate(result["groups"].items(), 1):
        story += [PageBreak(), Paragraph(f"{dev} Sensitivity", h1)]
        prefix = f"{index:02d}_{dev.lower()}"
        charts = [[svg_drawing(f"{prefix}_read_butterfly.svg", 125*mm, 48*mm),
                   svg_drawing(f"{prefix}_read_snm_sensitivity.svg", 125*mm, 48*mm)],
                  [svg_drawing(f"{prefix}_hold_snm_sensitivity.svg", 125*mm, 39*mm),
                   svg_drawing(f"{prefix}_write_snm_sensitivity.svg", 125*mm, 39*mm)],
                  [svg_drawing(f"{prefix}_model_read_vmin_sensitivity.svg", 125*mm, 39*mm),
                   svg_drawing(f"{prefix}_model_write_vmin_sensitivity.svg", 125*mm, 39*mm)]]
        chart_table = Table(charts, colWidths=[132*mm, 132*mm], hAlign="CENTER")
        chart_table.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("ALIGN", (0,0), (-1,-1), "CENTER"),
                                         ("LEFTPADDING", (0,0), (-1,-1), 2), ("RIGHTPADDING", (0,0), (-1,-1), 2),
                                         ("TOPPADDING", (0,0), (-1,-1), 2), ("BOTTOMPADDING", (0,0), (-1,-1), 2)]))
        baseline = items[0]["metrics"]
        rows = [["Scenario", "Cell ratio", "PU ratio", "Hold SNM", "Read SNM", "Write SNM*", "Delta WSNM", "Read Vmin", "Write Vmin"]]
        for item in items:
            m = item["metrics"]; w = item["wat"]
            rows.append([item["label"], f'{m["cell_ratio_beta"]:.3f}', f'{m["pull_up_ratio_beta"]:.3f}',
                         f'{m["hold_snm_mv"]:.2f}', f'{m["read_snm_mv"]:.2f}', f'{m["write_snm_mv"]:.2f}', f'{m["write_snm_mv"]-baseline["write_snm_mv"]:+.2f}',
                         _fmt(m["read_vmin_v"], 2), _fmt(m["write_vmin_v"], 2)])
        story += [chart_table, Spacer(1, 1*mm), data_table(rows, font_size=6.5),
                  Paragraph("* Write SNM is the compact-model bitline-noise margin proxy, not foundry/SPICE sign-off WSNM.", small)]

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return pdf_path


def _legacy_write_outputs(result: dict, out_dir: str | os.PathLike[str]) -> Path:
    out = Path(out_dir); out.mkdir(parents=True, exist_ok=True)
    image_dir = out/"images"; image_dir.mkdir(parents=True, exist_ok=True)
    image_manifest: list[dict[str, str]] = []

    def save_image(filename: str, svg: str, description: str, device: str = "CELL") -> str:
        """Publish a scalable SVG and a high-resolution PNG for the HTML report."""
        try:
            from reportlab.graphics import renderPM
            from svglib.svglib import svg2rlg
        except ImportError as exc:
            raise RuntimeError("PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc
        svg_path = image_dir/filename
        svg_path.write_text(svg, encoding="utf-8")
        png_name = str(Path(filename).with_suffix(".png"))
        png_path = image_dir/png_name
        drawing = svg2rlg(str(svg_path))
        if drawing is None:
            raise RuntimeError(f"Could not render chart image: {filename}")
        renderPM.drawToFile(drawing, str(png_path), fmt="PNG", dpi=180, backend="rlPyCairo")
        image_manifest.extend([
            {"filename": png_name, "format": "PNG", "role": "Primary image",
             "device": device, "description": description},
            {"filename": filename, "format": "SVG", "role": "Scalable chart source",
             "device": device, "description": description},
        ])
        return f'images/{png_name}'

    rows = [{"device": "6T Cell", "scenario": "Baseline", **result["wat"],
             **result["baseline_6t"]["metrics"]}]
    with open(out/"sram_wat_results.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0])); writer.writeheader(); writer.writerows(rows)
    with open(out/"sram_wat_results.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    if "wt_test_0bit" in result:
        wt_rows=[]
        for item in result["wt_test_0bit"]:
            wt_rows.append({"test":item["test"],"vmin_v":item["vmin_v"],
                            "source":item.get("source", "Model estimate"),
                            "zero_bit_pass":item.get("zero_bit_pass", ""),
                            "failed_phase_count":item.get("failed_phase_count", ""),
                            "phases_at_vmin":"; ".join(k for k,v in item.get("phases_at_vmin",{}).items() if v)})
        with open(out/"wt_test_0bit_vmin.csv","w",newline="",encoding="utf-8-sig") as f:
            writer=csv.DictWriter(f,fieldnames=list(wt_rows[0])); writer.writeheader(); writer.writerows(wt_rows)
    comparisons = result.get("target_comparisons", [])
    if comparisons:
        with open(out/"wat_target_comparison.csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(comparisons[0]))
            writer.writeheader(); writer.writerows(comparisons)
    if result.get("judgment"):
        judgment_rows = result["judgment"]["items"]
        with open(out/"parameter_judgment.csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(judgment_rows[0]))
            writer.writeheader(); writer.writerows(judgment_rows)
    if result.get("target_validation"):
        validation = result["target_validation"]
        with open(out/"wat_target_validation_rows.csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(validation["rows"][0]))
            writer.writeheader(); writer.writerows(validation["rows"])
        summary_row = {"verdict": validation["verdict"], **validation["counts"], **validation["statistics"]}
        with open(out/"wat_target_validation_summary.csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(summary_row))
            writer.writeheader(); writer.writerow(summary_row)
        with open(out/"wat_target_parameter_evidence.csv", "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(validation["parameter_evidence"][0]))
            writer.writeheader(); writer.writerows(validation["parameter_evidence"])

    cfg = result["config"]; wat = result["wat"]; tech = result["technology"]
    sections = []
    ordered_groups = list(result["groups"].items())
    for group_index, (dev, items) in enumerate(ordered_groups, 1):
        baseline = items[0]["metrics"]
        trs = "".join("<tr>" + "".join(f"<td>{html.escape(str(v))}</td>" for v in
            [x["label"], f'{x["wat"][dev.lower()+"_vt"]:.3f}', f'{x["wat"][dev.lower()+"_ids"]:.2f}',
             f'{x["metrics"]["hold_snm_mv"]:.2f}', f'{x["metrics"]["read_snm_mv"]:.2f}',
             f'{x["metrics"]["read_snm_mv"]-baseline["read_snm_mv"]:+.2f}',
             f'{x["metrics"]["write_snm_mv"]:.2f}',
             f'{x["metrics"]["write_snm_mv"]-baseline["write_snm_mv"]:+.2f}',
             f'{x["metrics"]["cell_ratio_beta"]:.3f}', f'{x["metrics"]["pull_up_ratio_beta"]:.3f}',
             _fmt(x["metrics"]["read_vmin_v"],2),
             "N/A" if x["metrics"]["read_vmin_v"] is None or baseline["read_vmin_v"] is None else f'{x["metrics"]["read_vmin_v"]-baseline["read_vmin_v"]:+.2f}',
             _fmt(x["metrics"]["write_vmin_v"],2),
             "N/A" if x["metrics"]["write_vmin_v"] is None or baseline["write_vmin_v"] is None else f'{x["metrics"]["write_vmin_v"]-baseline["write_vmin_v"]:+.2f}']) + "</tr>" for x in items)
        prefix = f"{group_index:02d}_{dev.lower()}"
        butterfly_path = save_image(f"{prefix}_read_butterfly.svg", butterfly_svg(items,cfg["nominal_vdd"]),
                                    f"{dev} read butterfly and SNM curves", dev)
        snm_path = save_image(f"{prefix}_read_snm_sensitivity.svg", bar_svg(items,"read_snm_mv",dev+" Read SNM","mV"),
                              f"{dev} read SNM sensitivity", dev)
        hold_snm_path = save_image(f"{prefix}_hold_snm_sensitivity.svg", bar_svg(items,"hold_snm_mv",dev+" Hold SNM","mV"),
                                   f"{dev} hold SNM sensitivity", dev)
        write_snm_path = save_image(f"{prefix}_write_snm_sensitivity.svg", bar_svg(items,"write_snm_mv",dev+" Write SNM (bitline-noise proxy)","mV"),
                                    f"{dev} compact-model write SNM sensitivity", dev)
        read_vmin_path = save_image(f"{prefix}_model_read_vmin_sensitivity.svg", bar_svg(items,"read_vmin_v",dev+" Model Read Vmin","V"),
                                    f"{dev} model-estimated read Vmin sensitivity", dev)
        write_vmin_path = save_image(f"{prefix}_model_write_vmin_sensitivity.svg", bar_svg(items,"write_vmin_v",dev+" Model Write Vmin","V"),
                                     f"{dev} model-estimated write Vmin sensitivity", dev)
        sections.append(f'''<section><h2>{dev} sensitivity</h2>
        <div class="grid"><div><img src="{butterfly_path}" alt="{dev} read butterfly"></div>
        <div><img src="{snm_path}" alt="{dev} read SNM sensitivity"></div>
        <div><img src="{hold_snm_path}" alt="{dev} hold SNM sensitivity"></div>
        <div><img src="{write_snm_path}" alt="{dev} write SNM sensitivity"></div>
        <div><img src="{read_vmin_path}" alt="{dev} model read Vmin sensitivity"></div>
        <div><img src="{write_vmin_path}" alt="{dev} model write Vmin sensitivity"></div></div>
        <table><thead><tr><th>Scenario</th><th>Vt (V)</th><th>Ids (uA)</th><th>Hold SNM (mV)</th><th>Read SNM (mV)</th><th>ΔRSNM</th><th>Write SNM* (mV)</th><th>ΔWSNM</th><th>Cell ratio β</th><th>Pull-up ratio β</th><th>Model Read Vmin (V)</th><th>ΔRVmin</th><th>Model Write Vmin (V)</th><th>ΔWVmin</th></tr></thead><tbody>{trs}</tbody></table>
        <p><small>* Compact-model bitline-noise margin proxy; not foundry/SPICE sign-off WSNM.</small></p></section>''')
    ratios = result.get("strength_ratios", Sram6T(WatPoint(**wat), Config(**cfg)).strength_ratios())
    ratio = (f'Model β Cell Ratio (PD/PG)={ratios["cell_ratio_beta"]:.3f}; '
             f'Model β Pull-up Ratio (PG/PU)={ratios["pull_up_ratio_beta"]:.3f}; '
             f'WAT Isat proxies={ratios["cell_ratio_ids_proxy"]:.3f} / {ratios["pull_up_ratio_ids_proxy"]:.3f}')
    wat_rows = "".join(f'<tr><td>{d}</td><td>{wat[d.lower()+"_vt"]:.3f}</td><td>{wat[d.lower()+"_ids"]:.2f}</td><td>{wat[d.lower()+"_ids"]/sum(wat[x+"_ids"] for x in ("pu","pg","pd")):.3f}</td></tr>' for d in ("PU","PG","PD"))
    individual_section = ""
    if "cell" in result:
        mos_rows = "".join(f'<tr><td>{name}</td><td>{values["vt"]:.3f}</td><td>{values["ids"]:.2f}</td></tr>' for name,values in result["cell"]["mos"].items())
        sensitivity_blocks = []
        for name, items in result.get("mos_sensitivity", {}).items():
            base = items[0]["metrics"]
            body = "".join("<tr>"+"".join(f"<td>{html.escape(str(v))}</td>" for v in [
                x["label"], f'{x["metrics"]["hold_snm_mv"]:.2f}', f'{x["metrics"]["read_snm_mv"]:.2f}',
                f'{x["metrics"]["write_snm_mv"]:.2f}', f'{x["metrics"]["cell_ratio_beta"]:.3f}',
                f'{x["metrics"]["pull_up_ratio_beta"]:.3f}',
                _fmt(x["metrics"]["read_vmin_v"],2), _fmt(x["metrics"]["write_vmin_v"],2)])+"</tr>" for x in items)
            sensitivity_blocks.append(f'<div><h3>{name}</h3><table><thead><tr><th>Scenario</th><th>Hold SNM</th><th>Read SNM</th><th>Write SNM*</th><th>Cell ratio β</th><th>Pull-up ratio β</th><th>Read Vmin</th><th>Write Vmin</th></tr></thead><tbody>{body}</tbody></table></div>')
        object_mode = result.get("object_mode", "6T Independent")
        individual_section = f'''<section><h2>{object_mode} objects</h2>
        <p>Object mode: {object_mode}. Model-estimated margins are kept separate from manually entered WT Vmin.</p>
        <table><thead><tr><th>MOS object</th><th>Vt (V)</th><th>Ids (µA)</th></tr></thead><tbody>{mos_rows}</tbody></table>
        <div class="grid">{''.join(sensitivity_blocks)}</div></section>'''
    target_section = ""
    if comparisons:
        target_rows = "".join(
            f'<tr><td>{x["object"]}</td><td>{x["device"]}</td><td>{x["target_vt_v"]:.3f}</td>'
            f'<td>{x["measured_vt_v"]:.3f}</td><td>{x["delta_vt_mv"]:+.1f}</td>'
            f'<td>{x["target_isat_ua"]:.2f}</td><td>{x["measured_isat_ua"]:.2f}</td>'
            f'<td>{x["delta_isat_ua"]:+.2f}</td><td>{x["delta_isat_pct"]:+.2f}%</td></tr>'
            for x in comparisons)
        target_section = f'''<section><h2>WAT Target vs WAT Measured</h2>
        <p>Target values are comparison references only. The compact model remains calibrated from the measured WAT values.</p>
        <table><thead><tr><th>MOS object</th><th>Type</th><th>Target Vt (V)</th><th>WAT Vt (V)</th><th>ΔVt (mV)</th><th>Target Isat (µA)</th><th>WAT Isat (µA)</th><th>ΔIsat (µA)</th><th>ΔIsat (%)</th></tr></thead><tbody>{target_rows}</tbody></table></section>'''
    wt_section = ""
    if "wt_test_0bit" in result:
        wt_defs={
            "Scan4N":"W0 → R0/W1 → R1/W0 → R0；全部 phase 通過",
            "Select_Write":"Write-0 與 Write-1 都達到 20%/80% rail",
            "Select_Read":"Read-0 與 Read-1 都保持 35%/65% rail 且符合 SNM 下限",
        }
        manual_source = result.get("vmin_source") == "manual"
        wt_rows="".join(f'<tr><td>{x["test"]}</td><td>{_fmt(x["vmin_v"],3)}</td><td>{html.escape(x.get("source", "Model estimate"))}</td><td>{html.escape(wt_defs[x["test"]])}</td></tr>' for x in result["wt_test_0bit"])
        wt_section=f'''<section><h2>WT Test 0-Bit Vmin</h2>
        <p>{"Values below were entered manually after WT measurement; the Python model did not generate them." if manual_source else "Model-estimated values; no manual WT values were supplied."}</p>
        <table><thead><tr><th>WT mode</th><th>Measured Vmin (V)</th><th>Source</th><th>Test definition</th></tr></thead><tbody>{wt_rows}</tbody></table></section>'''
    judgment_section = ""
    if result.get("judgment"):
        judgment = result["judgment"]
        judgment_rows = "".join(
            f'<tr><td>{item["label"]}</td><td>{item["value"]:.3f} {"" if item["unit"] == "ratio" else item["unit"]}</td>'
            f'<td>{item["target"]:.3f} {"" if item["unit"] == "ratio" else item["unit"]}</td>'
            f'<td>{item["margin"]:+.3f}</td><td><b>{item["status"]}</b></td>'
            f'<td>{html.escape(item["recommended_action"])}</td></tr>' for item in judgment["items"])
        judgment_section = f'''<section class="judgment"><h2>Parameter Judgment: {judgment["overall_status"]}</h2>
        <p>PASS requires the model value to meet the user-entered design target. MARGINAL means it is within {judgment["marginal_band_pct"]:.1f}% below target. These are not universal 28 nm limits.</p>
        <table><thead><tr><th>Parameter</th><th>Model value</th><th>Design target</th><th>Margin</th><th>Status</th><th>Recommended action</th></tr></thead><tbody>{judgment_rows}</tbody></table></section>'''
    validation_section = ""
    if result.get("target_validation"):
        validation = result["target_validation"]
        counts, stats, current = validation["counts"], validation["statistics"], validation["current_row"]
        stat_rows = "".join(
            f'<tr><td>{label}</td><td>{value}</td><td>{meaning}</td></tr>' for label, value, meaning in (
                ("WT-complete rows", counts["wt_complete"], "All three measured WT Vmin values are present"),
                ("Paired WAT + WT rows", counts["paired_wat_wt"], "Rows usable for target validation"),
                ("Inside / outside target band", f'{counts["inside_target_band"]} / {counts["outside_target_band"]}', "Both groups need at least 3 rows"),
                ("WT pass rate inside band", _fmt(stats["inside_wt_pass_rate_pct"], 1) + "%" if stats["inside_wt_pass_rate_pct"] is not None else "N/A", "Higher is better"),
                ("WT pass rate outside band", _fmt(stats["outside_wt_pass_rate_pct"], 1) + "%" if stats["outside_wt_pass_rate_pct"] is not None else "N/A", "Comparison population"),
                ("Pass-rate lift", (_fmt(stats["pass_rate_lift_pct_points"], 1) + " pp") if stats["pass_rate_lift_pct_points"] is not None else "N/A", "Inside minus outside"),
                ("Spearman correlation", _fmt(stats["target_distance_vs_worst_vmin_spearman"], 3), "Positive: farther from target tends to worsen Vmin"),
                ("Pearson correlation", _fmt(stats["target_distance_vs_worst_vmin_pearson"], 3), "Linear association; use with Spearman"),
            ))
        parameter_rows = "".join(
            f'<tr><td>{item["parameter"]}</td><td>{item["paired_n"]}</td>'
            f'<td>{item["inside_n"]} / {item["outside_n"]}</td>'
            f'<td>{(_fmt(item["pass_rate_lift_pct_points"], 1) + " pp") if item["pass_rate_lift_pct_points"] is not None else "N/A"}</td>'
            f'<td>{_fmt(item["deviation_vs_worst_vmin_spearman"], 3)}</td><td><b>{item["verdict"]}</b></td></tr>'
            for item in validation["parameter_evidence"])
        settings = validation["settings"]
        verdict_class = "ok" if validation["verdict"] == "SUPPORTED" else ("bad" if validation["verdict"] == "CONTRADICTED" else "warn")
        validation_section = f'''<section class="validation"><div class="section-heading"><div><h2>WAT Target Validation</h2><p>Primary decision: do lots near the WAT target actually achieve acceptable measured WT Vmin?</p></div><span class="verdict {verdict_class}">{validation["verdict"]}</span></div>
        <div class="validation-cards"><div><span>Current Lot/Wafer</span><b>{html.escape(str(current["lot_wafer"]))}</b></div><div><span>WAT band result</span><b>{"IN BAND" if current["target_band_pass"] else "OUT OF BAND"}</b></div><div><span>WT result</span><b>{"PASS" if current["wt_pass"] else "FAIL"}</b></div><div><span>Screen outcome</span><b>{current["consistency"]}</b></div></div>
        <p><b>Conclusion:</b> {html.escape(validation["explanation"])}</p>
        <p>Target band: Vt ±{settings["vt_tolerance_mv"]:.1f} mV and Isat ±{settings["idsat_tolerance_pct"]:.1f}%. WT pass limits: Scan4N ≤ {settings["scan4n_vmin_max"]:.3f} V, Select_Write ≤ {settings["select_write_vmin_max"]:.3f} V, Select_Read ≤ {settings["select_read_vmin_max"]:.3f} V.</p>
        <table><thead><tr><th>Evidence</th><th>Value</th><th>Interpretation</th></tr></thead><tbody>{stat_rows}</tbody></table>
        <h3>Individual WAT Target Evidence</h3><p>Each parameter uses every row where that WAT item and all three WT Vmin results are available; other missing WAT items do not discard the row.</p>
        <table><thead><tr><th>WAT target</th><th>Paired N</th><th>Inside / outside N</th><th>Pass-rate lift</th><th>Deviation–Vmin Spearman</th><th>Evidence</th></tr></thead><tbody>{parameter_rows}</tbody></table>
        <p class="footnote">WT-only rows remain in coverage counts, while correlation and band discrimination use paired WAT+WT rows only. Association supports screening usefulness but does not prove transistor-level causality.</p></section>'''
    snm_overview_path = save_image("01_6t_cell_snm_butterfly.svg", snm_overview_svg(result),
                                   "6T cell Hold, Read and Write SNM butterfly analysis", "6T CELL")
    snm_metrics = result["baseline_6t"]["metrics"]
    snm_section = f'''<section><h2>6T Cell SNM Analysis</h2>
    <p>SNM is evaluated at complete-cell level. PU, PG and PD are not assigned separate SNM values.</p>
    <p>Both axes span 0 to VDD, so the butterfly is entirely in the first quadrant. Hold and Read SNM evaluate the largest square in each lobe and use the smaller result. The two squares coincide in size for this symmetric 3T-merged input. Write is displayed as a write-noise proxy bracket because it is not calculated by the same symmetric butterfly-square rule.</p>
    <img src="{snm_overview_path}" alt="6T SRAM Hold Read Write SNM butterfly analysis">
    <table><thead><tr><th>Cell-level metric</th><th>Value (mV)</th><th>Operating condition</th></tr></thead><tbody>
    <tr><td>Hold SNM</td><td>{snm_metrics["hold_snm_mv"]:.2f}</td><td>WL off; data retention</td></tr>
    <tr><td>Read SNM</td><td>{snm_metrics["read_snm_mv"]:.2f}</td><td>WL on; BL and BLB precharged</td></tr>
    <tr><td>Write SNM*</td><td>{snm_metrics["write_snm_mv"]:.2f}</td><td>BL=0 / BLB=VDD compact-model proxy</td></tr>
    </tbody></table></section>'''
    with open(image_dir/"image_manifest.csv", "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["filename", "format", "role", "device", "description"])
        writer.writeheader(); writer.writerows(sorted(image_manifest, key=lambda row: row["filename"]))

    doc = f'''<!doctype html><html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>HV28 SRAM Analysis</title>
    <style>
    :root{{font:100%/1.5 Calibri,"Microsoft JhengHei",Arial,sans-serif;color:#1d1d1f;background:#f5f5f7}}
    *{{box-sizing:border-box}} body{{margin:0;padding:clamp(1.25rem,4vw,3rem);background:radial-gradient(circle at 85% 0,#e8f2ff 0,transparent 28rem),#f5f5f7}}
    main{{max-width:1760px;margin:auto}} h1{{font-size:clamp(2.2rem,4vw,4rem);line-height:1.05;letter-spacing:-.035em;margin:.5rem 0 .35rem}} h2{{font-size:1.7rem;letter-spacing:-.018em;margin-top:0}} h3{{font-size:1.3rem;letter-spacing:-.01em}}
    .note,.summary,section{{background:rgba(255,255,255,.82);backdrop-filter:blur(22px) saturate(160%);border:1px solid rgba(255,255,255,.75);border-radius:1.25rem;box-shadow:0 1px 2px #0000000a,0 12px 36px #0000000d;padding:clamp(1rem,2.4vw,1.65rem);margin:1rem 0}}
    .note{{border-left:4px solid #007aff}} .summary{{color:#3a3a3c;font-size:1.05rem}} .grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(min(100%,42rem),1fr));gap:1.5rem}}
    .target-badge{{display:inline-block;vertical-align:middle;margin-left:.55rem;padding:.28rem .55rem;border-radius:999px;background:#e5f1ff;color:#007aff;font-size:.72rem;letter-spacing:.04em}}
    .section-heading{{display:flex;align-items:flex-start;justify-content:space-between;gap:1rem}} .section-heading h2{{margin-bottom:.2rem}} .section-heading p{{margin:.1rem 0;color:#6e6e73}}
    .verdict{{white-space:nowrap;border-radius:999px;padding:.45rem .75rem;font-weight:700;font-size:.82rem}} .verdict.ok{{background:#eaf8ee;color:#167a36}} .verdict.warn{{background:#fff4df;color:#946200}} .verdict.bad{{background:#ffebe9;color:#c5221f}}
    .validation-cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(12rem,1fr));gap:.75rem;margin:1rem 0}} .validation-cards>div{{background:#f5f5f7;border-radius:.85rem;padding:.85rem 1rem}} .validation-cards span{{display:block;color:#6e6e73;font-size:.85rem;margin-bottom:.25rem}} .validation-cards b{{font-size:1.05rem}} .footnote{{color:#6e6e73;font-size:.9rem}}
    svg,img{{display:block;width:100%;min-height:24rem;height:auto;border:1px solid #e5e5ea;border-radius:1rem;overflow:hidden}} table{{border-collapse:separate;border-spacing:0;width:100%;margin-top:1rem;font-size:1rem;line-height:1.55;overflow:hidden}}
    th,td{{padding:.9rem .8rem;border-bottom:1px solid #e5e5ea;text-align:right;font-variant-numeric:tabular-nums}} th{{color:#6e6e73;font-size:.9rem;font-weight:680;letter-spacing:.015em}} th:first-child,td:first-child{{text-align:left}} tbody tr:last-child td{{border-bottom:0}} code{{background:#f2f2f7;padding:.18rem .38rem;border-radius:.35rem}}
    @media(prefers-reduced-transparency:reduce){{.note,.summary,section{{background:#fff;backdrop-filter:none;border-color:#d2d2d7}}}}
    @media(prefers-contrast:more){{.note,.summary,section{{background:#fff;border:2px solid #1d1d1f}}}}
    @page{{size:A4 landscape;margin:11mm}}
    @media print{{
      :root{{font-size:10pt}} body{{padding:0;background:#fff}} main{{max-width:none}}
      h1{{font-size:28pt}} h2{{font-size:18pt;break-after:avoid}} h3{{font-size:14pt;break-after:avoid}}
      .note,.summary,section{{background:#fff;border:1px solid #d2d2d7;box-shadow:none;backdrop-filter:none;margin:7mm 0;padding:6mm}}
      .summary,.note,table,.grid>div{{break-inside:avoid}} .grid{{grid-template-columns:1fr 1fr;gap:6mm}}
      svg,img{{min-height:0;max-height:155mm;object-fit:contain;border-radius:3mm}}
      table{{font-size:8.5pt;line-height:1.25}} th,td{{padding:5pt 4pt}}
    }}
    </style></head><body><main>
    <h1>HV28 SRAM Analysis</h1><p>Lot/Wafer: <b>{html.escape(wat["corner"])}</b> · Object mode: <b>{html.escape(result.get("object_mode","Grouped"))}</b> · SRAM VDD={cfg["nominal_vdd"]:.3f} V</p>
    <div class="summary"><b>Data meaning:</b> PU / PG / PD WAT Target Vt and Isat are references compared with measured WAT values. Scan4N, Select_Write and Select_Read Vmin are manually entered measured values. Model Read/Write Vmin is kept separate.</div>
    <div class="summary"><b>WT sweep setup:</b> Start={cfg["vmin_start"]:.3f} V · Stop={cfg["vmin_stop"]:.3f} V · Step={cfg["vmin_step"]:.3f} V. These are actual tester recipe inputs and also define the model search range and resolution.</div>
    {target_section}
    {wt_section}
    {validation_section}
    {judgment_section}
    {snm_section}
    {individual_section}
    <section><h2>WAT Vt / Ids comparison</h2><table><thead><tr><th>Device</th><th>Vt (V)</th><th>Ids (uA)</th><th>Normalized Ids</th></tr></thead><tbody>{wat_rows}</tbody></table><p>{ratio}</p></section>
    <p>Raw data: <code>sram_wat_results.csv</code>, <code>wat_target_comparison.csv</code>, <code>wat_target_validation_rows.csv</code>, <code>wat_target_validation_summary.csv</code>, <code>wat_target_parameter_evidence.csv</code>, <code>sram_wat_results.json</code>. Standalone charts: <code>images/</code> with <code>image_manifest.csv</code>.</p></main></body></html>'''
    report = out/"sram_wat_report.html"; report.write_text(doc, encoding="utf-8")
    return report


def multi_chip_vtc_svg(analysis: dict, mode: str, width: int = 1180, height: int = 735) -> str:
    """Overlay wafer VTCs and show the correct state-specific limiting margins."""
    if mode not in {"read", "write"}:
        raise ValueError("mode must be read or write")
    vdd, axis = analysis["vdd_v"], SNM_PLOT_AXIS_MAX_V
    # Equal physical X/Y scale is required so every electrical square is
    # rendered as a square, not a visually misleading rectangle.
    # Deliberately compact square plot.  The right-hand information card has
    # a fixed safe width, so long chip IDs never extend beyond the SVG canvas.
    left, top, plot_w, plot_h = 150, 185, 430, 430
    card_x, card_y, card_w, card_h = 650, 170, 470, 430
    worst = analysis["worst_rsnm"] if mode == "read" else analysis["worst_wsnm"]
    metric = "RSNM" if mode == "read" else "WSNM"
    upper_worst = analysis.get("worst_rsnm_upper") if mode == "read" else None
    lower_worst = analysis.get("worst_rsnm_lower") if mode == "read" else None
    def same_cell(left_row: dict, right_row: dict) -> bool:
        return (str(left_row.get("lot_wafer", "")) ==
                str(right_row.get("lot_wafer", "")) and
                str(left_row.get("chip_id", "")) ==
                str(right_row.get("chip_id", "")))

    def cell_label(row: dict) -> str:
        return f'{row.get("lot_wafer", "Wafer")} / {row.get("chip_id", "Unknown")}'

    def xy(x: float, y: float) -> tuple[float, float]:
        return left + x / axis * plot_w, top + (1 - y / axis) * plot_h
    parts = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" style="font-family:Calibri,Arial,sans-serif">',
             '<rect width="100%" height="100%" fill="#FFFFFF"/>',
             f'<text x="54" y="55" fill="#1D1D1F" font-size="31" font-weight="700">Wafer Multi-Cell {metric} VTC Overlay</text>',
             f'<text x="54" y="84" fill="#6E6E73" font-size="16">{html.escape(str(analysis["lot_wafer"]))} · {len(analysis["rows"])} chips · Model VDD = {vdd:.3f} V</text>',
             '<path d="M54 112 h30" stroke="#007AFF" stroke-width="4"/><text x="94" y="117" fill="#3A3A3C" font-size="14">All chip direct VTC</text>',
             '<path d="M265 112 h30" stroke="#AF52DE" stroke-width="4" stroke-dasharray="9 6"/><text x="305" y="117" fill="#3A3A3C" font-size="14">All chip mirrored / paired VTC</text>',
             '<path d="M535 112 h30" stroke="#FF9500" stroke-width="5"/><text x="575" y="117" fill="#3A3A3C" font-size="14">State-limit chip VTC pair</text>',
             f'<rect x="{card_x}" y="{card_y}" width="{card_w}" height="{card_h}" rx="16" fill="#F5F5F7" stroke="#E5E5EA"/>']
    for voltage in (0, .3, .6, .9, 1.2):
        x, y = xy(voltage, voltage)
        parts += [f'<path d="M{x:.1f} {top} V{top+plot_h} M{left} {y:.1f} H{left+plot_w}" stroke="#E5E5EA"/>',
                  f'<text x="{x:.1f}" y="{top+plot_h+28}" text-anchor="middle" fill="#6E6E73" font-size="14">{voltage:.2f}</text>',
                  f'<text x="{left-12}" y="{y+5:.1f}" text-anchor="end" fill="#6E6E73" font-size="14">{voltage:.2f}</text>']
    for row in analysis["rows"]:
        if mode == "read":
            direct, mirrored = _read_vtc_pair(row["read"])
        else:
            direct, mirrored = row["write"]["write_1"]["curve"], row["write"]["write_0"]["curve"]
        if mode == "read":
            # Upper and lower wafer minima can belong to different chips.
            # Highlight both complete VTC pairs so each drawn square remains
            # associated with physically valid curves from the same chip.
            is_upper_limit = same_cell(row, upper_worst)
            is_lower_limit = same_cell(row, lower_worst)
            is_limit = is_upper_limit or is_lower_limit
        else:
            is_upper_limit = is_lower_limit = False
            is_limit = same_cell(row, worst)
        width_px, opacity = (4.5, 1.0) if is_limit else (1.3, .18)
        direct_points = " ".join(f'{xy(x,y)[0]:.1f},{xy(x,y)[1]:.1f}' for x, y in direct)
        mirror_points = " ".join(f'{xy(x,y)[0]:.1f},{xy(x,y)[1]:.1f}' for x, y in mirrored)
        color = ("#FF9500" if is_upper_limit and not is_lower_limit else
                 "#FF3B30" if is_limit else "#007AFF")
        paired_color = "#AF52DE" if is_limit else "#C21FD4"
        parts += [f'<polyline points="{direct_points}" fill="none" stroke="{color}" stroke-width="{width_px}" opacity="{opacity}"/>',
                  f'<polyline points="{mirror_points}" fill="none" stroke="{paired_color}" stroke-width="{width_px}" opacity="{opacity}" stroke-dasharray="9 6"/>']
    if mode == "read":
        # Wafer Upper and Lower minima are reported independently.  Each
        # square is always selected from the matching VTC pair of its own
        # chip; we never combine a direct curve from one chip with a mirrored
        # curve from another chip to form a fictitious butterfly.
        def state_square(row: dict, state_key: str) -> dict | None:
            return next((item for item in row["read"]["read_butterfly"].get("squares", [])
                         if item.get("state_key") == state_key), None)
        upper_square = state_square(upper_worst, "upper_left")
        lower_square = state_square(lower_worst, "lower_right")
        square_specs = [
            ("Upper minimum", cell_label(upper_worst), upper_square),
            ("Lower minimum", cell_label(lower_worst), lower_square),
        ]
    else:
        square = worst["write"].get("write_square")
        square_specs = [("WSNM", cell_label(worst), square)] if square else []
    for state_label, chip_id, square in square_specs:
        if square.get("side_v", 0) <= 0:
            continue
        x0, y0, side = square["x_v"], square["y_v"], square["side_v"]
        x, y_top = xy(x0, y0 + side)
        side_x, side_y = side / axis * plot_w, side / axis * plot_h
        parts += [f'<rect x="{x:.1f}" y="{y_top:.1f}" width="{side_x:.1f}" height="{side_y:.1f}" fill="#EFFAF2" fill-opacity="0.75" stroke="#34C759" stroke-width="4"/>',
        ]
    value = worst["rsnm_mv"] if mode == "read" else worst["wsnm_mv"]
    if mode == "read":
        summary_rows = [
            ("Cell minimum", f'{value:.1f} mV · {cell_label(worst)}', "#FF3B30"),
            ("Upper minimum", f'{upper_worst["upper_rsnm_mv"]:.1f} mV · {cell_label(upper_worst)}', "#FF9500"),
            ("Lower minimum", f'{lower_worst["lower_rsnm_mv"]:.1f} mV · {cell_label(lower_worst)}', "#FF3B30"),
        ]
    else:
        summary_rows = [("Wafer WSNM", f'{value:.1f} mV · {cell_label(worst)}', "#FF9500")]
    card_text = [f'<text x="{card_x+24}" y="{card_y+36}" fill="#1D1D1F" font-size="20" font-weight="700">Wafer margin summary</text>']
    for index, (label, detail, color) in enumerate(summary_rows):
        y = card_y + 75 + index * 43
        card_text += [f'<circle cx="{card_x+30}" cy="{y-5}" r="5" fill="{color}"/>',
                      f'<text x="{card_x+45}" y="{y}" fill="#3A3A3C" font-size="15">{label}</text>',
                      f'<text x="{card_x+45}" y="{y+20}" fill="#1D1D1F" font-size="16" font-weight="700">{html.escape(detail)}</text>']
    # Keep the electrical inputs physically tied to the same cell that owns
    # the minimum reported SNM.  This lets the VTC overlay be audited without
    # accidentally combining a square from one cell with WAT data from another.
    source_cell = worst["cell"]
    raw_idsat_ua = worst.get("raw_idsat_ua", {})
    source_values = (
        ("PUL", source_cell.pu1), ("PUR", source_cell.pu2),
        ("PGL", source_cell.pg1), ("PGR", source_cell.pg2),
        ("PDL", source_cell.pd1), ("PDR", source_cell.pd2),
    )
    table_top = card_y + 218 if mode == "read" else card_y + 132
    card_text += [
        f'<path d="M{card_x+24} {table_top-18} H{card_x+card_w-24}" stroke="#D2D2D7"/>',
        f'<text x="{card_x+24}" y="{table_top}" fill="#1D1D1F" font-size="15" font-weight="700">Minimum {metric} source 6T WAT values</text>',
        f'<text x="{card_x+24}" y="{table_top+20}" fill="#6E6E73" font-size="13">{html.escape(cell_label(worst))} · Vt (V) / Idsat (uA)</text>',
        f'<text x="{card_x+28}" y="{table_top+45}" fill="#6E6E73" font-size="13" font-weight="700">MOS</text>',
        f'<text x="{card_x+145}" y="{table_top+45}" fill="#6E6E73" font-size="13" font-weight="700">Vt (V)</text>',
        f'<text x="{card_x+270}" y="{table_top+45}" fill="#6E6E73" font-size="13" font-weight="700">Idsat (uA)</text>',
    ]
    for index, (name, mos) in enumerate(source_values):
        y = table_top + 70 + index * 22
        ids_display = raw_idsat_ua.get(name.lower(), mos.ids)
        card_text += [
            f'<text x="{card_x+28}" y="{y}" fill="#1D1D1F" font-size="14" font-weight="700">{name}</text>',
            f'<text x="{card_x+145}" y="{y}" fill="#3A3A3C" font-size="14">{mos.vt:.4f}</text>',
            f'<text x="{card_x+270}" y="{y}" fill="#3A3A3C" font-size="14">{ids_display:.2f}</text>',
        ]
    parts += [f'<text x="{left+plot_w/2:.1f}" y="{top+plot_h+82}" text-anchor="middle" fill="#1D1D1F" font-size="19">Vin (V)</text>',
              f'<text x="70" y="{top+plot_h/2:.1f}" transform="rotate(-90 70 {top+plot_h/2:.1f})" text-anchor="middle" fill="#1D1D1F" font-size="19">Vout (V)</text>',
              *card_text,
              '</svg>']
    return "".join(parts)


def _multi_chip_summary_export_rows(analysis: dict) -> list[dict[str, object]]:
    """Build the common per-Cell CSV rows for full and fast Multi-VDD runs."""
    def family_average(row: dict, family: str, attribute: str) -> float:
        cell = row["cell"]
        return (float(getattr(getattr(cell, f"{family}1"), attribute)) +
                float(getattr(getattr(cell, f"{family}2"), attribute))) / 2.0

    def same_output_cell(left_row: dict, right_row: dict) -> bool:
        return (str(left_row.get("lot_wafer", "")) ==
                str(right_row.get("lot_wafer", "")) and
                str(left_row.get("chip_id", "")) ==
                str(right_row.get("chip_id", "")))

    def device_idsat(row: dict, label: str, field: str) -> float:
        raw_value = row.get("raw_idsat_ua", {}).get(label)
        return (float(raw_value) if raw_value is not None
                else float(getattr(row["cell"], field).ids))

    device_map = {
        "pul": "pu1", "pur": "pu2", "pgl": "pg1",
        "pgr": "pg2", "pdl": "pd1", "pdr": "pd2",
    }

    return [{"lot_wafer": row["lot_wafer"], "chip_id": row["chip_id"],
             "model_vdd_v": analysis["vdd_v"], "rsnm_mv": row["rsnm_mv"],
             "upper_rsnm_mv": row["upper_rsnm_mv"],
             "lower_rsnm_mv": row["lower_rsnm_mv"],
             "wsnm_mv": row["wsnm_mv"],
             "write_margin_mv": row["write_margin_mv"],
             "cell_ratio_beta": row["cell_ratio_beta"],
             "pull_up_ratio_beta": row["pull_up_ratio_beta"],
             "pu_vt_v": family_average(row, "pu", "vt"),
             "pu_idsat_ua": family_average(row, "pu", "ids"),
             "pg_vt_v": family_average(row, "pg", "vt"),
             "pg_idsat_ua": family_average(row, "pg", "ids"),
             "pd_vt_v": family_average(row, "pd", "vt"),
             "pd_idsat_ua": family_average(row, "pd", "ids"),
             **{f"{label}_vt_v": float(getattr(row["cell"], field).vt)
                for label, field in device_map.items()},
             **{f"{label}_idsat_ua": device_idsat(row, label, field)
                for label, field in device_map.items()},
             "is_worst_rsnm": same_output_cell(row, analysis["worst_rsnm"]),
             "is_worst_rsnm_upper": same_output_cell(
                 row, analysis["worst_rsnm_upper"]),
             "is_worst_rsnm_lower": same_output_cell(
                 row, analysis["worst_rsnm_lower"]),
             "is_worst_wsnm": same_output_cell(row, analysis["worst_wsnm"]),
             "is_worst_write_margin": same_output_cell(
                 row, analysis["worst_write_margin"])}
            for row in analysis["rows"]]


def write_multi_chip_outputs(analysis: dict, out_dir: str | os.PathLike[str],
                             input_excel_path: str | os.PathLike[str] | None = None) -> Path:
    """Export batch wafer VTC overlays, per-chip margin table and HTML summary."""
    out = Path(out_dir); image_dir = out / "images"; image_dir.mkdir(parents=True, exist_ok=True)
    backup_name = f'imported_6t_vt_idsat_data_{float(analysis["vdd_v"]):.3f}V.xlsx'
    if input_excel_path is not None:
        source = Path(input_excel_path)
        if not source.is_file():
            raise FileNotFoundError(f"Imported Multi-Cell Excel was not found: {source}")
        # Preserve the exact imported workbook, including any user metadata,
        # as a run-local audit backup.  The analysis always reads Vt in V and
        # Idsat in uA from its accepted 6T Multi-Cell sheet.
        shutil.copy2(source, out / backup_name)
    export_rows = _multi_chip_summary_export_rows(analysis)
    with open(out / "multi_chip_snm_summary.csv", "w", newline="",
              encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(export_rows[0]))
        writer.writeheader(); writer.writerows(export_rows)
    if not analysis.get("shmoo_enabled", True):
        body = "".join(
            f'<tr><td>{html.escape(str(row["lot_wafer"]))}</td>'
            f'<td>{html.escape(str(row["chip_id"]))}</td>'
            f'<td>{float(row["rsnm_mv"]):.2f}</td>'
            f'<td>{float(row["wsnm_mv"]):.2f}</td>'
            f'<td>{float(row["write_margin_mv"]):.2f}</td></tr>'
            for row in export_rows)
        report = out / "multi_cell_wafer_report.html"
        report.write_text(f'''<!doctype html><html><head><meta charset="utf-8"><title>HV28 SRAM Multi-Cell Summary</title><style>body{{font-family:Calibri,"Microsoft JhengHei",Arial,sans-serif;background:#f5f5f7;color:#1d1d1f;margin:32px}}main{{max-width:1200px;margin:auto}}section{{background:#fff;border-radius:16px;padding:24px}}table{{width:100%;border-collapse:collapse}}th,td{{padding:10px;border-bottom:1px solid #e5e5ea;text-align:right}}th:first-child,td:first-child,th:nth-child(2),td:nth-child(2){{text-align:left}}.note{{color:#6e6e73}}</style></head><body><main><h1>Multi-Cell VDD Summary</h1><p>Lot/Wafer: {html.escape(str(analysis["lot_wafer"]))} · Model VDD {float(analysis["vdd_v"]):.3f} V</p><section><h2>Fast curve mode</h2><p class="note">Shmoo and Drive Advisor were disabled. The per-Cell SNM and BL Write Margin results remain available in <code>multi_chip_snm_summary.csv</code>.</p><table><thead><tr><th>Lot/Wafer</th><th>Chip ID</th><th>RSNM (mV)</th><th>WSNM (mV)</th><th>BL Vtrip (mV)</th></tr></thead><tbody>{body}</tbody></table></section></main></body></html>''', encoding="utf-8")
        return report
    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError("PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc
    charts = [("01_multi_chip_read_vtc.svg", "01_multi_chip_read_vtc.png", "read"),
              ("02_multi_chip_write_vtc.svg", "02_multi_chip_write_vtc.png", "write")]
    for svg_name, png_name, mode in charts:
        path = image_dir / svg_name; path.write_text(multi_chip_vtc_svg(analysis, mode), encoding="utf-8")
        drawing = svg2rlg(str(path))
        if drawing is None:
            raise RuntimeError("Could not render multi-chip VTC overlay")
        renderPM.drawToFile(drawing, str(image_dir / png_name), fmt="PNG", dpi=180, backend="rlPyCairo")
    relative_shmoo = analysis["relative_shmoo"]
    shmoo_svg_name = "03_multi_cell_wafer_relative_shmoo.svg"
    shmoo_png_name = "03_multi_cell_wafer_relative_shmoo.png"
    shmoo_svg_path = image_dir / shmoo_svg_name
    shmoo_svg_path.write_text(estimate_vmin_ratio_shmoo_svg(relative_shmoo), encoding="utf-8")
    shmoo_drawing = svg2rlg(str(shmoo_svg_path))
    if shmoo_drawing is None:
        raise RuntimeError("Could not render Multi-Cell wafer-relative shmoo")
    renderPM.drawToFile(shmoo_drawing, str(image_dir / shmoo_png_name), fmt="PNG",
                        dpi=180, backend="rlPyCairo")
    relative_fields = [
        "lot_wafer", "chip_id", "rsnm_mv", "write_margin_mv",
        "cell_ratio_beta", "pull_up_ratio_beta", "read_percentile",
        "write_percentile", "cr_percentile", "pr_percentile",
        "performance_grade_score", "wafer_grade_score", "wafer_grade",
        "robust_low_outlier",
    ]
    with (out / "multi_cell_wafer_relative_grades.csv").open(
            "w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=relative_fields)
        writer.writeheader()
        writer.writerows({key: sample.get(key, "") for key in relative_fields}
                         for sample in relative_shmoo["samples"])
    with (out / "multi_cell_wafer_distribution_statistics.csv").open(
            "w", newline="", encoding="utf-8-sig") as stream:
        fields = ["metric", "p05", "q1", "median", "q3", "p95", "mad"]
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        for metric, statistics in relative_shmoo["distributions"].items():
            writer.writerow({"metric": metric, **statistics})
    for name, shmoo in (("read", analysis["median_target_read_shmoo"]),
                         ("write", analysis["median_target_write_shmoo"])):
        with open(out / f"median_target_{name}_shmoo.csv", "w", newline="", encoding="utf-8-sig") as stream:
            writer = csv.DictWriter(stream, fieldnames=list(shmoo["rows"][0]))
            writer.writeheader(); writer.writerows(shmoo["rows"])
    relative_by_chip = {(str(sample["lot_wafer"]), str(sample["chip_id"])): sample
                        for sample in relative_shmoo["samples"]}
    body = "".join(
        f'<tr><td>{html.escape(row["lot_wafer"])} / {html.escape(row["chip_id"])}</td><td>{row["upper_rsnm_mv"]:.2f}</td>'
        f'<td>{row["lower_rsnm_mv"]:.2f}</td><td>{row["rsnm_mv"]:.2f}</td>'
        f'<td>{row["write_margin_mv"]:.2f}</td><td>{row["cell_ratio_beta"]:.3f}</td>'
        f'<td>{row["pull_up_ratio_beta"]:.3f}</td>'
        f'<td>{100.0 * float(relative_by_chip[(row["lot_wafer"], row["chip_id"])]["cr_percentile"]):.1f}</td>'
        f'<td>{100.0 * float(relative_by_chip[(row["lot_wafer"], row["chip_id"])]["pr_percentile"]):.1f}</td>'
        f'<td>{html.escape(str(relative_by_chip[(row["lot_wafer"], row["chip_id"])]["wafer_grade"]).upper())}</td></tr>'
        for row in analysis["rows"])
    median = analysis["median_cell"]

    def recommendation_rows(shmoo: dict) -> str:
        if not shmoo["recommendations"]:
            return '<tr><td colspan="6">No one-factor 0–100% move toward the median reached this target.</td></tr>'
        return "".join(
            f'<tr><td>{row["family"]}</td><td>{row["parameter"]}</td><td>{row["toward_median_pct"]}%</td>'
            f'<td>{row["rsnm_mv"]:.2f}</td><td>{row["write_margin_mv"]:.2f}</td>'
            f'<td>CR={row["cell_ratio_beta"]:.3f}; PR={row["pull_up_ratio_beta"]:.3f}</td></tr>'
            for row in shmoo["recommendations"][:4])

    read_shmoo = analysis["median_target_read_shmoo"]
    write_shmoo = analysis["median_target_write_shmoo"]
    report = out / "multi_cell_wafer_report.html"
    backup_note = (f'<p class="note">Imported 6T Vt/Idsat backup: <code>{backup_name}</code>. '
                   'This is the original workbook used for this run.</p>'
                   if input_excel_path is not None else '')
    grade_counts = {
        grade: sum(sample["wafer_grade"] == grade for sample in relative_shmoo["samples"])
        for grade in ("preferred", "monitor", "low")
    }
    statistics_rows = "".join(
        f'<tr><td>{html.escape(metric)}</td><td>{values["p05"]:.4g}</td>'
        f'<td>{values["q1"]:.4g}</td><td>{values["median"]:.4g}</td>'
        f'<td>{values["q3"]:.4g}</td><td>{values["p95"]:.4g}</td>'
        f'<td>{values["mad"]:.4g}</td></tr>'
        for metric, values in relative_shmoo["distributions"].items())
    report.write_text(f'''<!doctype html><html><head><meta charset="utf-8"><title>HV28 SRAM Wafer Multi-Cell Analysis</title><style>body{{font-family:Calibri,Arial,sans-serif;background:#f5f5f7;color:#1d1d1f;margin:32px}}main{{max-width:1400px;margin:auto}}section{{background:#fff;border-radius:16px;padding:24px;margin:18px 0}}img{{width:100%;height:auto}}table{{width:100%;border-collapse:collapse}}th,td{{padding:10px;border-bottom:1px solid #e5e5ea;text-align:right}}th:first-child,td:first-child{{text-align:left}}.note{{color:#6e6e73}}.warn{{color:#c2410c;font-weight:bold}}</style></head><body><main><h1>HV28 SRAM Wafer Multi-Cell Analysis</h1><p>Lot/Wafer: {html.escape(str(analysis["lot_wafer"]))} · {len(analysis["rows"])} measured cells · Model VDD={analysis["vdd_v"]:.3f} V</p>{backup_note}<section><h2>Conservative wafer reference</h2><p><b>Minimum cell RSNM:</b> {analysis["worst_rsnm"]["rsnm_mv"]:.2f} mV ({html.escape(analysis["worst_rsnm"]["chip_id"])})<br><b>Minimum Write Margin:</b> {analysis["worst_write_margin"]["write_margin_mv"]:.2f} mV ({html.escape(analysis["worst_write_margin"]["chip_id"])})<br><b>Minimum WSNM:</b> {analysis["worst_wsnm"]["wsnm_mv"]:.2f} mV ({html.escape(analysis["worst_wsnm"]["chip_id"])})</p></section><section><h2>Wafer-relative quartile Shmoo</h2><p>At this fixed VDD, Read drive is ranked from CR and Write drive from PR. The lower drive percentile controls the grade: <b>Preferred</b> ≥ P50, <b>Monitor</b> P25–P50, and <b>Low</b> &lt; P25. RSNM and BL Write Margin percentiles are retained separately for correlation. This is an intra-wafer screening reference, not absolute silicon Pass/Fail.</p><p><b>Preferred:</b> {grade_counts["preferred"]} · <b>Monitor:</b> {grade_counts["monitor"]} · <b>Low:</b> {grade_counts["low"]}</p><img src="images/{shmoo_png_name}" alt="Wafer-relative CR PR quartile shmoo"><h3>Robust distribution statistics</h3><table><tr><th>Metric</th><th>P5</th><th>Q1</th><th>Median</th><th>Q3</th><th>P95</th><th>MAD</th></tr>{statistics_rows}</table><p class="note">Detailed outputs: <code>multi_cell_wafer_relative_grades.csv</code> and <code>multi_cell_wafer_distribution_statistics.csv</code>.</p></section><section><h2>Synthetic median reference cell</h2><p>The median cell is built from the per-device median Vt and Idsat values. It is a statistical reference and may not be a physically measured cell.</p><table><tr><th>RSNM</th><th>Write Margin</th><th>Cell Ratio (CR)</th><th>Pull-up Ratio (PR)</th></tr><tr><td>{median["rsnm_mv"]:.2f} mV</td><td>{median["write_margin_mv"]:.2f} mV</td><td>{median["cell_ratio_beta"]:.3f}</td><td>{median["pull_up_ratio_beta"]:.3f}</td></tr></table></section><section><h2>Worst-to-median adjustment screening</h2><p class="note">Each shmoo moves both devices of one family, one parameter at a time, from the worst cell toward its physical-MOS median in 10% steps. It is a direction-finding compact-model screening, not a simultaneous process correction.</p><h3>Read target: median RSNM = {read_shmoo["target_value_mv"]:.2f} mV; worst cell = {html.escape(analysis["worst_rsnm"]["chip_id"])}</h3><table><tr><th>Family</th><th>Parameter</th><th>Move toward median</th><th>RSNM</th><th>Write Margin</th><th>Ratios after move</th></tr>{recommendation_rows(read_shmoo)}</table><h3>Write target: median Write Margin = {write_shmoo["target_value_mv"]:.2f} mV; worst cell = {html.escape(analysis["worst_write_margin"]["chip_id"])}</h3><table><tr><th>Family</th><th>Parameter</th><th>Move toward median</th><th>RSNM</th><th>Write Margin</th><th>Ratios after move</th></tr>{recommendation_rows(write_shmoo)}</table><p class="note">Detailed sweep data: <code>median_target_read_shmoo.csv</code> and <code>median_target_write_shmoo.csv</code>.</p></section><section><h2>Read VTC / Mirror VTC</h2><p>Upper and Lower squares are each taken from their own state-limiting cell. Their direct/mirrored VTC pair is highlighted together; the two states are not combined into one artificial cell. The overlay card lists the complete six-MOS Vt/Idsat set for the cell that owns the minimum RSNM.</p><img src="images/01_multi_chip_read_vtc.png"></section><section><h2>Write W=1 / W=0 VTC</h2><p>The overlay card lists the complete six-MOS Vt/Idsat set for the cell that owns the minimum WSNM.</p><img src="images/02_multi_chip_write_vtc.png"></section><section><h2>Per-cell margins and wafer-relative grade</h2><table><tr><th>Cell / Chip ID</th><th>Upper RSNM</th><th>Lower RSNM</th><th>RSNM</th><th>Write Margin</th><th>CR</th><th>PR</th><th>CR pct.</th><th>PR pct.</th><th>Grade</th></tr>{body}</table></section></main></body></html>''', encoding="utf-8")
    return report


def process_multi_vdd_6t_excel(
        source_path: str | os.PathLike[str], cfg: Config,
        output_base: str | os.PathLike[str],
        vdd_groups: list[dict[str, object]] | None = None,
        include_shmoo: bool = True) -> dict[str, object]:
    """Run per-sheet Multi-Cell analysis, then aggregate all VDDs into Vmin curves."""
    source = Path(source_path)
    groups = (vdd_groups if vdd_groups is not None else
              read_multi_chip_6t_excel_vdd_sheets(source, cfg.nominal_vdd))
    analyses: list[dict] = []
    lot_wafers: set[str] = set()
    for group in groups:
        vdd = float(group["vdd_v"])
        chips = list(group["chips"])
        lot_wafers.update(chip.lot_wafer for chip in chips)
        model_config = replace(cfg, nominal_vdd=vdd, wat_vdd=vdd)
        analysis = analyze_multi_chip_wafer(
            chips, model_config, include_shmoo=include_shmoo)
        analysis["source_sheet_names"] = list(group["sheet_names"])
        analyses.append(analysis)

    wafer_label = (next(iter(lot_wafers)) if len(lot_wafers) == 1
                   else f"{len(lot_wafers)}_Lot_Wafer_groups")
    run_dir = create_run_output_dir(
        output_base, wafer_label, "multi_vdd_excel_estimate_vmin")
    per_vdd_root = run_dir / "multi_cell_by_vdd"
    per_vdd_root.mkdir(parents=True, exist_ok=True)
    per_vdd_reports: list[Path] = []
    for analysis in analyses:
        vdd = float(analysis["vdd_v"])
        vdd_dir = per_vdd_root / f"Model_VDD_{vdd:.3f}V"
        vdd_dir.mkdir(parents=True, exist_ok=False)
        per_vdd_reports.append(
            write_multi_chip_outputs(analysis, vdd_dir, source))

    source_rows = estimate_rows_from_multi_chip_analyses(analyses)
    estimate_analysis = analyze_estimate_vmin_curves(
        source_rows, include_shmoo=include_shmoo)
    estimate_analysis["multi_vdd_excel"] = True
    estimate_analysis["per_vdd_output_count"] = len(analyses)
    estimate_analysis["source_sheets"] = [
        name for analysis in analyses for name in analysis["source_sheet_names"]]
    estimate_dir = run_dir / "estimate_vmin"
    report = write_estimate_vmin_outputs(
        estimate_analysis, estimate_dir, [source])
    return {
        "analysis": estimate_analysis, "report": report,
        "run_dir": run_dir, "per_vdd_reports": per_vdd_reports,
    }


def write_outputs(result: dict, out_dir: str | os.PathLike[str]) -> Path:
    """Write the focused Read SNM current-versus-target HTML report."""
    out = Path(out_dir); out.mkdir(parents=True, exist_ok=True)
    image_dir = out / "images"; image_dir.mkdir(parents=True, exist_ok=True)
    has_target = bool(result.get("datasheet_targets") and result.get("target_6t"))

    # These files belonged to the removed WT/Vmin workflow; do not leave stale results behind.
    for name in ("sram_wat_results.csv", "wt_test_0bit_vmin.csv", "parameter_judgment.csv",
                 "wat_target_validation_rows.csv", "wat_target_validation_summary.csv",
                 "wat_target_parameter_evidence.csv", "analytical_read_snm_eq_3_36.csv",
                 "write_snm_vs_bitline.csv", "single_wat_write_snm_geometry.csv",
                 "write_butterfly_curve.csv"):
        path = out / name
        if path.exists():
            path.unlink()
    if not has_target:
        for name in ("snm_target_comparison.csv", "wat_target_comparison.csv"):
            path = out / name
            if path.exists():
                path.unlink()
    for path in image_dir.iterdir():
        if path.is_file() and path.suffix.lower() in {".png", ".svg", ".csv"}:
            path.unlink()

    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError("PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc

    svg_name = "01_read_snm_target_comparison.svg"
    png_name = "01_read_snm_target_comparison.png"
    butterfly_svg_name = "02_read_snm_butterfly.svg"
    butterfly_png_name = "02_read_snm_butterfly.png"
    write_svg_name = "03_w0_w1_wsnm_analysis.svg"
    write_png_name = "03_w0_w1_wsnm_analysis.png"
    svg_path = image_dir / svg_name
    svg_path.write_text(snm_overview_svg(result), encoding="utf-8")
    drawing = svg2rlg(str(svg_path))
    if drawing is None:
        raise RuntimeError("Could not render Read SNM target comparison")
    renderPM.drawToFile(drawing, str(image_dir / png_name), fmt="PNG", dpi=180, backend="rlPyCairo")
    butterfly_svg_path = image_dir / butterfly_svg_name
    butterfly_svg_path.write_text(read_snm_butterfly_svg(result), encoding="utf-8")
    butterfly_drawing = svg2rlg(str(butterfly_svg_path))
    if butterfly_drawing is None:
        raise RuntimeError("Could not render Read SNM butterfly chart")
    renderPM.drawToFile(butterfly_drawing, str(image_dir / butterfly_png_name),
                        fmt="PNG", dpi=180, backend="rlPyCairo")
    write_svg_path = image_dir / write_svg_name
    write_svg_path.write_text(write_wsnm_window_svg(result), encoding="utf-8")
    write_drawing = svg2rlg(str(write_svg_path))
    if write_drawing is None:
        raise RuntimeError("Could not render W0/W1 Write SNM analysis")
    renderPM.drawToFile(write_drawing, str(image_dir / write_png_name), fmt="PNG", dpi=180,
                        backend="rlPyCairo")

    comparison_rows = result.get("snm_target_comparison", [])
    comparison_export_rows = [{
        "mode": row["mode"],
        "lot_wafer_snm_mv": row["current_snm_mv"],
        "target_snm_mv": row["target_snm_mv"],
        "lot_wafer_minus_target_mv": row["delta_mv"],
        "difference_pct": row["delta_pct"],
    } for row in comparison_rows]
    if comparison_export_rows:
        with open(out / "snm_target_comparison.csv", "w", newline="", encoding="utf-8-sig") as target_file:
            writer = csv.DictWriter(target_file, fieldnames=list(comparison_export_rows[0]))
            writer.writeheader(); writer.writerows(comparison_export_rows)

    write_snm_rows = []
    for dataset, modeled in [("Lot/Wafer", result["baseline_6t"])] + ([("WAT Target", result["target_6t"])] if has_target else []):
        write_model = modeled["write_wsnm"]
        for state_key, state_label in (("write_1", "W=1 upper VTC"), ("write_0", "W=0 lower VTC")):
            state = write_model[state_key]
            for vin, vout in state["curve"]:
                write_snm_rows.append({
                    "dataset": dataset, "write_state": state_label,
                    "sram_vdd_v": result["config"]["nominal_vdd"], "vin_v": vin,
                    "write_vtc_vout_v": vout,
                    "wsnm_mv": write_model["snm_mv"],
                    "wordline_v": state["write_bias"]["wordline_v"],
                    "bl_v": state["write_bias"]["bl_v"], "blb_v": state["write_bias"]["blb_v"],
                })
    with open(out / "w0_w1_wsnm_analysis.csv", "w", newline="", encoding="utf-8-sig") as write_file:
        writer = csv.DictWriter(write_file, fieldnames=list(write_snm_rows[0]))
        writer.writeheader(); writer.writerows(write_snm_rows)

    state_rows = []
    state_datasets = [("Lot/Wafer", result["baseline_6t"])]
    if has_target:
        state_datasets.append(("WAT Target", result["target_6t"]))
    for label, data in state_datasets:
        butterfly = data["read_butterfly"]
        squares = butterfly["squares"]
        upper = butterfly.get("snm_upper_left_mv", squares[0]["side_mv"])
        lower = butterfly.get("snm_lower_right_mv", squares[1]["side_mv"])
        mean = (upper + lower) / 2.0
        state_rows.append({
            "dataset": label,
            "upper_left_state_snm_mv": upper,
            "lower_right_state_snm_mv": lower,
            "cell_read_snm_min_mv": min(upper, lower),
            "upper_minus_lower_mv": upper - lower,
            "mismatch_index_pct": abs(upper - lower) / mean * 100.0 if mean > 0 else None,
        })
    with open(out / "read_snm_state_mismatch.csv", "w", newline="", encoding="utf-8-sig") as state_file:
        writer = csv.DictWriter(state_file, fieldnames=list(state_rows[0]))
        writer.writeheader(); writer.writerows(state_rows)

    analytical = result.get("analytical_read_snm_comparison", {})
    analytical_rows = []
    for dataset in (("current", "target") if has_target else ("current",)):
        values = analytical.get(dataset, {})
        analytical_rows.append({
            "dataset": "Lot/Wafer" if dataset == "current" else "WAT Target",
            "valid": values.get("valid"),
            "reason": values.get("reason"),
            "snm_mv": values.get("snm_mv"),
            "vdd_v": values.get("vdd_v"),
            "vth_eff_v": values.get("vth_eff_v"),
            "q_beta_p_over_beta_a": values.get("q_beta_p_over_beta_a"),
            "r_beta_d_over_beta_a": values.get("r_beta_d_over_beta_a"),
            "vs_v": values.get("vs_v"),
            "vr_v": values.get("vr_v"),
            "k": values.get("k"),
            "term_a_v": values.get("term_a_v"),
            "term_b_v": values.get("term_b_v"),
        })
    with open(out / "analytical_read_snm.csv", "w", newline="", encoding="utf-8-sig") as analytical_file:
        writer = csv.DictWriter(analytical_file, fieldnames=list(analytical_rows[0]))
        writer.writeheader(); writer.writerows(analytical_rows)

    electrical_snm_rows = wat_electrical_snm_rows(result)
    with open(out / "wat_electrical_snm_table.csv", "w", newline="", encoding="utf-8-sig") as electrical_file:
        writer = csv.DictWriter(electrical_file, fieldnames=list(electrical_snm_rows[0]))
        writer.writeheader(); writer.writerows(electrical_snm_rows)

    assumption_rows = generic_28nm_assumption_rows(result)
    with open(out / "cell_geometry_reference.csv", "w", newline="", encoding="utf-8-sig") as assumption_file:
        writer = csv.DictWriter(assumption_file, fieldnames=list(assumption_rows[0]))
        writer.writeheader(); writer.writerows(assumption_rows)

    comparisons = result.get("target_comparisons", [])
    if comparisons:
        with open(out / "wat_target_comparison.csv", "w", newline="", encoding="utf-8-sig") as target_file:
            writer = csv.DictWriter(target_file, fieldnames=list(comparisons[0]))
            writer.writeheader(); writer.writerows(comparisons)
    with open(out / "sram_wat_results.json", "w", encoding="utf-8") as json_file:
        json.dump(result, json_file, ensure_ascii=False, indent=2)

    manifest_rows = [
        {"filename": png_name, "format": "PNG", "role": "Primary image", "device": "6T CELL",
         "description": ("Read SNM Lot/Wafer WAT versus WAT Target curves" if has_target
                         else "Read SNM Lot/Wafer WAT curves; Target reference disabled")},
        {"filename": svg_name, "format": "SVG", "role": "Scalable chart source", "device": "6T CELL",
         "description": ("Read SNM Lot/Wafer WAT versus WAT Target curves" if has_target
                         else "Read SNM Lot/Wafer WAT curves; Target reference disabled")},
        {"filename": butterfly_png_name, "format": "PNG", "role": "Read SNM butterfly", "device": "6T CELL",
         "description": "Read VTC with maximum squares 1 and 2 for Lot/Wafer and Target"},
        {"filename": butterfly_svg_name, "format": "SVG", "role": "Scalable Read SNM butterfly source", "device": "6T CELL",
         "description": "Read VTC with maximum squares 1 and 2 for Lot/Wafer and Target"},
        {"filename": write_png_name, "format": "PNG", "role": "W0/W1 Write SNM analysis", "device": "6T CELL",
         "description": "Separate write-zero and write-one VTC pairs with their maximum WSNM squares"},
        {"filename": write_svg_name, "format": "SVG", "role": "Scalable W0/W1 Write SNM source", "device": "6T CELL",
         "description": "Separate write-zero and write-one VTC pairs with their maximum WSNM squares"},
    ]
    with open(image_dir / "image_manifest.csv", "w", newline="", encoding="utf-8-sig") as manifest:
        writer = csv.DictWriter(manifest, fieldnames=list(manifest_rows[0]))
        writer.writeheader(); writer.writerows(manifest_rows)

    target_rows = "".join(
        f'<tr><td>{x["object"]}</td><td>{x["device"]}</td><td>{x["measured_vt_v"]:.3f}</td>'
        f'<td>{x["target_vt_v"]:.3f}</td><td>{x["delta_vt_mv"]:+.1f}</td>'
        f'<td>{x["measured_isat_ua"]:.2f}</td><td>{x["target_isat_ua"]:.2f}</td>'
        f'<td>{x["delta_isat_ua"]:+.2f}</td><td>{x["delta_isat_pct"]:+.2f}%</td></tr>'
        for x in comparisons)
    target_section = (f'''<section><h2>PU / PG / PD WAT vs WAT Target</h2>
    <p>The target values build a separate 6T model at the same SRAM VDD and WAT bias; they are not substituted into the Lot/Wafer WAT model.</p>
    <table><thead><tr><th>MOS object</th><th>Type</th><th>Lot/Wafer Vt (V)</th><th>WAT Target Vt (V)</th><th>ΔVt (mV)</th><th>Lot/Wafer Isat (µA)</th><th>WAT Target Isat (µA)</th><th>ΔIsat (µA)</th><th>ΔIsat (%)</th></tr></thead><tbody>{target_rows}</tbody></table></section>'''
                      if comparisons else "")

    snm_rows = "".join(
        f'<tr><td>{row["mode"]}</td><td>{row["current_snm_mv"]:.2f}</td>'
        f'<td>{row["target_snm_mv"]:.2f}</td><td>{row["delta_mv"]:+.2f}</td>'
        f'<td>{_fmt(row["delta_pct"], 2)}%</td></tr>' for row in comparison_rows)
    if has_target:
        read_overview_section = f'''<section><h2>Read SNM Target Comparison</h2><p>The plotted curves use the WAT-calibrated compact VTC model. X-axis is Vin and Y-axis is Vout, both expressed in volts and fixed at 0 to 1.20 V. Limiting squares are not drawn in this comparison view.</p>
        <img src="images/{png_name}" alt="Read SNM Lot/Wafer WAT versus WAT Target comparison">
        <table><thead><tr><th>Mode</th><th>Lot/Wafer SNM (mV)</th><th>WAT Target SNM (mV)</th><th>Lot/Wafer - WAT Target (mV)</th><th>Difference (%)</th></tr></thead><tbody>{snm_rows}</tbody></table></section>'''
    else:
        read_overview_section = f'''<section><h2>Read SNM Analysis</h2><p>WAT Target reference is disabled. This chart contains only the Lot/Wafer WAT-calibrated VTC and mirrored VTC. X-axis is Vin and Y-axis is Vout, both expressed in volts and fixed at 0 to 1.20 V.</p>
        <img src="images/{png_name}" alt="Read SNM Lot/Wafer WAT analysis"></section>'''
    state_table_rows = "".join(
        f'<tr><td>{row["dataset"]}</td><td>{row["upper_left_state_snm_mv"]:.2f}</td>'
        f'<td>{row["lower_right_state_snm_mv"]:.2f}</td><td>{row["cell_read_snm_min_mv"]:.2f}</td>'
        f'<td>{row["upper_minus_lower_mv"]:+.2f}</td><td>{_fmt(row["mismatch_index_pct"], 2)}%</td></tr>'
        for row in state_rows)
    analytical_table_rows = "".join(
        f'<tr><td>{row["dataset"]}</td><td>{"VALID" if row["valid"] else "N/A"}</td>'
        f'<td>{_fmt(row["snm_mv"], 2)}</td><td>{_fmt(row["vth_eff_v"], 4)}</td>'
        f'<td>{_fmt(row["q_beta_p_over_beta_a"], 4)}</td><td>{_fmt(row["r_beta_d_over_beta_a"], 4)}</td>'
        f'<td>{_fmt(row["k"], 4)}</td><td>{html.escape(str(row["reason"] or ""))}</td></tr>'
        for row in analytical_rows)
    analytical_section = f'''<section><h2>Analytical Read SNM Reference</h2>
    <p>This independent analytical estimate uses WAT-derived strength ratios and a common effective threshold mapped from PU/PG/PD Vt. It is reported separately from the geometric butterfly result.</p>
    <table><thead><tr><th>Dataset</th><th>Status</th><th>Analytical RSNM (mV)</th><th>VTH,eff (V)</th><th>q = βPU/βPG</th><th>r = βPD/βPG</th><th>k</th><th>Applicability</th></tr></thead><tbody>{analytical_table_rows}</tbody></table>
    <p>If the analytical square-root domain is non-positive, the result is shown as N/A instead of forcing a complex value.</p></section>'''
    electrical_snm_table_rows = "".join(
        f'<tr><td>{html.escape(row["dataset"])}</td>'
        f'<td>{row["pu_vt_v"]:.3f}</td><td>{row["pu_idsat_ua"]:.2f}</td>'
        f'<td>{row["pg_vt_v"]:.3f}</td><td>{row["pg_idsat_ua"]:.2f}</td>'
        f'<td>{row["pd_vt_v"]:.3f}</td><td>{row["pd_idsat_ua"]:.2f}</td>'
        f'<td>{row["q_beta_pu_over_pg"]:.4f}</td><td>{row["r_beta_pd_over_pg"]:.4f}</td>'
        f'<td>{row["idsat_pd_over_pg"]:.4f}</td><td>{row["idsat_pg_over_pu"]:.4f}</td>'
        f'<td>{row["read_snm_geometric_mv"]:.2f}</td>'
        f'<td>{_fmt(row["read_snm_eq_3_36_mv"], 2)}</td><td>{row["eq_3_36_status"]}</td></tr>'
        for row in electrical_snm_rows)
    electrical_snm_section = f'''<section><h2>WAT Electrical Parameters → SNM Table</h2>
    <p>This table uses only electrical values that can be entered from WAT measurement: PU/PG/PD Vt and Idsat at the stated WAT VDD. The β values, q/r and current ratios are mathematical proxies derived from those measurements. No W/L, Cox, mobility, BSIM coefficients, parasitics or other PDK/model-card-only parameters are required.</p>
    <div class="table-scroll"><table><thead><tr><th>Dataset</th><th>PU Vt<br>(V)</th><th>PU Idsat<br>(µA)</th><th>PG Vt<br>(V)</th><th>PG Idsat<br>(µA)</th><th>PD Vt<br>(V)</th><th>PD Idsat<br>(µA)</th><th>q<br>βPU/βPG</th><th>r<br>βPD/βPG</th><th>Idsat<br>PD/PG</th><th>Idsat<br>PG/PU</th><th>Read SNM<br>(mV)</th><th>Analytical<br>RSNM (mV)</th><th>Analytical<br>Status</th></tr></thead><tbody>{electrical_snm_table_rows}</tbody></table></div>
    <p><b>Interpretation:</b> geometric Read SNM comes from the WAT-calibrated butterfly curves; the analytical RSNM is an independent long-channel reference. These are correlation and trend metrics, not foundry sign-off values.</p></section>'''
    assumption_table_rows = "".join(
        f'<tr><td>{html.escape(str(row["parameter"]))}</td>'
        f'<td>{html.escape(str(row["value"]))}</td><td>{html.escape(str(row["unit"]))}</td>'
        f'<td>{html.escape(str(row["source"]))}</td><td>{html.escape(str(row["used_by"]))}</td>'
        f'<td>{row["active"]}</td></tr>' for row in assumption_rows)
    assumption_section = f'''<section><h2>6T Cell Geometry Reference</h2>
    <p>Channel length and PU/PG/PD widths are retained as known architecture references. The report derives geometry Cell Ratio = WPD/WPG and Pull-up Ratio = WPG/WPU from them.</p>
    <table><thead><tr><th>Parameter</th><th>Value</th><th>Unit</th><th>Source class</th><th>Used by</th><th>Role</th></tr></thead><tbody>{assumption_table_rows}</tbody></table>
    <p><b>Calculation policy:</b> VTC and SNM continue to use beta calibrated directly from measured WAT Vt/Idsat. Because the entered Idsat already represents measured device drive, W/L is not multiplied into beta a second time. Geometry ratios are comparison references unless future input supplies current normalized by device width.</p></section>'''
    object_section = ""
    if "cell" in result:
        mos_rows = "".join(
            f'<tr><td>{name}</td><td>{values["vt"]:.3f}</td><td>{values["ids"]:.2f}</td></tr>'
            for name, values in result["cell"]["mos"].items())
        object_section = f'''<section><h2>{html.escape(result.get("object_mode", "6T"))} WAT Objects</h2>
        <table><thead><tr><th>MOS object</th><th>Vt (V)</th><th>Isat / Ids (µA)</th></tr></thead><tbody>{mos_rows}</tbody></table></section>'''

    cfg, wat = result["config"], result["wat"]
    scope_reference_text = (
        "Blue curves use Lot/Wafer WAT inputs; orange curves use the enabled PU/PG/PD WAT Target reference."
        if has_target else
        "WAT Target reference is disabled; all charts and tables use Lot/Wafer WAT inputs only."
    )
    document = f'''<!doctype html><html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>HV28 SRAM Analysis</title>
    <style>
    :root{{font:100%/1.5 Calibri,"Microsoft JhengHei",Arial,sans-serif;color:#1d1d1f;background:#f5f5f7}}
    *{{box-sizing:border-box}} body{{margin:0;padding:clamp(1.25rem,4vw,3rem);background:radial-gradient(circle at 85% 0,#e8f2ff 0,transparent 28rem),#f5f5f7}}
    main{{max-width:1760px;margin:auto}} h1{{font-size:clamp(2.2rem,4vw,4rem);line-height:1.05;letter-spacing:-.035em;margin:.5rem 0 .35rem}} h2{{font-size:1.7rem;letter-spacing:-.018em;margin-top:0}}
    .summary,section{{background:rgba(255,255,255,.84);backdrop-filter:blur(22px) saturate(160%);border:1px solid rgba(255,255,255,.78);border-radius:1.25rem;box-shadow:0 1px 2px #0000000a,0 12px 36px #0000000d;padding:clamp(1rem,2.4vw,1.65rem);margin:1rem 0}}
    .summary{{color:#3a3a3c;font-size:1.05rem;border-left:4px solid #007aff}} img{{display:block;width:100%;height:auto;border:1px solid #e5e5ea;border-radius:1rem}} .formula{{background:#f2f2f7;border-radius:.8rem;padding:1rem;line-height:1.8;overflow:auto}} .table-scroll{{overflow-x:auto}}
    table{{border-collapse:separate;border-spacing:0;width:100%;margin-top:1rem;font-size:1rem;line-height:1.55}} th,td{{padding:.9rem .8rem;border-bottom:1px solid #e5e5ea;text-align:right;font-variant-numeric:tabular-nums}} th{{color:#6e6e73;font-size:.9rem;font-weight:680}} th:first-child,td:first-child{{text-align:left}} tbody tr:last-child td{{border-bottom:0}} code{{background:#f2f2f7;padding:.18rem .38rem;border-radius:.35rem}}
    @media(prefers-reduced-transparency:reduce){{.summary,section{{background:#fff;backdrop-filter:none;border-color:#d2d2d7}}}} @media(prefers-contrast:more){{.summary,section{{background:#fff;border:2px solid #1d1d1f}}}}
    </style></head><body><main>
    <h1>HV28 SRAM Analysis</h1><p>Lot/Wafer: <b>{html.escape(wat["corner"])}</b> · Object mode: <b>{html.escape(result.get("object_mode", "Grouped"))}</b> · SRAM VDD={cfg["nominal_vdd"]:.3f} V</p>
    <div class="summary"><b>Analysis scope:</b> Read SNM plus a W=1/W=0 Write SNM butterfly. WSNM is the largest valid square constrained to the Vin=Vout diagonal. {scope_reference_text}</div>
    {read_overview_section}
    <section><h2>Read SNM Butterfly and Left/Right Mismatch</h2><p>The measured 6T model keeps PUL/PGL/PDL and PUR/PGR/PDR independent. The upper-left and lower-right eyes represent opposite stored states. Cell RSNM is the smaller state margin; a larger difference or mismatch index indicates stronger left/right imbalance. X-axis is Vin and Y-axis is Vout, both expressed in volts and fixed at 0 to 1.20 V.</p>
    <img src="images/{butterfly_png_name}" alt="Asymmetric Read SNM butterfly with two state margins">
    <table><thead><tr><th>Dataset</th><th>Upper-left state SNM (mV)</th><th>Lower-right state SNM (mV)</th><th>Cell RSNM = min (mV)</th><th>Upper - Lower (mV)</th><th>Mismatch index</th></tr></thead><tbody>{state_table_rows}</tbody></table></section>
    <section><h2>Write SNM Butterfly Analysis</h2><p>The upper W=1 VTC is evaluated with BLB=VDD; the lower W=0 VTC is evaluated with BL=0. Both are plotted on the same Vin/Vout axes. WSNM is the side of the largest valid square whose diagonal is constrained to Vin=Vout; the diagonal guide itself is not displayed.</p>
    <img src="images/{write_png_name}" alt="W1 and W0 Write SNM butterfly analysis"></section>
    {electrical_snm_section}
    {assumption_section}
    {analytical_section}
    {target_section}
    <section><h2>Model Settings</h2><table><tbody><tr><td>Technology node</td><td>28 nm generic compact model</td></tr><tr><td>SRAM analysis VDD</td><td>{cfg["nominal_vdd"]:.3f} V</td></tr><tr><td>WAT calibration VDD</td><td>{cfg["wat_vdd"]:.3f} V</td></tr></tbody></table></section>
    {object_section}
    <p>Raw data: <code>snm_target_comparison.csv</code>, <code>w0_w1_wsnm_analysis.csv</code>, <code>read_snm_state_mismatch.csv</code>, <code>wat_electrical_snm_table.csv</code>, <code>cell_geometry_reference.csv</code>, <code>analytical_read_snm.csv</code>, <code>wat_target_comparison.csv</code>, <code>sram_wat_results.json</code>. Standalone charts are stored in <code>images/</code>.</p>
    </main></body></html>'''
    report = out / "sram_wat_report.html"
    report.write_text(document, encoding="utf-8")
    return report


def write_excel_sweep_outputs(entries: list[dict], out_dir: str | os.PathLike[str]) -> Path:
    """Write one HTML report plus combined butterfly and SNM trend plots for Excel VDD sweeps."""
    out = Path(out_dir); out.mkdir(parents=True, exist_ok=True)
    image_dir = out / "images"; image_dir.mkdir(parents=True, exist_ok=True)
    try:
        from reportlab.graphics import renderPM
        from svglib.svglib import svg2rlg
    except ImportError as exc:
        raise RuntimeError("PNG export packages are missing. Run: python -m pip install -r requirements.txt") from exc

    analyzed_entries = [item for item in entries if item.get("result") is not None]
    if not analyzed_entries:
        raise ValueError("Excel contains no Model VDD above the imported MOS threshold voltages")
    charts = [
        ("04_model_vdd_read_snm_butterfly", model_vdd_butterfly_svg(analyzed_entries), "Read SNM butterfly panels by model VDD"),
        ("05_snm_vs_model_vdd", snm_by_model_vdd_svg(analyzed_entries), "Read SNM versus model VDD"),
        ("06_all_vdd_read_snm_overlay", all_model_vdd_butterfly_overlay_svg(analyzed_entries), "All operating VDD Read SNM butterfly overlay"),
    ]
    manifest_rows = []
    for stem, svg, description in charts:
        svg_path = image_dir / f"{stem}.svg"
        png_path = image_dir / f"{stem}.png"
        svg_path.write_text(svg, encoding="utf-8")
        drawing = svg2rlg(str(svg_path))
        if drawing is None:
            raise RuntimeError(f"Could not render {stem}")
        renderPM.drawToFile(drawing, str(png_path), fmt="PNG", dpi=180, backend="rlPyCairo")
        manifest_rows.extend([
            {"filename": png_path.name, "format": "PNG", "description": description},
            {"filename": svg_path.name, "format": "SVG", "description": description},
        ])
    rows = []
    for item in entries:
        result = item.get("result")
        metrics = result["baseline_6t"]["metrics"] if result is not None else {}
        has_target = bool(result and result.get("datasheet_targets"))
        target_metrics = result["target_6t"]["metrics"] if has_target else {}
        read_snm = metrics.get("read_snm_mv")
        target_read_snm = target_metrics.get("read_snm_mv")
        rows.append({
            "lot_wafer": item["lot_wafer"], "model_vdd_v": item["model_vdd_v"],
            "read_snm_mv": read_snm,
            "read_snm_upper_left_mv": metrics.get("read_snm_upper_left_mv"),
            "read_snm_lower_right_mv": metrics.get("read_snm_lower_right_mv"),
            "read_snm_state_delta_mv": metrics.get("read_snm_delta_mv"),
            "read_snm_mismatch_index_pct": metrics.get("read_snm_mismatch_index_pct"),
            "target_read_snm_mv": target_read_snm,
            "read_snm_delta_mv": (read_snm - target_read_snm
                                  if read_snm is not None and target_read_snm is not None else None),
            "analytical_read_snm_mv": metrics.get("analytical_read_snm_mv"),
            "analysis_status": item.get("analysis_status", "Analyzed"),
        })
    with open(out / "excel_model_vdd_snm.csv", "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0])); writer.writeheader(); writer.writerows(rows)
    statistics_rows = []
    for item in entries:
        for mos_name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2"):
            stats = item.get("statistics", {}).get(mos_name)
            if stats is None:
                continue
            statistics_rows.append({
                "lot_wafer": item["lot_wafer"],
                "model_vdd_v": item["model_vdd_v"],
                "mos": DISPLAY_MOS_NAMES[mos_name],
                "valid_count": stats.valid_count,
                "total_count": stats.total_count,
                "coverage_pct": 100.0 * stats.valid_count / stats.total_count if stats.total_count else 0.0,
                "vt_mean_v": stats.vt_mean,
                "vt_median_v": stats.vt_median,
                "vt_stdev_v": stats.vt_stdev,
                "vt_min_v": stats.vt_min,
                "vt_max_v": stats.vt_max,
                "idsat_mean_ua": stats.ids_mean,
                "idsat_median_ua": stats.ids_median,
                "idsat_stdev_ua": stats.ids_stdev,
                "idsat_min_ua": stats.ids_min,
                "idsat_max_ua": stats.ids_max,
            })
    if statistics_rows:
        with open(out / "excel_wat_site_statistics.csv", "w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=list(statistics_rows[0])); writer.writeheader(); writer.writerows(statistics_rows)
    with open(image_dir / "excel_sweep_image_manifest.csv", "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(manifest_rows[0])); writer.writeheader(); writer.writerows(manifest_rows)
    table_rows = "".join(
        f'<tr><td>{html.escape(row["lot_wafer"])}</td><td>{row["model_vdd_v"]:.3f}</td>'
        f'<td>{_fmt(row["read_snm_upper_left_mv"], 2)}</td><td>{_fmt(row["read_snm_lower_right_mv"], 2)}</td>'
        f'<td>{_fmt(row["read_snm_mv"], 2)}</td><td>{_fmt(row["read_snm_state_delta_mv"], 2)}</td>'
        f'<td>{_fmt(row["read_snm_mismatch_index_pct"], 2)}</td><td>{_fmt(row["target_read_snm_mv"], 2)}</td><td>{_fmt(row["read_snm_delta_mv"], 2)}</td>'
        f'<td>{_fmt(row["analytical_read_snm_mv"], 2)}</td><td>{html.escape(row["analysis_status"])}</td></tr>' for row in rows)
    statistics_table_rows = "".join(
        f'<tr><td>{html.escape(row["lot_wafer"])}</td><td>{row["model_vdd_v"]:.3f}</td><td>{row["mos"]}</td>'
        f'<td>{row["valid_count"]}/{row["total_count"]}</td><td>{row["coverage_pct"]:.1f}%</td>'
        f'<td>{row["vt_mean_v"]:.4f}</td><td>{row["vt_stdev_v"]:.4f}</td>'
        f'<td>{row["idsat_mean_ua"]:.3f}</td><td>{row["idsat_stdev_ua"]:.3f}</td></tr>'
        for row in statistics_rows)
    statistics_section = (f'''<section><h2>Wafer-Site WAT Statistics</h2>
    <p>The 6T model uses the arithmetic mean of all valid sites at each Lot/Wafer, model VDD and physical MOS. Median, standard deviation, minimum and maximum are retained in <code>excel_wat_site_statistics.csv</code>.</p>
    <table><thead><tr><th>Lot/Wafer</th><th>Model VDD (V)</th><th>MOS</th><th>Valid/Total</th><th>Coverage</th><th>Mean Vt (V)</th><th>Vt σ (V)</th><th>Mean Idsat (µA)</th><th>Idsat σ (µA)</th></tr></thead><tbody>{statistics_table_rows}</tbody></table></section>'''
        if statistics_rows else "")
    document = f'''<!doctype html><html><head><meta charset="utf-8"><title>HV28 SRAM Excel VDD Sweep</title>
    <style>body{{font-family:Calibri,Arial,sans-serif;background:#F5F5F7;color:#1D1D1F;margin:0}}main{{max-width:1500px;margin:auto;padding:42px}}section{{background:#fff;border-radius:18px;padding:24px;margin:16px 0}}img{{width:100%;height:auto;border:1px solid #E5E5EA;border-radius:12px}}table{{border-collapse:collapse;width:100%}}th,td{{padding:10px;border-bottom:1px solid #E5E5EA;text-align:left}}th{{color:#6E6E73}}.note{{color:#6E6E73}}</style></head><body><main>
    <h1>HV28 SRAM Excel Model-VDD Sweep</h1><p class="note">PU, PG and PD worksheets are combined by Lot/Wafer and model VDD. Repeated wafer-site records are converted to V/uA and averaged per physical MOS before 6T analysis.</p>
    <section><h2>Read SNM Butterfly by Operating VDD</h2><p>Each panel compares measured-WAT and WAT-target VTCs, mirrored VTCs and SNM squares under the same operating voltage. Vin/Vout axes are fixed at 0 to 1.20 V.</p><img src="images/04_model_vdd_read_snm_butterfly.png" alt="Measured and target Read SNM butterflies by model VDD"></section>
    <section><h2>Read SNM Trend</h2><p>Measured and target Read SNM are plotted against imported model VDD.</p><img src="images/05_snm_vs_model_vdd.png" alt="Measured and target Read SNM versus model VDD"></section>
    <section><h2>All Operating Voltages Overlay</h2><p>All analyzed VDD butterfly curves are integrated in one Vin/Vout plot. Curve color identifies operating VDD; solid lines are measured WAT and dashed lines are WAT Target.</p><img src="images/06_all_vdd_read_snm_overlay.png" alt="All operating VDD Read SNM butterfly overlay"></section>
    <section><h2>Model-VDD Results</h2><table><thead><tr><th>Lot/Wafer</th><th>Model VDD (V)</th><th>Upper-left RSNM (mV)</th><th>Lower-right RSNM (mV)</th><th>Cell RSNM (mV)</th><th>State delta (mV)</th><th>Mismatch (%)</th><th>Target RSNM (mV)</th><th>WAT - Target (mV)</th><th>Analytical RSNM (mV)</th><th>Status</th></tr></thead><tbody>{table_rows}</tbody></table></section>
    {statistics_section}
    <p class="note">Raw results: <code>excel_model_vdd_snm.csv</code> and <code>excel_wat_site_statistics.csv</code>; images: <code>images/04_model_vdd_read_snm_butterfly.png</code>, <code>images/05_snm_vs_model_vdd.png</code> and <code>images/06_all_vdd_read_snm_overlay.png</code>.</p></main></body></html>'''
    report = out / "excel_wat_sweep_report.html"
    report.write_text(document, encoding="utf-8")
    return report


def analyze_excel_wat_sweep(path: str | os.PathLike[str], out_dir: str | os.PathLike[str], cfg: Config,
                            targets: DatasheetTargets | None = None,
                            archive_run: bool = False) -> Path:
    """Analyze each imported model voltage and create one combined Excel sweep report."""
    validate_config(cfg)
    samples = read_wat_excel(path, cfg.nominal_vdd)
    entries = []
    for sample in samples:
        highest_vt = max(getattr(sample.cell, name).vt for name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2"))
        if sample.model_vdd_v <= highest_vt:
            result = None
            status = f"WAT statistics only: VDD ≤ highest MOS Vt ({highest_vt:.3f} V)"
        else:
            model_cfg = replace(cfg, nominal_vdd=sample.model_vdd_v)
            result = analyze_six_mos(sample.cell, model_cfg, targets) if targets else analyze_six_mos(sample.cell, model_cfg, None)
            status = "Analyzed"
        entries.append({"lot_wafer": sample.lot_wafer, "model_vdd_v": sample.model_vdd_v,
                        "statistics": sample.statistics, "analysis_status": status, "result": result})
    if archive_run:
        wafer_ids = sorted({sample.lot_wafer for sample in samples})
        wafer_label = (wafer_ids[0] if len(wafer_ids) == 1 else
                       f"{wafer_ids[0]}_plus_{len(wafer_ids) - 1}")
        out_dir = create_run_output_dir(out_dir, wafer_label, "excel_vdd_sweep")
    return write_excel_sweep_outputs(entries, out_dir)


def run_analysis(csv_path: str, out_dir: str, cfg: Config, corner: str | None = None,
                 targets: DatasheetTargets | None = None) -> list[Path]:
    validate_config(cfg)
    if Path(csv_path).suffix.lower() == ".xlsx":
        if corner:
            raise ValueError("--corner is not used for Excel model-VDD sweep import")
        return [analyze_excel_wat_sweep(csv_path, out_dir, cfg, targets, archive_run=True)]
    points = read_wat_csv(csv_path)
    if corner:
        points = [p for p in points if p.corner.lower() == corner.lower()]
        if not points:
            raise ValueError(f"corner not found: {corner}")
    reports = []
    for p in points:
        if targets:
            cell = ThreeTWatCell(p.corner, MosWat(p.pu_vt, p.pu_ids),
                                 MosWat(p.pg_vt, p.pg_ids), MosWat(p.pd_vt, p.pd_ids))
            result = analyze_three_mos(cell, cfg, targets)
        else:
            result = analyze(p, cfg)
        target = create_run_output_dir(out_dir, p.corner, "6t_analysis")
        reports.append(write_outputs(result, target))
    return reports


def _launch_legacy_gui() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    root = tk.Tk(); root.title("HV28 SRAM Analysis"); root.geometry("820x720"); root.minsize(760, 680)
    saved_state = load_gui_state()
    def saved_text(group: str, key: str, default: object) -> str:
        return str(saved_state.get(group, {}).get(key, default)) if isinstance(saved_state.get(group), dict) else str(default)

    values = {"out": tk.StringVar(value=saved_text("values", "out", Path.cwd()/"output")),
              "corner": tk.StringVar(value=saved_text("values", "corner", "DEMO28_TT_W01")),
              "excel_input": tk.StringVar(value=saved_text("values", "excel_input", ""))}
    defaults={"pu":("0.385","44"),"pg":("0.365","82"),"pd":("0.355","124")}
    wat_values={}
    for dev in ("pu1","pu2","pg1","pg2","pd1","pd2"):
        vt,ids=defaults[dev[:2]]
        wat_values[f"{dev}_vt"]=tk.StringVar(value=vt); wat_values[f"{dev}_ids"]=tk.StringVar(value=ids)
    numeric = {k: tk.StringVar(value=str(v)) for k,v in asdict(Config()).items() if k != "grid_points"}
    frame = ttk.Frame(root, padding=18); frame.pack(fill="both", expand=True)
    ttk.Label(frame, text="HV28 SRAM Analysis", font=("Calibri", 17, "bold")).grid(row=0,column=0,columnspan=4,sticky="w")
    ttk.Label(frame, text="固定 28 nm 6T 架構；手動輸入 WAT，Python 產生 SNM 與 R/W Vmin（不使用 SPICE）",
              foreground="#475569").grid(row=1,column=0,columnspan=4,sticky="w",pady=(3,14))

    wat_box = ttk.LabelFrame(frame, text="WAT 手動輸入", padding=12)
    wat_box.grid(row=2,column=0,columnspan=4,sticky="ew",pady=5)
    ttk.Label(wat_box,text="Lot / Wafer ID").grid(row=0,column=0,sticky="w",padx=(0,8))
    ttk.Entry(wat_box,textvariable=values["corner"],width=16).grid(row=0,column=1,sticky="w")
    ttk.Label(wat_box,text="Q side",font=("Calibri",10,"bold")).grid(row=1,column=0,columnspan=3,pady=(10,3))
    ttk.Label(wat_box,text="QB side",font=("Calibri",10,"bold")).grid(row=1,column=3,columnspan=3,pady=(10,3))
    # Cards are placed like the physical cell: PU top, PG center, PD bottom.
    for row,(kind,desc) in enumerate((("pu","PMOS"),("pg","access"),("pd","NMOS")),2):
        for side,col in (("1",0),("2",3)):
            name=kind+side
            card=ttk.LabelFrame(wat_box,text=name.upper(),padding=5); card.grid(row=row,column=col,columnspan=3,padx=6,pady=4,sticky="ew")
            ttk.Label(card,text="Vt (V)").grid(row=0,column=0); ttk.Entry(card,textvariable=wat_values[f"{name}_vt"],width=9).grid(row=0,column=1,padx=(3,9))
            ttk.Label(card,text="Ids (µA)").grid(row=0,column=2); ttk.Entry(card,textvariable=wat_values[f"{name}_ids"],width=9).grid(row=0,column=3,padx=3)
        ttk.Label(wat_box,text={"pu":"VDD — pull-up","pg":"BL/BLB — WL access","pd":"pull-down — GND"}[kind],foreground="#64748b").grid(row=row,column=6,sticky="w")
    ttk.Label(wat_box,text="Q  ↔  cross-coupled  ↔  QB　　PU Vt 可輸入負值，計算使用 |Vtp|。",foreground="#7c3aed").grid(row=5,column=0,columnspan=7,pady=(8,0))

    out_box = ttk.Frame(frame); out_box.grid(row=3,column=0,columnspan=4,sticky="ew",pady=8)
    ttk.Label(out_box,text="報表輸出目錄").grid(row=0,column=0,sticky="w")
    ttk.Entry(out_box,textvariable=values["out"],width=70).grid(row=0,column=1,sticky="ew",padx=8)
    def pick_out():
        value=filedialog.askdirectory()
        if value: values["out"].set(value)
    ttk.Button(out_box,text="瀏覽",command=pick_out).grid(row=0,column=2)
    out_box.columnconfigure(1,weight=1)

    cfg_box = ttk.LabelFrame(frame,text="分析條件",padding=12)
    cfg_box.grid(row=4,column=0,columnspan=4,sticky="ew",pady=5)
    labels = [("wat_vdd","WAT Ids 測試 VDD (V)"),("nominal_vdd","SRAM 分析 VDD (V)"),("vt_step","Vt 調整量 (V)"),("ids_step_pct","Ids 調整量 (%)"),("vmin_start","Vmin 起點 (V)"),("vmin_stop","Vmin 終點 (V)"),("vmin_step","Vmin 步階 (V)"),("read_snm_limit","Read SNM 下限 (V)")]
    for i,(key,label) in enumerate(labels):
        r=i//2; c=(i%2)*2
        ttk.Label(cfg_box,text=label).grid(row=r,column=c,sticky="w",pady=6,padx=(0,8))
        ttk.Entry(cfg_box,textvariable=numeric[key],width=14).grid(row=r,column=c+1,sticky="w",padx=(0,28))
    status = tk.StringVar(value="待命")
    def execute():
        try:
            kwargs={k:float(v.get()) for k,v in numeric.items()}
            cfg=Config(**kwargs)
            validate_config(cfg)
            mos={}
            for name in ("pu1","pu2","pg1","pg2","pd1","pd2"):
                mos[name]=MosWat(_positive(wat_values[f"{name}_vt"].get(),f"{name}_vt"),
                                 _positive(wat_values[f"{name}_ids"].get(),f"{name}_ids"))
            point=SixTWatCell(values["corner"].get().strip() or "Manual",**mos)
            status.set("分析中…"); root.update_idletasks()
            run_dir=create_run_output_dir(values["out"].get(),point.corner,"6t_analysis")
            report=write_outputs(analyze_six_mos(point,cfg),run_dir)
            status.set(f"完成：{point.corner}；{report}")
            webbrowser.open(report.resolve().as_uri())
        except Exception as exc:
            status.set("失敗"); messagebox.showerror("分析失敗",str(exc))
    ttk.Button(frame,text="產生並開啟分析報表",command=execute).grid(row=5,column=0,columnspan=4,pady=(18,8),ipadx=24,ipady=6)
    ttk.Label(frame,textvariable=status,wraplength=740).grid(row=6,column=0,columnspan=4)
    ttk.Label(frame,text="輸出：SNM、R/W Vmin、WT 0-Bit Scan4N / Select_Write / Select_Read、CSV、JSON。",
              foreground="#475569").grid(row=7,column=0,columnspan=4,pady=12)
    frame.columnconfigure(1,weight=1); frame.columnconfigure(3,weight=1); root.mainloop()


def launch_gui() -> None:
    """Apple-inspired desktop UI with direct manipulation of the 6T diagram."""
    import queue
    import threading
    import tkinter as tk
    from tkinter import filedialog, font as tkfont, messagebox, ttk

    # Airbnb-inspired workspace tokens: generous white space, quiet neutral
    # surfaces and one deliberate Rausch-red action color.  Only presentation
    # changes here; all 6T WAT calculation paths remain untouched.
    BG, CARD, TEXT, SECONDARY = "#FFFFFF", "#FFFFFF", "#222222", "#6A6A6A"
    BLUE, BLUE_DARK, BORDER, GREEN, RED = "#FF385C", "#E00B41", "#DDDDDD", "#460479", "#C13515"
    ACCENT, SURFACE, HOVER = "#92174D", "#F7F7F7", "#F2F2F2"
    root = tk.Tk()
    root.title("HV28 SRAM Analysis")
    root.geometry("1180x940")
    root.minsize(1080, 860)
    root.configure(bg=BG)
    if sys.platform == "win32":
        root.after(0, lambda: root.state("zoomed"))
    for named_font in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont",
                       "TkCaptionFont", "TkSmallCaptionFont", "TkIconFont", "TkTooltipFont"):
        try:
            tkfont.nametofont(named_font).configure(family="Calibri")
        except tk.TclError:
            pass

    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("Root.TFrame", background=BG)
    style.configure("Card.TFrame", background=CARD, relief="flat", borderwidth=0)
    style.configure("Title.TLabel", background=BG, foreground=TEXT, font=("Calibri", 22, "bold"))
    style.configure("Subtitle.TLabel", background=BG, foreground=SECONDARY, font=("Calibri", 10))
    style.configure("Section.TLabel", background=CARD, foreground=TEXT, font=("Calibri", 13, "bold"))
    style.configure("ChartTitle.TLabel", background=CARD, foreground=TEXT,
                    font=("Calibri", 18, "bold"))
    style.configure("Body.TLabel", background=CARD, foreground=TEXT, font=("Calibri", 10))
    style.configure("Meta.TLabel", background=CARD, foreground=SECONDARY, font=("Calibri", 9))
    style.configure("Apple.TEntry", fieldbackground=CARD, foreground=TEXT, bordercolor=BORDER,
                    lightcolor=BORDER, darkcolor=BORDER, padding=(11, 8))
    style.map("Apple.TEntry", bordercolor=[("focus", BLUE)], lightcolor=[("focus", BLUE)])
    style.configure("Accent.TButton", background=BLUE, foreground="white", borderwidth=0,
                    font=("Calibri", 11, "bold"), padding=(22, 12))
    style.map("Accent.TButton", background=[("pressed", BLUE_DARK), ("active", "#FF5A78"), ("disabled", "#FFD1DA")])
    style.configure("Quiet.TButton", background=CARD, foreground=TEXT, borderwidth=1,
                    relief="solid", padding=(12, 8))
    style.map("Quiet.TButton", background=[("pressed", SURFACE), ("active", SURFACE)],
              bordercolor=[("active", "#C1C1C1")])
    style.configure("Apple.Horizontal.TProgressbar", background=BLUE, troughcolor="#EBEBEB", borderwidth=0)
    style.configure("Apple.TNotebook", background=BG, borderwidth=0, tabmargins=(0, 0, 0, 16))
    style.configure("Apple.TNotebook.Tab", background=BG, foreground=SECONDARY,
                    borderwidth=0, padding=(18, 12), font=("Calibri", 10, "bold"))
    style.map("Apple.TNotebook.Tab", background=[("selected", BG), ("active", BG)],
              foreground=[("selected", TEXT), ("active", TEXT)],
              font=[("selected", ("Calibri", 12, "bold")),
                    ("!selected", ("Calibri", 10, "bold"))],
              padding=[("selected", (22, 12)), ("!selected", (18, 12))])

    saved_state = load_gui_state()

    def saved_text(group: str, key: str, default: object) -> str:
        group_values = saved_state.get(group, {})
        return str(group_values.get(key, default)) if isinstance(group_values, dict) else str(default)

    def saved_bool(group: str, key: str, default: bool) -> bool:
        group_values = saved_state.get(group, {})
        if not isinstance(group_values, dict):
            return default
        value = group_values.get(key, default)
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "on"}

    values = {
        "out": tk.StringVar(value=saved_text("values", "out", Path.cwd() / "output")),
        "corner": tk.StringVar(value=saved_text("values", "corner", "DEMO28_TT_W01")),
        "excel_input": tk.StringVar(value=saved_text("values", "excel_input", "")),
    }
    defaults = {"pu": ("0.385", "44"), "pg": ("0.365", "82"), "pd": ("0.355", "124")}
    target_defaults = {"pu": ("0.380", "45"), "pg": ("0.370", "80"), "pd": ("0.360", "120")}
    wat_values: dict[str, tk.StringVar] = {}
    for dev in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2"):
        vt, ids = defaults[dev[:2]]
        wat_values[f"{dev}_vt"] = tk.StringVar(value=saved_text("wat", f"{dev}_vt", vt))
        wat_values[f"{dev}_ids"] = tk.StringVar(value=saved_text("wat", f"{dev}_ids", ids))
    target_values: dict[str, tk.StringVar] = {}
    for dev in ("pu", "pg", "pd"):
        vt, ids = target_defaults[dev]
        target_values[f"{dev}_vt"] = tk.StringVar(value=saved_text("targets", f"{dev}_vt", vt))
        target_values[f"{dev}_ids"] = tk.StringVar(value=saved_text("targets", f"{dev}_ids", ids))
    use_wat_target_reference = tk.BooleanVar(
        value=saved_bool("options", "use_wat_target_reference", True))
    config_defaults = Config()
    numeric = {
        "nominal_vdd": tk.StringVar(value=saved_text("numeric", "nominal_vdd", config_defaults.nominal_vdd)),
        "wat_vdd": tk.StringVar(value=saved_text("numeric", "wat_vdd", config_defaults.wat_vdd)),
    }
    assumption_specs = (
        ("channel_length_nm", "Channel length L", "nm"),
        ("pu_width_nm", "PU width", "nm"),
        ("pg_width_nm", "PG width", "nm"),
        ("pd_width_nm", "PD width", "nm"),
    )
    assumption_values = {
        key: tk.StringVar(value=saved_text("assumptions", key, getattr(config_defaults, key)))
        for key, _label, _unit in assumption_specs
    }

    shell = ttk.Frame(root, style="Root.TFrame", padding=(40, 24, 40, 30)); shell.pack(fill="both", expand=True)
    header = ttk.Frame(shell, style="Root.TFrame"); header.pack(fill="x", pady=(0, 22))
    ttk.Label(header, text="HV28 SRAM Analysis", style="Title.TLabel").pack(side="left")
    badge = tk.Label(header, text="  WAT STUDIO  ", bg="#FFF1F3", fg=BLUE,
                     font=("Calibri", 9, "bold"), padx=10, pady=6)
    badge.pack(side="left", padx=14, pady=(4, 0))
    ttk.Label(header, text="6T WAT-calibrated analysis workspace", style="Subtitle.TLabel").pack(side="right", pady=(7, 0))

    notebook = ttk.Notebook(shell, style="Apple.TNotebook")
    notebook.pack(fill="both", expand=True)
    bitcell_tab = ttk.Frame(notebook, style="Root.TFrame")
    curve_tab = ttk.Frame(notebook, style="Root.TFrame")
    advisor_tab = ttk.Frame(notebook, style="Root.TFrame")
    write_margin_tab = ttk.Frame(notebook, style="Root.TFrame")
    training_tab = ttk.Frame(notebook, style="Root.TFrame")
    notebook.add(bitcell_tab, text="6T Bitcell Analysis")
    notebook.add(curve_tab, text="Estimate Vmin Curve")
    notebook.add(advisor_tab, text="Lot/Wafer Advisor")
    notebook.add(training_tab, text="6T Drive Monitor")

    # Interactive educational view.  The controls intentionally use the same
    # compact WAT-calibrated model as the main analysis, while making it clear
    # that the displayed values are trends rather than sign-off specifications.
    training_tab.columnconfigure(0, weight=4)
    training_tab.columnconfigure(1, weight=6)
    training_tab.rowconfigure(0, weight=1)
    training_input_card = ttk.Frame(training_tab, style="Card.TFrame", padding=20)
    training_output_card = ttk.Frame(training_tab, style="Card.TFrame", padding=20)
    training_input_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    training_output_card.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
    ttk.Label(training_input_card, text="6T Drive Monitor", style="Section.TLabel").pack(anchor="w")
    ttk.Label(
        training_input_card,
        text="Adjust PU / PG / PD Vt or Idsat to monitor the 6T read/write drive-balance trend in real time.",
        style="Meta.TLabel", wraplength=390).pack(anchor="w", pady=(3, 16))

    training_defaults = {
        "vdd": 0.90, "pu_vt": 0.385, "pu_ids": 44.0,
        "pg_vt": 0.365, "pg_ids": 82.0, "pd_vt": 0.355, "pd_ids": 124.0,
    }
    training_values = {
        key: tk.DoubleVar(value=float(saved_text("training", key, default)))
        for key, default in training_defaults.items()
    }
    training_display: dict[str, tk.StringVar] = {
        key: tk.StringVar() for key in training_values
    }

    def training_slider(parent, label: str, key: str, low: float, high: float,
                        resolution: float, unit: str, color: str) -> None:
        row = ttk.Frame(parent, style="Card.TFrame")
        row.pack(fill="x", pady=4)
        ttk.Label(row, text=label, style="Body.TLabel", width=12).pack(side="left")
        value_label = tk.Label(row, textvariable=training_display[key], bg=CARD, fg=color,
                               font=("Calibri", 10, "bold"), width=10, anchor="e")
        value_label.pack(side="right")
        tk.Scale(row, variable=training_values[key], from_=low, to=high,
                 resolution=resolution, orient="horizontal", showvalue=False,
                 length=190, highlightthickness=0, bd=0, bg=CARD, fg=TEXT,
                 activebackground=color, troughcolor="#E5E5EA").pack(side="right", padx=(8, 4))

    ttk.Label(training_input_card, text="Operating condition", style="Section.TLabel").pack(anchor="w")
    training_slider(training_input_card, "Model VDD", "vdd", 0.50, 1.20, 0.01, "V", BLUE)
    for device, title, color in (("pu", "PU / pull-up", RED), ("pg", "PG / access", GREEN),
                                 ("pd", "PD / pull-down", BLUE)):
        ttk.Label(training_input_card, text=title, style="Section.TLabel").pack(anchor="w", pady=(12, 0))
        training_slider(training_input_card, "Vt", f"{device}_vt", 0.20, 0.55, 0.001, "V", color)
        training_slider(training_input_card, "Idsat", f"{device}_ids", 5.0, 220.0, 1.0, "µA", color)

    ttk.Label(
        training_input_card,
        text="Trend monitor only: results show relative 6T drive changes. Use PDK simulation and measured WT for sign-off.",
        style="Meta.TLabel", wraplength=390).pack(anchor="w", pady=(16, 0))

    ttk.Label(training_output_card, text="Live 6T Trend", style="ChartTitle.TLabel").pack(anchor="w")
    training_summary = tk.Label(training_output_card, bg=CARD, fg=SECONDARY,
                                font=("Calibri", 10), anchor="w", justify="left")
    training_summary.pack(anchor="w", pady=(2, 8))
    training_canvas = tk.Canvas(training_output_card, bg=CARD, highlightthickness=0, height=300)
    training_canvas.pack(fill="x")
    ttk.Separator(training_output_card, orient="horizontal").pack(fill="x", pady=(4, 10))
    ttk.Label(training_output_card, text="Statistical CR / PR Shmoo",
              style="ChartTitle.TLabel").pack(anchor="w")
    training_shmoo_summary = tk.Label(
        training_output_card, bg=CARD, fg=SECONDARY, font=("Calibri", 9),
        anchor="w", justify="left")
    training_shmoo_summary.pack(anchor="w", pady=(2, 4))
    training_shmoo_canvas = tk.Canvas(
        training_output_card, bg=CARD, highlightthickness=0, height=300)
    training_shmoo_canvas.pack(fill="both", expand=True)

    def draw_drive_monitor_shmoo(data: dict[str, object]) -> None:
        reference = drive_monitor_shmoo_reference(training_values["vdd"].get())
        cr_stats = reference["cr"]
        pr_stats = reference["pr"]
        current_cr = float(data["cell_ratio"])
        current_pr = float(data["pull_up_ratio"])
        if current_cr >= cr_stats["median"] and current_pr >= pr_stats["median"]:
            grade, grade_color = "PREFERRED", GREEN
        elif current_cr >= cr_stats["q1"] and current_pr >= pr_stats["q1"]:
            grade, grade_color = "MONITOR", "#B77900"
        else:
            grade, grade_color = "LOW", RED
        training_shmoo_summary.configure(
            text=(f"Auto target @ {reference['vdd_v']:.2f} V · "
                  f"Median CR {cr_stats['median']:.3f} / PR {pr_stats['median']:.3f} · "
                  f"Current {grade}"), fg=grade_color)

        canvas = training_shmoo_canvas
        canvas.delete("all")
        width = max(canvas.winfo_width(), 620)
        height = max(canvas.winfo_height(), 280)
        left, top, right, bottom = 72.0, 22.0, width - 28.0, height - 54.0
        x_min = min(pr_stats["p05"], current_pr) * .94
        x_max = max(pr_stats["p95"], current_pr) * 1.06
        y_min = min(cr_stats["p05"], current_cr) * .94
        y_max = max(cr_stats["p95"], current_cr) * 1.06
        if x_max <= x_min:
            x_max = x_min + .1
        if y_max <= y_min:
            y_max = y_min + .1

        def plot_xy(pr_value: float, cr_value: float) -> tuple[float, float]:
            x = left + (pr_value - x_min) / (x_max - x_min) * (right - left)
            y = bottom - (cr_value - y_min) / (y_max - y_min) * (bottom - top)
            return x, y

        # Nested rectangles intentionally require both axes to reach the same
        # quartile. This keeps upper-left and lower-right trade-off cells from
        # being marked green on the strength of only one ratio.
        canvas.create_rectangle(left, top, right, bottom, fill="#FBE4E1", outline=BORDER)
        q1_x, q1_y = plot_xy(pr_stats["q1"], cr_stats["q1"])
        med_x, med_y = plot_xy(pr_stats["median"], cr_stats["median"])
        canvas.create_rectangle(q1_x, top, right, q1_y, fill="#FFF0C2", outline="")
        canvas.create_rectangle(med_x, top, right, med_y, fill="#DDF3E2", outline="")

        for fraction in (0.0, .25, .50, .75, 1.0):
            gx = left + fraction * (right - left)
            gy = top + fraction * (bottom - top)
            x_value = x_min + fraction * (x_max - x_min)
            y_value = y_max - fraction * (y_max - y_min)
            canvas.create_line(gx, top, gx, bottom, fill="#D8DEE4")
            canvas.create_line(left, gy, right, gy, fill="#D8DEE4")
            canvas.create_text(gx, bottom + 18, text=f"{x_value:.2f}",
                               fill=SECONDARY, font=("Calibri", 9))
            canvas.create_text(left - 10, gy, text=f"{y_value:.2f}", anchor="e",
                               fill=SECONDARY, font=("Calibri", 9))

        canvas.create_line(q1_x, top, q1_x, bottom, fill="#B77900", dash=(5, 4), width=2)
        canvas.create_line(left, q1_y, right, q1_y, fill="#B77900", dash=(5, 4), width=2)
        canvas.create_line(med_x, top, med_x, bottom, fill="#248A3D", dash=(6, 4), width=2)
        canvas.create_line(left, med_y, right, med_y, fill="#248A3D", dash=(6, 4), width=2)
        canvas.create_text(med_x + 8, top + 14, text="Auto target median",
                           anchor="w", fill="#1B6E35", font=("Calibri", 9, "bold"))

        current_x, current_y = plot_xy(current_pr, current_cr)
        canvas.create_oval(current_x - 7, current_y - 7, current_x + 7, current_y + 7,
                           fill=BLUE, outline="#FFFFFF", width=2)
        label_anchor = "e" if current_x > (left + right) / 2 else "w"
        label_x = current_x - 11 if label_anchor == "e" else current_x + 11
        canvas.create_text(
            label_x, current_y - 10,
            text=f"Current  PR {current_pr:.3f} / CR {current_cr:.3f}",
            anchor=label_anchor, fill=TEXT, font=("Calibri", 10, "bold"))
        canvas.create_text((left + right) / 2, height - 18,
                           text="Pull-up Ratio  PR = MOSdrive(PG) / MOSdrive(PU)  (right = easier write)",
                           fill=TEXT, font=("Calibri", 10, "bold"))
        canvas.create_text(18, (top + bottom) / 2,
                           text="CR = MOSdrive(PD) / MOSdrive(PG)  (up = stronger read)", angle=90,
                           fill=TEXT, font=("Calibri", 10, "bold"))
        canvas.create_text(right, bottom + 34,
                           text="Green: both ≥ Median   Yellow: both ≥ Q1   Red: either < Q1",
                           anchor="e", fill=SECONDARY, font=("Calibri", 8))

    def draw_training_trend(*_args) -> None:
        try:
            data = drive_monitor_metrics(
                WatPoint("DriveMonitor", training_values["pu_vt"].get(), training_values["pu_ids"].get(),
                         training_values["pg_vt"].get(), training_values["pg_ids"].get(),
                         training_values["pd_vt"].get(), training_values["pd_ids"].get()),
                training_values["vdd"].get())
        except Exception as exc:
            training_summary.configure(text=f"Input error: {exc}", fg=RED)
            return
        for key, variable in training_values.items():
            unit = "V" if key.endswith("vt") or key == "vdd" else "µA"
            training_display[key].set(f"{variable.get():.3f} {unit}" if unit == "V" else f"{variable.get():.0f} {unit}")
        training_summary.configure(
            text=(f"CR = MOSdrive(PD) / MOSdrive(PG) = {data['cell_ratio']:.2f}    ·    "
                  f"PR = MOSdrive(PG) / MOSdrive(PU) = {data['pull_up_ratio']:.2f}\n"
                  "Higher CR generally helps read stability; higher PR generally helps write-0 ability."),
            fg=SECONDARY)
        canvas = training_canvas
        canvas.delete("all")
        width = max(canvas.winfo_width(), 620)
        height = max(canvas.winfo_height(), 290)
        canvas.create_text(28, 16, text="Effective WAT-calibrated MOSdrive (relative)", anchor="w",
                           fill=TEXT, font=("Calibri", 12, "bold"))
        betas = [("PU", data["beta_pu"], RED), ("PG", data["beta_pg"], GREEN), ("PD", data["beta_pd"], BLUE)]
        max_beta = max(value for _, value, _ in betas) or 1.0
        x0, bar_w, gap = 70, 95, 54
        chart_bottom, chart_top = 118, 38
        for index, (label, beta, color) in enumerate(betas):
            x = x0 + index * (bar_w + gap)
            h = (chart_bottom - chart_top) * beta / max_beta
            canvas.create_rectangle(x, chart_bottom - h, x + bar_w, chart_bottom,
                                    fill=color, outline="")
            canvas.create_text(x + bar_w / 2, chart_bottom + 15, text=label,
                               fill=TEXT, font=("Calibri", 11, "bold"))
            canvas.create_text(x + bar_w / 2, chart_bottom - h - 12, text=f"{beta:.0f}",
                               fill=TEXT, font=("Calibri", 10, "bold"))
        canvas.create_line(48, chart_bottom, 48 + 3 * (bar_w + gap) - gap, chart_bottom,
                           fill=BORDER, width=1)
        canvas.create_text(28, 154, text="Compact-model margins at selected VDD", anchor="w",
                           fill=TEXT, font=("Calibri", 12, "bold"))
        gauges = [
            ("Read SNM", data["read_snm_mv"], 0.5 * training_values["vdd"].get() * 1000.0, BLUE,
             "Higher is more read-stable"),
            ("Write margin", data["write_margin_mv"], training_values["vdd"].get() * 1000.0, GREEN,
             "Higher means PG can overcome PU with more bitline tolerance"),
        ]
        for index, (label, value, scale, color, note) in enumerate(gauges):
            y = 190 + index * 50
            fraction = max(0.0, min(1.0, value / max(scale, 1.0)))
            canvas.create_text(48, y, text=label, anchor="w", fill=TEXT, font=("Calibri", 11, "bold"))
            canvas.create_text(48, y + 15, text=note, anchor="w", fill=SECONDARY, font=("Calibri", 8))
            left, right = 270, width - 35
            canvas.create_rectangle(left, y - 8, right, y + 12, fill="#E5E5EA", outline="")
            canvas.create_rectangle(left, y - 8, left + (right-left)*fraction, y + 12, fill=color, outline="")
            canvas.create_text(right, y + 2, text=f"{value:.1f} mV", anchor="e", fill=TEXT,
                               font=("Calibri", 11, "bold"))
        canvas.create_text(48, height - 8,
                           text="Read trend: PD↑ or PG↓ tends to raise CR.  Write trend: PG↑ or PU↓ tends to raise PR.",
                           anchor="w", fill=SECONDARY, font=("Calibri", 10))
        draw_drive_monitor_shmoo(data)

    for variable in training_values.values():
        variable.trace_add("write", draw_training_trend)
    training_canvas.bind("<Configure>", draw_training_trend)
    training_shmoo_canvas.bind("<Configure>", draw_training_trend)
    root.after_idle(draw_training_trend)

    content = ttk.Frame(bitcell_tab, style="Root.TFrame"); content.pack(fill="both", expand=True)
    content.columnconfigure(0, weight=7); content.columnconfigure(1, weight=4); content.rowconfigure(0, weight=1)
    left = ttk.Frame(content, style="Card.TFrame", padding=18); left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    right_card = ttk.Frame(content, style="Card.TFrame"); right_card.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
    # Keep the primary action visible while the analysis inputs scroll independently.
    right_footer = ttk.Frame(right_card, style="Card.TFrame", padding=(18, 8, 18, 18))
    right_footer.pack(side="bottom", fill="x")
    right_view = tk.Canvas(right_card, bg=CARD, highlightthickness=0, bd=0)
    right_scroll = ttk.Scrollbar(right_card, orient="vertical", command=right_view.yview)
    right_scroll.pack(side="right", fill="y", pady=(12, 6))
    right_view.pack(side="left", fill="both", expand=True)
    right_view.configure(yscrollcommand=right_scroll.set)
    right = ttk.Frame(right_view, style="Card.TFrame", padding=(18, 18, 12, 8))
    right_window = right_view.create_window((0, 0), window=right, anchor="nw")

    def sync_right_scroll(_event=None) -> None:
        right_view.configure(scrollregion=right_view.bbox("all"))

    def fit_right_width(event) -> None:
        right_view.itemconfigure(right_window, width=event.width)

    def scroll_right(event) -> None:
        if event.delta:
            right_view.yview_scroll(-1 if event.delta > 0 else 1, "units")

    def enable_right_scroll(_event) -> None:
        root.bind_all("<MouseWheel>", scroll_right)

    def disable_right_scroll(_event) -> None:
        root.unbind_all("<MouseWheel>")

    right.bind("<Configure>", sync_right_scroll)
    right_view.bind("<Configure>", fit_right_width)
    right_view.bind("<Enter>", enable_right_scroll)
    right_view.bind("<Leave>", disable_right_scroll)
    root.after_idle(sync_right_scroll)

    ttk.Label(left, text="Bitcell WAT", style="Section.TLabel").pack(anchor="w")
    ttk.Label(left, text="Import a 6T Excel set or enter Vt and Idsat beside each physical MOS.", style="Meta.TLabel").pack(anchor="w", pady=(2, 8))
    top_fields = ttk.Frame(left, style="Card.TFrame"); top_fields.pack(fill="x", pady=(0, 4))
    ttk.Label(top_fields, text="Lot / Wafer", style="Body.TLabel").pack(side="left")
    ttk.Entry(top_fields, textvariable=values["corner"], width=14, style="Apple.TEntry").pack(side="left", padx=10)
    tk.Label(top_fields, text="  6T INDEPENDENT  ", bg="#EEF6FF", fg=BLUE,
             font=("Calibri", 9, "bold"), padx=8, pady=5).pack(side="right")

    excel_row = ttk.Frame(left, style="Card.TFrame"); excel_row.pack(fill="x", pady=(2, 7))
    ttk.Label(excel_row, text="Excel 6T WAT", style="Body.TLabel").pack(side="left")
    excel_label = ttk.Label(excel_row, textvariable=values["excel_input"], style="Meta.TLabel")
    excel_label.pack(side="left", fill="x", expand=True, padx=(10, 6))

    def pick_excel() -> None:
        selected = filedialog.askopenfilename(
            title="Import 6T WAT Excel", filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if not selected:
            return
        try:
            samples = read_wat_excel(selected, float(numeric["nominal_vdd"].get()))
            first = samples[0]
            values["excel_input"].set(selected)
            values["corner"].set(first.lot_wafer)
            numeric["nominal_vdd"].set(f"{first.model_vdd_v:.3f}")
            for mos_name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2"):
                mos = getattr(first.cell, mos_name)
                wat_values[f"{mos_name}_vt"].set(f"{mos.vt:.6g}")
                wat_values[f"{mos_name}_ids"].set(f"{mos.ids:.6g}")
            status.set(f"Excel loaded: {len(samples)} model-VDD point(s); first point copied to the 6T inputs")
            status_label.configure(fg=GREEN)
        except Exception as exc:
            messagebox.showerror("Excel import", str(exc))

    def save_current_excel() -> None:
        try:
            mos6 = {
                name: MosWat(
                    _positive(wat_values[f"{name}_vt"].get(), f"{name}_vt"),
                    _positive(wat_values[f"{name}_ids"].get(), f"{name}_ids"),
                )
                for name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2")
            }
            wafer_id = values["corner"].get().strip() or "Manual_6T_WAT"
            default_name = re.sub(r'[^A-Za-z0-9._-]+', '_', wafer_id).strip('._') or "Manual_6T_WAT"
            selected = filedialog.asksaveasfilename(
                title="Save Current 6T WAT Excel",
                initialfile=f"{default_name}_6T_WAT.xlsx",
                defaultextension=".xlsx",
                filetypes=[("Excel workbook", "*.xlsx")],
            )
            if not selected:
                return
            cell = SixTWatCell(wafer_id, **mos6)
            saved_path = write_single_6t_wat_excel(
                selected, cell, float(numeric["nominal_vdd"].get()))
            values["excel_input"].set(str(saved_path))
            status.set(f"Current 6T WAT values saved: {saved_path.name}")
            status_label.configure(fg=GREEN)
        except Exception as exc:
            messagebox.showerror("Excel export", str(exc))

    def import_multi_chip_excel() -> None:
        selected = filedialog.askopenfilename(
            title="Import Wafer Multi-Cell 6T Excel",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if not selected:
            return
        try:
            chips = read_multi_chip_6t_excel(selected, float(numeric["nominal_vdd"].get()))
            assumptions = {key: (getattr(config_defaults, key) if not assumption_values[key].get().strip()
                                 else float(assumption_values[key].get()))
                           for key, _label, _unit in assumption_specs}
            cfg = Config(nominal_vdd=chips[0].model_vdd_v, wat_vdd=chips[0].model_vdd_v, **assumptions)
            analysis = analyze_multi_chip_wafer(chips, cfg)
            run_dir = create_run_output_dir(
                values["out"].get(), chips[0].lot_wafer,
                f'multi_cell_wafer_{chips[0].model_vdd_v:.3f}V')
            report = write_multi_chip_outputs(analysis, run_dir, selected)
            values["corner"].set(chips[0].lot_wafer)
            status.set(f"Wafer multi-cell complete: {len(chips)} cells; minimum RSNM={analysis['worst_rsnm']['rsnm_mv']:.1f} mV, WSNM={analysis['worst_wsnm']['wsnm_mv']:.1f} mV")
            status_label.configure(fg=GREEN)
            webbrowser.open(report.resolve().as_uri())
        except Exception as exc:
            messagebox.showerror("Wafer multi-cell import", str(exc))

    ttk.Button(excel_row, text="Save Current...", style="Quiet.TButton",
               command=save_current_excel).pack(side="right")
    ttk.Button(excel_row, text="Import Excel...", style="Quiet.TButton",
               command=pick_excel).pack(side="right", padx=(0, 6))
    ttk.Button(excel_row, text="Import Multi-Cell...", style="Quiet.TButton",
               command=import_multi_chip_excel).pack(side="right", padx=(0, 6))

    schematic = tk.Canvas(left, bg=CARD, highlightthickness=0, height=540)
    schematic.pack(fill="both", expand=True)
    # Standard 6T arrangement: WL above, BLB/BL at the sides, cross-coupled Q/QB.
    schematic.create_line(45, 48, 605, 48, fill=BORDER, width=2)
    schematic.create_text(325, 30, text="WL", fill=TEXT, font=("Calibri", 12, "bold"))
    schematic.create_line(68, 48, 68, 474, fill=BORDER, width=2)
    schematic.create_line(582, 48, 582, 474, fill=BORDER, width=2)
    schematic.create_text(68, 498, text="BLB", fill=SECONDARY, font=("Calibri", 10, "bold"))
    schematic.create_text(582, 498, text="BL", fill=SECONDARY, font=("Calibri", 10, "bold"))
    schematic.create_line(205, 105, 445, 105, fill=BORDER, width=2)
    schematic.create_text(325, 91, text="VDD", fill=SECONDARY, font=("Calibri", 9, "bold"))
    schematic.create_line(205, 442, 445, 442, fill=BORDER, width=2)
    schematic.create_text(325, 463, text="GND", fill=SECONDARY, font=("Calibri", 9, "bold"))
    schematic.create_line(265, 105, 265, 257, fill=BORDER, width=2)
    schematic.create_line(385, 105, 385, 257, fill=BORDER, width=2)
    schematic.create_line(265, 277, 265, 442, fill=BORDER, width=2)
    schematic.create_line(385, 277, 385, 442, fill=BORDER, width=2)
    schematic.create_line(68, 267, 202, 267, fill=GREEN, width=2)
    schematic.create_line(448, 267, 582, 267, fill=GREEN, width=2)
    schematic.create_line(265, 215, 385, 318, fill="#AF52DE", width=2, dash=(5, 4))
    schematic.create_line(385, 215, 265, 318, fill="#AF52DE", width=2, dash=(5, 4))
    schematic.create_oval(255, 257, 275, 277, fill=TEXT, outline="")
    schematic.create_oval(375, 257, 395, 277, fill=TEXT, outline="")
    schematic.create_text(265, 245, text="QB", fill=TEXT, font=("Calibri", 11, "bold"))
    schematic.create_text(385, 245, text="Q", fill=TEXT, font=("Calibri", 11, "bold"))

    def mos_panel(name: str, accent: str, source: dict[str, tk.StringVar]) -> ttk.Frame:
        panel = tk.Frame(schematic, bg="#FAFAFC", highlightbackground=BORDER, highlightthickness=1, padx=7, pady=6)
        display_name = DISPLAY_MOS_NAMES.get(name, name.upper())
        tk.Label(panel, text=display_name, bg="#FAFAFC", fg=accent,
                 font=("Calibri", 9, "bold")).grid(row=0, column=0, columnspan=4, sticky="w")
        tk.Label(panel, text="Vt", bg="#FAFAFC", fg=SECONDARY, font=("Calibri", 8)).grid(row=1, column=0)
        ttk.Entry(panel, textvariable=source[f"{name}_vt"], width=6, style="Apple.TEntry").grid(row=1, column=1, padx=(3, 6), pady=(4, 0))
        tk.Label(panel, text="Ids", bg="#FAFAFC", fg=SECONDARY, font=("Calibri", 8)).grid(row=1, column=2)
        ttk.Entry(panel, textvariable=source[f"{name}_ids"], width=6, style="Apple.TEntry").grid(row=1, column=3, padx=(3, 0), pady=(4, 0))
        return panel

    positions = {"pu1": (135, 120), "pu2": (350, 120), "pg1": (2, 230), "pg2": (475, 230),
                 "pd1": (135, 335), "pd2": (350, 335)}
    for name, (x, y) in positions.items():
        accent = RED if name.startswith("pu") else GREEN if name.startswith("pg") else BLUE
        schematic.create_window(x, y, anchor="nw", window=mos_panel(name, accent, wat_values))
    schematic.create_text(325, 526, text="Vt in V · Isat / Ids in µA", fill=SECONDARY, font=("Calibri", 8))

    ttk.Label(right, text="Analysis", style="Section.TLabel").pack(anchor="w")
    ttk.Label(right, text="Compare the entered Lot/Wafer WAT with WAT Target Read SNM curves.", style="Meta.TLabel").pack(anchor="w", pady=(2, 12))

    target_header = ttk.Frame(right, style="Card.TFrame")
    target_header.pack(fill="x")
    ttk.Label(target_header, text="WAT Target", style="Body.TLabel").pack(side="left")
    ttk.Checkbutton(target_header, text="Use as reference",
                    variable=use_wat_target_reference).pack(side="right")
    target_reference_note = tk.StringVar()
    footer_reference_text = tk.StringVar()
    output_scope_note = tk.StringVar()
    ttk.Label(right, textvariable=target_reference_note,
              style="Meta.TLabel", wraplength=350).pack(anchor="w", pady=(2, 5))
    target_grid = ttk.Frame(right, style="Card.TFrame"); target_grid.pack(fill="x", pady=(0, 12))
    ttk.Label(target_grid, text="Type", style="Meta.TLabel").grid(row=0, column=0, sticky="w")
    ttk.Label(target_grid, text="Vt (V)", style="Meta.TLabel").grid(row=0, column=1, sticky="w", padx=(8, 0))
    ttk.Label(target_grid, text="Isat (µA)", style="Meta.TLabel").grid(row=0, column=2, sticky="w", padx=(8, 0))
    target_entries: list[ttk.Entry] = []
    for row, (name, color) in enumerate((("PU", RED), ("PG", GREEN), ("PD", BLUE)), 1):
        tk.Label(target_grid, text=name, bg=CARD, fg=color,
                 font=("Calibri", 10, "bold")).grid(row=row, column=0, sticky="w", pady=3)
        vt_entry = ttk.Entry(target_grid, textvariable=target_values[f"{name.lower()}_vt"], width=9,
                             style="Apple.TEntry")
        ids_entry = ttk.Entry(target_grid, textvariable=target_values[f"{name.lower()}_ids"], width=10,
                              style="Apple.TEntry")
        vt_entry.grid(row=row, column=1, padx=(8, 0), pady=3)
        ids_entry.grid(row=row, column=2, padx=(8, 0), pady=3)
        target_entries.extend((vt_entry, ids_entry))

    def sync_target_reference_state(*_args) -> None:
        enabled = use_wat_target_reference.get()
        for entry in target_entries:
            entry.state(["!disabled"] if enabled else ["disabled"])
        target_reference_note.set(
            "Enabled: build a separate Target model and include comparison curves / deltas."
            if enabled else
            "Disabled: retain the entered values, but analyze and report Lot/Wafer only."
        )
        footer_reference_text.set(
            "Read SNM / W0-W1 WSNM · Lot/Wafer vs WAT Target" if enabled
            else "Read SNM / W0-W1 WSNM · Lot/Wafer only"
        )
        output_scope_note.set(
            "Output: Read SNM butterfly plus W0/W1 Write SNM comparisons against WAT Target."
            if enabled else
            "Output: Lot/Wafer Read SNM butterfly plus W0/W1 Write SNM; no Target comparison."
        )

    use_wat_target_reference.trace_add("write", sync_target_reference_state)
    sync_target_reference_state()

    ttk.Label(right, text="Model settings", style="Body.TLabel").pack(anchor="w", pady=(2, 0))
    config_grid = ttk.Frame(right, style="Card.TFrame"); config_grid.pack(fill="x")
    labels = [("nominal_vdd", "SRAM VDD", "V"), ("wat_vdd", "WAT VDD", "V")]
    for row, (key, label, unit) in enumerate(labels):
        ttk.Label(config_grid, text=label, style="Body.TLabel").grid(row=row, column=0, sticky="w", pady=5)
        ttk.Entry(config_grid, textvariable=numeric[key], width=10, style="Apple.TEntry").grid(row=row, column=1, sticky="e", padx=(12, 5))
        ttk.Label(config_grid, text=unit, style="Meta.TLabel").grid(row=row, column=2, sticky="w")
    config_grid.columnconfigure(0, weight=1)

    assumption_header = ttk.Frame(right, style="Card.TFrame")
    assumption_header.pack(fill="x", pady=(15, 0))
    ttk.Label(assumption_header, text="6T Cell Geometry Reference",
              style="Body.TLabel").pack(side="left")

    def restore_assumption_defaults() -> None:
        for key, _label, _unit in assumption_specs:
            assumption_values[key].set(str(getattr(config_defaults, key)))

    ttk.Button(assumption_header, text="Restore Defaults", style="Quiet.TButton",
               command=restore_assumption_defaults).pack(side="right")
    ttk.Label(right,
              text="Keep the known L and PU/PG/PD widths here. Blank fields use the generic 28 nm defaults; ratios are reported as references without double-counting measured Idsat.",
              style="Meta.TLabel", wraplength=350).pack(anchor="w", pady=(3, 5))
    assumption_grid = ttk.Frame(right, style="Card.TFrame")
    assumption_grid.pack(fill="x")
    for row, (key, label, unit) in enumerate(assumption_specs):
        ttk.Label(assumption_grid, text=label, style="Body.TLabel").grid(row=row, column=0, sticky="w", pady=4)
        ttk.Entry(assumption_grid, textvariable=assumption_values[key], width=10,
                  style="Apple.TEntry").grid(row=row, column=1, sticky="e", padx=(12, 5), pady=2)
        ttk.Label(assumption_grid, text=unit, style="Meta.TLabel").grid(row=row, column=2, sticky="w")
    assumption_grid.columnconfigure(0, weight=1)
    ttk.Label(right,
              textvariable=output_scope_note,
              style="Meta.TLabel", wraplength=350).pack(anchor="w", pady=(7, 0))

    ttk.Separator(right).pack(fill="x", pady=16)
    ttk.Label(right, text="Report destination", style="Body.TLabel").pack(anchor="w")
    out_row = ttk.Frame(right, style="Card.TFrame"); out_row.pack(fill="x", pady=(6, 14))
    ttk.Entry(out_row, textvariable=values["out"], style="Apple.TEntry").pack(side="left", fill="x", expand=True)
    def pick_out() -> None:
        selected = filedialog.askdirectory()
        if selected: values["out"].set(selected)
    ttk.Button(out_row, text="Choose…", style="Quiet.TButton", command=pick_out).pack(side="left", padx=(7, 0))

    def open_out() -> None:
        try:
            opened = open_output_directory(values["out"].get())
            status.set(f"Opened output folder: {opened}")
            status_label.configure(fg=SECONDARY)
        except Exception as exc:
            messagebox.showerror("Output folder", str(exc))

    ttk.Button(out_row, text="Open Folder", style="Quiet.TButton", command=open_out).pack(side="left", padx=(7, 0))
    ttk.Label(right,
              text="Each run is archived as YYYY-MM-DD / WaferID / HHMMSS_analysis.",
              style="Meta.TLabel", wraplength=350).pack(anchor="w", pady=(0, 10))

    status = tk.StringVar(value="Ready to analyze")
    status_label = tk.Label(right, textvariable=status, bg=CARD, fg=SECONDARY,
                            font=("Calibri", 9), anchor="w", justify="left", wraplength=330)
    status_label.pack(fill="x", pady=(0, 7))
    progress = ttk.Progressbar(right, mode="indeterminate", style="Apple.Horizontal.TProgressbar")
    progress.pack(fill="x", pady=(0, 12))
    result_queue: queue.Queue = queue.Queue()

    def collect_inputs() -> tuple[SixTWatCell, Config, DatasheetTargets | None]:
        resolved_assumptions: dict[str, float] = {}
        for key, _label, _unit in assumption_specs:
            raw = assumption_values[key].get().strip()
            resolved_assumptions[key] = getattr(config_defaults, key) if not raw else float(raw)
        cfg = Config(
            nominal_vdd=float(numeric["nominal_vdd"].get()),
            wat_vdd=float(numeric["wat_vdd"].get()),
            **resolved_assumptions,
        )
        validate_config(cfg)
        corner = values["corner"].get().strip() or "Manual"
        mos6 = {name: MosWat(_positive(wat_values[f"{name}_vt"].get(), f"{name}_vt"),
                             _positive(wat_values[f"{name}_ids"].get(), f"{name}_ids"))
                for name in ("pu1", "pu2", "pg1", "pg2", "pd1", "pd2")}
        cell = SixTWatCell(corner, **mos6)
        targets = None
        if use_wat_target_reference.get():
            targets = DatasheetTargets(**{
                name: MosWat(_positive(target_values[f"{name}_vt"].get(), f"{name} target Vt"),
                             _positive(target_values[f"{name}_ids"].get(), f"{name} target Isat"))
                for name in ("pu", "pg", "pd")
            })
        return cell, cfg, targets

    def worker(cell: SixTWatCell, cfg: Config,
               targets: DatasheetTargets | None, out_path: str) -> None:
        try:
            result = analyze_six_mos(cell, cfg, targets)
            run_dir = create_run_output_dir(out_path, cell.corner, "6t_analysis")
            report = write_outputs(result, run_dir)
            result_queue.put((True, cell, report))
        except Exception as exc:
            result_queue.put((False, None, exc))

    def excel_worker(excel_path: str, cfg: Config, targets: DatasheetTargets | None, out_path: str) -> None:
        try:
            report = analyze_excel_wat_sweep(
                excel_path, out_path, cfg, targets, archive_run=True)
            result_queue.put((True, None, report))
        except Exception as exc:
            result_queue.put((False, None, exc))

    def poll_result() -> None:
        try: ok, cell, payload = result_queue.get_nowait()
        except queue.Empty:
            root.after(80, poll_result); return
        progress.stop(); analyze_button.state(["!disabled"]); excel_analyze_button.state(["!disabled"])
        if ok:
            label = "Excel model-VDD sweep" if cell is None else cell.corner
            status.set(f"Complete - {label}; saved to {Path(payload).parent}")
            status_label.configure(fg=GREEN)
            webbrowser.open(Path(payload).resolve().as_uri())
        else:
            status.set("Analysis could not be completed")
            status_label.configure(fg=RED)
            messagebox.showerror("Analysis error", str(payload))

    def execute() -> None:
        try: cell, cfg, targets = collect_inputs()
        except Exception as exc:
            status.set("Check the highlighted input values")
            status_label.configure(fg=RED)
            messagebox.showerror("Invalid input", str(exc)); return
        status.set("Analyzing 6T Independent · Read SNM and W0/W1 Write SNM…")
        status_label.configure(fg=BLUE)
        analyze_button.state(["disabled"]); progress.start(10)
        threading.Thread(target=worker, args=(cell, cfg, targets, values["out"].get()), daemon=True).start()
        root.after(80, poll_result)

    def execute_excel_sweep() -> None:
        excel_path = values["excel_input"].get().strip()
        if not excel_path:
            messagebox.showerror("Excel import", "Choose a 6T WAT Excel workbook first.")
            return
        try:
            _cell, cfg, targets = collect_inputs()
        except Exception as exc:
            status.set("Check the input values")
            status_label.configure(fg=RED)
            messagebox.showerror("Invalid input", str(exc)); return
        status.set("Analyzing Excel model-VDD sweep…")
        status_label.configure(fg=BLUE)
        analyze_button.state(["disabled"]); excel_analyze_button.state(["disabled"]); progress.start(10)
        threading.Thread(target=excel_worker, args=(excel_path, cfg, targets, values["out"].get()), daemon=True).start()
        root.after(80, poll_result)

    ttk.Label(right_footer, textvariable=footer_reference_text,
              style="Meta.TLabel").pack(pady=(0, 7))
    analyze_button = ttk.Button(right_footer, text="Analyze & Open HTML", style="Accent.TButton", command=execute)
    analyze_button.pack(fill="x")
    excel_analyze_button = ttk.Button(right_footer, text="Analyze Excel VDD Sweep", style="Quiet.TButton", command=execute_excel_sweep)
    excel_analyze_button.pack(fill="x", pady=(7, 0))

    # Estimate Vmin tab: consumes conservative results exported by the
    # Multi-Cell 6T analysis rather than a second manual/IV input path.
    curve_tab.columnconfigure(0, weight=4)
    curve_tab.columnconfigure(1, weight=9)
    curve_tab.rowconfigure(0, weight=1)
    curve_input_card = ttk.Frame(curve_tab, style="Card.TFrame", padding=18)
    curve_input_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    curve_chart_card = ttk.Frame(curve_tab, style="Card.TFrame", padding=18)
    curve_chart_card.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
    curve_chart_card.columnconfigure(0, weight=1)
    curve_chart_card.rowconfigure(3, weight=1)

    ttk.Label(curve_input_card, text="Multi-Cell Estimate Vmin", style="Section.TLabel").pack(anchor="w")
    ttk.Label(curve_input_card,
              text="Select generated multi_chip_snm_summary.csv files or raw 6T Multi-Cell Excel workbooks (.xlsx/.xlsm). For one-step multi-VDD analysis, name Excel sheets by Model VDD (for example 0.90V, 0.80V). Each sheet receives its own output folder before all VDDs are combined into W/R Estimate Vmin curves.",
              style="Meta.TLabel", wraplength=560).pack(anchor="w", pady=(2, 12))
    selected_summary_paths: list[str] = []
    curve_run_shmoo = tk.BooleanVar(
        value=bool(saved_state.get("options", {}).get(
            "estimate_vmin_run_shmoo", False)))
    selected_summary_text = tk.StringVar(value="No Multi-Cell summary selected")
    tk.Label(curve_input_card, textvariable=selected_summary_text, bg=CARD, fg=SECONDARY,
             font=("Calibri", 10), anchor="w", justify="left", wraplength=560).pack(fill="x", pady=(0, 10))
    saved_summary_paths = saved_state.get("estimate_vmin_summary_paths", [])
    if isinstance(saved_summary_paths, list):
        selected_summary_paths.extend(
            str(item) for item in saved_summary_paths if Path(str(item)).is_file())
    if selected_summary_paths:
        selected_summary_text.set(
            f"{len(selected_summary_paths)} previously selected summary file(s)\n" +
            "\n".join(Path(item).name for item in selected_summary_paths[:5]))
    ttk.Checkbutton(
        curve_input_card,
        text="Run Shmoo analysis (slower)",
        variable=curve_run_shmoo).pack(anchor="w", pady=(0, 4))
    ttk.Label(
        curve_input_card,
        text="Off: calculate SNM / BL Write Margin curves only. On: also run per-VDD CR/PR Shmoo and Drive Advisor.",
        style="Meta.TLabel", wraplength=560).pack(anchor="w", pady=(0, 10))
    comparison_summary_paths: list[str] = []
    saved_comparison_paths = saved_state.get("estimate_vmin_comparison_paths", [])
    if isinstance(saved_comparison_paths, list):
        comparison_summary_paths.extend(
            str(item) for item in saved_comparison_paths
            if Path(str(item)).is_file() or Path(str(item)).is_dir())
    comparison_file_text = tk.StringVar(
        value="Comparison View: no Multi-VDD source selected")
    if comparison_summary_paths:
        comparison_file_text.set(
            f"Comparison View: {len(comparison_summary_paths)} source(s)\n" +
            "\n".join(Path(item).name for item in comparison_summary_paths[:5]))
    tk.Label(curve_input_card, textvariable=comparison_file_text, bg=CARD, fg=SECONDARY,
             font=("Calibri", 9), anchor="w", justify="left", wraplength=560).pack(
                 fill="x", pady=(0, 8))

    curve_status = tk.StringVar(value="Ready to analyze the VDD sweep")
    curve_status_label = tk.Label(curve_input_card, textvariable=curve_status, bg=CARD, fg=SECONDARY,
                                  font=("Calibri", 9), anchor="w", justify="left", wraplength=560)
    curve_status_label.pack(fill="x", pady=(3, 6))
    curve_progress = ttk.Progressbar(curve_input_card, mode="indeterminate",
                                     style="Apple.Horizontal.TProgressbar")
    curve_progress.pack(fill="x", pady=(0, 9))

    curve_kind = tk.StringVar(value="rsnm_mv")
    curve_title = tk.StringVar(value="Estimate Vmin Curve")
    ttk.Label(curve_chart_card, textvariable=curve_title, style="ChartTitle.TLabel").grid(row=0, column=0, sticky="w")
    curve_switch = ttk.Frame(curve_chart_card, style="Card.TFrame")
    curve_switch.grid(row=1, column=0, sticky="w", pady=(4, 6))
    def select_curve_kind() -> None:
        draw_curve_chart()
        if curve_result:
            if curve_result.get("mode") == "shmoo_only":
                curve_summary.set("Single-VDD input: Shmoo and Advisor outputs only.")
                return
            if curve_result.get("mode") == "single_vdd":
                curve_summary.set(
                    "Single-VDD fast mode: summary data exported; enable Shmoo "
                    "to generate drive-balance analysis.")
                return
            if curve_kind.get() == "stacked":
                curve_summary.set("Comparison view: Read/Write SNM and BL Write Margin versus Model VDD.")
                return
            curve = curve_result["curves"][curve_kind.get()]
            closure = curve.get("eye_closure")
            estimate_kind = "extrapolated" if closure and closure.get("extrapolated") else "estimated"
            curve_summary.set(
                f'{curve["label"]} {estimate_kind} eye-closure VDD: {closure["estimated_vdd_v"]:.4f} V'
                if closure else f'{curve["label"]} eye-closure VDD not bracketed by imported points')
    for key, short_label, _label, _color in _ESTIMATE_VMIN_METRICS:
        ttk.Radiobutton(curve_switch, text=short_label, value=key, variable=curve_kind,
                        command=select_curve_kind).pack(side="left", padx=(0, 12))
    ttk.Radiobutton(curve_switch, text="Stacked", value="stacked", variable=curve_kind,
                    command=select_curve_kind).pack(side="left", padx=(0, 12))
    curve_summary = tk.StringVar(value="Import one VDD for Shmoo analysis, or two or more VDD points for Estimate Vmin curves.")
    curve_summary_label = tk.Label(curve_chart_card, textvariable=curve_summary, bg=CARD, fg=SECONDARY,
                                   font=("Calibri", 10, "bold"), anchor="w", justify="left")
    curve_summary_label.grid(row=2, column=0, sticky="ew", pady=(0, 6))
    curve_canvas = tk.Canvas(curve_chart_card, bg=CARD, highlightthickness=0, bd=0,
                             width=760, height=620)
    curve_canvas.grid(row=3, column=0, sticky="nsew")
    curve_result: dict | None = None
    curve_report_path: Path | None = None

    def draw_curve_chart(_event=None) -> None:
        curve_canvas.delete("all")
        width = max(curve_canvas.winfo_width(), 520)
        height = max(curve_canvas.winfo_height(), 420)
        left_margin, right_margin, top_margin, bottom_margin = 78, 24, 42, 142
        plot_width = width - left_margin - right_margin
        plot_height = height - top_margin - bottom_margin
        if not curve_result:
            curve_canvas.create_text(width / 2, height / 2, text="Estimate Vmin curve will appear here",
                                     fill=SECONDARY, font=("Calibri", 13))
            return
        if curve_result.get("mode") in {"shmoo_only", "single_vdd"}:
            shmoo_enabled = curve_result.get("mode") == "shmoo_only"
            curve_title.set("Multi-Cell Shmoo Analysis" if shmoo_enabled
                            else "Multi-Cell Fast Analysis")
            vdds = [float(row["vdd_v"]) for row in curve_result["rows"]]
            vdd_text = (f"{vdds[0]:.3f} V" if len(vdds) == 1 else
                        f"{vdds[0]:.3f}–{vdds[-1]:.3f} V ({len(vdds)} points)")
            curve_canvas.create_text(
                width / 2, height / 2 - 18,
                text=(f"Model VDD {vdd_text} · " +
                      ("Shmoo output complete" if shmoo_enabled
                       else "fast summary output complete")),
                fill=TEXT, font=("Calibri", 16, "bold"))
            curve_canvas.create_text(
                width / 2, height / 2 + 18,
                text=("Open HTML Result to inspect the CR/PR Shmoo and Drive-to-Preferred Advisor."
                      if shmoo_enabled else
                      "Enable Run Shmoo analysis when CR/PR grading is required."),
                fill=SECONDARY, font=("Calibri", 11))
            return
        if curve_kind.get() == "stacked":
            curve_title.set("Estimate Vmin Curves - Comparison View")
            groups = (
                ("Read / Write SNM", ("rsnm_mv", "wsnm_mv"), "SNM (mV)"),
                ("BL Write Margin", ("write_margin_mv",), "Vtrip (mV)"),
            )
            panel_gap = 74
            panel_height = max(112, (height - 64 - panel_gap) / 2)
            for panel, (group_label, keys, y_axis_label) in enumerate(groups):
                panel_top = 38 + panel * (panel_height + panel_gap)
                panel_bottom = panel_top + panel_height - 28
                curves = [curve_result["curves"][key] for key in keys]
                maximum = max((row["margin_mv"] for curve in curves for row in curve["rows"]), default=50.0)
                y_max = max(50.0, math.ceil(maximum / 50.0) * 50.0)

                def stacked_xy(vdd: float, margin: float) -> tuple[float, float]:
                    return (left_margin + vdd / SNM_PLOT_AXIS_MAX_V * plot_width,
                            panel_top + (1 - margin / y_max) * (panel_bottom - panel_top))

                curve_canvas.create_text(left_margin, panel_top - 20, text=group_label, anchor="w",
                                         fill=TEXT, font=("Calibri", 11, "bold"))
                curve_canvas.create_text(15, (panel_top + panel_bottom) / 2,
                                         text=y_axis_label, angle=90, fill=TEXT,
                                         font=("Calibri", 9, "bold"))
                legend_x = left_margin + 190
                for curve in curves:
                    curve_canvas.create_line(legend_x, panel_top - 21, legend_x + 16, panel_top - 21,
                                             fill=curve["color"], width=3)
                    curve_canvas.create_text(legend_x + 21, panel_top - 21, text=curve["label"], anchor="w",
                                             fill=SECONDARY, font=("Calibri", 9, "bold"))
                    legend_x += 140
                for step in range(5):
                    margin = y_max * step / 4
                    _x, y = stacked_xy(0, margin)
                    curve_canvas.create_line(left_margin, y, left_margin + plot_width, y, fill="#E5E5EA")
                    curve_canvas.create_text(left_margin - 8, y, text=f"{margin:.0f}", anchor="e",
                                             fill=SECONDARY, font=("Calibri", 8))
                for vdd_step in range(7):
                    x, _y = stacked_xy(vdd_step * .2, 0)
                    curve_canvas.create_line(x, panel_top, x, panel_bottom, fill="#F1F1F4")
                    curve_canvas.create_text(x, panel_bottom + 13, text=f"{vdd_step*.2:.1f}",
                                             fill=SECONDARY, font=("Calibri", 8))
                for curve_index, curve in enumerate(curves):
                    points = [stacked_xy(row["vdd_v"], row["margin_mv"]) for row in curve["rows"]]
                    if len(points) >= 2:
                        curve_canvas.create_line(
                            *[coordinate for point in points for coordinate in point],
                            fill=curve["color"], width=2)
                    for point_index, (_row, (x, y)) in enumerate(zip(curve["rows"], points)):
                        curve_canvas.create_oval(x-3, y-3, x+3, y+3,
                                                 fill=CARD, outline=curve["color"], width=2)
                        label_y = y - 11 if curve_index == 0 else y + 13
                        if label_y < panel_top + 9:
                            label_y = y + 13
                        elif label_y > panel_bottom - 4:
                            label_y = y - 11
                        label_x = x + (-3 if point_index % 2 else 3)
                        curve_canvas.create_text(
                            label_x, label_y, text=f'{_row["margin_mv"]:.1f} mV',
                            fill=curve["color"], font=("Calibri", 8, "bold"))
                if "rsnm_mv" in keys:
                    rows = curve_result["curves"]["rsnm_mv"]["rows"]
                    if len(rows) >= 2:
                        _low, marker = max(zip(rows, rows[1:]), key=lambda pair: abs((pair[1]["margin_mv"] - pair[0]["margin_mv"]) / (pair[1]["vdd_v"] - pair[0]["vdd_v"])))
                        x, _ = stacked_xy(marker["vdd_v"], marker["margin_mv"])
                        curve_canvas.create_line(x, panel_top, x, panel_bottom, fill="#FF385C", dash=(4, 4))
                        curve_canvas.create_text(x + 5, panel_top + 8, text=f"Largest slope\n{marker['vdd_v']:.2f} V", anchor="nw", fill="#C13515", font=("Calibri", 8, "bold"))
                        curve_canvas.create_rectangle(
                            x - 25, panel_bottom + 22, x + 25, panel_bottom + 40,
                            fill="#FFF1F3", outline="#FF385C")
                        curve_canvas.create_text(
                            x, panel_bottom + 31, text=f'{marker["vdd_v"]:.2f} V',
                            fill="#C13515", font=("Calibri", 8, "bold"))
            curve_canvas.create_text(left_margin + plot_width / 2, height - 8, text="Model VDD (V)", fill=TEXT, font=("Calibri", 11, "bold"))
            return
        curve = curve_result["curves"][curve_kind.get()]
        rows = curve["rows"]
        max_rsnm = max((row["margin_mv"] for row in rows), default=50.0)
        y_max = max(50.0, math.ceil(max_rsnm / 50.0) * 50.0)
        color = curve["color"]
        curve_title.set(f'Estimated {curve["label"]} Curve')

        def xy(vcc_v: float, rsnm_mv: float) -> tuple[float, float]:
            return (left_margin + vcc_v / SNM_PLOT_AXIS_MAX_V * plot_width,
                    top_margin + (1.0 - rsnm_mv / y_max) * plot_height)

        for step in range(7):
            value = step * .2
            x, _ = xy(value, 0)
            curve_canvas.create_line(x, top_margin, x, top_margin + plot_height, fill="#E5E5EA")
            curve_canvas.create_text(x, top_margin + plot_height + 20, text=f"{value:.1f}",
                                     fill=SECONDARY, font=("Calibri", 9))
        for step in range(6):
            value = y_max * step / 5
            _, y = xy(0, value)
            curve_canvas.create_line(left_margin, y, left_margin + plot_width, y, fill="#E5E5EA")
            curve_canvas.create_text(left_margin - 10, y, text=f"{value:.0f}", anchor="e",
                                     fill=SECONDARY, font=("Calibri", 9))
        curve_canvas.create_line(left_margin, top_margin, left_margin,
                                 top_margin + plot_height, fill=TEXT, width=2)
        curve_canvas.create_line(left_margin, top_margin + plot_height,
                                 left_margin + plot_width, top_margin + plot_height, fill=TEXT, width=2)

        display_points: list[tuple[float, float]] = []
        closure = curve.get("eye_closure")
        if closure:
            display_points.append(xy(closure["estimated_vdd_v"], 0.0))
        display_points.extend(xy(row["vdd_v"], row["margin_mv"]) for row in rows)
        baseline_y = top_margin + plot_height
        voltage_labels = [(xy(row["vdd_v"], 0.0)[0], f'{row["vdd_v"]:.2f} V') for row in rows]
        voltage_label_rows = _stagger_label_rows(
            voltage_labels, character_width=7.2, minimum_gap=7.0)
        voltage_label_y = [baseline_y + 44 + label_row * 18
                           for label_row in voltage_label_rows]
        for row, label_y in zip(rows, voltage_label_y):
            guide_x, guide_y = xy(row["vdd_v"], row["margin_mv"])
            curve_canvas.create_line(guide_x, guide_y + 5, guide_x, label_y - 13,
                                     fill="#B9D7FF", width=1, dash=(3, 4))
            curve_canvas.create_text(
                guide_x, label_y, text=f'{row["vdd_v"]:.2f} V',
                fill="#0062CC", font=("Calibri", 11, "bold"))
        if len(display_points) >= 2:
            curve_canvas.create_line(*[coordinate for point in display_points for coordinate in point],
                                     fill=color, width=3, smooth=False)

        curve_segment_boxes = [
            (int(min(first[0], second[0]) - 5), int(min(first[1], second[1]) - 5),
             int(max(first[0], second[0]) + 5), int(max(first[1], second[1]) + 5))
            for first, second in zip(display_points, display_points[1:])
        ]

        placed_label_boxes: list[tuple[int, int, int, int]] = []

        def boxes_overlap(first: tuple[int, int, int, int],
                          second: tuple[int, int, int, int], padding: int = 5) -> bool:
            return not (first[2] + padding < second[0] or
                        first[0] - padding > second[2] or
                        first[3] + padding < second[1] or
                        first[1] - padding > second[3])

        label_candidates = (
            (-15, -13, "se"), (15, -13, "sw"),
            (-15, 13, "ne"), (15, 13, "nw"),
            (0, -22, "s"), (0, 22, "n"),
            (-22, 0, "e"), (22, 0, "w"),
        )
        for index, row in enumerate(rows):
            x, y = xy(row["vdd_v"], row["margin_mv"])
            curve_canvas.create_oval(x - 4, y - 4, x + 4, y + 4, fill=CARD, outline=color, width=2)
            if row["margin_mv"] <= 0:
                continue
            ordered_candidates = label_candidates[index % 4:] + label_candidates[:index % 4]
            chosen_item = None
            for label_dx, label_dy, label_anchor in ordered_candidates:
                label_item = curve_canvas.create_text(
                    x + label_dx, y + label_dy,
                    text=f'{row["margin_mv"]:.1f} mV', fill=TEXT,
                    anchor=label_anchor, font=("Calibri", 11, "bold"))
                bbox = curve_canvas.bbox(label_item)
                within_plot = bool(bbox and bbox[0] >= left_margin + 3 and
                                   bbox[2] <= left_margin + plot_width - 3 and
                                   bbox[1] >= top_margin + 3 and
                                   bbox[3] <= baseline_y - 3)
                collision = bool(bbox and any(boxes_overlap(bbox, used)
                                              for used in placed_label_boxes))
                curve_collision = bool(bbox and any(
                    boxes_overlap(bbox, segment, padding=1)
                    for segment in curve_segment_boxes))
                if bbox and within_plot and not collision and not curve_collision:
                    placed_label_boxes.append(bbox)
                    chosen_item = label_item
                    label_background = curve_canvas.create_rectangle(
                        bbox[0] - 2, bbox[1] - 1, bbox[2] + 2, bbox[3] + 1,
                        fill=CARD, outline="")
                    curve_canvas.tag_lower(label_background, label_item)
                    break
                curve_canvas.delete(label_item)
            if chosen_item is None:
                # Close VDD sweeps can put several values in the same few
                # pixels.  Prefer a clearly separated label over a label that
                # overlaps another value or hides the curve.
                fallback_candidates = (
                    (0, -38, "s"), (0, 38, "n"),
                    (-38, -26, "se"), (38, -26, "sw"),
                    (-38, 26, "ne"), (38, 26, "nw"),
                    (0, -58, "s"), (0, 58, "n"),
                )
                for label_dx, label_dy, label_anchor in fallback_candidates:
                    label_item = curve_canvas.create_text(
                        x + label_dx, y + label_dy,
                        text=f'{row["margin_mv"]:.1f} mV', fill=TEXT,
                        anchor=label_anchor, font=("Calibri", 11, "bold"))
                    fallback_bbox = curve_canvas.bbox(label_item)
                    within_plot = bool(fallback_bbox and fallback_bbox[0] >= left_margin + 3 and
                                       fallback_bbox[2] <= left_margin + plot_width - 3 and
                                       fallback_bbox[1] >= top_margin + 3 and
                                       fallback_bbox[3] <= baseline_y - 3)
                    collision = bool(fallback_bbox and any(
                        boxes_overlap(fallback_bbox, used) for used in placed_label_boxes))
                    if fallback_bbox and within_plot and not collision:
                        placed_label_boxes.append(fallback_bbox)
                        label_background = curve_canvas.create_rectangle(
                            fallback_bbox[0] - 2, fallback_bbox[1] - 1,
                            fallback_bbox[2] + 2, fallback_bbox[3] + 1,
                            fill=CARD, outline="")
                        curve_canvas.tag_lower(label_background, label_item)
                        chosen_item = label_item
                        break
                    curve_canvas.delete(label_item)
        for row in rows:
            if row["valid"]:
                continue
            x, y = xy(row["vdd_v"], 0.0)
            curve_canvas.create_line(x - 4, y - 4, x + 4, y + 4, fill=SECONDARY, width=2)
            curve_canvas.create_line(x + 4, y - 4, x - 4, y + 4, fill=SECONDARY, width=2)
        # Highlight the operating-voltage interval where RSNM changes most
        # rapidly.  The right-hand point marks the requested VDD node.
        if curve["key"] == "rsnm_mv" and len(rows) >= 2:
            _left_row, slope_row = max(
                zip(rows, rows[1:]),
                key=lambda pair: abs((pair[1]["margin_mv"] - pair[0]["margin_mv"]) /
                                     (pair[1]["vdd_v"] - pair[0]["vdd_v"])))
            marker_x, _ = xy(slope_row["vdd_v"], slope_row["margin_mv"])
            curve_canvas.create_line(marker_x, top_margin, marker_x, baseline_y,
                                     fill="#FF385C", width=2, dash=(5, 5))
            curve_canvas.create_text(marker_x + 7, top_margin + 12,
                                     text=f"Largest RSNM slope\n{slope_row['vdd_v']:.2f} V",
                                     anchor="nw", fill="#C13515",
                                     font=("Calibri", 10, "bold"), justify="left")
        if closure:
            x, y = xy(closure["estimated_vdd_v"], 0.0)
            curve_canvas.create_line(x, top_margin, x, top_margin + plot_height,
                                     fill="#FF9500", width=2, dash=(6, 4))
            estimate_kind = "Extrapolated" if closure.get("extrapolated") else "Estimated"
            curve_canvas.create_text(x + 8, top_margin + 12,
                                     text=f'{estimate_kind} eye-closure VDD {closure["estimated_vdd_v"]:.4f} V',
                                     anchor="w", fill="#C56A00", font=("Calibri", 12, "bold"))
        curve_canvas.create_text(left_margin + plot_width / 2, height - 24, text="Model VDD (V)",
                                 fill=TEXT, font=("Calibri", 12, "bold"))
        curve_canvas.create_text(18, top_margin + plot_height / 2, text=f'{curve["label"]} (mV)', angle=90,
                                 fill=TEXT, font=("Calibri", 12, "bold"))

    curve_canvas.bind("<Configure>", draw_curve_chart)
    draw_curve_chart()

    curve_result_queue: queue.Queue = queue.Queue()
    comparison_result_queue: queue.Queue = queue.Queue()

    def import_multi_cell_summaries() -> None:
        selected = filedialog.askopenfilenames(
            title="Import Multi-Cell CSV or Excel",
            initialdir=values["out"].get(),
            filetypes=[("Multi-Cell CSV / Excel", "*.csv *.xlsx *.xlsm"),
                       ("CSV summary", "*.csv"),
                       ("Excel workbook", "*.xlsx *.xlsm"),
                       ("All files", "*.*")])
        if not selected: return
        selected_summary_paths[:] = list(selected)
        selected_summary_text.set(f"{len(selected_summary_paths)} Multi-Cell input file(s) selected\n" + "\n".join(Path(item).name for item in selected_summary_paths[:5]))
        curve_status.set("Ready to analyze imported Multi-Cell CSV / Excel data")
        curve_status_label.configure(fg=SECONDARY)

    def curve_worker(summary_paths: list[str], out_path: Path, wafer_id: str,
                     model_config: Config, include_shmoo: bool) -> None:
        try:
            if len(summary_paths) == 1 and Path(summary_paths[0]).suffix.lower() in {".xlsx", ".xlsm"}:
                vdd_groups = read_multi_chip_6t_excel_vdd_sheets(
                    summary_paths[0], model_config.nominal_vdd,
                    allow_no_vdd_sheets=True)
                if vdd_groups:
                    result = process_multi_vdd_6t_excel(
                        summary_paths[0], model_config, out_path, vdd_groups,
                        include_shmoo=include_shmoo)
                    curve_result_queue.put(
                        (True, result["analysis"], result["report"]))
                    return
            source_rows = read_multi_chip_snm_summary(
                summary_paths, model_config.nominal_vdd, model_config,
                include_shmoo=include_shmoo)
            analysis = analyze_estimate_vmin_curves(
                source_rows,
                force_shmoo_only=len(summary_paths) == 1 and include_shmoo,
                include_shmoo=include_shmoo)
            run_dir = create_run_output_dir(out_path, wafer_id, "estimate_vmin_curve")
            report = write_estimate_vmin_outputs(analysis, run_dir, summary_paths)
            curve_result_queue.put((True, analysis, report))
        except Exception as exc:
            curve_result_queue.put((False, None, exc))

    def comparison_worker(summary_files: list[str], out_path: Path) -> None:
        try:
            datasets = read_estimate_vmin_combined_files(summary_files)
            run_dir = create_run_output_dir(
                out_path, "Combined_Summaries", "estimate_vmin_curve_comparison")
            report = write_estimate_vmin_combined_comparison_outputs(datasets, run_dir)
            comparison_result_queue.put((True, datasets, report))
        except Exception as exc:
            comparison_result_queue.put((False, None, exc))

    def import_comparison_combined_files() -> None:
        selected = filedialog.askopenfilenames(
            title="Import multi_chip_snm_summary_combined.csv files",
            initialdir=values["out"].get(),
            filetypes=[("Estimate Vmin combined summary", "*.csv"),
                       ("All files", "*.*")])
        if not selected:
            return
        comparison_summary_paths[:] = [str(Path(item).resolve()) for item in selected]
        comparison_file_text.set(
            f"Comparison View: {len(comparison_summary_paths)} source(s)\n" +
            "\n".join(Path(item).name for item in comparison_summary_paths[:5]))
        curve_status.set("Ready to compare selected combined summary files")
        curve_status_label.configure(fg=SECONDARY)

    def import_comparison_multi_vdd_folder() -> None:
        selected = filedialog.askdirectory(
            title="Import Multi-VDD output folder",
            initialdir=values["out"].get())
        if not selected:
            return
        resolved = str(Path(selected).resolve())
        if resolved not in comparison_summary_paths:
            comparison_summary_paths.append(resolved)
        comparison_file_text.set(
            f"Comparison View: {len(comparison_summary_paths)} source(s)\n" +
            "\n".join(Path(item).name for item in comparison_summary_paths[:5]))
        curve_status.set(
            "Ready to combine SNM and BL Write Margin from the Multi-VDD folder")
        curve_status_label.configure(fg=SECONDARY)

    def clear_comparison_summary_files() -> None:
        comparison_summary_paths.clear()
        comparison_file_text.set("Comparison View: no Multi-VDD source selected")

    def poll_comparison_result() -> None:
        try:
            ok, datasets, payload = comparison_result_queue.get_nowait()
        except queue.Empty:
            root.after(80, poll_comparison_result)
            return
        curve_progress.stop()
        curve_compare_button.state(["!disabled"])
        if ok:
            curve_status.set(
                f"Comparison View complete: {len(datasets)} combined summaries")
            curve_status_label.configure(fg=GREEN)
            webbrowser.open(Path(payload).resolve().as_uri())
        else:
            curve_status.set("Combined-summary comparison could not be completed")
            curve_status_label.configure(fg=RED)
            messagebox.showerror("Multi Curve Comparison", str(payload))

    def execute_combined_file_comparison() -> None:
        if not comparison_summary_paths:
            curve_status.set("Select a Multi-VDD output folder or combined summary CSV")
            curve_status_label.configure(fg=RED)
            messagebox.showerror(
                "Multi Curve Comparison",
                "Select at least one Multi-VDD output folder or "
                "multi_chip_snm_summary_combined.csv file.")
            return
        curve_status.set("Reading combined summaries and drawing Comparison View...")
        curve_status_label.configure(fg=BLUE)
        curve_compare_button.state(["disabled"])
        curve_progress.start(10)
        threading.Thread(
            target=comparison_worker,
            args=(list(comparison_summary_paths), Path(values["out"].get())),
            daemon=True).start()
        root.after(80, poll_comparison_result)

    def poll_curve_result() -> None:
        nonlocal curve_result, curve_report_path
        try:
            ok, analysis, payload = curve_result_queue.get_nowait()
        except queue.Empty:
            root.after(80, poll_curve_result)
            return
        curve_progress.stop()
        curve_analyze_button.state(["!disabled"])
        if ok:
            curve_result = analysis
            curve_report_path = Path(payload)
            if analysis.get("mode") == "shmoo_only":
                summary = (f'Shmoo-only analysis complete for {len(analysis["rows"])} '
                           'VDD point(s); Estimate Vmin curves were not generated.')
                curve_summary_label.configure(fg=SECONDARY)
            elif analysis.get("mode") == "single_vdd":
                summary = ("Single-VDD fast analysis complete; summary exported "
                           "without Shmoo analysis or an Estimate Vmin trend.")
                curve_summary_label.configure(fg=SECONDARY)
            elif curve_kind.get() == "stacked":
                summary = "Comparison view: Read/Write SNM and BL Write Margin versus Model VDD."
                curve_summary_label.configure(fg=SECONDARY)
            else:
                curve = analysis["curves"][curve_kind.get()]
                closure = curve.get("eye_closure")
                if closure:
                    estimate_kind = "extrapolated" if closure.get("extrapolated") else "estimated"
                    summary = f'{curve["label"]} {estimate_kind} eye-closure VDD: {closure["estimated_vdd_v"]:.4f} V'
                    curve_summary_label.configure(fg="#C56A00")
                else:
                    summary = f'{curve["label"]} eye-closure VDD not bracketed by imported points'
                    curve_summary_label.configure(fg=SECONDARY)
            curve_summary.set(summary)
            if analysis.get("multi_vdd_excel"):
                curve_status.set(
                    f"Complete - {len(analysis['rows'])} VDD point(s); "
                    f"{analysis['per_vdd_output_count']} per-VDD Multi-Cell folder(s); "
                    f"saved to {Path(payload).parent.parent}")
            else:
                curve_status.set(
                    f"Complete - {len(analysis['rows'])} VDD point(s); "
                    f"saved to {Path(payload).parent}")
            curve_status_label.configure(fg=GREEN)
            curve_open_button.state(["!disabled"])
            if analysis.get("mode") in {"shmoo_only", "single_vdd"}:
                curve_stacked_button.state(["disabled"])
                curve_transparent_button.state(["disabled"])
            else:
                curve_stacked_button.state(["!disabled"])
                curve_transparent_button.state(["!disabled"])
            draw_curve_chart()
        else:
            curve_status.set("Estimate Vmin analysis could not be completed")
            curve_status_label.configure(fg=RED)
            messagebox.showerror("Estimate Vmin Curve", str(payload))

    def execute_curve_analysis() -> None:
        if not selected_summary_paths:
            curve_status.set("Select at least one Multi-Cell CSV or Excel file")
            curve_status_label.configure(fg=RED)
            messagebox.showerror(
                "Estimate Vmin Curve", "Select at least one Multi-Cell CSV or Excel file.")
            return
        try:
            _cell, model_config, _targets = collect_inputs()
        except Exception as exc:
            curve_status.set("Check the Model Settings used for Excel analysis")
            curve_status_label.configure(fg=RED)
            messagebox.showerror("Estimate Vmin Curve", str(exc))
            return
        curve_status.set("Combining minimum RSNM / WSNM / Write Margin by Model VDD...")
        curve_status_label.configure(fg=BLUE)
        curve_analyze_button.state(["disabled"])
        curve_open_button.state(["disabled"])
        curve_stacked_button.state(["disabled"])
        curve_transparent_button.state(["disabled"])
        curve_progress.start(10)
        wafer_id = values["corner"].get().strip() or "Multi_Cell"
        output_path = Path(values["out"].get())
        threading.Thread(target=curve_worker,
                         args=(list(selected_summary_paths), output_path, wafer_id,
                               model_config, bool(curve_run_shmoo.get())),
                         daemon=True).start()
        root.after(80, poll_curve_result)

    def open_curve_report() -> None:
        if curve_report_path and curve_report_path.exists():
            webbrowser.open(curve_report_path.resolve().as_uri())

    def open_stacked_curve_png() -> None:
        if curve_report_path and curve_report_path.exists():
            # Open the responsive report wrapper instead of a raw PNG.  Chrome
            # and Edge on Windows otherwise present a large PNG as a reduced
            # image preview, which makes the chart look like a thumbnail.
            webbrowser.open(curve_report_path.resolve().as_uri() + "#stacked-trends")

    def open_transparent_stacked_curve_png() -> None:
        if not curve_report_path:
            return
        transparent_png = (curve_report_path.parent / "images" /
                           "05_estimate_vmin_stacked_transparent.png")
        if transparent_png.exists():
            webbrowser.open(transparent_png.resolve().as_uri())

    curve_action_row = ttk.Frame(curve_input_card, style="Card.TFrame")
    curve_action_row.pack(side="bottom", fill="x", pady=(8, 0))
    ttk.Button(curve_action_row, text="Import Multi-Cell CSV / Excel...", style="Quiet.TButton",
               command=import_multi_cell_summaries).pack(fill="x", pady=(0, 7))
    curve_analyze_button = ttk.Button(curve_action_row, text="Analyze Estimate Vmin Curves",
                                      style="Accent.TButton", command=execute_curve_analysis)
    curve_analyze_button.pack(fill="x")
    curve_open_button = ttk.Button(curve_action_row, text="Open HTML Result",
                                   style="Quiet.TButton", command=open_curve_report)
    curve_open_button.pack(fill="x", pady=(7, 0))
    curve_open_button.state(["disabled"])
    curve_stacked_button = ttk.Button(curve_action_row, text="Open Full-Size Stacked View",
                                      style="Quiet.TButton", command=open_stacked_curve_png)
    curve_stacked_button.pack(fill="x", pady=(7, 0))
    curve_stacked_button.state(["disabled"])
    curve_transparent_button = ttk.Button(
        curve_action_row, text="Open Transparent PNG", style="Quiet.TButton",
        command=open_transparent_stacked_curve_png)
    curve_transparent_button.pack(fill="x", pady=(7, 0))
    curve_transparent_button.state(["disabled"])
    ttk.Separator(curve_action_row, orient="horizontal").pack(fill="x", pady=(12, 7))
    ttk.Button(curve_action_row, text="Import Multi-VDD Output Folder...",
               style="Quiet.TButton",
               command=import_comparison_multi_vdd_folder).pack(fill="x")
    ttk.Button(curve_action_row, text="Import Combined Summaries...", style="Quiet.TButton",
               command=import_comparison_combined_files).pack(fill="x", pady=(7, 0))
    ttk.Button(curve_action_row, text="Clear Comparison Files", style="Quiet.TButton",
               command=clear_comparison_summary_files).pack(fill="x", pady=(7, 0))
    curve_compare_button = ttk.Button(
        curve_action_row, text="Create Comparison View", style="Accent.TButton",
        command=execute_combined_file_comparison)
    curve_compare_button.pack(fill="x", pady=(7, 0))

    # Dedicated Lot/Wafer Advisor. This keeps batch recommendations and
    # group-distribution work separate from the Estimate Vmin workflow.
    advisor_tab.columnconfigure(0, weight=4)
    advisor_tab.columnconfigure(1, weight=9)
    advisor_tab.rowconfigure(0, weight=1)
    advisor_input_card = ttk.Frame(advisor_tab, style="Card.TFrame", padding=20)
    advisor_chart_card = ttk.Frame(advisor_tab, style="Card.TFrame", padding=20)
    advisor_input_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    advisor_chart_card.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
    advisor_chart_card.columnconfigure(0, weight=1)
    advisor_chart_card.rowconfigure(2, weight=1)

    ttk.Label(advisor_input_card, text="Lot/Wafer Drive Advisor",
              style="Section.TLabel").pack(anchor="w")
    ttk.Label(
        advisor_input_card,
        text=("Import one or more multi_chip_snm_summary.csv files. Rows with the "
              "same Lot/Wafer name are treated as one group at each Model VDD."),
        style="Meta.TLabel", wraplength=430).pack(anchor="w", pady=(3, 14))
    advisor_explanation = tk.Frame(advisor_input_card, bg="#F5F9FF", padx=14, pady=12)
    advisor_explanation.pack(fill="x", pady=(0, 14))
    tk.Label(advisor_explanation, text="Distribution comparison", bg="#F5F9FF",
             fg=TEXT, font=("Calibri", 11, "bold"), anchor="w").pack(fill="x")
    tk.Label(
        advisor_explanation,
        text=("Box plots compare Read SNM, BL Write Trip Margin and Balanced Drive "
              "Score. The CR–PR map uses one color and marker per Lot/Wafer to expose "
              "group concentration, shift and outliers."),
        bg="#F5F9FF", fg=SECONDARY, font=("Calibri", 10), justify="left",
        wraplength=400, anchor="w").pack(fill="x", pady=(4, 0))
    advisor_selected_paths: list[str] = []
    saved_advisor_paths = saved_state.get("lot_wafer_advisor_paths", [])
    if isinstance(saved_advisor_paths, list):
        advisor_selected_paths.extend(
            str(item) for item in saved_advisor_paths if Path(str(item)).is_file())
    advisor_file_text = tk.StringVar(value="No Multi-Cell summary selected")
    if advisor_selected_paths:
        advisor_file_text.set(
            f"{len(advisor_selected_paths)} previously selected summary file(s)\n" +
            "\n".join(Path(item).name for item in advisor_selected_paths[:6]))
    tk.Label(advisor_input_card, textvariable=advisor_file_text, bg=CARD, fg=SECONDARY,
             font=("Calibri", 10), anchor="w", justify="left", wraplength=430).pack(
                 fill="x", pady=(0, 10))
    advisor_status = tk.StringVar(value="Ready for Lot/Wafer distribution analysis")
    advisor_status_label = tk.Label(
        advisor_input_card, textvariable=advisor_status, bg=CARD, fg=SECONDARY,
        font=("Calibri", 9), anchor="w", justify="left", wraplength=430)
    advisor_status_label.pack(fill="x", pady=(3, 6))
    advisor_progress = ttk.Progressbar(
        advisor_input_card, mode="indeterminate", style="Apple.Horizontal.TProgressbar")
    advisor_progress.pack(fill="x", pady=(0, 9))

    ttk.Label(advisor_chart_card, text="Lot/Wafer CR–PR Distribution Preview",
              style="ChartTitle.TLabel").grid(row=0, column=0, sticky="w")
    advisor_summary = tk.StringVar(
        value="Import summary files to compare same-VDD Lot/Wafer distributions.")
    tk.Label(advisor_chart_card, textvariable=advisor_summary, bg=CARD, fg=SECONDARY,
             font=("Calibri", 10), anchor="w", justify="left").grid(
                 row=1, column=0, sticky="ew", pady=(3, 7))
    advisor_canvas = tk.Canvas(
        advisor_chart_card, bg=CARD, highlightthickness=0, bd=0,
        width=850, height=590)
    advisor_canvas.grid(row=2, column=0, sticky="nsew")
    advisor_canvas.create_text(
        425, 285, text="Lot/Wafer drive distribution will appear here",
        fill=SECONDARY, font=("Calibri", 13))
    advisor_result_queue: queue.Queue = queue.Queue()
    advisor_report_path: Path | None = None
    advisor_chart_image = None

    def import_advisor_summaries() -> None:
        selected = filedialog.askopenfilenames(
            title="Import Multi-Cell Summary CSV for Lot/Wafer Advisor",
            initialdir=values["out"].get(),
            filetypes=[("Multi-Cell summary CSV", "*.csv"), ("All files", "*.*")])
        if not selected:
            return
        advisor_selected_paths[:] = [str(Path(item).resolve()) for item in selected]
        advisor_file_text.set(
            f"{len(advisor_selected_paths)} summary file(s) selected\n" +
            "\n".join(Path(item).name for item in advisor_selected_paths[:6]))
        advisor_status.set("Ready to compare Lot/Wafer distributions")
        advisor_status_label.configure(fg=SECONDARY)

    def advisor_worker(summary_paths: list[str], out_path: Path) -> None:
        try:
            source_rows = read_multi_chip_snm_summary(summary_paths)
            result = analyze_lot_wafer_drive_advisor(source_rows)
            run_dir = create_run_output_dir(
                out_path, "Lot_Wafer", "lot_wafer_drive_advisor")
            report = write_lot_wafer_drive_advisor_outputs(
                result, run_dir, summary_paths)
            advisor_result_queue.put((True, result, report))
        except Exception as exc:
            advisor_result_queue.put((False, None, exc))

    def poll_advisor_result() -> None:
        nonlocal advisor_report_path, advisor_chart_image
        try:
            ok, result, payload = advisor_result_queue.get_nowait()
        except queue.Empty:
            root.after(80, poll_advisor_result)
            return
        advisor_progress.stop()
        advisor_analyze_button.state(["!disabled"])
        if ok:
            advisor_report_path = Path(payload)
            total_cells = sum(int(item["sample_count"]) for item in result["vdds"])
            advisor_summary.set(
                f'{len(result["lot_wafers"])} Lot/Wafer group(s) · '
                f'{len(result["vdds"])} Model VDD point(s) · {total_cells} same-VDD Cell records')
            advisor_status.set(f"Complete - saved to {advisor_report_path.parent}")
            advisor_status_label.configure(fg=GREEN)
            advisor_open_button.state(["!disabled"])
            preview_paths = sorted((advisor_report_path.parent / "images").glob(
                "*_lot_wafer_drive_scatter.png"))
            if preview_paths:
                image = tk.PhotoImage(file=str(preview_paths[0]))
                divisor = max(1, math.ceil(image.width()/900), math.ceil(image.height()/620))
                advisor_chart_image = image.subsample(divisor, divisor)
                advisor_canvas.delete("all")
                advisor_canvas.create_image(
                    max(advisor_canvas.winfo_width(), 850)/2,
                    max(advisor_canvas.winfo_height(), 590)/2,
                    image=advisor_chart_image, anchor="center")
        else:
            advisor_status.set("Lot/Wafer Advisor could not be completed")
            advisor_status_label.configure(fg=RED)
            messagebox.showerror("Lot/Wafer Advisor", str(payload))

    def execute_advisor_analysis() -> None:
        if not advisor_selected_paths:
            advisor_status.set("Select at least one Multi-Cell summary CSV file")
            advisor_status_label.configure(fg=RED)
            messagebox.showerror(
                "Lot/Wafer Advisor", "Select at least one multi_chip_snm_summary.csv file.")
            return
        advisor_status.set("Grouping Lot/Wafer and calculating Read / Write distributions...")
        advisor_status_label.configure(fg=BLUE)
        advisor_analyze_button.state(["disabled"])
        advisor_open_button.state(["disabled"])
        advisor_progress.start(10)
        threading.Thread(
            target=advisor_worker,
            args=(list(advisor_selected_paths), Path(values["out"].get())),
            daemon=True).start()
        root.after(80, poll_advisor_result)

    def open_advisor_report() -> None:
        if advisor_report_path and advisor_report_path.exists():
            webbrowser.open(advisor_report_path.resolve().as_uri())

    advisor_actions = ttk.Frame(advisor_input_card, style="Card.TFrame")
    advisor_actions.pack(side="bottom", fill="x", pady=(8, 0))
    ttk.Button(
        advisor_actions, text="Import Multi-Cell Summary CSV...", style="Quiet.TButton",
        command=import_advisor_summaries).pack(fill="x", pady=(0, 7))
    advisor_analyze_button = ttk.Button(
        advisor_actions, text="Analyze Lot/Wafer Advisor", style="Accent.TButton",
        command=execute_advisor_analysis)
    advisor_analyze_button.pack(fill="x")
    advisor_open_button = ttk.Button(
        advisor_actions, text="Open HTML Result", style="Quiet.TButton",
        command=open_advisor_report)
    advisor_open_button.pack(fill="x", pady=(7, 0))
    advisor_open_button.state(["disabled"])

    # Independent write-trip analysis. It intentionally reuses the same manual
    # VDD/WAT rows so Read and Write trends are compared from identical inputs.
    write_margin_tab.columnconfigure(0, weight=4)
    write_margin_tab.columnconfigure(1, weight=9)
    write_margin_tab.rowconfigure(0, weight=1)
    wtm_input_card = ttk.Frame(write_margin_tab, style="Card.TFrame", padding=20)
    wtm_input_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    wtm_chart_card = ttk.Frame(write_margin_tab, style="Card.TFrame", padding=18)
    wtm_chart_card.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
    wtm_chart_card.columnconfigure(0, weight=1)
    wtm_chart_card.rowconfigure(3, weight=1)

    ttk.Label(wtm_input_card, text="Write Trip Margin", style="Section.TLabel").pack(anchor="w")
    ttk.Label(
        wtm_input_card,
        text=("Uses the VDD / PU / PG / PD rows from the RSNM tab. "
              "The result estimates how far the low write bitline may rise while PG can still overcome PU."),
        style="Meta.TLabel", wraplength=410).pack(anchor="w", pady=(3, 16))
    wtm_definition = tk.Frame(wtm_input_card, bg="#F5F9FF", padx=14, pady=12)
    wtm_definition.pack(fill="x", pady=(0, 14))
    tk.Label(wtm_definition, text="Interpretation", bg="#F5F9FF", fg=TEXT,
             font=("Calibri", 11, "bold"), anchor="w").pack(fill="x")
    tk.Label(
        wtm_definition,
        text=("Larger positive WTM means more write voltage tolerance. "
              "WTM near 0 mV indicates the model write boundary. It is not measured Select_Write Vmin."),
        bg="#F5F9FF", fg=SECONDARY, font=("Calibri", 10),
        justify="left", wraplength=380, anchor="w").pack(fill="x", pady=(4, 0))
    ttk.Button(wtm_input_card, text="Edit shared VDD inputs", style="Quiet.TButton",
               command=lambda: notebook.select(curve_tab)).pack(fill="x", pady=(0, 14))

    wtm_status = tk.StringVar(value="Ready to estimate Write Trip Margin")
    wtm_status_label = tk.Label(
        wtm_input_card, textvariable=wtm_status, bg=CARD, fg=SECONDARY,
        font=("Calibri", 9), anchor="w", justify="left", wraplength=410)
    wtm_status_label.pack(fill="x", pady=(3, 6))
    wtm_progress = ttk.Progressbar(
        wtm_input_card, mode="indeterminate", style="Apple.Horizontal.TProgressbar")
    wtm_progress.pack(fill="x", pady=(0, 9))

    ttk.Label(wtm_chart_card, text="Estimated Write Trip Margin Curve",
              style="ChartTitle.TLabel").grid(row=0, column=0, sticky="w")
    ttk.Label(
        wtm_chart_card,
        text="X: Model VDD (V)  /  Y: Write Trip Margin (mV). X marks indicate no positive modeled write margin.",
        style="Meta.TLabel").grid(row=1, column=0, sticky="w", pady=(2, 6))
    wtm_summary = tk.StringVar(value="Analyze at least two shared VDD rows to display the curve.")
    wtm_summary_label = tk.Label(
        wtm_chart_card, textvariable=wtm_summary, bg=CARD, fg=SECONDARY,
        font=("Calibri", 10, "bold"), anchor="w", justify="left")
    wtm_summary_label.grid(row=2, column=0, sticky="ew", pady=(0, 6))
    wtm_canvas = tk.Canvas(wtm_chart_card, bg=CARD, highlightthickness=0, bd=0,
                           width=720, height=550)
    wtm_canvas.grid(row=3, column=0, sticky="nsew")
    wtm_canvas.create_text(360, 260, text="Write Trip Margin curve will appear here",
                           fill=SECONDARY, font=("Calibri", 13))
    wtm_result_queue: queue.Queue = queue.Queue()
    wtm_report_path: Path | None = None
    wtm_chart_image = None

    def wtm_worker(points: list[RsnmVccPoint], cfg: Config,
                   out_path: Path, wafer_id: str) -> None:
        try:
            analysis = analyze_write_trip_margin_curve(points, cfg)
            run_dir = create_run_output_dir(out_path, wafer_id, "write_trip_margin_vdd_curve")
            report = write_write_trip_margin_outputs(analysis, run_dir)
            wtm_result_queue.put((True, analysis, report))
        except Exception as exc:
            wtm_result_queue.put((False, None, exc))

    def poll_wtm_result() -> None:
        nonlocal wtm_report_path, wtm_chart_image
        try:
            ok, analysis, payload = wtm_result_queue.get_nowait()
        except queue.Empty:
            root.after(80, poll_wtm_result)
            return
        wtm_progress.stop()
        wtm_analyze_button.state(["!disabled"])
        if ok:
            wtm_report_path = Path(payload)
            boundary = analysis.get("write_boundary")
            if boundary:
                wtm_summary.set(
                    f'Estimated write boundary VDD: {boundary["estimated_vdd_v"]:.4f} V')
                wtm_summary_label.configure(fg="#C56A00")
            else:
                wtm_summary.set("Write boundary VDD not bracketed by the entered rows")
                wtm_summary_label.configure(fg=SECONDARY)
            wtm_status.set(
                f"Complete - {len(analysis['rows'])} VDD point(s); saved to {wtm_report_path.parent}")
            wtm_status_label.configure(fg=GREEN)
            wtm_open_button.state(["!disabled"])
            png_path = wtm_report_path.parent / "images" / "01_write_trip_margin_vs_model_vdd.png"
            image = tk.PhotoImage(file=str(png_path))
            wtm_chart_image = image.subsample(2, 2)
            wtm_canvas.delete("all")
            wtm_canvas.create_image(
                max(wtm_canvas.winfo_width(), 720) / 2,
                max(wtm_canvas.winfo_height(), 550) / 2,
                image=wtm_chart_image, anchor="center")
        else:
            wtm_status.set("Write Trip Margin analysis could not be completed")
            wtm_status_label.configure(fg=RED)
            messagebox.showerror("Write Trip Margin analysis", str(payload))

    def execute_wtm_analysis() -> None:
        try:
            points, cfg = collect_curve_inputs()
        except Exception as exc:
            wtm_status.set("Check the shared VDD sweep input values")
            wtm_status_label.configure(fg=RED)
            messagebox.showerror("Invalid VDD sweep input", str(exc))
            return
        wtm_status.set("Calculating Write Trip Margin at each VDD point...")
        wtm_status_label.configure(fg=BLUE)
        wtm_analyze_button.state(["disabled"])
        wtm_open_button.state(["disabled"])
        wtm_progress.start(10)
        wafer_id = values["corner"].get().strip() or "Manual"
        output_path = Path(values["out"].get())
        threading.Thread(target=wtm_worker,
                         args=(points, cfg, output_path, wafer_id), daemon=True).start()
        root.after(80, poll_wtm_result)

    def open_wtm_report() -> None:
        if wtm_report_path and wtm_report_path.exists():
            webbrowser.open(wtm_report_path.resolve().as_uri())

    wtm_action_row = ttk.Frame(wtm_input_card, style="Card.TFrame")
    wtm_action_row.pack(side="bottom", fill="x", pady=(8, 0))
    wtm_analyze_button = ttk.Button(
        wtm_action_row, text="Analyze Write Trip Margin vs VDD",
        style="Accent.TButton", command=execute_wtm_analysis)
    wtm_analyze_button.pack(fill="x")
    wtm_open_button = ttk.Button(
        wtm_action_row, text="Open HTML Result", style="Quiet.TButton",
        command=open_wtm_report)
    wtm_open_button.pack(fill="x", pady=(7, 0))
    wtm_open_button.state(["disabled"])

    def persist_and_close() -> None:
        state = {
            "values": {key: variable.get() for key, variable in values.items()},
            "wat": {key: variable.get() for key, variable in wat_values.items()},
            "targets": {key: variable.get() for key, variable in target_values.items()},
            "options": {
                "use_wat_target_reference": use_wat_target_reference.get(),
                "estimate_vmin_run_shmoo": curve_run_shmoo.get(),
            },
            "numeric": {key: variable.get() for key, variable in numeric.items()},
            "assumptions": {key: variable.get() for key, variable in assumption_values.items()},
            "training": {key: variable.get() for key, variable in training_values.items()},
            "estimate_vmin_summary_paths": list(selected_summary_paths),
            "estimate_vmin_comparison_paths": list(comparison_summary_paths),
            "lot_wafer_advisor_paths": list(advisor_selected_paths),
        }
        try:
            save_gui_state(state)
        except OSError:
            pass
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", persist_and_close)
    root.mainloop()


def parse_args(argv: list[str]) -> argparse.Namespace:
    p=argparse.ArgumentParser(description="6T SRAM / WAT cell-level SNM analyzer")
    p.add_argument("--input",help="WAT CSV or six-MOS WAT Excel (.xlsx); omit to open GUI")
    p.add_argument("--output",default="output",help="output directory")
    p.add_argument("--corner",help="analyze only this corner")
    p.add_argument("--vdd",type=float,default=.80,help="nominal SRAM VDD")
    p.add_argument("--wat-vdd",type=float,default=1.20,help="WAT Ids test voltage")
    p.add_argument("--pu-target-vt",type=float,default=.380)
    p.add_argument("--pu-target-ids",type=float,default=45.0)
    p.add_argument("--pg-target-vt",type=float,default=.370)
    p.add_argument("--pg-target-ids",type=float,default=80.0)
    p.add_argument("--pd-target-vt",type=float,default=.360)
    p.add_argument("--pd-target-ids",type=float,default=120.0)
    return p.parse_args(argv)


def main(argv: list[str] | None=None) -> int:
    args=parse_args(sys.argv[1:] if argv is None else argv)
    if not args.input:
        launch_gui(); return 0
    cfg=Config(wat_vdd=args.wat_vdd, nominal_vdd=args.vdd)
    targets = DatasheetTargets(MosWat(args.pu_target_vt, args.pu_target_ids),
                               MosWat(args.pg_target_vt, args.pg_target_ids),
                               MosWat(args.pd_target_vt, args.pd_target_ids))
    try:
        reports=run_analysis(args.input,args.output,cfg,args.corner,targets)
        for report in reports: print(report.resolve())
        return 0
    except Exception as exc:
        print(f"error: {exc}",file=sys.stderr); return 2


if __name__ == "__main__":
    raise SystemExit(main())
