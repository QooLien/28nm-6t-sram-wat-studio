import csv
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest import mock

from sram_wat_analyzer import (
    AsymmetricSram6T, Config, DatasheetTargets, Device, MosWat, RsnmVccPoint,
    SixTWatCell, Sram6T, ThreeTWatCell, WaferChipWat,
    WatPoint, analyze, analyze_six_mos, analyze_three_mos,
    _read_wat_excel_rows, analyze_estimate_vmin_curves, analyze_multi_chip_wafer, analyze_rsnm_vcc_curve, estimate_vmin_combined_comparison_svg, estimate_vmin_curve_svg, estimate_vmin_ratio_shmoo_svg, estimate_vmin_stacked_svg, read_estimate_vmin_combined_files,
    analyze_write_trip_margin_curve,
    analyze_lot_wafer_drive_advisor,
    build_batch_drive_to_preferred_advice, build_drive_to_preferred_advice,
    _drive_advisor_html,
    generic_28nm_assumption_rows,
    drive_monitor_metrics, drive_monitor_shmoo_reference,
    create_run_output_dir, load_gui_state, model_vdd_butterfly_svg, multi_chip_vtc_svg, open_output_directory,
    read_iv_curve_excel, read_multi_chip_6t_excel, read_multi_chip_snm_summary, read_wat_csv, read_wat_excel, rsnm_vcc_curve_svg,
    lot_wafer_boxplot_svg, lot_wafer_drive_scatter_svg,
    lot_wafer_grade_counts_svg,
    save_gui_state,
    validate_config, wat_electrical_snm_rows,
    write_iv_curve_excel_template, write_multi_chip_6t_excel_template, write_multi_chip_outputs, write_outputs, write_rsnm_vcc_curve_outputs, write_single_6t_wat_excel,
    write_estimate_vmin_outputs, write_lot_wafer_drive_advisor_outputs,
    write_trip_margin_curve_svg, write_write_trip_margin_outputs,
)


class AnalyzerTests(unittest.TestCase):
    def setUp(self):
        self.cfg = Config(grid_points=101)
        self.targets = DatasheetTargets(MosWat(.380, 45), MosWat(.370, 80), MosWat(.360, 120))
        self.cell = ThreeTWatCell("LOT_W01", MosWat(.385, 44), MosWat(.365, 82), MosWat(.355, 124))

    def test_smooth_overdrive_keeps_current_continuous_near_threshold(self):
        device = Device(.385, 44.0, .90)
        below = device.current(.384, .90)
        at_threshold = device.current(.385, .90)
        above = device.current(.386, .90)
        self.assertGreater(below, 0.0)
        self.assertLess(below, at_threshold)
        self.assertLess(at_threshold, above)
        self.assertAlmostEqual(device.current(.90, .90), 44.0, places=7)

    def test_drive_monitor_metrics_use_wat_calibrated_beta_ratios(self):
        baseline = drive_monitor_metrics(
            WatPoint("TRAIN", .385, 44.0, .365, 82.0, .355, 124.0), .90)
        stronger_pg = drive_monitor_metrics(
            WatPoint("TRAIN", .385, 44.0, .365, 110.0, .355, 124.0), .90)
        self.assertGreater(baseline["read_snm_mv"], 0.0)
        self.assertGreaterEqual(baseline["write_margin_mv"], 0.0)
        self.assertGreater(baseline["cell_ratio"], 0.0)
        self.assertGreater(baseline["pull_up_ratio"], 0.0)
        self.assertGreater(stronger_pg["beta_pg"], baseline["beta_pg"])
        self.assertGreater(stronger_pg["pull_up_ratio"], baseline["pull_up_ratio"])

    def test_drive_monitor_shmoo_reference_uses_stable_quartile_targets(self):
        low_vdd = drive_monitor_shmoo_reference(.68)
        high_vdd = drive_monitor_shmoo_reference(.90)
        self.assertEqual(low_vdd["sample_count"], 729)
        for axis in ("cr", "pr"):
            self.assertLess(low_vdd[axis]["q1"], low_vdd[axis]["median"])
            self.assertLess(low_vdd[axis]["median"], low_vdd[axis]["q3"])
        self.assertNotAlmostEqual(low_vdd["cr"]["median"],
                                  high_vdd["cr"]["median"])

    def test_run_output_directory_uses_date_time_wafer_and_never_overwrites(self):
        stamp = datetime(2026, 8, 1, 14, 30, 25)
        with tempfile.TemporaryDirectory() as td:
            first = create_run_output_dir(td, "LOT/W01:17", "6t analysis", stamp)
            second = create_run_output_dir(td, "LOT/W01:17", "6t analysis", stamp)
            self.assertEqual(first.relative_to(td).parts,
                             ("2026-08-01", "LOT_W01_17", "143025_6t_analysis"))
            self.assertEqual(second.name, "143025_6t_analysis_02")
            manifest = (first / "run_info.json").read_text(encoding="utf-8")
            self.assertIn('"wafer_id": "LOT/W01:17"', manifest)
            self.assertIn('"created_local": "2026-08-01T14:30:25"', manifest)

    def test_open_output_directory_creates_and_opens_selected_folder(self):
        with tempfile.TemporaryDirectory() as td:
            target = Path(td) / "new output"
            with mock.patch("sram_wat_analyzer.sys.platform", "win32"), \
                    mock.patch("sram_wat_analyzer.os.startfile", create=True) as startfile:
                opened = open_output_directory(target)
            self.assertEqual(opened, target.resolve())
            self.assertTrue(target.is_dir())
            startfile.assert_called_once_with(str(target.resolve()))

    def test_open_output_directory_rejects_blank_path(self):
        with self.assertRaisesRegex(ValueError, "Choose an output folder"):
            open_output_directory("  ")

    def test_read_snm_is_bounded(self):
        model = Sram6T(self.cell.representative(), self.cfg)
        result = model.butterfly_squares(self.cfg.nominal_vdd, "read", points=601)
        self.assertTrue(result["valid"])
        self.assertEqual(len(result["squares"]), 2)
        value = result["snm_v"]
        self.assertGreaterEqual(value, 0)
        self.assertLessEqual(value, self.cfg.nominal_vdd / 2)

    def test_figure_3_15_geometric_read_snm(self):
        wat = WatPoint("CURRENT", .35, 29.2, .27, 40.3, .27, 47.6)
        result = Sram6T(wat, self.cfg).butterfly_squares(.9, "read", points=1201)
        self.assertTrue(result["valid"])
        self.assertAlmostEqual(result["snm_mv"], 156.0, delta=1.0)
        self.assertAlmostEqual(result["squares"][0]["side_mv"],
                               result["squares"][1]["side_mv"], delta=1.0)

    def test_metrics_include_write_snm(self):
        metrics = analyze(self.cell.representative(), self.cfg)["baseline_6t"]["metrics"]
        self.assertNotIn("hold_snm_mv", metrics)
        self.assertIn("read_snm_mv", metrics)
        self.assertIn("write_snm_mv", metrics)
        self.assertGreater(metrics["write_snm_mv"], 0.0)
        self.assertNotIn("read_vmin_v", metrics)
        self.assertNotIn("write_vmin_v", metrics)

    def test_target_model_and_delta(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        self.assertIn("target_6t", result)
        self.assertEqual([row["mode"] for row in result["snm_target_comparison"]],
                         ["Read SNM"])
        self.assertTrue(any(abs(row["delta_mv"]) > 0 for row in result["snm_target_comparison"]))

    def test_wat_target_reference_can_be_disabled(self):
        result = analyze_six_mos(self.cell.to_six_t(), self.cfg, None)
        self.assertIsNone(result["datasheet_targets"])
        self.assertNotIn("target_6t", result)
        self.assertEqual(result["snm_target_comparison"], [])
        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "output"
            report = write_outputs(result, output)
            overview = (output / "images" / "01_read_snm_target_comparison.svg").read_text(
                encoding="utf-8")
            write_svg = (output / "images" / "03_w0_w1_wsnm_analysis.svg").read_text(
                encoding="utf-8")
            html_text = report.read_text(encoding="utf-8")
            self.assertIn("Read SNM Analysis", overview)
            self.assertNotIn("WAT Target VTC", overview)
            self.assertNotIn("WAT Target pair", write_svg)
            self.assertIn("WAT Target reference is disabled", html_text)
            self.assertFalse((output / "snm_target_comparison.csv").exists())
            self.assertFalse((output / "wat_target_comparison.csv").exists())
            with (output / "w0_w1_wsnm_analysis.csv").open(
                    newline="", encoding="utf-8-sig") as source:
                self.assertEqual({row["dataset"] for row in csv.DictReader(source)},
                                 {"Lot/Wafer"})

    def test_write_snm_contains_w1_w0_window_and_square(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        states = result["baseline_6t"]["write_wsnm"]
        self.assertGreater(len(states["write_0"]["curve"]), 100)
        self.assertGreater(len(states["write_1"]["curve"]), 100)
        self.assertGreater(states["snm_mv"], 0.0)
        self.assertIsNotNone(states["write_square"])

    def test_pdf_equation_3_36_with_given_wat_values(self):
        current = WatPoint("CURRENT", .35, 29.2, .27, 40.3, .27, 47.6)
        target = WatPoint("TARGET", .33, 19.5, .28, 39.0, .29, 44.9)
        current_eq = Sram6T(current, self.cfg).analytical_read_snm_eq_3_36(.9)
        target_eq = Sram6T(target, self.cfg).analytical_read_snm_eq_3_36(.9)
        self.assertTrue(current_eq["valid"])
        self.assertTrue(target_eq["valid"])
        self.assertAlmostEqual(current_eq["vth_eff_v"], (.35 + .27 + .27) / 3)
        self.assertAlmostEqual(current_eq["snm_mv"], 189.7967615, places=5)
        self.assertAlmostEqual(target_eq["snm_mv"], 183.5971209, places=5)

    def test_pdf_equation_3_36_reports_domain_failure(self):
        equation = Sram6T(WatPoint(), self.cfg).analytical_read_snm_eq_3_36(.9)
        self.assertFalse(equation["valid"])
        self.assertIsNone(equation["snm_mv"])
        self.assertIn("square-root domain", equation["reason"])

    def test_six_independent_objects_remain_supported(self):
        cell = SixTWatCell("MISMATCH", MosWat(.38, 45), MosWat(.40, 42),
                           MosWat(.37, 80), MosWat(.39, 75),
                           MosWat(.36, 120), MosWat(.38, 110))
        result = analyze_six_mos(cell, self.cfg, self.targets)
        self.assertEqual(set(result["cell"]["mos"]), {"PUL", "PUR", "PGL", "PGR", "PDL", "PDR"})
        self.assertEqual(len(result["target_comparisons"]), 6)
        self.assertEqual(len(result["snm_target_comparison"]), 1)
        metrics = result["baseline_6t"]["metrics"]
        self.assertIn("read_snm_upper_left_mv", metrics)
        self.assertIn("read_snm_lower_right_mv", metrics)
        self.assertAlmostEqual(metrics["read_snm_mv"],
                               min(metrics["read_snm_upper_left_mv"],
                                   metrics["read_snm_lower_right_mv"]))

    def test_asymmetric_read_snm_detects_and_swaps_left_right_mismatch(self):
        cell = SixTWatCell(
            "MISMATCH",
            MosWat(.35, 29.2), MosWat(.39, 20.0),
            MosWat(.27, 40.3), MosWat(.31, 30.0),
            MosWat(.27, 47.6), MosWat(.33, 35.0),
        )
        original = AsymmetricSram6T(cell, self.cfg).read_butterfly(.9, 1201)["read_butterfly"]
        swapped_cell = SixTWatCell(
            "SWAPPED", cell.pu2, cell.pu1, cell.pg2, cell.pg1, cell.pd2, cell.pd1)
        swapped = AsymmetricSram6T(swapped_cell, self.cfg).read_butterfly(.9, 1201)["read_butterfly"]
        self.assertGreater(original["mismatch_index_pct"], 1.0)
        self.assertAlmostEqual(original["snm_upper_left_mv"],
                               swapped["snm_lower_right_mv"], delta=1.0)
        self.assertAlmostEqual(original["snm_lower_right_mv"],
                               swapped["snm_upper_left_mv"], delta=1.0)
        self.assertAlmostEqual(original["snm_mv"], swapped["snm_mv"], delta=1.0)
        self.assertAlmostEqual(original["delta_snm_mv"], -swapped["delta_snm_mv"], delta=1.0)

    def test_html_png_and_csv_include_read_and_w0_w1_wsnm(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "output"
            output.mkdir()
            (output / "wt_test_0bit_vmin.csv").write_text("old", encoding="utf-8")
            (output / "sram_wat_results.csv").write_text("old vmin schema", encoding="utf-8")
            report = write_outputs(result, output)
            image_dir = output / "images"
            self.assertTrue((image_dir / "01_read_snm_target_comparison.png").exists())
            self.assertTrue((image_dir / "01_read_snm_target_comparison.svg").exists())
            self.assertTrue((image_dir / "02_read_snm_butterfly.png").exists())
            self.assertTrue((image_dir / "02_read_snm_butterfly.svg").exists())
            self.assertTrue((image_dir / "03_w0_w1_wsnm_analysis.png").exists())
            self.assertTrue((image_dir / "03_w0_w1_wsnm_analysis.svg").exists())
            self.assertEqual(len(list(image_dir.glob("*.png"))), 3)
            self.assertEqual(len(list(image_dir.glob("*.svg"))), 3)
            svg = (image_dir / "01_read_snm_target_comparison.svg").read_text(encoding="utf-8")
            self.assertIn("LOT_W01 WAT VTC", svg)
            self.assertNotIn("Lot/Wafer WAT VTC", svg)
            self.assertNotIn("Current WAT VTC", svg)
            self.assertIn("WAT Target VTC", svg)
            self.assertIn("Vin (V)", svg)
            self.assertIn("Vout (V)", svg)
            self.assertIn("0.30", svg)
            self.assertIn("1.20", svg)
            self.assertNotIn("SNM squares in both butterfly lobes", svg)
            butterfly_svg = (image_dir / "02_read_snm_butterfly.svg").read_text(encoding="utf-8")
            self.assertIn("Maximum squares 1 and 2", butterfly_svg)
            self.assertNotIn("Geometric RSNM", butterfly_svg)
            self.assertNotIn("Analytical RSNM", butterfly_svg)
            self.assertIn("cell RSNM is the smaller value", butterfly_svg)
            self.assertNotIn("QB=0 / Q=1", butterfly_svg)
            self.assertIn("mV</text>", butterfly_svg)
            self.assertNotIn("QB=1 / Q=0", butterfly_svg)
            for dataset in (result["baseline_6t"], result["target_6t"]):
                for square in dataset["read_butterfly"]["squares"]:
                    self.assertIn(f'>{square["side_mv"]:.1f} mV</text>', butterfly_svg)
            self.assertIn("Vin (V)", butterfly_svg)
            self.assertIn("Vout (V)", butterfly_svg)
            self.assertIn("1.20", butterfly_svg)
            self.assertNotIn("Figure 3.15", butterfly_svg)
            write_svg = (image_dir / "03_w0_w1_wsnm_analysis.svg").read_text(encoding="utf-8")
            self.assertIn("Write SNM Butterfly Analysis", write_svg)
            self.assertIn("W=1 VTC (upper)", write_svg)
            self.assertIn("W=0 VTC (lower)", write_svg)
            self.assertIn("Vin=Vout diagonal-constrained WSNM square", write_svg)
            self.assertNotIn("stroke-dasharray=\"8 6\"", write_svg)
            html = report.read_text(encoding="utf-8")
            self.assertIn("Read SNM Target Comparison", html)
            self.assertIn("Lot/Wafer SNM", html)
            self.assertNotIn("Current SNM", html)
            self.assertNotIn("Hold SNM", html)
            self.assertIn("Write SNM Butterfly Analysis", html)
            self.assertIn("constrained to Vin=Vout", html)
            self.assertNotIn("WT Test 0-Bit Vmin", html)
            self.assertNotIn("Vmin", html)
            self.assertFalse((output / "wt_test_0bit_vmin.csv").exists())
            self.assertFalse((output / "sram_wat_results.csv").exists())
            self.assertTrue((output / "snm_target_comparison.csv").exists())
            self.assertTrue((output / "w0_w1_wsnm_analysis.csv").exists())
            self.assertFalse((output / "write_snm_vs_bitline.csv").exists())
            self.assertFalse((output / "single_wat_write_snm_geometry.csv").exists())
            self.assertTrue((output / "read_snm_state_mismatch.csv").exists())
            self.assertTrue((output / "analytical_read_snm.csv").exists())
            self.assertFalse((output / "analytical_read_snm_eq_3_36.csv").exists())
            self.assertTrue((output / "wat_electrical_snm_table.csv").exists())
            self.assertTrue((output / "cell_geometry_reference.csv").exists())
            self.assertFalse((output / "generic_28nm_assumptions.csv").exists())
            self.assertIn("Analytical Read SNM Reference", html)
            self.assertNotIn("PDF", html)
            self.assertNotIn("Figure 3.15", html)
            self.assertNotIn("Equation 3.36", html)
            self.assertIn("WAT Electrical Parameters", html)
            self.assertIn("No W/L, Cox, mobility", html)
            self.assertIn("6T Cell Geometry Reference", html)
            self.assertIn("VTH,eff", html)
            self.assertIn("Read SNM Butterfly and Left/Right Mismatch", html)
            self.assertIn("Mismatch index", html)
            with open(output / "snm_target_comparison.csv", encoding="utf-8-sig") as source:
                rows = list(csv.DictReader(source))
            self.assertEqual(len(rows), 1)
            self.assertEqual([row["mode"] for row in rows], ["Read SNM"])
            self.assertIn("lot_wafer_snm_mv", rows[0])
            self.assertNotIn("current_snm_mv", rows[0])

            sweep_svg = model_vdd_butterfly_svg([
                {"model_vdd_v": self.cfg.nominal_vdd, "result": result}
            ])
            measured_square = result["baseline_6t"]["read_butterfly"]["squares"][0]
            target_square = result["target_6t"]["read_butterfly"]["squares"][0]
            self.assertIn(f'WAT {measured_square["side_mv"]:.1f} mV', sweep_svg)
            self.assertIn(f'Target {target_square["side_mv"]:.1f} mV', sweep_svg)

    def test_wat_electrical_snm_table_uses_measured_inputs(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        rows = wat_electrical_snm_rows(result)
        self.assertEqual([row["dataset"] for row in rows],
                         ["Lot/Wafer", "WAT Target"])
        current = rows[0]
        self.assertAlmostEqual(current["pu_vt_v"], .385)
        self.assertAlmostEqual(current["pg_idsat_ua"], 82)
        self.assertAlmostEqual(current["idsat_pd_over_pg"], 124 / 82)
        self.assertGreater(current["q_beta_pu_over_pg"], 0)
        self.assertGreater(current["r_beta_pd_over_pg"], 0)
        self.assertNotIn("hold_snm_geometric_mv", current)
        self.assertIn("read_snm_geometric_mv", current)
        self.assertEqual(
            current["evidence_scope"],
            "WAT Vt + Idsat; no PDK/model-card-only parameters",
        )

    def test_geometry_references_are_explicit_and_do_not_override_wat(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        rows = {row["parameter"]: row for row in generic_28nm_assumption_rows(result)}
        self.assertEqual(rows["Channel length L"]["value"], 28.0)
        self.assertEqual(rows["Channel length L"]["active"], "REFERENCE")
        self.assertEqual(rows["Geometry Cell Ratio"]["value"], 1.4)
        self.assertAlmostEqual(rows["Geometry Pull-up Ratio"]["value"], 1.4286, places=4)
        self.assertEqual(len(rows), 6)

    def test_editable_technology_assumptions_propagate(self):
        cfg = Config(
            grid_points=101,
            technology_node_nm=27.5,
            channel_length_nm=29.0,
            pu_width_nm=72.0,
            pg_width_nm=104.0,
            pd_width_nm=146.0,
            nominal_temperature_c=30.0,
            read_wordline_over_vdd=0.95,
            read_bitline_over_vdd=0.98,
        )
        result = analyze_three_mos(self.cell, cfg, self.targets)
        tech = result["technology"]
        self.assertEqual(tech["node_nm"], 28)
        self.assertEqual(tech["channel_length_nm"], 29.0)
        self.assertEqual(tech["pu_width_nm"], 72.0)
        self.assertEqual(tech["pg_width_nm"], 104.0)
        self.assertEqual(tech["pd_width_nm"], 146.0)
        self.assertEqual(tech["nominal_temperature_c"], 30.0)
        self.assertEqual(tech["read_wordline_over_vdd"], 0.95)
        self.assertEqual(tech["read_bitline_over_vdd"], 0.98)

    def test_geometry_reference_rejects_zero_width(self):
        with self.assertRaisesRegex(ValueError, "PG width must be a positive"):
            analyze_three_mos(self.cell, Config(grid_points=101, pg_width_nm=0.0), self.targets)

    def test_read_bias_assumptions_affect_read_snm(self):
        baseline = analyze_three_mos(self.cell, self.cfg, self.targets)["baseline_6t"]["metrics"]["read_snm_mv"]
        adjusted_cfg = Config(grid_points=101, read_wordline_over_vdd=0.90,
                              read_bitline_over_vdd=0.95)
        adjusted = analyze_three_mos(self.cell, adjusted_cfg, self.targets)["baseline_6t"]["metrics"]["read_snm_mv"]
        self.assertNotAlmostEqual(baseline, adjusted, places=4)

    def test_wat_target_deltas(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        pu = result["target_comparisons"][0]
        self.assertAlmostEqual(pu["delta_vt_mv"], 5.0)
        self.assertAlmostEqual(pu["delta_isat_ua"], -1.0)

    def test_csv_input_and_config(self):
        validate_config(self.cfg)
        with tempfile.TemporaryDirectory() as td:
            source = Path(td) / "wat.csv"
            source.write_text(
                "corner,pu_vt,pu_ids,pg_vt,pg_ids,pd_vt,pd_ids\nTT,.42,45,.40,80,.39,120\n",
                encoding="utf-8",
            )
            self.assertEqual(read_wat_csv(source)[0].corner, "TT")

    def test_excel_long_form_converts_units_and_groups_six_mos(self):
        headers = ["Lot/Wafer", "Model VDD", "VDD Unit", "MOS", "Vt", "Vt Unit", "Idsat", "Idsat Unit"]
        rows = [
            ["W01", 900, "mV", name, value, "mV", current, unit]
            for name, value, current, unit in (
                ("PUL", 380, 0.045, "mA"), ("PUR", 382, 46, "uA"),
                ("PGL", 365, 0.082, "mA"), ("PGR", 366, 83, "uA"),
                ("PDL", 355, 0.124, "mA"), ("PDR", 356, 125, "uA"),
            )
        ]
        samples = _read_wat_excel_rows(headers, rows, .9)
        self.assertEqual(len(samples), 1)
        self.assertEqual(samples[0].lot_wafer, "W01")
        self.assertAlmostEqual(samples[0].model_vdd_v, .9)
        self.assertAlmostEqual(samples[0].cell.pu1.vt, .380)
        self.assertAlmostEqual(samples[0].cell.pu1.ids, 45.0)
        self.assertAlmostEqual(samples[0].cell.pd2.ids, 125.0)

    def test_excel_wide_form_and_gui_state(self):
        headers = ["Lot/Wafer", "Model VDD (V)", "PUL Vt (mV)", "PUL Idsat (uA)",
                   "PUR Vt (mV)", "PUR Idsat (uA)", "PGL Vt (mV)", "PGL Idsat (uA)",
                   "PGR Vt (mV)", "PGR Idsat (uA)", "PDL Vt (mV)", "PDL Idsat (uA)",
                   "PDR Vt (mV)", "PDR Idsat (uA)"]
        row = ["W02", 1.2, 380, 45, 381, 46, 365, 82, 366, 83, 355, 124, 356, 125]
        samples = _read_wat_excel_rows(headers, [row], .9)
        self.assertEqual(len(samples), 1)
        self.assertAlmostEqual(samples[0].model_vdd_v, 1.2)
        self.assertAlmostEqual(samples[0].cell.pg2.vt, .366)
        with tempfile.TemporaryDirectory() as td:
            state_path = Path(td) / "state.json"
            save_gui_state({"values": {"corner": "W02"}}, state_path)
            self.assertEqual(load_gui_state(state_path)["values"]["corner"], "W02")

    def test_single_6t_wat_excel_round_trip(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "single_6t.xlsx"
            written = write_single_6t_wat_excel(path, self.cell.to_six_t(), 0.9)
            self.assertEqual(written, path)
            workbook = load_workbook(path, read_only=False, data_only=False)
            self.assertEqual(workbook.sheetnames, ["6T WAT Input", "Instructions"])
            sheet = workbook["6T WAT Input"]
            self.assertEqual(sheet.auto_filter.ref, "A1:I7")
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual([cell.value for cell in sheet[1]], [
                "Lot/Wafer", "Model VDD", "VDD Unit", "MOS", "Vt", "Vt Unit",
                "Idsat", "Idsat Unit", "Notes",
            ])
            self.assertEqual([sheet.cell(row=row, column=4).value for row in range(2, 8)],
                             ["PUL", "PUR", "PGL", "PGR", "PDL", "PDR"])
            workbook.close()
            samples = read_wat_excel(path)
            self.assertEqual(len(samples), 1)
            self.assertEqual(samples[0].lot_wafer, self.cell.corner)
            self.assertAlmostEqual(samples[0].model_vdd_v, 0.9)
            self.assertAlmostEqual(samples[0].cell.pg2.ids, self.cell.pg.ids)

    def test_excel_repeated_wafer_sites_are_aggregated_with_coverage(self):
        headers = ["Lot/Wafer", "Site", "Model VDD", "MOS", "Vt", "Vt Unit",
                   "Idsat", "Idsat Unit", "Notes"]
        rows = []
        for mos in ("PUL", "PUR", "PGL", "PGR", "PDL", "PDR"):
            for site in range(1, 18):
                if site <= 12:
                    rows.append(["W03", f"S{site:02d}", 1.2, mos,
                                 0.30 + site * 0.001, "V", 40.0 + site, "uA", "Measured"])
                else:
                    rows.append(["W03", f"S{site:02d}", 1.2, mos,
                                 None, "V", None, "uA", "Unavailable"])
        sample = _read_wat_excel_rows(headers, rows, .9)[0]
        stats = sample.statistics["pu1"]
        self.assertEqual(stats.valid_count, 12)
        self.assertEqual(stats.total_count, 17)
        self.assertAlmostEqual(stats.vt_mean, 0.3065)
        self.assertAlmostEqual(sample.cell.pu1.ids, 46.5)

    def test_manual_rsnm_vcc_curve_brackets_eye_closure_and_exports(self):
        base_vt = {"pu": .385, "pg": .365, "pd": .355}
        base_ids = {"pu": 44.0, "pg": 82.0, "pd": 124.0}

        def scaled(kind, vcc):
            return base_ids[kind] * (max(vcc - base_vt[kind], 0) /
                                     (.9 - base_vt[kind])) ** 2

        points = [
            RsnmVccPoint(vcc,
                         MosWat(base_vt["pu"], scaled("pu", vcc)),
                         MosWat(base_vt["pg"], scaled("pg", vcc)),
                         MosWat(base_vt["pd"], scaled("pd", vcc)))
            for vcc in (.34, .36, .38, .40, .60, .90)
        ]
        analysis = analyze_rsnm_vcc_curve(points, self.cfg, fit_points=401)
        self.assertFalse(analysis["rows"][0]["valid_eye"])
        self.assertTrue(analysis["rows"][-1]["valid_eye"])
        self.assertIsNotNone(analysis["eye_closure"])
        self.assertGreater(analysis["eye_closure"]["estimated_vcc_v"], .36)
        self.assertLess(analysis["eye_closure"]["estimated_vcc_v"], .38)
        svg = rsnm_vcc_curve_svg(analysis)
        self.assertIn("Estimated RSNM versus Model VDD", svg)
        self.assertIn("Estimated eye-closure VDD", svg)
        self.assertIn("Read SNM (mV)", svg)
        self.assertIn('data-vdd-guide="0.38"', svg)
        self.assertIn(">0.38 V</text>", svg)
        self.assertEqual(svg.count("data-vdd-guide="),
                         sum(row["valid_eye"] for row in analysis["rows"]))
        self.assertNotIn(">VDD 0.38 V<", svg)
        self.assertNotIn(">RSNM ", svg)
        self.assertRegex(svg, r">\d+\.\d mV</text>")
        with tempfile.TemporaryDirectory() as td:
            report = write_rsnm_vcc_curve_outputs(analysis, Path(td) / "curve")
            self.assertTrue(report.exists())
            self.assertTrue((report.parent / "rsnm_vcc_curve.csv").exists())
            self.assertTrue((report.parent / "images" / "01_rsnm_vs_model_vcc.png").exists())
            self.assertIn("Estimated eye-closure VDD", report.read_text(encoding="utf-8"))

    def test_iv_curve_excel_extracts_idsat_at_model_vdd(self):
        from openpyxl import Workbook
        headers = ["Lot/Wafer", "Model VDD", "VDD Unit", "Vg", "Vg Unit", "Idsat", "Idsat Unit", "Vt", "Vt Unit"]
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "raw_iv.xlsx"
            book = Workbook()
            book.remove(book.active)
            for family, vt, multiplier in (("PU", .38, 10), ("PG", .36, 20), ("PD", .35, 30)):
                sheet = book.create_sheet(family)
                sheet.append(headers)
                for vdd in (.5, .6):
                    for vg in (.4, .5, .6):
                        sheet.append(["LOT_IV", vdd, "V", vg, "V", multiplier * vg, "uA", vt, "V"])
            book.save(path)
            lot, points, extraction = read_iv_curve_excel(path)
            self.assertEqual(lot, "LOT_IV")
            self.assertEqual(len(points), 2)
            at_six = next(point for point in points if point.vcc_v == .6)
            self.assertAlmostEqual(at_six.pu.ids, 6.0)
            self.assertAlmostEqual(at_six.pg.ids, 12.0)
            self.assertAlmostEqual(at_six.pd.ids, 18.0)
            self.assertEqual(len(extraction), 6)
            template = write_iv_curve_excel_template(Path(td) / "iv_template.xlsx")
            self.assertTrue(template.exists())

    def test_multi_chip_6t_excel_template_and_wafer_analysis(self):
        with tempfile.TemporaryDirectory() as td:
            template = write_multi_chip_6t_excel_template(Path(td) / "wafer.xlsx", chip_count=2)
            from openpyxl import load_workbook
            workbook = load_workbook(template)
            workbook["6T Multi-Cell"]["D2"] = -44.0  # Signed PMOS WAT current
            workbook["6T Multi-Cell"]["A3"] = "DEMO28_TT_W02"
            workbook.save(template)
            chips = read_multi_chip_6t_excel(template)
            self.assertEqual([item.chip_id for item in chips], ["CHIP_01", "CHIP_02"])
            self.assertEqual([item.lot_wafer for item in chips],
                             ["DEMO28_TT_W01", "DEMO28_TT_W02"])
            self.assertEqual(chips[0].raw_idsat_ua["pul"], -44.0)
            analysis = analyze_multi_chip_wafer(chips, Config(grid_points=101), fit_points=201)
            self.assertEqual(len(analysis["rows"]), 2)
            self.assertEqual(analysis["lot_wafers"],
                             ["DEMO28_TT_W01", "DEMO28_TT_W02"])
            self.assertEqual([row["lot_wafer"] for row in analysis["rows"]],
                             ["DEMO28_TT_W01", "DEMO28_TT_W02"])
            self.assertEqual(
                [sample["lot_wafer"] for sample in analysis["relative_shmoo"]["samples"]],
                ["DEMO28_TT_W01", "DEMO28_TT_W02"])
            self.assertGreater(analysis["worst_rsnm"]["rsnm_mv"], 0)
            self.assertGreater(analysis["worst_wsnm"]["wsnm_mv"], 0)
            self.assertEqual(analysis["median_cell"]["chip_id"], "MEDIAN_CELL")
            self.assertEqual(len(analysis["median_target_read_shmoo"]["rows"]), 66)
            self.assertEqual(len(analysis["median_target_write_shmoo"]["rows"]), 66)
            self.assertIn("relative_shmoo", analysis)
            self.assertEqual(analysis["relative_shmoo"]["samples"][0]["wafer_grade"],
                             "preferred")
            report = write_multi_chip_outputs(analysis, Path(td) / "batch", template)
            self.assertTrue(report.exists())
            self.assertTrue((report.parent / "images" / "01_multi_chip_read_vtc.png").exists())
            self.assertTrue((report.parent / "median_target_read_shmoo.csv").exists())
            self.assertTrue((report.parent / "images" /
                             "03_multi_cell_wafer_relative_shmoo.png").exists())
            self.assertTrue((report.parent /
                             "multi_cell_wafer_relative_grades.csv").exists())
            self.assertTrue((report.parent /
                             "multi_cell_wafer_distribution_statistics.csv").exists())
            self.assertTrue((report.parent / "imported_6t_vt_idsat_data.xlsx").exists())
            with (report.parent / "multi_chip_snm_summary.csv").open(
                    newline="", encoding="utf-8-sig") as stream:
                summary_rows = list(csv.DictReader(stream))
            self.assertEqual([row["lot_wafer"] for row in summary_rows],
                             ["DEMO28_TT_W01", "DEMO28_TT_W02"])
            self.assertIn("Minimum RSNM source 6T WAT values",
                          (report.parent / "images" / "01_multi_chip_read_vtc.svg").read_text(encoding="utf-8"))
            self.assertIn(">-44.00</text>",
                          (report.parent / "images" / "01_multi_chip_read_vtc.svg").read_text(encoding="utf-8"))

    def test_multi_cell_summary_builds_three_estimate_vmin_curves(self):
        with tempfile.TemporaryDirectory() as td:
            fields = ["lot_wafer", "chip_id", "model_vdd_v", "rsnm_mv", "wsnm_mv", "write_margin_mv"]
            paths = []
            for vdd, rsnm, wsnm, write_margin in ((.40, 0.0, 0.0, 0.0), (.60, 80.0, 50.0, 40.0)):
                path = Path(td) / f"summary_{vdd:.2f}.csv"
                with path.open("w", newline="", encoding="utf-8-sig") as stream:
                    writer = csv.DictWriter(stream, fieldnames=fields); writer.writeheader()
                    writer.writerow({"lot_wafer": "W01", "chip_id": "C01", "model_vdd_v": vdd,
                                     "rsnm_mv": rsnm + 5, "wsnm_mv": wsnm + 5, "write_margin_mv": write_margin + 5})
                    writer.writerow({"lot_wafer": "W01", "chip_id": "C02", "model_vdd_v": vdd,
                                     "rsnm_mv": rsnm, "wsnm_mv": wsnm, "write_margin_mv": write_margin})
                paths.append(path)
            result = analyze_estimate_vmin_curves(read_multi_chip_snm_summary(paths))
            self.assertEqual(set(result["curves"]), {"rsnm_mv", "wsnm_mv", "write_margin_mv"})
            self.assertEqual(result["curves"]["rsnm_mv"]["rows"][0]["chip_id"], "C02")
            self.assertAlmostEqual(result["curves"]["write_margin_mv"]["eye_closure"]["estimated_vdd_v"], .40)

    def test_estimate_vmin_summary_rejects_excel_with_clear_message(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "wrong_input.xlsx"
            path.write_bytes(b"PK\x03\x04\x98 binary Excel data")
            with self.assertRaisesRegex(ValueError, "select multi_chip_snm_summary.csv"):
                read_multi_chip_snm_summary([path])

    def test_estimate_vmin_marks_largest_rsnm_slope_and_renders_stacked_view(self):
        rows = []
        for vdd, rsnm in ((.40, 10.0), (.50, 90.0), (.80, 120.0)):
            row = {"vdd_v": vdd, "sample_count": 1}
            for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
                row[key] = rsnm
                row[f"{key}_chip_id"] = "C01"
                row[f"{key}_lot_wafer"] = "W01"
            rows.append(row)
        result = analyze_estimate_vmin_curves(rows)
        svg = estimate_vmin_curve_svg(result["curves"]["rsnm_mv"])
        self.assertIn("Largest RSNM slope", svg)
        self.assertIn("0.50 V", svg)
        stacked = estimate_vmin_stacked_svg(result)
        self.assertIn("Estimate Vmin Curves - Comparison View", stacked)
        self.assertIn("BL Write Margin", stacked)
        self.assertNotIn("WL Write Margin", stacked)
        self.assertIn("SNM (mV)", stacked)
        self.assertIn("Vtrip (mV)", stacked)
        self.assertIn('class="curve-data-label"', stacked)
        self.assertIn('>90.0 mV</text>', stacked)
        self.assertIn('class="vertical-vdd-label"', stacked)
        self.assertIn('>0.50 V</text>', stacked)
        self.assertIn('Left: Read / Write SNM', stacked)
        self.assertEqual(stacked.count('>Model VDD (V)</text>'), 2)
        self.assertEqual(stacked.count('class="measured-vdd-guide"'), 6)
        self.assertIn('stroke="#8E8E93" stroke-width="1.4" stroke-dasharray="5 5"', stacked)
        self.assertNotIn('data-extrapolated-to-zero="true"', stacked)
        self.assertNotIn("Largest RSNM slope:", stacked)
        transparent = estimate_vmin_stacked_svg(result, transparent_background=True)
        self.assertNotIn('<rect width="100%" height="100%" fill="#FFFFFF"/>', transparent)

    def test_combined_summary_comparison_reads_two_files(self):
        with tempfile.TemporaryDirectory() as td:
            paths = []
            fields = ["vdd_v", "sample_count", "rsnm_mv", "rsnm_mv_lot_wafer",
                      "rsnm_mv_chip_id", "wsnm_mv", "wsnm_mv_lot_wafer",
                      "wsnm_mv_chip_id", "write_margin_mv",
                      "write_margin_mv_lot_wafer", "write_margin_mv_chip_id"]
            for folder_index, lot in enumerate(("LOT_A_W01", "LOT_B_W02")):
                folder = Path(td) / f"output_{folder_index+1}"
                folder.mkdir()
                path = folder / "multi_chip_snm_summary_combined.csv"
                with path.open(
                        "w", newline="", encoding="utf-8-sig") as stream:
                    writer = csv.DictWriter(stream, fieldnames=fields)
                    writer.writeheader()
                    for vdd, base in ((.5, 50.0), (.7, 90.0)):
                        writer.writerow({"vdd_v": vdd, "sample_count": 8,
                                         "rsnm_mv": base + folder_index * 10,
                                         "rsnm_mv_lot_wafer": lot,
                                         "rsnm_mv_chip_id": "C01",
                                         "wsnm_mv": base + 20 + folder_index * 10,
                                         "wsnm_mv_lot_wafer": lot,
                                         "wsnm_mv_chip_id": "C02",
                                         "write_margin_mv": base / 2 + folder_index * 5,
                                         "write_margin_mv_lot_wafer": lot,
                                         "write_margin_mv_chip_id": "C03"})
                paths.append(path)
            datasets = read_estimate_vmin_combined_files(paths)
            self.assertEqual([item["lot_wafer"] for item in datasets],
                             ["LOT_A_W01", "LOT_B_W02"])
            self.assertEqual(len(datasets[0]["rows"]), 2)
            svg = estimate_vmin_combined_comparison_svg(datasets)
            self.assertIn("Estimate Vmin Curves - Comparison View", svg)
            self.assertIn("LOT_A_W01", svg)
            self.assertIn("LOT_B_W02", svg)
            self.assertIn("Read SNM", svg)
            self.assertIn("Write SNM", svg)
            self.assertIn("BL Write Margin", svg)
            self.assertIn('class="multi-lot-data-label"', svg)
            self.assertIn('<text x="56" y="46" fill="#1D1D1F" font-size="34"', svg)
            self.assertEqual(svg.count('y="156.0" fill="#1D1D1F" font-size="20"'), 2)
            self.assertEqual(svg.count('>Model VDD (V)</text>'), 2)
            self.assertIn('opacity=".42"', svg)
            self.assertIn('>50.0 mV</text>', svg)
            self.assertNotIn('>R 50.0', svg)
            self.assertNotIn('>W 70.0', svg)
            self.assertNotIn('>BL 25.0', svg)
            self.assertIn('stroke="#8E8E93" stroke-width="1.4" stroke-dasharray="5 5"', svg)
            self.assertEqual(svg.count('class="measured-vdd-guide"'), 4)
            self.assertEqual(svg.count('>0.50 V</text>'), 2)
            self.assertIn('stroke-dasharray="8 5"', svg)
            transparent = estimate_vmin_combined_comparison_svg(
                datasets, transparent_background=True)
            self.assertNotIn('<rect width="100%" height="100%" fill="#FFFFFF"/>',
                             transparent)
            self.assertIn('fill="none" stroke="#D8DDE3"', transparent)

    def test_estimate_vmin_extrapolates_two_lowest_vdd_points_to_zero(self):
        rows = []
        for vdd, margin in ((.60, 40.0), (.80, 80.0), (1.00, 95.0)):
            row = {"vdd_v": vdd, "sample_count": 1}
            for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
                row[key] = margin
                row[f"{key}_chip_id"] = "C01"
                row[f"{key}_lot_wafer"] = "W01"
            rows.append(row)
        result = analyze_estimate_vmin_curves(rows)
        closure = result["curves"]["rsnm_mv"]["eye_closure"]
        self.assertTrue(closure["extrapolated"])
        self.assertAlmostEqual(closure["slope_mv_per_v"], 200.0)
        self.assertAlmostEqual(closure["estimated_vdd_v"], .40)
        self.assertEqual((closure["low_vdd_v"], closure["high_vdd_v"]), (.60, .80))
        svg = estimate_vmin_curve_svg(result["curves"]["rsnm_mv"])
        self.assertIn("Extrapolated eye-closure VDD 0.4000 V", svg)
        self.assertIn("Two-lowest-VDD slope: 200.00 mV/V", svg)
        self.assertIn('data-extrapolated-to-zero="true"', svg)
        self.assertIn('stroke-dasharray="8 6"', svg)

    def test_estimate_vmin_rejects_nonpositive_low_vdd_slope(self):
        rows = []
        for vdd, margin in ((.60, 80.0), (.80, 60.0), (1.00, 90.0)):
            row = {"vdd_v": vdd, "sample_count": 1}
            for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
                row[key] = margin
                row[f"{key}_chip_id"] = "C01"
                row[f"{key}_lot_wafer"] = "W01"
            rows.append(row)
        result = analyze_estimate_vmin_curves(rows)
        self.assertIsNone(result["curves"]["rsnm_mv"]["eye_closure"])

    def test_estimate_vmin_builds_per_vdd_cr_pr_and_family_wat_shmoo(self):
        rows = []
        for vdd in (.60, .80):
            samples = []
            for index, scale in enumerate((.75, 1.0), 1):
                samples.append({
                    "lot_wafer": "W01", "chip_id": f"C{index:02d}",
                    "rsnm_mv": 80*scale, "wsnm_mv": 65*scale,
                    "write_margin_mv": 55*scale,
                    "cell_ratio_beta": 1.2+index*.1, "pull_up_ratio_beta": 1.5+index*.1,
                    "pu_vt_v": .38, "pu_idsat_ua": 44*scale,
                    "pg_vt_v": .36, "pg_idsat_ua": 82*scale,
                    "pd_vt_v": .35, "pd_idsat_ua": 124*scale,
                })
            row = {"vdd_v": vdd, "sample_count": 2, "samples": samples}
            for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
                row[key] = min(item[key] for item in samples)
                row[f"{key}_chip_id"] = "C01"; row[f"{key}_lot_wafer"] = "W01"
            rows.append(row)
        analysis = analyze_estimate_vmin_curves(rows)
        self.assertEqual(len(analysis["ratio_shmoos"]), 2)
        shmoo = analysis["ratio_shmoos"][0]
        self.assertTrue(shmoo["has_family_wat"])
        self.assertFalse(shmoo["samples"][0]["best_region"])
        self.assertTrue(shmoo["samples"][1]["best_region"])
        self.assertEqual(shmoo["best"]["chip_id"], "C02")
        self.assertAlmostEqual(shmoo["samples"][0]["read_score"], .75)
        self.assertAlmostEqual(shmoo["samples"][0]["write_score"], .75)
        self.assertAlmostEqual(shmoo["samples"][0]["delta_vs_median_pu_idsat_ua_pct"],
                               -14.285714, places=5)
        self.assertAlmostEqual(shmoo["samples"][1]["write_balance_vs_median_pct"],
                               100 * (1.7 - 1.65) / 1.65, places=5)
        self.assertEqual(shmoo["samples"][0]["wafer_grade"], "monitor")
        self.assertEqual(shmoo["samples"][1]["wafer_grade"], "preferred")
        self.assertAlmostEqual(shmoo["samples"][0]["wafer_grade_score"], .25)
        self.assertAlmostEqual(shmoo["samples"][1]["wafer_grade_score"], .75)
        self.assertAlmostEqual(shmoo["distributions"]["rsnm_mv"]["median"], 70.0)
        svg = estimate_vmin_ratio_shmoo_svg(shmoo)
        self.assertAlmostEqual(shmoo["samples"][1]["pull_up_ratio_beta"], 1.7)
        self.assertIn("Read / Write Drive-Balance Shmoo", svg)
        self.assertIn("HOW TO READ", svg)
        self.assertIn("Wafer-relative grade", svg)
        self.assertIn("Dynamic wafer thresholds", svg)
        self.assertIn("COLOR SCALE", svg)
        self.assertNotIn("Preferred drive direction", svg)
        self.assertIn("Weakest: C01", svg)
        self.assertIn("Best measured: C02", svg)
        self.assertIn("Best measured cell", svg)
        self.assertIn("#FFC447", svg)
        self.assertIn("P50/P50 relative boundary", svg)
        self.assertIn('font-family="Times New Roman">MOS</tspan>', svg)
        self.assertIn('baseline-shift="sub"', svg)
        self.assertIn("drive</tspan>", svg)
        self.assertIn("right = easier write", svg)
        self.assertNotIn("1/PR", svg)
        self.assertIn("#DDF3E2", svg)
        self.assertIn("#FFF0C2", svg)
        self.assertIn("#F7C9C2", svg)
        self.assertIn('class="measured-cell"', svg)
        self.assertIn('data-cell-tooltip=', svg)
        self.assertIn('aria-label="Cell: C01', svg)
        self.assertNotIn("<title>", svg)
        self.assertIn('class="special-cell-highlights" pointer-events="none"', svg)
        self.assertIn("Cell: C01 (1/2)", svg)
        self.assertIn("RSNM: 60.0 mV", svg)
        self.assertIn("BL Write Vtrip: 41.2 mV", svg)
        self.assertIn("Read balance vs median:", svg)
        self.assertIn("Write balance vs median:", svg)
        self.assertIn("Δmed -14.3%", svg)
        self.assertIn("PU: Vt", svg)
        self.assertIn("PG: Vt", svg)
        self.assertIn("PD: Vt", svg)

    def test_drive_to_preferred_advisor_uses_same_vdd_p55_and_holds_pg(self):
        samples = []
        for index, ratio in enumerate((1.0, 1.2, 1.4, 1.6, 1.8), 1):
            samples.append({
                "lot_wafer": "W01", "chip_id": f"C{index:02d}",
                "rsnm_mv": 40.0 + index, "wsnm_mv": 30.0 + index,
                "write_margin_mv": 20.0 + index,
                "cell_ratio_beta": ratio,
                "pull_up_ratio_beta": ratio,
                "pu_vt_v": .38, "pu_idsat_ua": 50.0,
                "pg_vt_v": .36, "pg_idsat_ua": 80.0,
                "pd_vt_v": .35, "pd_idsat_ua": 110.0,
            })
        row = {"vdd_v": .68, "sample_count": len(samples), "samples": samples}
        for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
            row[key] = min(sample[key] for sample in samples)
            row[f"{key}_chip_id"] = "C01"
            row[f"{key}_lot_wafer"] = "W01"
        second = {**row, "vdd_v": .70,
                  "samples": [dict(sample) for sample in samples]}
        shmoo = analyze_estimate_vmin_curves([row, second])["ratio_shmoos"][0]
        advice = build_drive_to_preferred_advice(shmoo, "C01", .55, "W01")
        self.assertAlmostEqual(advice["target"]["cr"], 1.44)
        self.assertAlmostEqual(advice["target"]["pr"], 1.44)
        devices = {item["family"]: item for item in advice["devices"]}
        self.assertAlmostEqual(devices["PG"]["beta_change_pct"], 0.0)
        self.assertGreater(devices["PD"]["beta_change_pct"], 0.0)
        self.assertLess(devices["PU"]["beta_change_pct"], 0.0)
        self.assertAlmostEqual(devices["PD"]["idsat_target_fixed_vt_ua"], 158.4)
        self.assertAlmostEqual(devices["PU"]["idsat_target_fixed_vt_ua"],
                               50.0 / 1.44)
        self.assertEqual(advice["predicted"]["grade"], "preferred")
        batch = build_batch_drive_to_preferred_advice(shmoo, .55)
        self.assertEqual(batch["affected_count"], 2)
        self.assertAlmostEqual(batch["devices"][0]["drive_multiplier"], 1.44)
        self.assertAlmostEqual(batch["devices"][1]["drive_multiplier"], 1.0)
        self.assertAlmostEqual(batch["devices"][2]["drive_multiplier"], 1.0 / 1.44)
        self.assertAlmostEqual(batch["affected_coverage_after_pct"], 100.0)
        self.assertIn("preserves CR/PR rank ordering", batch["caution"])
        section, rows = _drive_advisor_html(shmoo, 1)
        self.assertIn("Drive-to-Preferred Advisor", section)
        self.assertIn("P55 guardband", section)
        self.assertNotIn("Low / Monitor Batch Adjustment", section)
        self.assertIn("MOS<sub>drive</sub>", section)
        self.assertEqual(len(rows), 15)

    def test_lot_wafer_advisor_groups_names_and_exports_boxplots(self):
        samples = []
        for lot, ratios, base in (
                ("LOT_A_W01", (1.0, 1.1, 1.2, 1.3), 50.0),
                ("LOT_B_W02", (1.5, 1.6, 1.7, 1.8), 80.0)):
            for index, ratio in enumerate(ratios, 1):
                samples.append({
                    "lot_wafer": lot, "chip_id": f"C{index:02d}",
                    "rsnm_mv": base + index, "wsnm_mv": base*.7 + index,
                    "write_margin_mv": base*.6 + index,
                    "cell_ratio_beta": ratio, "pull_up_ratio_beta": ratio,
                    "pu_vt_v": .38, "pu_idsat_ua": 50.0,
                    "pg_vt_v": .36, "pg_idsat_ua": 80.0,
                    "pd_vt_v": .35, "pd_idsat_ua": 110.0,
                })
        row = {"vdd_v": .68, "sample_count": len(samples), "samples": samples}
        for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
            winner = min(samples, key=lambda sample: sample[key])
            row[key] = winner[key]
            row[f"{key}_chip_id"] = winner["chip_id"]
            row[f"{key}_lot_wafer"] = winner["lot_wafer"]
        result = analyze_lot_wafer_drive_advisor([row])
        self.assertEqual(result["lot_wafers"], ["LOT_A_W01", "LOT_B_W02"])
        vdd = result["vdds"][0]
        self.assertEqual(vdd["lot_count"], 2)
        group_a = vdd["groups"][0]
        self.assertEqual(group_a["sample_count"], 4)
        self.assertAlmostEqual(group_a["metrics"]["rsnm_mv"]["median"], 52.5)
        self.assertGreater(group_a["batch_advice"]["devices"][0]["drive_change_pct"], 0)
        self.assertEqual(group_a["batch_advice"]["devices"][1]["drive_change_pct"], 0)
        scatter = lot_wafer_drive_scatter_svg(vdd, result["styles"])
        boxes = lot_wafer_boxplot_svg(vdd, result["styles"])
        grades = lot_wafer_grade_counts_svg(vdd)
        self.assertIn("LOT_A_W01", scatter)
        self.assertIn("central 50%", scatter)
        self.assertIn("Read SNM", boxes)
        self.assertIn("Balanced Drive Score", boxes)
        self.assertIn("Preferred, Monitor and Low Counts", grades)
        self.assertIn("labels show Cell count", grades)
        self.assertIn("LOT_A_W01", grades)
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "multi_chip_snm_summary.csv"
            source.write_text("QA source", encoding="utf-8")
            report = write_lot_wafer_drive_advisor_outputs(
                result, root / "advisor", [source])
            text = report.read_text(encoding="utf-8")
            self.assertIn("Lot/Wafer Drive Advisor", text)
            self.assertIn("All-Cell Preferred / Monitor / Low Shmoo", text)
            self.assertIn("Drive-to-Preferred Batch Sensitivity", text)
            self.assertTrue((report.parent / "lot_wafer_distribution_statistics.csv").exists())
            self.assertTrue((report.parent / "lot_wafer_cell_drive_scores.csv").exists())
            self.assertTrue((report.parent / "lot_wafer_batch_drive_advisor.csv").exists())
            self.assertTrue(any((report.parent / "images").glob("*_lot_wafer_boxplots.png")))
            self.assertTrue(any((report.parent / "images").glob("*_lot_wafer_grade_counts.png")))
            self.assertTrue(any((report.parent / "images").glob(
                "*_all_cell_drive_balance_shmoo.png")))

    def test_single_vdd_multi_cell_summary_outputs_shmoo_only(self):
        fieldnames = [
            "lot_wafer", "chip_id", "model_vdd_v", "rsnm_mv", "wsnm_mv",
            "write_margin_mv", "cell_ratio_beta", "pull_up_ratio_beta",
            "pu_vt_v", "pu_idsat_ua", "pg_vt_v", "pg_idsat_ua",
            "pd_vt_v", "pd_idsat_ua",
        ]
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "multi_chip_snm_summary.csv"
            with source.open("w", newline="", encoding="utf-8-sig") as stream:
                writer = csv.DictWriter(stream, fieldnames=fieldnames)
                writer.writeheader()
                for index, ratio in enumerate((1.1, 1.3, 1.5, 1.7), 1):
                    writer.writerow({
                        "lot_wafer": "W01", "chip_id": f"C{index:02d}",
                        "model_vdd_v": .68, "rsnm_mv": 50 + index,
                        "wsnm_mv": 40 + index, "write_margin_mv": 30 + index,
                        "cell_ratio_beta": ratio,
                        "pull_up_ratio_beta": ratio + .2,
                        "pu_vt_v": .38, "pu_idsat_ua": 50,
                        "pg_vt_v": .36, "pg_idsat_ua": 80,
                        "pd_vt_v": .35, "pd_idsat_ua": 110,
                    })
            rows = read_multi_chip_snm_summary([source])
            self.assertEqual(len(rows), 1)
            analysis = analyze_estimate_vmin_curves(rows)
            self.assertEqual(analysis["mode"], "shmoo_only")
            report = write_estimate_vmin_outputs(analysis, root / "result", [source])
            report_text = report.read_text(encoding="utf-8")
            self.assertIn("Shmoo-only mode", report_text)
            self.assertIn("Drive-to-Preferred Advisor", report_text)
            self.assertTrue((report.parent / "estimate_vmin_cr_pr_shmoo.csv").exists())
            self.assertTrue((report.parent / "estimate_vmin_drive_to_preferred_advisor.csv").exists())
            self.assertFalse((report.parent / "estimate_vmin_batch_drive_advisor.csv").exists())
            self.assertFalse((report.parent / "multi_chip_snm_summary_combined.csv").exists())
            self.assertFalse((report.parent / "images" / "05_estimate_vmin_stacked.png").exists())
            self.assertFalse((report.parent / "images" / "01_rsnm_mv_estimate_vmin.png").exists())

    def test_single_source_can_force_multi_vdd_rows_to_shmoo_only(self):
        rows = []
        for vdd in (.68, .80):
            samples = [{
                "lot_wafer": "W01", "chip_id": "C01", "rsnm_mv": 60.0,
                "wsnm_mv": 50.0, "write_margin_mv": 40.0,
                "cell_ratio_beta": 1.4, "pull_up_ratio_beta": 1.6,
            }]
            row = {"vdd_v": vdd, "sample_count": 1, "samples": samples}
            for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
                row[key] = samples[0][key]
                row[f"{key}_chip_id"] = "C01"
                row[f"{key}_lot_wafer"] = "W01"
            rows.append(row)
        analysis = analyze_estimate_vmin_curves(rows, force_shmoo_only=True)
        self.assertEqual(analysis["mode"], "shmoo_only")
        self.assertEqual(len(analysis["ratio_shmoos"]), 2)

    def test_estimate_vmin_balance_does_not_reward_large_residual_wsnm(self):
        samples = [
            {"lot_wafer": "W01", "chip_id": "GOOD", "rsnm_mv": 90.0,
             "wsnm_mv": 10.0, "write_margin_mv": 80.0,
             "cell_ratio_beta": 1.5, "pull_up_ratio_beta": 1.8},
            {"lot_wafer": "W01", "chip_id": "RESIDUAL_EYE", "rsnm_mv": 90.0,
             "wsnm_mv": 999.0, "write_margin_mv": 20.0,
             "cell_ratio_beta": 1.4, "pull_up_ratio_beta": 1.6},
        ]
        row = {"vdd_v": .8, "sample_count": 2, "samples": samples}
        for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
            row[key] = min(float(item.get(key, item["write_margin_mv"])) for item in samples)
            row[f"{key}_chip_id"] = "GOOD"
            row[f"{key}_lot_wafer"] = "W01"
        second_row = {**row, "vdd_v": .9, "samples": [dict(item) for item in samples]}
        shmoo = analyze_estimate_vmin_curves([row, second_row])["ratio_shmoos"][0]
        by_chip = {sample["chip_id"]: sample for sample in shmoo["samples"]}
        self.assertTrue(by_chip["GOOD"]["best_region"])
        self.assertFalse(by_chip["RESIDUAL_EYE"]["best_region"])
        self.assertAlmostEqual(by_chip["RESIDUAL_EYE"]["balanced_score"], .25)

    def test_wafer_relative_grade_follows_both_cr_and_pr_quartiles(self):
        samples = []
        for index, (ratio, margin) in enumerate(
                ((1.0, 100.0), (1.2, 80.0), (1.4, 60.0), (1.6, 40.0)), 1):
            samples.append({
                "lot_wafer": "W01", "chip_id": f"C{index:02d}",
                "rsnm_mv": margin, "wsnm_mv": margin,
                "write_margin_mv": margin,
                "cell_ratio_beta": ratio, "pull_up_ratio_beta": ratio,
            })
        base = {"vdd_v": .68, "sample_count": len(samples), "samples": samples}
        for key in ("rsnm_mv", "wsnm_mv", "write_margin_mv"):
            base[key] = min(sample[key] for sample in samples)
            base[f"{key}_chip_id"] = "C04"
            base[f"{key}_lot_wafer"] = "W01"
        second = {**base, "vdd_v": .70, "samples": [dict(item) for item in samples]}
        shmoo = analyze_estimate_vmin_curves([base, second])["ratio_shmoos"][0]
        by_chip = {sample["chip_id"]: sample for sample in shmoo["samples"]}
        self.assertEqual(by_chip["C01"]["wafer_grade"], "low")
        self.assertEqual(by_chip["C04"]["wafer_grade"], "preferred")
        self.assertGreater(by_chip["C01"]["performance_grade_score"],
                           by_chip["C01"]["wafer_grade_score"])
        self.assertAlmostEqual(shmoo["target"]["cell_ratio_beta"], 1.3)
        self.assertAlmostEqual(shmoo["target"]["pull_up_ratio_beta"], 1.3)

    def test_single_and_multi_chip_use_identical_snm_calculation(self):
        """One multi-chip row must numerically match the 6T single-cell result."""
        cfg = Config(nominal_vdd=.90, wat_vdd=.90, grid_points=101)
        pu, pg, pd = MosWat(.385, 44.0), MosWat(.365, 82.0), MosWat(.355, 124.0)
        cell = SixTWatCell("SYNC", pu, pu, pg, pg, pd, pd)
        single = analyze_six_mos(cell, cfg)["baseline_6t"]["metrics"]
        multi = analyze_multi_chip_wafer(
            [WaferChipWat("SYNC", "CHIP_01", .90, cell)], cfg)["rows"][0]
        self.assertAlmostEqual(multi["rsnm_mv"], single["read_snm_mv"], places=9)
        self.assertAlmostEqual(multi["wsnm_mv"], single["write_snm_mv"], places=9)

    def test_multi_chip_read_chart_uses_independent_upper_and_lower_limits(self):
        """The summary card must show each state limit; squares stay text-free."""
        cfg = Config(nominal_vdd=.90, wat_vdd=.90)
        pu, pg, pd = MosWat(.385, 44.0), MosWat(.365, 82.0), MosWat(.355, 124.0)
        cell = SixTWatCell("SYNC", pu, pu, pg, pg, pd, pd)
        chips = [WaferChipWat("SYNC", "CHIP_A", .90, cell),
                 WaferChipWat("SYNC", "CHIP_B", .90, cell)]
        analysis = analyze_multi_chip_wafer(chips, cfg)
        # Deliberately select different rows to verify that the SVG uses the
        # state-specific references rather than the cell-RSNM winner twice.
        upper, lower = analysis["rows"]
        analysis["worst_rsnm_upper"] = upper
        analysis["worst_rsnm_lower"] = lower
        analysis["worst_rsnm"] = lower
        svg = multi_chip_vtc_svg(analysis, "read")
        self.assertIn(">Upper minimum</text>", svg)
        self.assertIn(">Lower minimum</text>", svg)
        self.assertIn(f'>{upper["upper_rsnm_mv"]:.1f} mV · SYNC / CHIP_A</text>', svg)
        self.assertIn(f'>{lower["lower_rsnm_mv"]:.1f} mV · SYNC / CHIP_B</text>', svg)
        self.assertNotIn("Upper minimum 204.0 mV", svg)

    def test_rsnm_vdd_row_matches_main_symmetric_6t_analysis(self):
        """The curve and primary report must share one RSNM calculation path."""
        pu, pg, pd = MosWat(.385, 44.0), MosWat(.365, 82.0), MosWat(.355, 124.0)
        cfg = Config(nominal_vdd=.90, wat_vdd=.90, grid_points=401)
        cell = SixTWatCell("SYNC", pu, pu, pg, pg, pd, pd)
        main_rsnm = analyze_six_mos(cell, cfg)["baseline_6t"]["metrics"]["read_snm_mv"]
        curve = analyze_rsnm_vcc_curve([
            RsnmVccPoint(.80, pu, pg, pd),
            RsnmVccPoint(.90, pu, pg, pd),
        ], cfg)
        at_ninety = next(row for row in curve["rows"] if row["vcc_v"] == .90)
        self.assertTrue(at_ninety["valid_eye"])
        self.assertAlmostEqual(at_ninety["rsnm_mv"], main_rsnm, places=8)

    def test_write_trip_margin_curve_brackets_boundary_and_exports(self):
        base_vt = {"pu": .385, "pg": .365, "pd": .355}
        base_ids = {"pu": 44.0, "pg": 82.0, "pd": 124.0}

        def scaled(kind, vdd):
            return base_ids[kind] * (max(vdd - base_vt[kind], 0) /
                                     (.9 - base_vt[kind])) ** 2

        points = [
            RsnmVccPoint(vdd,
                         MosWat(base_vt["pu"], scaled("pu", vdd)),
                         MosWat(base_vt["pg"], scaled("pg", vdd)),
                         MosWat(base_vt["pd"], scaled("pd", vdd)))
            for vdd in (.34, .38, .40, .50, .90)
        ]
        analysis = analyze_write_trip_margin_curve(points, self.cfg, fit_points=401)
        self.assertFalse(analysis["rows"][0]["writable"])
        self.assertTrue(analysis["rows"][-1]["writable"])
        self.assertIsNotNone(analysis["write_boundary"])
        self.assertGreater(analysis["write_boundary"]["estimated_vdd_v"], .38)
        self.assertLess(analysis["write_boundary"]["estimated_vdd_v"], .40)
        svg = write_trip_margin_curve_svg(analysis)
        self.assertIn("Estimated Write Trip Margin versus Model VDD", svg)
        self.assertIn("Write Trip Margin (mV)", svg)
        self.assertIn("Estimated write boundary VDD", svg)
        with tempfile.TemporaryDirectory() as td:
            report = write_write_trip_margin_outputs(analysis, Path(td) / "wtm")
            self.assertTrue(report.exists())
            self.assertTrue((report.parent / "write_trip_margin_curve.csv").exists())
            self.assertTrue((report.parent / "images" /
                             "01_write_trip_margin_vs_model_vdd.png").exists())

if __name__ == "__main__":
    unittest.main()
