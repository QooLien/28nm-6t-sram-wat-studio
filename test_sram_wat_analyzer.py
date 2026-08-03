import csv
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest import mock

from sram_wat_analyzer import (
    AsymmetricSram6T, Config, DatasheetTargets, MosWat, RsnmVccPoint,
    SixTWatCell, Sram6T, ThreeTWatCell,
    WatPoint, analyze, analyze_six_mos, analyze_three_mos,
    _read_wat_excel_rows, analyze_mismatch_rsnm_boundaries, analyze_rsnm_vcc_curve,
    analyze_write_trip_margin_curve,
    generic_28nm_assumption_rows,
    create_run_output_dir, load_gui_state, model_vdd_butterfly_svg, open_output_directory,
    read_wat_csv, rsnm_vcc_curve_svg,
    save_gui_state,
    validate_config, wat_electrical_snm_rows, write_mismatch_boundary_outputs,
    write_outputs, write_rsnm_vcc_curve_outputs,
    write_trip_margin_curve_svg, write_write_trip_margin_outputs,
)


class AnalyzerTests(unittest.TestCase):
    def setUp(self):
        self.cfg = Config(grid_points=101)
        self.targets = DatasheetTargets(MosWat(.380, 45), MosWat(.370, 80), MosWat(.360, 120))
        self.cell = ThreeTWatCell("LOT_W01", MosWat(.385, 44), MosWat(.365, 82), MosWat(.355, 124))

    def test_run_output_directory_uses_date_time_wafer_and_never_overwrites(self):
        stamp = datetime(2026, 8, 1, 14, 30, 25)
        with tempfile.TemporaryDirectory() as td:
            first = create_run_output_dir(td, "LOT/W01:17", "6t analysis", stamp)
            second = create_run_output_dir(td, "LOT/W01:17", "6t analysis", stamp)
            self.assertEqual(first.relative_to(td).parts,
                             ("2026-08-01", "LOT_W01_17", "143025_6t_analysis"))
            self.assertEqual(second.name, "143025_6t_analysis_02")
            manifest = (first / "run_info.json").read_text(encoding="utf-8")
            self.assertIn('"wafer_id": "LOT/W01:17"', manifest)
            self.assertIn('"created_local": "2026-08-01T14:30:25"', manifest)

    def test_open_output_directory_creates_and_opens_selected_folder(self):
        with tempfile.TemporaryDirectory() as td:
            target = Path(td) / "new output"
            with mock.patch("sram_wat_analyzer.sys.platform", "win32"), \
                    mock.patch("sram_wat_analyzer.os.startfile") as startfile:
                opened = open_output_directory(target)
            self.assertEqual(opened, target.resolve())
            self.assertTrue(target.is_dir())
            startfile.assert_called_once_with(str(target.resolve()))

    def test_open_output_directory_rejects_blank_path(self):
        with self.assertRaisesRegex(ValueError, "Choose an output folder"):
            open_output_directory("  ")

    def test_read_snm_is_bounded(self):
        model = Sram6T(self.cell.representative(), self.cfg)
        result = model.butterfly_squares(self.cfg.nominal_vdd, "read", points=601)
        self.assertTrue(result["valid"])
        self.assertEqual(len(result["squares"]), 2)
        value = result["snm_v"]
        self.assertGreaterEqual(value, 0)
        self.assertLessEqual(value, self.cfg.nominal_vdd / 2)

    def test_figure_3_15_geometric_read_snm(self):
        wat = WatPoint("CURRENT", .35, 29.2, .27, 40.3, .27, 47.6)
        result = Sram6T(wat, self.cfg).butterfly_squares(.9, "read", points=1201)
        self.assertTrue(result["valid"])
        self.assertAlmostEqual(result["snm_mv"], 156.0, delta=1.0)
        self.assertAlmostEqual(result["squares"][0]["side_mv"],
                               result["squares"][1]["side_mv"], delta=1.0)

    def test_metrics_are_read_snm_only(self):
        metrics = analyze(self.cell.representative(), self.cfg)["baseline_6t"]["metrics"]
        self.assertNotIn("hold_snm_mv", metrics)
        self.assertIn("read_snm_mv", metrics)
        self.assertNotIn("write_snm_mv", metrics)
        self.assertNotIn("read_vmin_v", metrics)
        self.assertNotIn("write_vmin_v", metrics)

    def test_target_model_and_delta(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        self.assertIn("target_6t", result)
        self.assertEqual([row["mode"] for row in result["snm_target_comparison"]],
                         ["Read SNM"])
        self.assertTrue(any(abs(row["delta_mv"]) > 0 for row in result["snm_target_comparison"]))

    def test_wat_target_reference_can_be_disabled(self):
        result = analyze_six_mos(self.cell.to_six_t(), self.cfg, None)
        self.assertIsNone(result["datasheet_targets"])
        self.assertNotIn("target_6t", result)
        self.assertEqual(result["snm_target_comparison"], [])
        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "output"
            report = write_outputs(result, output)
            overview = (output / "images" / "01_read_snm_target_comparison.svg").read_text(
                encoding="utf-8")
            write_svg = (output / "images" / "03_write_snm_vs_bitline.svg").read_text(
                encoding="utf-8")
            html_text = report.read_text(encoding="utf-8")
            self.assertIn("Read SNM Analysis", overview)
            self.assertNotIn("WAT Target VTC", overview)
            self.assertNotIn("WAT Target model", write_svg)
            self.assertIn("WAT Target reference is disabled", html_text)
            self.assertFalse((output / "snm_target_comparison.csv").exists())
            self.assertFalse((output / "wat_target_comparison.csv").exists())
            with (output / "write_snm_vs_bitline.csv").open(
                    newline="", encoding="utf-8-sig") as source:
                self.assertEqual({row["dataset"] for row in csv.DictReader(source)},
                                 {"Lot/Wafer"})

    def test_textbook_write_snm_sweep_closes_and_is_monotonic(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        sweep = result["baseline_6t"]["write_snm_vs_bitline"]
        self.assertEqual(sweep["status"], "VALID")
        self.assertGreater(sweep["write_trip_bl_v"], 0.0)
        self.assertLess(sweep["write_trip_bl_v"], self.cfg.nominal_vdd)
        self.assertAlmostEqual(
            sweep["required_bl_swing_v"],
            self.cfg.nominal_vdd - sweep["write_trip_bl_v"], places=10)
        values = [row["cell_write_snm_mv"] for row in sweep["points"]]
        self.assertTrue(all(a <= b + 1e-9 for a, b in zip(values, values[1:])))
        self.assertEqual(values[0], 0.0)
        self.assertGreater(values[-1], 0.0)

    def test_single_wat_write_snm_geometry_uses_vtc_diagonal_crossing(self):
        result = analyze_six_mos(self.cell.to_six_t(), self.cfg, self.targets)
        geometry = result["baseline_6t"]["single_wat_write_snm_geometry"]
        self.assertTrue(geometry["valid"])
        self.assertGreater(geometry["wsnm_v"], 0.0)
        self.assertLess(geometry["wsnm_v"], self.cfg.nominal_vdd)
        self.assertAlmostEqual(geometry["intersection"][0], geometry["intersection"][1], places=10)
        self.assertAlmostEqual(geometry["wsnm_mv"], geometry["wsnm_v"] * 1000.0, places=8)
        polarity_values = [row["wsnm_v"] for row in geometry["polarity_results"]
                           if row["wsnm_v"] is not None]
        self.assertAlmostEqual(geometry["wsnm_v"], min(polarity_values), places=10)

    def test_pdf_equation_3_36_with_given_wat_values(self):
        current = WatPoint("CURRENT", .35, 29.2, .27, 40.3, .27, 47.6)
        target = WatPoint("TARGET", .33, 19.5, .28, 39.0, .29, 44.9)
        current_eq = Sram6T(current, self.cfg).analytical_read_snm_eq_3_36(.9)
        target_eq = Sram6T(target, self.cfg).analytical_read_snm_eq_3_36(.9)
        self.assertTrue(current_eq["valid"])
        self.assertTrue(target_eq["valid"])
        self.assertAlmostEqual(current_eq["vth_eff_v"], (.35 + .27 + .27) / 3)
        self.assertAlmostEqual(current_eq["snm_mv"], 189.7967615, places=5)
        self.assertAlmostEqual(target_eq["snm_mv"], 183.5971209, places=5)

    def test_pdf_equation_3_36_reports_domain_failure(self):
        equation = Sram6T(WatPoint(), self.cfg).analytical_read_snm_eq_3_36(.9)
        self.assertFalse(equation["valid"])
        self.assertIsNone(equation["snm_mv"])
        self.assertIn("square-root domain", equation["reason"])

    def test_six_independent_objects_remain_supported(self):
        cell = SixTWatCell("MISMATCH", MosWat(.38, 45), MosWat(.40, 42),
                           MosWat(.37, 80), MosWat(.39, 75),
                           MosWat(.36, 120), MosWat(.38, 110))
        result = analyze_six_mos(cell, self.cfg, self.targets)
        self.assertEqual(set(result["cell"]["mos"]), {"PUL", "PUR", "PGL", "PGR", "PDL", "PDR"})
        self.assertEqual(len(result["target_comparisons"]), 6)
        self.assertEqual(len(result["snm_target_comparison"]), 1)
        metrics = result["baseline_6t"]["metrics"]
        self.assertIn("read_snm_upper_left_mv", metrics)
        self.assertIn("read_snm_lower_right_mv", metrics)
        self.assertAlmostEqual(metrics["read_snm_mv"],
                               min(metrics["read_snm_upper_left_mv"],
                                   metrics["read_snm_lower_right_mv"]))

    def test_asymmetric_read_snm_detects_and_swaps_left_right_mismatch(self):
        cell = SixTWatCell(
            "MISMATCH",
            MosWat(.35, 29.2), MosWat(.39, 20.0),
            MosWat(.27, 40.3), MosWat(.31, 30.0),
            MosWat(.27, 47.6), MosWat(.33, 35.0),
        )
        original = AsymmetricSram6T(cell, self.cfg).read_butterfly(.9, 1201)["read_butterfly"]
        swapped_cell = SixTWatCell(
            "SWAPPED", cell.pu2, cell.pu1, cell.pg2, cell.pg1, cell.pd2, cell.pd1)
        swapped = AsymmetricSram6T(swapped_cell, self.cfg).read_butterfly(.9, 1201)["read_butterfly"]
        self.assertGreater(original["mismatch_index_pct"], 1.0)
        self.assertAlmostEqual(original["snm_upper_left_mv"],
                               swapped["snm_lower_right_mv"], delta=1.0)
        self.assertAlmostEqual(original["snm_lower_right_mv"],
                               swapped["snm_upper_left_mv"], delta=1.0)
        self.assertAlmostEqual(original["snm_mv"], swapped["snm_mv"], delta=1.0)
        self.assertAlmostEqual(original["delta_snm_mv"], -swapped["delta_snm_mv"], delta=1.0)

    def test_mismatch_boundary_search_exports_complete_six_mos_values(self):
        cell = SixTWatCell(
            "BOUNDARY_W01", MosWat(.385, 44), MosWat(.385, 44),
            MosWat(.365, 82), MosWat(.365, 82),
            MosWat(.355, 124), MosWat(.355, 124))
        analysis = analyze_mismatch_rsnm_boundaries(
            cell, self.cfg, fit_points=101, scan_steps=4, bisection_steps=6)
        self.assertEqual(len(analysis["rows"]), 24)
        found = [row for row in analysis["rows"] if row["status"] == "BOUNDARY FOUND"]
        self.assertTrue(found)
        self.assertTrue(any(row["upper_rsnm_mv"] == 0 for row in found))
        self.assertTrue(any(row["lower_rsnm_mv"] == 0 for row in found))
        for key in ("pul_vt_v", "pul_idsat_ua", "pgr_vt_v", "pgr_idsat_ua",
                    "pdr_vt_v", "pdr_idsat_ua"):
            self.assertIn(key, found[0])
        with tempfile.TemporaryDirectory() as td:
            report = write_mismatch_boundary_outputs(analysis, td)
            self.assertTrue(report.exists())
            self.assertTrue((Path(td) / "rsnm_mismatch_boundaries.csv").exists())
            html_text = report.read_text(encoding="utf-8")
            self.assertIn("Complete 6T Boundary Values", html_text)
            self.assertIn("PUL Vt (V)", html_text)

    def test_html_png_and_csv_include_read_and_write_snm(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / "output"
            output.mkdir()
            (output / "wt_test_0bit_vmin.csv").write_text("old", encoding="utf-8")
            (output / "sram_wat_results.csv").write_text("old vmin schema", encoding="utf-8")
            report = write_outputs(result, output)
            image_dir = output / "images"
            self.assertTrue((image_dir / "01_read_snm_target_comparison.png").exists())
            self.assertTrue((image_dir / "01_read_snm_target_comparison.svg").exists())
            self.assertTrue((image_dir / "02_read_snm_butterfly.png").exists())
            self.assertTrue((image_dir / "02_read_snm_butterfly.svg").exists())
            self.assertTrue((image_dir / "03_write_snm_vs_bitline.png").exists())
            self.assertTrue((image_dir / "03_write_snm_vs_bitline.svg").exists())
            self.assertTrue((image_dir / "04_single_wat_write_snm_geometry.png").exists())
            self.assertTrue((image_dir / "04_single_wat_write_snm_geometry.svg").exists())
            self.assertEqual(len(list(image_dir.glob("*.png"))), 4)
            self.assertEqual(len(list(image_dir.glob("*.svg"))), 4)
            svg = (image_dir / "01_read_snm_target_comparison.svg").read_text(encoding="utf-8")
            self.assertIn("LOT_W01 WAT VTC", svg)
            self.assertNotIn("Lot/Wafer WAT VTC", svg)
            self.assertNotIn("Current WAT VTC", svg)
            self.assertIn("WAT Target VTC", svg)
            self.assertIn("Vin (V)", svg)
            self.assertIn("Vout (V)", svg)
            self.assertIn("0.30", svg)
            self.assertIn("1.20", svg)
            self.assertNotIn("SNM squares in both butterfly lobes", svg)
            butterfly_svg = (image_dir / "02_read_snm_butterfly.svg").read_text(encoding="utf-8")
            self.assertIn("Maximum squares 1 and 2", butterfly_svg)
            self.assertNotIn("Geometric RSNM", butterfly_svg)
            self.assertNotIn("Analytical RSNM", butterfly_svg)
            self.assertIn("cell RSNM is the smaller value", butterfly_svg)
            self.assertIn("QB=0 / Q=1", butterfly_svg)
            self.assertIn("QB=1 / Q=0", butterfly_svg)
            for dataset in (result["baseline_6t"], result["target_6t"]):
                for square in dataset["read_butterfly"]["squares"]:
                    self.assertIn(f'>{square["side_mv"]:.1f} mV</text>', butterfly_svg)
            self.assertIn("Vin (V)", butterfly_svg)
            self.assertIn("Vout (V)", butterfly_svg)
            self.assertIn("1.20", butterfly_svg)
            self.assertNotIn("Figure 3.15", butterfly_svg)
            write_svg = (image_dir / "03_write_snm_vs_bitline.svg").read_text(encoding="utf-8")
            self.assertIn("Write SNM versus Write-Bitline Voltage", write_svg)
            self.assertIn("Write Trip BL", write_svg)
            self.assertIn("Limiting Write SNM (mV)", write_svg)
            self.assertNotIn("WSNM proxy", write_svg)
            geometry_svg = (image_dir / "04_single_wat_write_snm_geometry.svg").read_text(encoding="utf-8")
            self.assertIn("6T Single-WAT Write SNM", geometry_svg)
            self.assertIn("WSNM geometry", geometry_svg)
            self.assertIn("Vin (V)", geometry_svg)
            self.assertIn("Vout (V)", geometry_svg)
            self.assertNotIn("QB (V)", geometry_svg)
            html = report.read_text(encoding="utf-8")
            self.assertIn("Read SNM Target Comparison", html)
            self.assertIn("Lot/Wafer SNM", html)
            self.assertNotIn("Current SNM", html)
            self.assertNotIn("Hold SNM", html)
            self.assertIn("Write SNM versus Write-Bitline Voltage", html)
            self.assertIn("Write Trip BL", html)
            self.assertNotIn("Write SNM Proxy", html)
            self.assertNotIn("WT Test 0-Bit Vmin", html)
            self.assertNotIn("Vmin", html)
            self.assertFalse((output / "wt_test_0bit_vmin.csv").exists())
            self.assertFalse((output / "sram_wat_results.csv").exists())
            self.assertTrue((output / "snm_target_comparison.csv").exists())
            self.assertTrue((output / "write_snm_vs_bitline.csv").exists())
            self.assertTrue((output / "single_wat_write_snm_geometry.csv").exists())
            self.assertTrue((output / "read_snm_state_mismatch.csv").exists())
            self.assertTrue((output / "analytical_read_snm.csv").exists())
            self.assertFalse((output / "analytical_read_snm_eq_3_36.csv").exists())
            self.assertTrue((output / "wat_electrical_snm_table.csv").exists())
            self.assertTrue((output / "cell_geometry_reference.csv").exists())
            self.assertFalse((output / "generic_28nm_assumptions.csv").exists())
            self.assertIn("Analytical Read SNM Reference", html)
            self.assertNotIn("PDF", html)
            self.assertNotIn("Figure 3.15", html)
            self.assertNotIn("Equation 3.36", html)
            self.assertIn("WAT Electrical Parameters", html)
            self.assertIn("No W/L, Cox, mobility", html)
            self.assertIn("6T Cell Geometry Reference", html)
            self.assertIn("VTH,eff", html)
            self.assertIn("Read SNM Butterfly and Left/Right Mismatch", html)
            self.assertIn("Mismatch index", html)
            with open(output / "snm_target_comparison.csv", encoding="utf-8-sig") as source:
                rows = list(csv.DictReader(source))
            self.assertEqual(len(rows), 1)
            self.assertEqual([row["mode"] for row in rows], ["Read SNM"])
            self.assertIn("lot_wafer_snm_mv", rows[0])
            self.assertNotIn("current_snm_mv", rows[0])

            sweep_svg = model_vdd_butterfly_svg([
                {"model_vdd_v": self.cfg.nominal_vdd, "result": result}
            ])
            measured_square = result["baseline_6t"]["read_butterfly"]["squares"][0]
            target_square = result["target_6t"]["read_butterfly"]["squares"][0]
            self.assertIn(f'WAT {measured_square["side_mv"]:.1f} mV', sweep_svg)
            self.assertIn(f'Target {target_square["side_mv"]:.1f} mV', sweep_svg)

    def test_wat_electrical_snm_table_uses_measured_inputs(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        rows = wat_electrical_snm_rows(result)
        self.assertEqual([row["dataset"] for row in rows],
                         ["Lot/Wafer", "WAT Target"])
        current = rows[0]
        self.assertAlmostEqual(current["pu_vt_v"], .385)
        self.assertAlmostEqual(current["pg_idsat_ua"], 82)
        self.assertAlmostEqual(current["idsat_pd_over_pg"], 124 / 82)
        self.assertGreater(current["q_beta_pu_over_pg"], 0)
        self.assertGreater(current["r_beta_pd_over_pg"], 0)
        self.assertNotIn("hold_snm_geometric_mv", current)
        self.assertIn("read_snm_geometric_mv", current)
        self.assertEqual(
            current["evidence_scope"],
            "WAT Vt + Idsat; no PDK/model-card-only parameters",
        )

    def test_geometry_references_are_explicit_and_do_not_override_wat(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        rows = {row["parameter"]: row for row in generic_28nm_assumption_rows(result)}
        self.assertEqual(rows["Channel length L"]["value"], 28.0)
        self.assertEqual(rows["Channel length L"]["active"], "REFERENCE")
        self.assertEqual(rows["Geometry Cell Ratio"]["value"], 1.4)
        self.assertAlmostEqual(rows["Geometry Pull-up Ratio"]["value"], 1.4286, places=4)
        self.assertEqual(len(rows), 6)

    def test_editable_technology_assumptions_propagate(self):
        cfg = Config(
            grid_points=101,
            technology_node_nm=27.5,
            channel_length_nm=29.0,
            pu_width_nm=72.0,
            pg_width_nm=104.0,
            pd_width_nm=146.0,
            nominal_temperature_c=30.0,
            read_wordline_over_vdd=0.95,
            read_bitline_over_vdd=0.98,
        )
        result = analyze_three_mos(self.cell, cfg, self.targets)
        tech = result["technology"]
        self.assertEqual(tech["node_nm"], 28)
        self.assertEqual(tech["channel_length_nm"], 29.0)
        self.assertEqual(tech["pu_width_nm"], 72.0)
        self.assertEqual(tech["pg_width_nm"], 104.0)
        self.assertEqual(tech["pd_width_nm"], 146.0)
        self.assertEqual(tech["nominal_temperature_c"], 30.0)
        self.assertEqual(tech["read_wordline_over_vdd"], 0.95)
        self.assertEqual(tech["read_bitline_over_vdd"], 0.98)

    def test_geometry_reference_rejects_zero_width(self):
        with self.assertRaisesRegex(ValueError, "PG width must be a positive"):
            analyze_three_mos(self.cell, Config(grid_points=101, pg_width_nm=0.0), self.targets)

    def test_read_bias_assumptions_affect_read_snm(self):
        baseline = analyze_three_mos(self.cell, self.cfg, self.targets)["baseline_6t"]["metrics"]["read_snm_mv"]
        adjusted_cfg = Config(grid_points=101, read_wordline_over_vdd=0.90,
                              read_bitline_over_vdd=0.95)
        adjusted = analyze_three_mos(self.cell, adjusted_cfg, self.targets)["baseline_6t"]["metrics"]["read_snm_mv"]
        self.assertNotAlmostEqual(baseline, adjusted, places=4)

    def test_wat_target_deltas(self):
        result = analyze_three_mos(self.cell, self.cfg, self.targets)
        pu = result["target_comparisons"][0]
        self.assertAlmostEqual(pu["delta_vt_mv"], 5.0)
        self.assertAlmostEqual(pu["delta_isat_ua"], -1.0)

    def test_csv_input_and_config(self):
        validate_config(self.cfg)
        with tempfile.TemporaryDirectory() as td:
            source = Path(td) / "wat.csv"
            source.write_text(
                "corner,pu_vt,pu_ids,pg_vt,pg_ids,pd_vt,pd_ids\nTT,.42,45,.40,80,.39,120\n",
                encoding="utf-8",
            )
            self.assertEqual(read_wat_csv(source)[0].corner, "TT")

    def test_excel_long_form_converts_units_and_groups_six_mos(self):
        headers = ["Lot/Wafer", "Model VDD", "VDD Unit", "MOS", "Vt", "Vt Unit", "Idsat", "Idsat Unit"]
        rows = [
            ["W01", 900, "mV", name, value, "mV", current, unit]
            for name, value, current, unit in (
                ("PUL", 380, 0.045, "mA"), ("PUR", 382, 46, "uA"),
                ("PGL", 365, 0.082, "mA"), ("PGR", 366, 83, "uA"),
                ("PDL", 355, 0.124, "mA"), ("PDR", 356, 125, "uA"),
            )
        ]
        samples = _read_wat_excel_rows(headers, rows, .9)
        self.assertEqual(len(samples), 1)
        self.assertEqual(samples[0].lot_wafer, "W01")
        self.assertAlmostEqual(samples[0].model_vdd_v, .9)
        self.assertAlmostEqual(samples[0].cell.pu1.vt, .380)
        self.assertAlmostEqual(samples[0].cell.pu1.ids, 45.0)
        self.assertAlmostEqual(samples[0].cell.pd2.ids, 125.0)

    def test_excel_wide_form_and_gui_state(self):
        headers = ["Lot/Wafer", "Model VDD (V)", "PUL Vt (mV)", "PUL Idsat (uA)",
                   "PUR Vt (mV)", "PUR Idsat (uA)", "PGL Vt (mV)", "PGL Idsat (uA)",
                   "PGR Vt (mV)", "PGR Idsat (uA)", "PDL Vt (mV)", "PDL Idsat (uA)",
                   "PDR Vt (mV)", "PDR Idsat (uA)"]
        row = ["W02", 1.2, 380, 45, 381, 46, 365, 82, 366, 83, 355, 124, 356, 125]
        samples = _read_wat_excel_rows(headers, [row], .9)
        self.assertEqual(len(samples), 1)
        self.assertAlmostEqual(samples[0].model_vdd_v, 1.2)
        self.assertAlmostEqual(samples[0].cell.pg2.vt, .366)
        with tempfile.TemporaryDirectory() as td:
            state_path = Path(td) / "state.json"
            save_gui_state({"values": {"corner": "W02"}}, state_path)
            self.assertEqual(load_gui_state(state_path)["values"]["corner"], "W02")

    def test_excel_repeated_wafer_sites_are_aggregated_with_coverage(self):
        headers = ["Lot/Wafer", "Site", "Model VDD", "MOS", "Vt", "Vt Unit",
                   "Idsat", "Idsat Unit", "Notes"]
        rows = []
        for mos in ("PUL", "PUR", "PGL", "PGR", "PDL", "PDR"):
            for site in range(1, 18):
                if site <= 12:
                    rows.append(["W03", f"S{site:02d}", 1.2, mos,
                                 0.30 + site * 0.001, "V", 40.0 + site, "uA", "Measured"])
                else:
                    rows.append(["W03", f"S{site:02d}", 1.2, mos,
                                 None, "V", None, "uA", "Unavailable"])
        sample = _read_wat_excel_rows(headers, rows, .9)[0]
        stats = sample.statistics["pu1"]
        self.assertEqual(stats.valid_count, 12)
        self.assertEqual(stats.total_count, 17)
        self.assertAlmostEqual(stats.vt_mean, 0.3065)
        self.assertAlmostEqual(sample.cell.pu1.ids, 46.5)

    def test_distributed_excel_template_uses_compatible_filtered_ranges(self):
        from openpyxl import load_workbook

        template = Path(__file__).with_name(
            "HV28_6T_WAT_12Point_VDD_Sweep_Template.xlsx"
        )
        workbook = load_workbook(template, read_only=False, data_only=False)
        self.assertEqual(workbook.sheetnames, ["PU", "PG", "PD", "Instructions"])
        for sheet_name in ("PU", "PG", "PD"):
            sheet = workbook[sheet_name]
            self.assertEqual(len(sheet.tables), 0)
            self.assertEqual(sheet.auto_filter.ref, "A1:I145")
            self.assertEqual(
                [cell.value for cell in sheet[1]],
                ["Lot/Wafer", "Site", "Model VDD", "MOS", "Vt", "Vt Unit",
                 "Idsat", "Idsat Unit", "Notes"],
            )

    def test_manual_rsnm_vcc_curve_brackets_eye_closure_and_exports(self):
        base_vt = {"pu": .385, "pg": .365, "pd": .355}
        base_ids = {"pu": 44.0, "pg": 82.0, "pd": 124.0}

        def scaled(kind, vcc):
            return base_ids[kind] * (max(vcc - base_vt[kind], 0) /
                                     (.9 - base_vt[kind])) ** 2

        points = [
            RsnmVccPoint(vcc,
                         MosWat(base_vt["pu"], scaled("pu", vcc)),
                         MosWat(base_vt["pg"], scaled("pg", vcc)),
                         MosWat(base_vt["pd"], scaled("pd", vcc)))
            for vcc in (.34, .36, .38, .40, .60, .90)
        ]
        analysis = analyze_rsnm_vcc_curve(points, self.cfg, fit_points=401)
        self.assertFalse(analysis["rows"][0]["valid_eye"])
        self.assertTrue(analysis["rows"][-1]["valid_eye"])
        self.assertIsNotNone(analysis["eye_closure"])
        self.assertGreater(analysis["eye_closure"]["estimated_vcc_v"], .36)
        self.assertLess(analysis["eye_closure"]["estimated_vcc_v"], .38)
        svg = rsnm_vcc_curve_svg(analysis)
        self.assertIn("Estimated RSNM versus Model VDD", svg)
        self.assertIn("Estimated eye-closure VDD", svg)
        self.assertIn("Read SNM (mV)", svg)
        self.assertIn('data-vdd-guide="0.38"', svg)
        self.assertIn(">0.38 V</text>", svg)
        self.assertEqual(svg.count("data-vdd-guide="),
                         sum(row["valid_eye"] for row in analysis["rows"]))
        self.assertNotIn(">VDD 0.38 V<", svg)
        self.assertNotIn(">RSNM ", svg)
        self.assertRegex(svg, r">\d+\.\d mV</text>")
        with tempfile.TemporaryDirectory() as td:
            report = write_rsnm_vcc_curve_outputs(analysis, Path(td) / "curve")
            self.assertTrue(report.exists())
            self.assertTrue((report.parent / "rsnm_vcc_curve.csv").exists())
            self.assertTrue((report.parent / "images" / "01_rsnm_vs_model_vcc.png").exists())
            self.assertIn("Estimated eye-closure VDD", report.read_text(encoding="utf-8"))

    def test_write_trip_margin_curve_brackets_boundary_and_exports(self):
        base_vt = {"pu": .385, "pg": .365, "pd": .355}
        base_ids = {"pu": 44.0, "pg": 82.0, "pd": 124.0}

        def scaled(kind, vdd):
            return base_ids[kind] * (max(vdd - base_vt[kind], 0) /
                                     (.9 - base_vt[kind])) ** 2

        points = [
            RsnmVccPoint(vdd,
                         MosWat(base_vt["pu"], scaled("pu", vdd)),
                         MosWat(base_vt["pg"], scaled("pg", vdd)),
                         MosWat(base_vt["pd"], scaled("pd", vdd)))
            for vdd in (.34, .38, .40, .50, .90)
        ]
        analysis = analyze_write_trip_margin_curve(points, self.cfg, fit_points=401)
        self.assertFalse(analysis["rows"][0]["writable"])
        self.assertTrue(analysis["rows"][-1]["writable"])
        self.assertIsNotNone(analysis["write_boundary"])
        self.assertGreater(analysis["write_boundary"]["estimated_vdd_v"], .38)
        self.assertLess(analysis["write_boundary"]["estimated_vdd_v"], .40)
        svg = write_trip_margin_curve_svg(analysis)
        self.assertIn("Estimated Write Trip Margin versus Model VDD", svg)
        self.assertIn("Write Trip Margin (mV)", svg)
        self.assertIn("Estimated write boundary VDD", svg)
        with tempfile.TemporaryDirectory() as td:
            report = write_write_trip_margin_outputs(analysis, Path(td) / "wtm")
            self.assertTrue(report.exists())
            self.assertTrue((report.parent / "write_trip_margin_curve.csv").exists())
            self.assertTrue((report.parent / "images" /
                             "01_write_trip_margin_vs_model_vdd.png").exists())


if __name__ == "__main__":
    unittest.main()
